"""
Import Monthwise_legacy.xlsx → production_table

Usage (run from d:/opr-mis1 or d:/opr-mis1/backend):
    python import_legacy.py [--dry-run] [--xlsx PATH]

Options:
    --dry-run     Print rows without writing to DB
    --xlsx PATH   Path to the Excel file (default: Report_format/Monthwise_legacy.xlsx)

Excel layout (sheet "monthwise"):
    Row 7          : Header  — col A="Items", col B="Plant", col D="Year",
                               cols E-P = "Apr"…"Mar"
    Block header   : col A = item name,  col B = plant name  (row with no data)
    Data rows      : col C = "Actual",   col D = FY string "YYYY-YY",
                               cols E(5)–P(16) = monthly values Apr–Mar

Item → canonical DB name mapping:
    "Oven Pushing (Nos./day)"  → "Oven Pushing(nos/d)"
    "Sinter  ('000 T)"         → "Total Sinter"
    "Hot metal ('000 T)"       → "Hot Metal"
    "Crude Steel ('000 T)"     → "Total Crude Steel"
    "Saleable Steel ('000 T)"  → "Saleable Steel"

Values are already in '000 T (or nos/day) — no conversion applied.
Rows are inserted with ON CONFLICT … DO UPDATE SET (upsert).
"""

import os
import sys
import sqlite3
import openpyxl

# ── Configuration ──────────────────────────────────────────────────────────────

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_DEFAULT_XLSX = os.path.join(_SCRIPT_DIR, "Report_format", "Monthwise_legacy.xlsx")
_DEFAULT_DB   = os.path.join(_SCRIPT_DIR, "backend", "mis_reports.db")

# Canonical item name mapping  (normalised Excel label → DB item_name)
# Keys are normalised: stripped + internal newlines replaced with single space.
ITEM_MAP = {
    "Oven Pushing (Nos./day)":  "Oven Pushing(nos/d)",
    "Oven Pushing(Nos./day)":   "Oven Pushing(nos/d)",
    "Sinter  ('000 T)":         "Total Sinter",
    "Sinter ('000 T)":          "Total Sinter",
    "Hot metal ('000 T)":       "Hot Metal",
    "Hot Metal ('000 T)":       "Hot Metal",
    "Crude Steel ('000 T)":     "Total Crude Steel",
    "Saleable Steel ('000 T)":  "Saleable Steel",
}


def _normalise_label(raw: str) -> str:
    """Strip and replace internal newlines/multiple spaces for map lookup."""
    return " ".join(raw.split())

# Column index (1-based) → (month_abbr, offset_from_apr_year)
# Apr–Dec are in the FY start year; Jan–Mar are in start_year+1
MONTH_COLS = {
     5: ("Apr",  0),
     6: ("May",  1),
     7: ("Jun",  2),
     8: ("Jul",  3),
     9: ("Aug",  4),
    10: ("Sep",  5),
    11: ("Oct",  6),
    12: ("Nov",  7),
    13: ("Dec",  8),
    14: ("Jan",  9),
    15: ("Feb", 10),
    16: ("Mar", 11),
}

MONTH_TO_NUM = {
    "Apr": "04", "May": "05", "Jun": "06", "Jul": "07",
    "Aug": "08", "Sep": "09", "Oct": "10", "Nov": "11",
    "Dec": "12", "Jan": "01", "Feb": "02", "Mar": "03",
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def fy_to_start_year(fy_str: str) -> int:
    """'2025-26' → 2025"""
    return int(str(fy_str).split("-")[0])


def col_to_report_month(col: int, start_year: int) -> str:
    """column index + FY start year → 'YYYY-MM'"""
    abbr, offset = MONTH_COLS[col]
    month_num = MONTH_TO_NUM[abbr]
    # Jan/Feb/Mar belong to start_year+1
    year = start_year + (1 if offset >= 9 else 0)
    return f"{year}-{month_num}"


def clean(val):
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "-", "N/A", "nan", "None"):
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


# ── Main extraction ───────────────────────────────────────────────────────────

def extract_rows(xlsx_path: str):
    """Yield (report_month, plant_name, item_name, value) for all Actual rows."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    cur_item_raw = ""
    cur_item_db  = ""
    cur_plant    = ""

    for row in range(7, ws.max_row + 1):
        col_a = ws.cell(row, 1).value
        col_b = ws.cell(row, 2).value
        col_c = ws.cell(row, 3).value
        col_d = ws.cell(row, 4).value

        # Block header: item name and/or plant name present
        if col_a is not None:
            raw = str(col_a).strip()
            if raw and raw.lower() != "items":
                cur_item_raw = raw
                cur_item_db  = ITEM_MAP.get(_normalise_label(raw), "")
        if col_b is not None:
            b = str(col_b).strip()
            if b and b.lower() != "plant":
                cur_plant = b

        # Data row: col C = "Actual", col D = FY string
        if str(col_c or "").strip().lower() != "actual":
            continue
        if not col_d or not cur_plant or not cur_item_db:
            continue

        fy_str = str(col_d).strip()
        if "-" not in fy_str:
            continue

        try:
            start_year = fy_to_start_year(fy_str)
        except (ValueError, IndexError):
            continue

        for col_idx in MONTH_COLS:
            val = clean(ws.cell(row, col_idx).value)
            if val is None:
                continue
            report_month = col_to_report_month(col_idx, start_year)
            yield report_month, cur_plant, cur_item_db, val


def main():
    dry_run   = "--dry-run" in sys.argv
    xlsx_path = _DEFAULT_XLSX
    if "--xlsx" in sys.argv:
        i = sys.argv.index("--xlsx")
        xlsx_path = sys.argv[i + 1]

    print(f"Source : {xlsx_path}")
    print(f"DB     : {_DEFAULT_DB}")
    print(f"Mode   : {'DRY RUN (no writes)' if dry_run else 'LIVE INSERT'}")
    print()

    if not os.path.exists(xlsx_path):
        print(f"ERROR: file not found: {xlsx_path}")
        sys.exit(1)

    rows = list(extract_rows(xlsx_path))
    print(f"Rows extracted: {len(rows)}")

    if dry_run:
        print()
        print(f"{'report_month':<13} {'plant':<6} {'item':<30} {'value'}")
        print("-" * 70)
        for r in rows:
            print(f"{r[0]:<13} {r[1]:<6} {r[2]:<30} {r[3]}")
        return

    conn = sqlite3.connect(_DEFAULT_DB)
    cur  = conn.cursor()
    inserted = updated = 0

    for report_month, plant, item, val in rows:
        cur.execute(
            "SELECT month_actual FROM production_table "
            "WHERE report_month=? AND plant_name=? AND item_name=?",
            (report_month, plant, item),
        )
        existing = cur.fetchone()
        cur.execute("""
            INSERT INTO production_table (report_month, plant_name, item_name, month_actual)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(report_month, plant_name, item_name)
            DO UPDATE SET month_actual = excluded.month_actual
        """, (report_month, plant, item, val))
        if existing is None:
            inserted += 1
        else:
            updated += 1

    conn.commit()
    conn.close()

    print(f"Inserted : {inserted} new rows")
    print(f"Updated  : {updated} existing rows")
    print(f"Total    : {inserted + updated} rows written to production_table")


if __name__ == "__main__":
    main()
