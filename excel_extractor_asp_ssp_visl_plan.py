import openpyxl
import logging
import sqlite3
import os
from typing import Optional

logger = logging.getLogger("excel_extractor_plan")

DB_PATH = os.path.join(os.path.dirname(__file__), "backend", "mis_reports.db")

_MONTH_NAMES = {
    1: "January", 2: "February", 3: "March", 4: "April",
    5: "May", 6: "June", 7: "July", 8: "August",
    9: "September", 10: "October", 11: "November", 12: "December",
}


def clean_val(val) -> Optional[float]:
    if val is None or str(val).strip().lower() in ("nan", "###", "-", "#div/0!"):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def extract_and_save_excel_plan(file_path: str, financial_year: str) -> bool:
    """
    Extracts ABP plan data for ASP, SSP, and VISL from a combined Excel file.

    Expected sheet: 'APP 26-27' (or similar — first sheet is used as fallback).

    Layout:
      Row 1  : headers — col A & B empty, col C onward has datetime values
               (one per month Apr–Mar) then an annual total column (skipped).
      Col A  : plant name ('ASP', 'SSP', 'VISL') — present only on the first
               row of each plant block; carried forward for subsequent items.
      Col B  : item name ('Total Crude Steel', 'Saleable Steel', etc.)
      Col C+ : monthly plan values, already in '000 T.

    All three plants are extracted in a single upload.
    Inserts/upserts into production_plan_table.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames
        logger.info(f"ASP/SSP/VISL Plan: loading file. Sheets: {sheet_names}")

        # Prefer known sheet name, fall back to first sheet
        sheet_key = None
        for candidate in sheet_names:
            if "APP" in candidate.upper() or "PLAN" in candidate.upper():
                sheet_key = candidate
                break
        if not sheet_key:
            sheet_key = sheet_names[0]
        logger.info(f"ASP/SSP/VISL Plan: using sheet '{sheet_key}'")

        ws = wb[sheet_key]

        # --- Build month column map from row 1 ---
        # Each cell in row 1 (starting col C) is a datetime for the 1st of that month.
        # Stop when we hit a non-datetime cell (the annual total column).
        from datetime import datetime as dt
        month_cols = []  # list of (col_index_0based, "Month YYYY")
        col_idx = 2  # 0-based; col A=0, B=1, C=2, ...
        while True:
            cell = ws.cell(row=1, column=col_idx + 1)
            val = cell.value
            if val is None:
                break
            if isinstance(val, dt):
                month_label = f"{val.year}-{val.month:02d}"
                month_cols.append((col_idx + 1, month_label))  # 1-based column number
            else:
                # Non-datetime (e.g., "FY 26-27" annual total) — stop
                break
            col_idx += 1

        if not month_cols:
            raise ValueError(
                "Could not find datetime month headers in row 1 starting from column C. "
                "Ensure the sheet matches the expected ASP/SSP/VISL ABP format."
            )
        logger.info(f"Detected {len(month_cols)} month columns: {month_cols[0][1]} → {month_cols[-1][1]}")

        # --- Scan data rows ---
        # Col A: plant name (carried forward when blank)
        # Col B: item name (row skipped when blank)
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        vals_written = 0
        current_plant = None

        for row_num in range(2, ws.max_row + 1):
            a_val = ws.cell(row=row_num, column=1).value
            b_val = ws.cell(row=row_num, column=2).value

            if a_val is not None:
                current_plant = str(a_val).strip()

            item_name = str(b_val).strip() if b_val is not None else None
            if not item_name or current_plant is None:
                continue  # blank spacer row

            for col_num, report_month in month_cols:
                raw = ws.cell(row=row_num, column=col_num).value
                val = clean_val(raw)
                stored = round(val, 3) if val is not None else None

                cursor.execute(
                    """
                    INSERT INTO production_plan_table
                        (report_month, plant_name, item_name, month_actual)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT(report_month, plant_name, item_name)
                    DO UPDATE SET month_actual = excluded.month_actual
                    """,
                    (report_month, current_plant, item_name, stored),
                )
                vals_written += 1

        if vals_written == 0:
            raise ValueError(
                "No data rows found in the ASP/SSP/VISL Plan Excel. "
                "Check that the sheet has plant names in col A, item names in col B, "
                "and datetime headers in row 1 from col C onward."
            )

        conn.commit()
        conn.close()
        logger.info(
            f"ASP/SSP/VISL Plan extraction done: {vals_written} cells written "
            f"for FY {financial_year}."
        )
        return True

    except ValueError as ve:
        logger.error(f"ASP/SSP/VISL Plan validation error: {ve}")
        raise
    except Exception as e:
        logger.error(f"ASP/SSP/VISL Plan extraction error: {e}")
        return False
