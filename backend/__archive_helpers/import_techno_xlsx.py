"""
Extract techno data from Report_format/techno.xlsx into techno_actuals.

Sheet per FY: '2022-23', '2023-24', '2024-25', '2025-26'
Row 5  : header — col B='Parameters', col C='Plant',
         cols 5-16 = monthly dates (Apr-Mar),
         col 17   = blank separator,
         cols 18-28 = cumulative (Apr-May, Apr-Jun, ..., Apr-Mar)
Data rows: col C = plant label; monthly actual at col M; cumulative at col M+12 (May-Mar),
           April (col 5) has no separate cumulative — till_month_actual = actual.

Run: python import_techno_xlsx.py [path_to_techno.xlsx]
"""
import sys, os, sqlite3
sys.path.insert(0, os.path.dirname(__file__))
import openpyxl
import db
from db import get_or_create_techno_param

XLSX_PATH = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
    os.path.dirname(__file__), "..", "Report_format", "techno.xlsx"
)

# ── Param blocks (start_row, end_row_inclusive, canonical_name, unit, group_code) ──
# Rows are consistent across all sheets; 6 rows per block (one per plant).
# Complex blocks (Gross HM, Scrap, TMI rows 48-77) are excluded — irregular structure.
PARAM_BLOCKS = [
    ( 6, 11, "Coal to Hot Metal",          "Ratio",      "MAJOR"),
    (12, 17, "Coke Rate",                   "Kg/THM",     "MAJOR"),
    (18, 23, "Nut Coke Rate",               "Kg/THM",     "MAJOR"),
    (24, 29, "CDI Rate",                    "Kg/THM",     "MAJOR"),
    (30, 35, "Fuel Rate",                   "Kg/THM",     "MAJOR"),
    (36, 41, "Sinter in Burden",            "%",          "MAJOR"),
    (42, 47, "BF Productivity",             "T/m³/day",   "MAJOR"),
    (78, 83, "Specific Energy Consumption", "G.Cal/TCS",  "MAJOR"),
]

KNOWN_PLANTS = {"BSP", "DSP", "RSP", "BSL", "ISP", "ISP New", "SAIL"}
PLANT_NORM   = {"ISP New": "ISP"}
MONTH_COLS   = range(5, 17)   # cols 5-16: Apr-Mar
APR_COL      = 5              # April has no cumulative row
# cum_col = month_col + 12 for month_col 6-16 (May-Mar)


def _clean(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "-", "#REF!", "#DIV/0!", "#VALUE!", "#N/A", "nan"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


db.init_db()
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

# ── Phase 1: build param_id cache (each call opens/commits/closes its own conn) ──
print("Building param registry …")
param_id_cache = {}   # (group_code, param_name, plant) -> param_id
ws0 = wb[wb.sheetnames[0]]   # structure same across sheets
for start_row, end_row, param_name, unit, group_code in PARAM_BLOCKS:
    so = (start_row - 6) * 10
    for r in range(start_row, end_row + 1):
        plant_raw = ws0.cell(r, 3).value
        if not plant_raw or plant_raw not in KNOWN_PLANTS:
            continue
        plant = PLANT_NORM.get(plant_raw, plant_raw)
        key = (group_code, param_name, plant)
        if key not in param_id_cache:
            param_id_cache[key] = get_or_create_techno_param(
                group_code, param_name, plant, unit, so
            )
print(f"  {len(param_id_cache)} params registered")

# ── Phase 2: bulk insert actuals via a single connection ─────────────────────
conn = sqlite3.connect(db.DB_PATH)
total_written = 0

for sheet in wb.sheetnames:
    ws = wb[sheet]

    # Parse month columns from header row 5
    month_map = {}   # col_idx -> 'YYYY-MM'
    for c in MONTH_COLS:
        v = ws.cell(5, c).value
        if hasattr(v, "strftime"):
            month_map[c] = v.strftime("%Y-%m")

    if not month_map:
        print(f"  [{sheet}] Could not parse month headers — skipping.")
        continue

    sheet_written = 0

    for start_row, end_row, param_name, unit, group_code in PARAM_BLOCKS:
        for r in range(start_row, end_row + 1):
            plant_raw = ws.cell(r, 3).value
            if not plant_raw or plant_raw not in KNOWN_PLANTS:
                continue
            plant = PLANT_NORM.get(plant_raw, plant_raw)

            param_id = param_id_cache.get((group_code, param_name, plant))
            if param_id is None:
                continue

            for c, report_month in month_map.items():
                actual = _clean(ws.cell(r, c).value)

                if c == APR_COL:
                    till_month = actual          # April: no separate cumulative
                else:
                    till_month = _clean(ws.cell(r, c + 12).value)

                if actual is None and till_month is None:
                    continue

                conn.execute("""
                    INSERT INTO techno_actuals
                        (report_month, param_id, actual, till_month_actual, source)
                    VALUES (?, ?, ?, ?, 'excel')
                    ON CONFLICT(report_month, param_id) DO UPDATE SET
                        actual            = excluded.actual,
                        till_month_actual = COALESCE(excluded.till_month_actual,
                                                     techno_actuals.till_month_actual),
                        source            = excluded.source
                """, (report_month, param_id, actual, till_month))
                sheet_written += 1

    conn.commit()
    print(f"  [{sheet}]  {sheet_written} rows upserted")
    total_written += sheet_written

conn.close()
print(f"\nDone. Total rows upserted: {total_written}")
