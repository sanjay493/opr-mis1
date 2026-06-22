"""
Import legacy techno data from Report_format/techno.xlsx → techno_table.

Covers all 4 FY sheets: 2022-23, 2023-24, 2024-25, 2025-26
Plants: BSP, DSP, RSP, BSL, ISP, SAIL

Usage (run from h:/opr-mis1/backend):
    python import_techno_excel.py [--dry-run] [--xlsx PATH]
"""

import os
import sys
import re
import sqlite3
import openpyxl

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_DEFAULT_XLSX = os.path.join(_SCRIPT_DIR, "..", "Report_format", "techno.xlsx")
_DEFAULT_DB   = os.path.join(_SCRIPT_DIR, "mis_reports.db")

VALID_PLANTS = {"BSP", "DSP", "RSP", "BSL", "ISP", "SAIL"}

# Excel column indices (1-based)
COL_PARAM = 2   # B  – parameter / label
COL_PLANT = 3   # C  – plant name
# Monthly actuals: cols 5–16  → Apr … Mar
MONTHLY_COLS = {5: 0, 6: 1, 7: 2, 8: 3, 9: 4, 10: 5,
                11: 6, 12: 7, 13: 8, 14: 9, 15: 10, 16: 11}
# YTD cumulative: cols 18–28  → Apr-May … Apr-Mar
# Col 18 = YTD through May (month offset 1), col 28 = YTD through Mar (offset 11)
# April YTD = April monthly (no separate column)
YTD_COLS = {i: (i - 17) for i in range(18, 29)}   # col → month_offset(1..11)


def fy_sheet_to_start_year(sheet_name: str) -> int:
    """'2024-25' → 2024"""
    return int(sheet_name.split("-")[0])


def offset_to_report_month(start_year: int, offset: int) -> str:
    """offset 0=Apr, 1=May, … 11=Mar → 'YYYY-MM'"""
    if offset < 9:          # Apr(0)–Dec(8): start_year
        year = start_year
        month = offset + 4  # 4=Apr … 12=Dec
    else:                   # Jan(9)–Mar(11): start_year + 1
        year = start_year + 1
        month = offset - 8  # 1=Jan … 3=Mar
    return f"{year}-{month:02d}"


def is_unit_only(s: str) -> bool:
    """Returns True for strings like '(Kg/Thm)' or '( Working Volume )' that are
    unit/description labels, not parameter names."""
    s = s.strip()
    return bool(re.fullmatch(r'\(.*\)', s))


def clean_param(s: str) -> str:
    """Strip trailing unit in parentheses and extra whitespace."""
    s = s.strip()
    s = re.sub(r'\s*\([^)]*\)\s*$', '', s)
    return s.strip()


def has_ref_error(row_vals: tuple) -> bool:
    return any(str(v).strip() == '#REF!' for v in row_vals if v is not None)


def clean_float(val) -> float | None:
    if val is None:
        return None
    s = str(val).strip()
    if s in ('', '-', 'N/A', 'nan', 'None', '#REF!'):
        return None
    try:
        return float(str(s).replace(',', ''))
    except (ValueError, TypeError):
        return None


def extract_sheet(ws, start_year: int):
    """Yield (report_month, plant, param, month_actual, ytd_actual) for one sheet."""
    cur_param = ''

    for row_idx in range(6, ws.max_row + 1):
        vals = [ws.cell(row_idx, c).value for c in range(1, 29)]
        col_b = str(vals[COL_PARAM - 1] or '').strip()
        col_c = str(vals[COL_PLANT - 1] or '').strip().upper()

        # Update current parameter name when col B has a real name on a BSP row
        if col_b and col_c == 'BSP' and not is_unit_only(col_b):
            cur_param = clean_param(col_b)

        # Skip rows without a recognised plant
        if col_c not in VALID_PLANTS:
            continue

        # Skip rows with formula errors
        if has_ref_error(vals):
            continue

        # Skip if no parameter known yet
        if not cur_param:
            continue

        # Build monthly + YTD lookup for this row
        monthly_vals = {offset: clean_float(vals[col - 1])
                        for col, offset in MONTHLY_COLS.items()}
        ytd_vals     = {offset: clean_float(vals[col - 1])
                        for col, offset in YTD_COLS.items()}

        for offset in range(12):   # 0=Apr … 11=Mar
            month_actual = monthly_vals.get(offset)
            if month_actual is None:
                continue

            # April YTD = monthly; May onwards use the cumulative column
            ytd_actual = month_actual if offset == 0 else ytd_vals.get(offset)

            report_month = offset_to_report_month(start_year, offset)
            yield report_month, col_c, cur_param, month_actual, ytd_actual


def _load_workbook_safe(xlsx_path: str):
    """
    Read an xlsx that has a corrupt/bloated stylesheet by parsing
    the XML zip entries directly, bypassing openpyxl's style engine.
    Returns a simple dict: { sheet_name: [[row_values], ...] }
    wrapped in a minimal object with .sheetnames and [sheet_name] access.
    """
    import zipfile
    import xml.etree.ElementTree as ET

    NS = {
        'ss': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
        'r' : 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    }

    with zipfile.ZipFile(xlsx_path, 'r') as zf:
        names = zf.namelist()

        # ── shared strings ────────────────────────────────────────────────────
        shared = []
        if 'xl/sharedStrings.xml' in names:
            with zf.open('xl/sharedStrings.xml') as f:
                for event, elem in ET.iterparse(f):
                    if elem.tag == '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si':
                        parts = [t.text or '' for t in elem.iter(
                            '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t')]
                        shared.append(''.join(parts))
                        elem.clear()

        # ── workbook to get sheet names + rId ─────────────────────────────────
        with zf.open('xl/workbook.xml') as f:
            wb_root   = ET.parse(f).getroot()
        sheet_els   = wb_root.findall('.//ss:sheet', NS)
        sheet_names = [s.get('name') for s in sheet_els]
        sheet_rids  = [s.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                       for s in sheet_els]

        # ── rels to map rId → sheet file ──────────────────────────────────────
        with zf.open('xl/_rels/workbook.xml.rels') as f:
            rels_root = ET.parse(f).getroot()
        rid_to_target = {rel.get('Id'): rel.get('Target') for rel in rels_root}

        # ── helpers ───────────────────────────────────────────────────────────
        def col_letter_to_idx(col_str: str) -> int:
            idx = 0
            for ch in col_str:
                idx = idx * 26 + (ord(ch) - ord('A') + 1)
            return idx - 1

        _ref_re = re.compile(r'([A-Z]+)(\d+)')

        def parse_cell_ref(ref: str):
            m = _ref_re.match(ref)
            return col_letter_to_idx(m.group(1)), int(m.group(2)) - 1

        import datetime
        _epoch = datetime.datetime(1899, 12, 30)

        # ── parse each sheet with streaming iterparse ─────────────────────────
        sheets_data = {}
        for name, rid in zip(sheet_names, sheet_rids):
            target = rid_to_target.get(rid, '')
            path   = f'xl/{target}' if not target.startswith('xl/') else target
            if path not in names:
                sheets_data[name] = []
                continue

            max_col  = 0
            raw_rows = {}

            with zf.open(path) as f:
                for event, elem in ET.iterparse(f, events=('end',)):
                    tag = elem.tag
                    if tag != '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row':
                        continue
                    row_idx = int(elem.get('r', 0)) - 1
                    for c in elem:
                        if not c.tag.endswith('}c'):
                            continue
                        ref  = c.get('r', '')
                        t    = c.get('t', '')
                        v_el = c.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                        f_el = c.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}f')
                        if not ref:
                            continue
                        col_idx, _ = parse_cell_ref(ref)
                        max_col    = max(max_col, col_idx)

                        if v_el is not None and v_el.text is not None:
                            raw = v_el.text
                            if t == 's':
                                val = shared[int(raw)]
                            elif t == 'b':
                                val = bool(int(raw))
                            else:
                                try:
                                    fval = float(raw)
                                    if 30000 < fval < 60000:
                                        val = _epoch + datetime.timedelta(days=fval)
                                    else:
                                        val = fval
                                except (ValueError, OverflowError):
                                    val = raw
                        elif f_el is not None:
                            val = '#REF!'
                        else:
                            val = None

                        raw_rows.setdefault(row_idx, {})[col_idx] = val
                    elem.clear()

            if not raw_rows:
                sheets_data[name] = []
                continue

            max_row = max(raw_rows.keys())
            grid    = []
            for r in range(max_row + 1):
                row_data = raw_rows.get(r, {})
                grid.append([row_data.get(c) for c in range(max_col + 1)])
            sheets_data[name] = grid

    # ── minimal workbook-like wrapper ─────────────────────────────────────────
    class _Sheet:
        def __init__(self, grid):
            self._grid     = grid
            self.max_row   = len(grid)
            self.max_column = max((len(r) for r in grid), default=0)

        def cell(self, row, column):
            class _Cell:
                def __init__(self, value): self.value = value
            try:
                return _Cell(self._grid[row - 1][column - 1])
            except IndexError:
                return _Cell(None)

    class _Workbook:
        def __init__(self, data):
            self._data    = data
            self.sheetnames = list(data.keys())

        def __getitem__(self, name):
            return _Sheet(self._data[name])

    return _Workbook(sheets_data)


def main():
    dry_run   = '--dry-run' in sys.argv
    xlsx_path = _DEFAULT_XLSX
    if '--xlsx' in sys.argv:
        xlsx_path = sys.argv[sys.argv.index('--xlsx') + 1]

    if not os.path.exists(xlsx_path):
        print(f"ERROR: file not found: {xlsx_path}")
        sys.exit(1)

    print(f"Source : {xlsx_path}")
    print(f"DB     : {_DEFAULT_DB}")
    print(f"Mode   : {'DRY RUN (no writes)' if dry_run else 'LIVE INSERT'}")
    print()

    wb = _load_workbook_safe(xlsx_path)

    all_rows = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            start_year = fy_sheet_to_start_year(sheet_name)
        except (ValueError, IndexError):
            print(f"  [SKIP] unrecognised sheet name: {sheet_name}")
            continue

        sheet_rows = list(extract_sheet(ws, start_year))
        print(f"  Sheet {sheet_name}: {len(sheet_rows)} rows extracted")
        all_rows.extend(sheet_rows)

    print(f"\nTotal rows: {len(all_rows)}")

    if dry_run:
        print()
        print(f"{'report_month':<13} {'plant':<6} {'parameter':<35} {'actual':<12} {'ytd'}")
        print('-' * 80)
        for r in all_rows[:50]:
            print(f"{r[0]:<13} {r[1]:<6} {r[2]:<35} {str(r[3]):<12} {r[4]}")
        if len(all_rows) > 50:
            print(f"  ... and {len(all_rows) - 50} more rows (showing first 50 only)")
        return

    conn = sqlite3.connect(_DEFAULT_DB)
    cur  = conn.cursor()
    inserted = updated = 0

    for report_month, plant, param, month_actual, ytd_actual in all_rows:
        cur.execute(
            "SELECT month_actual FROM techno_table "
            "WHERE report_month=? AND plant_name=? AND parameter_name=?",
            (report_month, plant, param),
        )
        existing = cur.fetchone()
        cur.execute("""
            INSERT INTO techno_table
                (report_month, plant_name, parameter_name, unit, month_actual, ytd_actual)
            VALUES (?, ?, ?, '', ?, ?)
            ON CONFLICT(report_month, plant_name, parameter_name)
            DO UPDATE SET
                month_actual = excluded.month_actual,
                ytd_actual   = excluded.ytd_actual
        """, (report_month, plant, param, month_actual, ytd_actual))
        if existing is None:
            inserted += 1
        else:
            updated += 1

    conn.commit()
    conn.close()

    print(f"Inserted : {inserted} new rows")
    print(f"Updated  : {updated} existing rows")
    print(f"Total    : {inserted + updated} rows written to techno_table")


if __name__ == '__main__':
    main()
