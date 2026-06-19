"""
ASP (Alloy Steels Plant, Durgapur) Excel extractor.

Source file: Daily Performance Summary Report (asp.xlsx or asp.xls, sheet "md cell").

Row positions are found dynamically by scanning column A for keywords:
    "CRUDE STEEL"     → Total Crude Steel  (col F)
    "INGOT PRODUCTION"→ Ingot Steel         (col F)
    Concast           → Total Crude Steel − Ingot Steel  (computed)
    "SALEABLE STEEL"  → Saleable Steel      (col F, first match)
    "TOTAL PLANT ST." → Closing Stock       (col L)

Sheet "DAILY FLASH" additional cells (fixed addresses):
    AN45 → BARS
    AN46 → FORGINGS
    AN47 → PLATES
    Finished Steel = BARS + FORGINGS + PLATES  (computed)

Report month is auto-detected from cell E3 which contains the report date
(e.g. '30/04/2026' → 2026-04).  The user-supplied month is used as fallback
if E3 is blank or unparseable.

All values are in Tonnes in the source file; stored as '000T in the DB.
"""
import os
import re

# ---------------------------------------------------------------------------
# xlrd wrapper — makes .xls files look like openpyxl (address-based access)
# ---------------------------------------------------------------------------

try:
    import xlrd as _xlrd
    _XLRD_OK = True
except ImportError:
    _XLRD_OK = False


def _col_letter_to_idx(letters: str) -> int:
    """Convert column letters (e.g. 'AN') to 1-based column index."""
    idx = 0
    for ch in letters.upper():
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx


def _parse_addr(addr: str):
    """Split 'AN45' → (row=45, col=40) both 1-based."""
    m = re.match(r"([A-Za-z]+)(\d+)", addr)
    if not m:
        raise ValueError(f"Cannot parse cell address: {addr!r}")
    return int(m.group(2)), _col_letter_to_idx(m.group(1))


class _XlsCell:
    __slots__ = ("value",)
    def __init__(self, v):
        self.value = None if v == "" else v


class _XlsSheet:
    def __init__(self, sheet):
        self._s = sheet

    def __getitem__(self, addr: str) -> "_XlsCell":
        row, col = _parse_addr(addr)
        r, c = row - 1, col - 1
        if r < 0 or r >= self._s.nrows or c < 0 or c >= self._s.ncols:
            return _XlsCell(None)
        return _XlsCell(self._s.cell_value(r, c))

    @property
    def title(self):
        return self._s.name


class _XlsWb:
    def __init__(self, wb):
        self._wb = wb

    @property
    def sheetnames(self):
        return self._wb.sheet_names()

    def __getitem__(self, name: str) -> "_XlsSheet":
        return _XlsSheet(self._wb.sheet_by_name(name))

    def __contains__(self, name: str) -> bool:
        return name in self._wb.sheet_names()


def _open_wb(file_path: str):
    """Open .xls (xlrd) or .xlsx (openpyxl), return a unified workbook."""
    if file_path.lower().endswith(".xls"):
        if not _XLRD_OK:
            raise ImportError("xlrd is required for .xls files: pip install xlrd")
        return _XlsWb(_xlrd.open_workbook(file_path))
    import openpyxl
    return openpyxl.load_workbook(file_path, data_only=True)


def _cell_value(ws, row: int, col: int):
    """Read cell value by 1-based row/col for both xlrd wrapper and openpyxl sheet."""
    if isinstance(ws, _XlsSheet):
        r, c = row - 1, col - 1
        if r < 0 or r >= ws._s.nrows or c < 0 or c >= ws._s.ncols:
            return None
        v = ws._s.cell_value(r, c)
        return None if v == "" else v
    return ws.cell(row=row, column=col).value


def _find_row(ws, keyword: str, search_cols=range(1, 27), max_rows: int = 400) -> int:
    """Return 1-based row where any column A–Z contains keyword (case-insensitive).
    Strips whitespace from cell text before comparing. Returns None if not found."""
    kw = keyword.strip().upper()
    for row in range(1, max_rows + 1):
        for col in search_cols:
            val = _cell_value(ws, row, col)
            if val is not None and kw in str(val).strip().upper():
                return row
    return None


def _scan_sheet(ws, label: str, max_rows: int = 60, max_cols: int = 26):
    """Print non-blank cell content for debugging — called when a keyword search fails."""
    import sys
    print(f"[ASP Debug] Scanning sheet '{ws.title}' rows 1-{max_rows} cols 1-{max_cols} "
          f"(triggered by failed search for '{label}'):", flush=True, file=sys.stderr)
    for r in range(1, max_rows + 1):
        row_vals = []
        for c in range(1, max_cols + 1):
            v = _cell_value(ws, r, c)
            if v is not None and str(v).strip():
                row_vals.append(f"col{c}={repr(str(v).strip()[:30])}")
        if row_vals:
            print(f"  row {r:3d}: {' | '.join(row_vals)}", flush=True, file=sys.stderr)

# ---------------------------------------------------------------------------

PLANT       = "ASP"
SHEET_NAME  = "md cell"
SHEET_FLASH = "DAILY FLASH"
COL_F       = 6    # column F — monthly production value
COL_L       = 12   # column L — closing stock

_MONTHS = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN",
           "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]

# (cell address, item_name) — from "DAILY FLASH" sheet, values in Tonnes
_FLASH_CELL_MAP = [
    ("AN45", "BARS"),
    ("AN46", "FORGINGS"),
    ("AN47", "PLATES"),
]


def _parse_date_cell(raw):
    """Convert E3 value to 'YYYY-MM' string.

    Handles:
      - datetime / date objects  (openpyxl returns these for real date cells)
      - string like '30/04/2026' or '2026-04-30'
    Returns None if the value cannot be parsed.
    """
    if raw is None:
        return None
    # openpyxl datetime / date
    if hasattr(raw, "month") and hasattr(raw, "year"):
        return f"{raw.year}-{str(raw.month).zfill(2)}"
    s = str(raw).strip()
    # dd/mm/yyyy
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        return f"{m.group(3)}-{m.group(2).zfill(2)}"
    # yyyy-mm-dd
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    return None


def extract_preview(file_path: str, report_month: str, **_kwargs) -> dict:
    """Extract ASP production actuals from the Daily Performance Summary Excel.

    Returns a dict in the standard extract_preview() format — no DB writes.
    Month is auto-detected from cell E3; report_month is used as fallback.
    """
    import sys

    fname = os.path.basename(file_path)
    print(f"[ASP Excel] extract_preview: file={fname}  month={report_month}",
          flush=True, file=sys.stderr)

    try:
        wb = _open_wb(file_path)
    except Exception as exc:
        raise ValueError(f"Cannot open Excel file '{fname}': {exc}") from exc

    # Try named sheet first, fall back to first sheet
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        first = wb.sheetnames[0]
        ws = wb[first]
        print(f"[ASP Excel] Sheet '{SHEET_NAME}' not found, using first: '{first}'",
              flush=True, file=sys.stderr)

    # ── Auto-detect month from E3 ─────────────────────────────────────────
    date_val    = ws["E3"].value
    detected    = _parse_date_cell(date_val)
    actual_month = detected or report_month

    y, m   = int(actual_month[:4]), int(actual_month[5:7])
    want_mon = _MONTHS[m - 1]
    yy       = str(y)[2:]

    print(f"[ASP Excel] E3={repr(date_val)}  detected={detected}  using={actual_month}",
          flush=True, file=sys.stderr)

    # ── Keyword-based row search → column F / L ───────────────────────────
    def _read_row(keyword, col, label, search_cols=range(1, 27)):
        """Find row by keyword, read col value (T → '000T). Returns (row_num, stored, raw).

        search_cols: columns to scan for the keyword.  Pass (1,) to restrict to
        column A, which avoids false matches in the performance-summary or
        monthly-history columns that share the same label text.
        """
        rn = _find_row(ws, keyword, search_cols=search_cols)
        if rn is None:
            print(f"[ASP Excel] '{keyword}' not found in sheet", flush=True, file=sys.stderr)
            return None, None, None
        raw = _cell_value(ws, rn, col)
        if raw is not None and isinstance(raw, (int, float)):
            stored = round(float(raw) / 1000.0, 3)
            print(f"[ASP Excel] {label:20s} row={rn} col={col} = {raw} T → {stored} '000T",
                  flush=True, file=sys.stderr)
            return rn, stored, raw
        print(f"[ASP Excel] {label:20s} row={rn} col={col} = EMPTY (raw={repr(raw)})",
              flush=True, file=sys.stderr)
        return rn, None, None

    def _row_entry(item_name, stored, rn, col, keyword):
        addr = f"row {rn} col {col} ('{keyword}')" if rn else f"'{keyword}' not found"
        if stored is not None:
            return {"item_name": item_name, "value": stored, "unit": "'000T",
                    "cell": f"Excel [{ws.title}!{addr}] · {want_mon}'{yy}",
                    "pdf_label": addr, "status": "ok"}
        return {"item_name": f"(empty) {item_name}", "value": None, "unit": "T",
                "cell": f"Excel [{ws.title}!{addr}]", "pdf_label": addr, "status": "unmapped"}

    rows = []

    # Total Crude Steel — search col A only.
    # 2025 format: col A = "CRUDE STEEL"; 2026 format: col A = "CRUDE".
    # Searching all columns would hit the performance-summary (col J "CRUDE") or the
    # monthly-history section ("CRUDE STEEL" row with no col-F value) first.
    cs_rn, cs_val = None, None
    for kw in ("CRUDE STEEL", "CRUDE"):
        cs_rn, cs_val, _ = _read_row(kw, COL_F, "Total Crude Steel", search_cols=(1,))
        if cs_val is not None:
            break
    rows.append(_row_entry("Total Crude Steel", cs_val, cs_rn, COL_F, "CRUDE STEEL/CRUDE"))

    # Ingot Steel — col A only (same monthly-history duplication risk)
    ing_rn, ing_val, _ = _read_row("INGOT PRODUCTION", COL_F, "Ingot Steel", search_cols=(1,))
    rows.append(_row_entry("Ingot Steel", ing_val, ing_rn, COL_F, "INGOT PRODUCTION"))

    # Concast = Total Crude Steel − Ingot Steel
    if cs_val is not None and ing_val is not None:
        cc_val = round(cs_val - ing_val, 3)
        print(f"[ASP Excel] {'Concast':20s} computed = {cc_val} '000T (CS−Ingot)",
              flush=True, file=sys.stderr)
        rows.append({"item_name": "Total Caster", "value": cc_val, "unit": "'000T",
                     "cell": f"Excel [computed: CS−Ingot] · {want_mon}'{yy}",
                     "pdf_label": "CS-Ingot", "status": "ok"})
    else:
        rows.append({"item_name": "(empty) Total Caster", "value": None, "unit": "T",
                     "cell": "Excel [computed: CS−Ingot — missing inputs]",
                     "pdf_label": "CS-Ingot", "status": "unmapped"})

    # Saleable Steel — col A only (monthly-history section also has "SALEABLE STEEL" in col A
    # but at a higher row number, so col-A-only still finds the production row first)
    sal_rn, sal_val, _ = _read_row("SALEABLE STEEL", COL_F, "Saleable Steel", search_cols=(1,))
    rows.append(_row_entry("Saleable Steel", sal_val, sal_rn, COL_F, "SALEABLE STEEL"))

    # Closing Stock — "TOTAL PLANT ST." lives in col J, so search all columns
    cst_rn, cst_val, _ = _read_row("TOTAL PLANT ST.", COL_L, "Closing Stock")
    if cst_val is None:
        _scan_sheet(ws, "TOTAL PLANT ST.")
    rows.append(_row_entry("Closing Stock", cst_val, cst_rn, COL_L, "TOTAL PLANT ST."))

    # ── DAILY FLASH sheet — BARS / FORGINGS / PLATES ─────────────────────
    flash_vals = {}   # item_name → value in '000T
    if SHEET_FLASH in wb.sheetnames:
        wf = wb[SHEET_FLASH]
        cell_tag_f = f"Excel [{SHEET_FLASH}!{{addr}}] · {want_mon}'{yy}"
        for cell_addr, item_name in _FLASH_CELL_MAP:
            raw = wf[cell_addr].value
            if raw is not None and isinstance(raw, (int, float)):
                stored = round(float(raw) / 1000.0, 3)
                flash_vals[item_name] = stored
                rows.append({
                    "item_name": item_name,
                    "value":     stored,
                    "unit":      "'000T",
                    "cell":      cell_tag_f.format(addr=cell_addr),
                    "pdf_label": cell_addr,
                    "status":    "ok",
                })
                print(f"[ASP Flash] {item_name:15s} [{cell_addr}] = {raw} T → {stored} '000T",
                      flush=True, file=sys.stderr)
            else:
                rows.append({
                    "item_name": f"(empty) {item_name}",
                    "value":     None,
                    "unit":      "T",
                    "cell":      cell_tag_f.format(addr=cell_addr),
                    "pdf_label": cell_addr,
                    "status":    "unmapped",
                })
                print(f"[ASP Flash] {item_name:15s} [{cell_addr}] = EMPTY (raw={repr(raw)})",
                      flush=True, file=sys.stderr)

        # Finished Steel = BARS + FORGINGS + PLATES (only when all three found)
        fs_parts = [flash_vals.get(n) for n in ("BARS", "FORGINGS", "PLATES")]
        if all(v is not None for v in fs_parts):
            fs_val = round(sum(fs_parts), 3)
            rows.append({
                "item_name": "Finished Steel",
                "value":     fs_val,
                "unit":      "'000T",
                "cell":      f"Excel [computed: BARS+FORGINGS+PLATES] · {want_mon}'{yy}",
                "pdf_label": "AN45+AN46+AN47",
                "status":    "ok",
            })
            print(f"[ASP Flash] Finished Steel = {fs_val} '000T (sum of BARS+FORGINGS+PLATES)",
                  flush=True, file=sys.stderr)
    else:
        print(f"[ASP Excel] Sheet '{SHEET_FLASH}' not found — BARS/FORGINGS/PLATES skipped",
              flush=True, file=sys.stderr)

    ok = sum(1 for r in rows if r["status"] == "ok")
    print(f"[ASP Excel] {ok}/{len(rows)} rows ok", flush=True, file=sys.stderr)

    date_display = str(date_val) if date_val else "unknown date"

    return {
        "plant":              PLANT,
        "month":              actual_month,
        "source_type":        "ASP Daily Performance Summary (Excel)",
        "sheets":             f"Sheet '{ws.title}' · Report date: {date_display}",
        "workbook_sheets":    [ws.title],
        "report_type":        "EXCEL",
        "detected_month":     detected,
        "production_rows":    rows,
        "special_steel_rows": [],
        "special_steel_note": "",
        "techno_rows":        [],
        "techno_param_rows":  [],
    }
