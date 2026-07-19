import openpyxl
import logging
import sqlite3
import os
from typing import Optional
import db

logger = logging.getLogger("excel_extractor_plan")

DB_PATH = os.path.join(os.path.dirname(__file__), "..", "mis_reports.db")

_MONTH_NAMES = {
    1: "January", 2: "February", 3: "March", 4: "April",
    5: "May", 6: "June", 7: "July", 8: "August",
    9: "September", 10: "October", 11: "November", 12: "December",
}


def clean_val(val) -> Optional[float]:
    if val is None or str(val).strip().lower() in ("nan", "###", "-", "#div/0!"):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def extract_and_save_excel_plan(file_path: str, financial_year: str) -> bool:
    """
    Extracts ABP plan data for ASP, SSP, and VISL from a combined Excel file.

    Expected sheet: 'APP 26-27' (or similar — first sheet is used as fallback).

    Layout:
      Row 1  : headers — col A & B empty, col C onward has datetime values
               (one per month Apr–Mar) then an annual total column (skipped).
      Col A  : plant name ('ASP', 'SSP', 'VISL') — present only on the first
               row of each plant block; carried forward for subsequent items.
      Col B  : item name ('Total Crude Steel', 'Saleable Steel', etc.)
      Col C+ : monthly plan values, already in '000 T.

    All three plants are extracted in a single upload.
    Inserts/upserts into production_plan_table.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames
        logger.info(f"ASP/SSP/VISL Plan: loading file. Sheets: {sheet_names}")

        # Prefer known sheet name, fall back to first sheet
        sheet_key = None
        for candidate in sheet_names:
            if "APP" in candidate.upper() or "PLAN" in candidate.upper():
                sheet_key = candidate
                break
        if not sheet_key:
            sheet_key = sheet_names[0]
        logger.info(f"ASP/SSP/VISL Plan: using sheet '{sheet_key}'")

        ws = wb[sheet_key]

        # --- Build month column map from row 1 ---
        # Each cell in row 1 (starting col C) is a datetime for the 1st of that month.
        # Stop when we hit a non-datetime cell (the annual total column).
        from datetime import datetime as dt
        month_cols = []  # list of (col_index_0based, "Month YYYY")
        col_idx = 2  # 0-based; col A=0, B=1, C=2, ...
        while True:
            cell = ws.cell(row=1, column=col_idx + 1)
            val = cell.value
            if val is None:
                break
            if isinstance(val, dt):
                month_label = f"{val.year}-{val.month:02d}"
                month_cols.append((col_idx + 1, month_label))  # 1-based column number
            else:
                # Non-datetime (e.g., "FY 26-27" annual total) — stop
                break
            col_idx += 1

        if not month_cols:
            raise ValueError(
                "Could not find datetime month headers in row 1 starting from column C. "
                "Ensure the sheet matches the expected ASP/SSP/VISL ABP format."
            )
        logger.info(f"Detected {len(month_cols)} month columns: {month_cols[0][1]} → {month_cols[-1][1]}")

        # --- Scan data rows ---
        # Col A: plant name (carried forward when blank)
        # Col B: item name (row skipped when blank)
        conn = db.connect()
        cursor = conn.cursor()
        vals_written = 0
        current_plant = None

        for row_num in range(2, ws.max_row + 1):
            a_val = ws.cell(row=row_num, column=1).value
            b_val = ws.cell(row=row_num, column=2).value

            if a_val is not None:
                current_plant = str(a_val).strip()

            item_name = str(b_val).strip() if b_val is not None else None
            if not item_name or current_plant is None:
                continue  # blank spacer row

            for col_num, report_month in month_cols:
                raw = ws.cell(row=row_num, column=col_num).value
                val = clean_val(raw)
                stored = round(val, 3) if val is not None else None

                cursor.execute(
                    """
                    INSERT INTO production_plan_table
                        (report_month, plant_name, item_name, month_actual)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT(report_month, plant_name, item_name)
                    DO UPDATE SET month_actual = excluded.month_actual
                    """,
                    (report_month, current_plant, item_name, stored),
                )
                vals_written += 1

                # For SSP and VISL, Finished Steel = Saleable Steel.
                # Mirror the value so plan pages that query "Finished Steel" find data.
                if item_name == "Saleable Steel" and current_plant in ("SSP", "VISL"):
                    cursor.execute(
                        """
                        INSERT INTO production_plan_table
                            (report_month, plant_name, item_name, month_actual)
                        VALUES (?, ?, 'Finished Steel', ?)
                        ON CONFLICT(report_month, plant_name, item_name)
                        DO UPDATE SET month_actual = excluded.month_actual
                        """,
                        (report_month, current_plant, stored),
                    )
                    vals_written += 1

        if vals_written == 0:
            raise ValueError(
                "No data rows found in the ASP/SSP/VISL Plan Excel. "
                "Check that the sheet has plant names in col A, item names in col B, "
                "and datetime headers in row 1 from col C onward."
            )

        conn.commit()
        conn.close()
        logger.info(
            f"ASP/SSP/VISL Plan extraction done: {vals_written} cells written "
            f"for FY {financial_year}."
        )
        return True

    except ValueError as ve:
        logger.error(f"ASP/SSP/VISL Plan validation error: {ve}")
        raise
    except Exception as e:
        logger.error(f"ASP/SSP/VISL Plan extraction error: {e}")
        return False


# ---------------------------------------------------------------------------
# ASP ABP Plan PDF extractor — BARS, FORGINGS, PLATES + Finished Steel
# ---------------------------------------------------------------------------
#
# PDF layout (single page, UNIT - T):
#   Header:  ITEM APR MAY JUN Q-1 JUL AUG SEP Q-2 H-1 OCT NOV DEC Q-3 JAN FEB MAR Q-4 H-2 TOTAL
#   Data rows: item name followed by 19 numbers
#
# Column positions (0-indexed from the numbers on each row):
#   0=APR  1=MAY  2=JUN  3=Q-1(skip)
#   4=JUL  5=AUG  6=SEP  7=Q-2(skip)  8=H-1(skip)
#   9=OCT  10=NOV  11=DEC  12=Q-3(skip)
#   13=JAN  14=FEB  15=MAR  16=Q-4(skip)  17=H-2(skip)  18=TOTAL(skip)
#
# Values in PDF are in Tonnes → stored as '000T (÷ 1000).
# Finished Steel = BARS + FORGINGS + PLATES (computed per month).
# ---------------------------------------------------------------------------

_ASP_COL_MONTHS = [
    (0, 4),  (1, 5),  (2, 6),           # APR MAY JUN
    (4, 7),  (5, 8),  (6, 9),           # JUL AUG SEP
    (9, 10), (10, 11), (11, 12),        # OCT NOV DEC
    (13, 1), (14, 2), (15, 3),          # JAN FEB MAR
]

_ASP_PLAN_ITEMS = [
    ("^BARS",     "BARS"),
    ("^FORGINGS", "FORGINGS"),
    ("^PLATES",   "PLATES"),
]


def extract_preview_pdf(file_path: str, month: str) -> dict:
    """
    Extract ASP ABP Production Plan from PDF — all 12 months of the FY.

    Extracts BARS, FORGINGS, PLATES and computes Finished Steel per month.
    Values are in Tonnes in the PDF → stored as '000T (÷ 1000).

    month: any YYYY-MM in the target FY (e.g. '2026-04' for FY 2026-27).
    Returns a preview dict with plan_rows — no DB writes.
    """
    import pdfplumber
    import re

    y, m_num = int(month[:4]), int(month[5:7])
    fy_start = y if m_num >= 4 else y - 1
    fy_end   = fy_start + 1
    financial_year = f"{fy_start}-{str(fy_end)[2:]}"

    with pdfplumber.open(file_path) as pdf:
        n_pages = len(pdf.pages)
        full_text = "\n".join(pg.extract_text() or "" for pg in pdf.pages)

    lines = full_text.splitlines()

    def fy_year(mon_num):
        return fy_start if mon_num >= 4 else fy_end

    def parse_nums(line):
        """All integers on the line (including 0), excluding year-like 2000-2099."""
        result = []
        for tok in re.findall(r'\d[\d,]*', line):
            try:
                v = int(tok.replace(',', ''))
            except ValueError:
                continue
            if 2000 <= v <= 2099:
                continue
            result.append(v)
        return result

    def find_row(pattern):
        """Return first line matching pattern anchored near start of line."""
        rx = re.compile(r'^\s*' + pattern, re.IGNORECASE)
        for ln in lines:
            if rx.match(ln):
                return ln
        return None

    def extract_item_rows(pattern, item_name):
        ln = find_row(pattern)
        if ln is None:
            return []
        nums = parse_nums(ln)
        rows = []
        for col_idx, mon_num in _ASP_COL_MONTHS:
            val_t = nums[col_idx] if col_idx < len(nums) else None
            yr = fy_year(mon_num)
            rows.append({
                "item_name": item_name,
                "month":     f"{yr}-{mon_num:02d}",
                "value":     round(val_t / 1000.0, 4) if val_t is not None else None,
                "unit":      "'000T",
                "plant":     "ASP",
                "status":    "ok" if val_t is not None else "skip",
            })
        return rows

    bars_rows     = extract_item_rows(_ASP_PLAN_ITEMS[0][0], _ASP_PLAN_ITEMS[0][1])
    forgings_rows = extract_item_rows(_ASP_PLAN_ITEMS[1][0], _ASP_PLAN_ITEMS[1][1])
    plates_rows   = extract_item_rows(_ASP_PLAN_ITEMS[2][0], _ASP_PLAN_ITEMS[2][1])

    # Compute Finished Steel = BARS + FORGINGS + PLATES per month
    by_month = {}
    for rows in (bars_rows, forgings_rows, plates_rows):
        for r in rows:
            by_month.setdefault(r["month"], []).append(r["value"] or 0)

    fs_rows = [
        {
            "item_name": "Finished Steel",
            "month":     mon,
            "value":     round(sum(vals), 4),
            "unit":      "'000T",
            "plant":     "ASP",
            "status":    "ok",
        }
        for mon, vals in sorted(by_month.items())
    ]

    plan_rows = bars_rows + forgings_rows + plates_rows + fs_rows
    ok = sum(1 for r in plan_rows if r["status"] == "ok")
    logger.info("ASP plan PDF: %d/%d rows ok for FY %s", ok, len(plan_rows), financial_year)

    if ok == 0:
        raise ValueError(
            "No plan values found. Verify this is the ASP ABP Monthwise Production Plan PDF "
            "with rows BARS, FORGINGS, PLATES."
        )

    return {
        "plant":          "ASP",
        "month":          month,
        "financial_year": financial_year,
        "source_type":    f"ASP ABP Production Plan {financial_year} (PDF)",
        "sheets":         f"PDF ({n_pages} page) — 12 months, 4 items",
        "workbook_sheets": [f"PDF ({n_pages} page)"],
        "production_rows":    [],
        "plan_rows":          plan_rows,
        "special_steel_rows": [],
        "techno_rows":        [],
        "techno_param_rows":  [],
    }

