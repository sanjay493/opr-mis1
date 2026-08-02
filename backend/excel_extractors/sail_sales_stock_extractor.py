"""
SAIL "1 page report" extractor — Table A (Sales) and Table D (Stock -
8 Plants) from the monthly "<Mon>'YY PS CMO" workbook
(Report_format/1 page report for <Mon><YY>.xlsx style, single sheet).

Tables B (Production) and C (Techno-Economic Parameters) are NOT
extracted here — they're already derivable from production_table /
techno tables already in this app (see page_one_page_report.py).

Table A (Sales) — anchor row starts with 'A.' and contains 'SALES';
row layout: label (col B) | .. | month ABP (D) | month Actual (E) |
month %Ful (F) | month CPLY Actual (G) | month growth (H) | cum ABP
(I) | cum Actual (J) | cum %Ful (K) | cum CPLY Actual (L) | cum growth
(M). ALL ten columns are stored verbatim, as a single JSON blob per
(report_month, item) — this table is a pure archive of what the
source department reported that month, not a computed view. Two
reasons this matters:
  1. We don't have — and can't assume we'll ever backfill — the prior
     year's own report, so deriving CPLY/growth ourselves from our own
     stored history would silently go blank instead of showing the
     figure the department already computed and included.
  2. The department's own cumulative (Apr-month) figures are
     provisional and get revised between reports — they are NOT
     reliably equal to summing each month's own individually-reported
     Actual. Re-deriving "Apr-Jun" by summing Apr+May+Jun as separately
     stored would produce a different number than what was actually
     reported for Jun, silently diverging from the source.
So: extract exactly what's on the row, store it exactly, and the
generator (page_one_page_report.py) reads it back with zero
recalculation.

Table D (Stock - 8 Plants) — anchor row starts with a single letter +
'.' (the section letter varies by report vintage — seen as both 'D.'
and 'B.') and contains 'STOCK'; row layout: label | .. | one column
per historical snapshot date (e.g. '1.4.21', '01.06.26') | a trailing
'Var. w.r.t. ...' column (not a real date, skipped). Every date column
present is extracted (not just the latest) so a single file backfills
history — the same file that reports June's figures also carries
1.4.21 through 1.4.25 as reference columns.

Two report layouts are in circulation and this module auto-detects
which one it's looking at rather than hardcoding column positions:
  - "4-section" layout (A. Sales, B. Production, C. Techno, D. Stock):
    item label sits one column right of the anchor (col B when the
    anchor is in col A), with an empty column before data starts.
  - "2-section" layout (A. Sales, B. Stock only, newer "as on <date>"
    workbooks): item label sits in the SAME column as the anchor, and
    data starts immediately in the next column.
Rather than branch on a layout flag, each row's label is looked up in
column A first, then column B — whichever holds the known label text
wins — and the first Sales data column is located by searching the
anchor row for the literal header 'ABP' instead of assuming a fixed
column. The ten Sales fields and the Stock date columns are then read
as a fixed run of columns from that anchor, since column ORDER is
consistent across both layouts even though the starting column isn't.
"""
import re
from datetime import date, datetime

_SALES_LABELS = [
    "LP SALES", "FP SALES", "PET SALES", "TOTAL : HOME SALES",
    "SALES BY SPECIAL STEEL PLANTS", "EXPORTS", "TOTAL CMO SALES",
    "PLANT SALES", "TOTAL SALES",
]
_STOCK_LABELS = ["PLANTS", "STOCKYARDS", "STOCK IN TRANSIT", "TOTAL"]

# json keys in on-sheet column order, starting from wherever the 'ABP'
# header is found in the Sales anchor row (see module docstring)
_SALES_FIELD_ORDER = [
    "month_abp", "month_actual", "month_ful", "month_cply", "month_growth",
    "till_month_abp", "till_month_actual", "till_month_ful",
    "till_month_cply", "till_month_growth",
]
_MAX_SCAN_COL = 30

_DATE_RE = re.compile(r'^(\d{1,2})\.(\d{1,2})\.(\d{2})$')


def _clean(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _norm(s):
    return re.sub(r'\s+', ' ', str(s or "")).strip().upper()


def _load_grid(file_path: str):
    with open(file_path, "rb") as f:
        magic = f.read(4)
    if magic[:2] == b"PK":
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        grid = [list(row) for row in wb[wb.sheetnames[0]].iter_rows(values_only=True)]
        wb.close()
        return grid
    if magic == b"\xd0\xcf\x11\xe0":
        import xlrd
        sh = xlrd.open_workbook(file_path).sheet_by_index(0)
        return [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    raise ValueError(
        "Unrecognised file format — expected the '<Mon>'YY PS CMO' Excel "
        "workbook (.xlsx/.xls)."
    )


def _cell(grid, row_1b, col_1b):
    if row_1b < 1 or row_1b > len(grid):
        return None
    row = grid[row_1b - 1]
    if col_1b < 1 or col_1b > len(row):
        return None
    v = row[col_1b - 1]
    return None if v == "" else v


_SECTION_LETTER_RE = re.compile(r'^[A-Z]\.')


def _find_anchor(grid, contains):
    for r in range(1, len(grid) + 1):
        text = _norm(_cell(grid, r, 1))
        if _SECTION_LETTER_RE.match(text) and contains in text:
            return r
    return None


def _find_row_label(grid, r, remaining):
    """Item labels sit in col A (2-section layout) or col B (4-section
    layout) depending on report vintage — try both. Matched by prefix,
    not exact equality, since some months append annotations onto the
    label itself (e.g. 'TOTAL : HOME SALES incl NSL'); no two labels in
    _SALES_LABELS/_STOCK_LABELS share a prefix, so this stays unambiguous.
    Returns the canonical label (from `remaining`), not the raw cell text."""
    for col in (1, 2):
        text = _norm(_cell(grid, r, col))
        if not text:
            continue
        for candidate in remaining:
            if text.startswith(candidate):
                return candidate
    return None


def _parse_snapshot_date(header_text):
    if isinstance(header_text, datetime):
        return header_text.date()
    if isinstance(header_text, date):
        return header_text
    m = _DATE_RE.match(str(header_text or "").strip())
    if not m:
        return None
    d, mo, yy = (int(g) for g in m.groups())
    try:
        return date(2000 + yy, mo, d)
    except ValueError:
        return None


def extract_preview(file_path: str, report_month: str, **_kwargs) -> dict:
    """
    Extract Table A (Sales) + Table D (Stock) from the SAIL 1-page report.
    Returns {"sales_rows": [...], "stock_rows": [...]} — no DB writes.

    sales_rows: [{"item_name", "data": {month_abp, month_actual, month_ful,
                  month_cply, month_growth, till_month_abp, till_month_actual,
                  till_month_ful, till_month_cply, till_month_growth},
                  "cell", "status"}]
        Every field in "data" is exactly what's in the source cell — no
        %Ful/growth/cumulative is computed here (see the module docstring).
    stock_rows: [{"item_name", "snapshot_date" (YYYY-MM-DD), "value", "cell", "status"}]
    """
    grid = _load_grid(file_path)

    sales_rows = []
    sales_anchor = _find_anchor(grid, "SALES")
    if sales_anchor is None:
        raise ValueError(
            "Sales table not found — expected a row starting with a "
            "section letter (e.g. 'A.') and containing 'SALES' (e.g. "
            "\"A.    SALES ['000 T]\")."
        )
    abp_col = None
    for c in range(2, _MAX_SCAN_COL):
        if _norm(_cell(grid, sales_anchor, c)) == "ABP":
            abp_col = c
            break
    if abp_col is None:
        raise ValueError(
            "Sales table found but no 'ABP' column header on that row — "
            "cannot locate the data columns."
        )
    sales_cols = [(key, abp_col + i) for i, key in enumerate(_SALES_FIELD_ORDER)]

    remaining = list(_SALES_LABELS)
    last_sales_row = sales_anchor
    for r in range(sales_anchor + 1, min(sales_anchor + 15, len(grid)) + 1):
        label = _find_row_label(grid, r, remaining)
        if label is not None:
            remaining.remove(label)
            data = {key: _clean(_cell(grid, r, col)) for key, col in sales_cols}
            sales_rows.append({
                "item_name": label.title(),
                "data": data,
                "cell": f"row {r}",
                "status": "ok" if any(v is not None for v in data.values()) else "skip",
            })
            last_sales_row = r
        if not remaining:
            break

    # An asterisked remark (e.g. "*Jul25 & Apr-Jul25 fig incl NSL sales: 98 &
    # 482 respectively") sometimes follows the Sales table, in whichever
    # column happens to be free on that row — position varies by report
    # vintage same as everything else here, so just scan for a cell
    # starting with '*' in the few rows right after the last Sales row.
    sales_note = None
    for r in range(last_sales_row + 1, min(last_sales_row + 3, len(grid)) + 1):
        for c in range(1, _MAX_SCAN_COL):
            text = str(_cell(grid, r, c) or "").strip()
            if text.startswith("*"):
                sales_note = text
                break
        if sales_note:
            break
    for label in remaining:
        sales_rows.append({
            "item_name": label.title(),
            "data": {key: None for key in _SALES_FIELD_ORDER},
            "cell": "", "status": "not found",
        })

    stock_rows = []
    stock_anchor = _find_anchor(grid, "STOCK")
    if stock_anchor is None:
        raise ValueError(
            "Stock table not found — expected a row starting with a "
            "section letter (e.g. 'D.' or 'B.') and containing 'STOCK' "
            "(e.g. \"D.    STOCK  - 8 PLANTS ['000 T]\")."
        )
    header_row = stock_anchor
    date_cols = []
    for c in range(2, _MAX_SCAN_COL):
        snap = _parse_snapshot_date(_cell(grid, header_row, c))
        if snap is not None:
            date_cols.append((c, snap))
        elif date_cols:
            break

    remaining = list(_STOCK_LABELS)
    for r in range(stock_anchor + 1, min(stock_anchor + 10, len(grid)) + 1):
        label = _find_row_label(grid, r, remaining)
        if label is not None:
            remaining.remove(label)
            for col, snap in date_cols:
                val = _clean(_cell(grid, r, col))
                stock_rows.append({
                    "item_name": label.title(),
                    "snapshot_date": snap.isoformat(),
                    "value": val,
                    "cell": f"row {r} / {snap.isoformat()}",
                    "status": "ok" if val is not None else "skip",
                })
        if not remaining:
            break
    for label in remaining:
        stock_rows.append({
            "item_name": label.title(), "snapshot_date": None, "value": None,
            "cell": "", "status": "not found",
        })

    ok_sales = sum(1 for r in sales_rows if r["status"] == "ok")
    ok_stock = sum(1 for r in stock_rows if r["status"] == "ok")
    if ok_sales == 0 and ok_stock == 0:
        raise ValueError(
            "No values extracted from either table — verify this is the "
            "SAIL '1 page report' workbook."
        )

    return {
        "report_month": report_month,
        "source_type": "SAIL 1-Page Report (Sales & Stock)",
        "sales_rows": sales_rows,
        "stock_rows": stock_rows,
        "sales_note": sales_note,
    }
