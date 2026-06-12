"""
RSP Techno-Parameter extractor (e.g. 'technopara may-2026.xlsx').

Position-tolerant by design:
  1. Value COLUMNS are located by searching header cells for the report-month
     label ("May'26") and the cumulative label ("Apr-May'26") in any common
     spelling.  Falls back to the known defaults (X / AM) only if no header
     matches.
  2. Parameter ROWS are located by anchor text: first the shop heading
     ("Plate Mill", "HSM-II", ...) then the parameter label inside that shop's
     row range.  Default row hints (from the sample mapping) are used only as
     a last resort.

extract_techno() returns preview rows WITHOUT writing to the DB — insertion
happens separately after the user confirms.
"""
import re
import datetime
import openpyxl

PLANT = "RSP"
GROUP = "MILL_RSP"

# default column hints from the sample mapping (X = 24, AM = 39)
DEFAULT_MONTH_COL = 24
DEFAULT_CUM_COL   = 39
MAX_SCAN_ROWS = 600
MAX_SCAN_COLS = 60
LABEL_COLS    = 14          # parameter labels live in the first few columns

_MON = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
        'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def _norm(s):
    return re.sub(r'[^a-z0-9]+', ' ', str(s).lower()).strip()


def _clean(val):
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in ("", "nan", "-", "###", "#div/0!", "#value!", "cr", "na"):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


# ── shop / parameter configuration ───────────────────────────────────────────
# (shop section, [shop aliases], [(row_label, [param aliases], unit, sort, row_hint)])
# row_hint = sample-mapping row, used only when anchor search fails.

SHOPS = [
    ("Plate Mill", ["plate mill"], [
        ("Yield - Primes",      ["yield primes", "yield prime"],          "%",       1, 219),
        ("Yield - Total",       ["yield total"],                          "%",       2, 221),
        ("Average Slab Weight", ["average slab weight", "avg slab weight"], "Tons",  3, 222),
        ("Mill Availability",   ["mill availability"],                    "%",       4, 223),
        ("Mill Utilisation",    ["mill utilisation", "mill utilization"], "%Avl.",   5, 224),
        ("Rolling Rate",        ["rolling rate"],                         "T/Hr.",   6, None),
        ("Specific Heat",       ["specific heat", "sp heat cons"],        "M.Cal/T", 7, 309),
        ("Specific Power",      ["specific power", "sp power cons"],      "Kwh/T",   8, 327),
    ]),
    ("HSM-II", ["hsm ii", "hsm 2", "hot strip mill ii", "hot strip mill 2"], [
        ("H R Coil Yield",        ["h r coil yield", "hr coil yield"],     "%",      10, None),
        ("Average Slab Weight",   ["average slab weight"],                 "Tons",   11, None),
        ("Mill Availability",     ["mill availability"],                   "%",      12, None),
        ("Mill Utilisation",      ["mill utilisation", "mill utilization"],"%Avl.",  13, None),
        ("Rolling Rate -HR Coils",["rolling rate hr coils", "rolling rate"],"T/Hr.", 14, None),
        ("Specific Heat",         ["specific heat"],                       "M.Cal/T",15, None),
        ("Specific Power",        ["specific power"],                      "Kwh/T",  16, None),
    ]),
    ("SSM", ["ssm", "silicon steel mill"], [
        ("Yield from HRC-CRNO",  ["yield from hrc crno", "yield from hrc"], "%",    20, None),
        ("Acid cons. in AP line",["acid cons in ap line", "acid cons"],     "Kg/T", 21, None),
        ("BUST Availability",    ["bust availability"],                     "%",    22, None),
        ("Mill Utilisation",     ["mill utilisation", "mill utilization"],  "%Avl.",23, None),
        ("Rolling Rate(Rev Mill)",["rolling rate rev mill", "rolling rate"],"T/Hr", 24, None),
    ]),
    ("ERW Pipe Plant", ["erw pipe plant", "erw pipe"], [
        ("Yield from HR Coils", ["yield from hr coils"],                  "%",     30, None),
        ("Mill Availability",   ["mill availability"],                    "%",     31, None),
        ("Mill Utilisation",    ["mill utilisation", "mill utilization"], "%Avl.", 32, None),
        ("Rolling Rate",        ["rolling rate"],                         "%",     33, None),
    ]),
    ("SW Pipe Plant", ["sw pipe plant", "sw pipe", "spiral weld pipe"], [
        ("Yield from HR Coils", ["yield from hr coils"],                  "%",     40, None),
        ("Mill Availability",   ["mill availability"],                    "%",     41, None),
        ("Mill Utilisation",    ["mill utilisation", "mill utilization"], "%Avl.", 42, None),
        ("Rolling Rate",        ["rolling rate"],                         "T/Hr.", 43, None),
    ]),
    ("New Plate Mill", ["new plate mill", "npm"], [
        ("Yield - Primes",      ["yield primes", "yield prime"],          "%",      50, None),
        ("Yield - Total",       ["yield total"],                          "%",      51, None),
        ("Average Slab Weight", ["average slab weight"],                  "Tons",   52, None),
        ("Mill Availability",   ["mill availability"],                    "%",      53, None),
        ("Mill Utilisation",    ["mill utilisation", "mill utilization"], "%Avl.",  54, None),
        ("Rolling Rate",        ["rolling rate"],                         "T/Hr.",  55, None),
        ("Specific Heat",       ["specific heat"],                        "M.Cal/T",56, None),
        ("Specific Power",      ["specific power"],                       "Kwh/T",  57, None),
    ]),
]


# ── column detection ──────────────────────────────────────────────────────────

def _month_aliases(ym):
    y, m = int(ym[:4]), int(ym[5:7])
    mon, yy = _MON[m], str(y)[2:]
    return {_norm(x) for x in (
        f"{mon}'{yy}", f"{mon} {yy}", f"{mon}-{yy}", f"{mon}{yy}",
        f"{mon}' {yy}", f"{mon} 20{yy}", f"{mon}'20{yy}",
    )}

def _cum_aliases(ym):
    y, m = int(ym[:4]), int(ym[5:7])
    mon, yy = _MON[m], str(y)[2:]
    out = set()
    for apr in ("apr", "april"):
        for sep in ("-", " ", " - ", " to "):
            out.add(_norm(f"{apr}{sep}{mon}'{yy}"))
            out.add(_norm(f"{apr}{sep}{mon} {yy}"))
            out.add(_norm(f"{apr}{sep}{mon}-{yy}"))
            out.add(_norm(f"{apr}{sep}{mon} 20{yy}"))
    return out


def _find_value_columns(ws, report_month):
    """Locate the month-actual and cumulative columns by header text."""
    y, m = int(report_month[:4]), int(report_month[5:7])
    want_month = _month_aliases(report_month)
    want_cum   = _cum_aliases(report_month)
    month_col = cum_col = None

    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, MAX_SCAN_ROWS),
                            max_col=min(ws.max_column, MAX_SCAN_COLS)):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            if isinstance(v, datetime.datetime):
                if v.year == y and v.month == m and month_col is None:
                    month_col = cell.column
                continue
            t = _norm(v)
            if not t:
                continue
            if cum_col is None and t in want_cum:
                cum_col = cell.column
            elif month_col is None and t in want_month:
                month_col = cell.column
        if month_col and cum_col:
            break

    return (month_col or DEFAULT_MONTH_COL,
            cum_col   or DEFAULT_CUM_COL,
            month_col is not None,
            cum_col   is not None)


# ── row anchors ───────────────────────────────────────────────────────────────

def _label_at(ws, r):
    """Concatenated normalized text of the first LABEL_COLS cells of a row."""
    parts = []
    for c in range(1, LABEL_COLS + 1):
        v = ws.cell(row=r, column=c).value
        if v is not None and not isinstance(v, (int, float, datetime.datetime)):
            parts.append(str(v))
    return _norm(" ".join(parts))


def _find_shop_rows(ws, max_row):
    """shop section name → anchor row. Searched across the label columns."""
    anchors = {}
    for r in range(1, max_row + 1):
        t = _label_at(ws, r)
        if not t:
            continue
        for section, aliases, _ in SHOPS:
            if section in anchors:
                continue
            if any(t == a or t.startswith(a + " ") or (" " + a) in (" " + t)
                   for a in aliases):
                # avoid matching a parameter row that merely mentions the shop
                if len(t) <= len(max(aliases, key=len)) + 25:
                    anchors[section] = r
    return anchors


def _find_param_row(ws, aliases, start, end):
    for r in range(start, end + 1):
        t = _label_at(ws, r)
        if not t:
            continue
        for a in aliases:
            if t == a or t.startswith(a):
                return r
    return None


# ── main entry ────────────────────────────────────────────────────────────────

def _pick_sheet(wb):
    """Sheet with the most shop anchors wins."""
    best, best_hits = wb.worksheets[0], -1
    for ws in wb.worksheets:
        hits = len(_find_shop_rows(ws, min(ws.max_row, MAX_SCAN_ROWS)))
        if hits > best_hits:
            best, best_hits = ws, hits
    return best


def extract_techno(file_path: str, report_month: str) -> dict:
    """Extract RSP techno parameters. Returns a preview dict — NO DB writes."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = _pick_sheet(wb)
    max_row = min(ws.max_row, MAX_SCAN_ROWS)

    month_col, cum_col, mc_found, cc_found = _find_value_columns(ws, report_month)
    shop_rows = _find_shop_rows(ws, max_row)

    # shop ranges: anchor → next anchor (or sheet end)
    ordered = sorted(shop_rows.items(), key=lambda kv: kv[1])
    ranges = {}
    for i, (sec, r) in enumerate(ordered):
        end = ordered[i + 1][1] - 1 if i + 1 < len(ordered) else max_row
        ranges[sec] = (r, end)

    def col_letter(c):
        return openpyxl.utils.get_column_letter(c)

    rows = []
    for section, aliases, params in SHOPS:
        rng = ranges.get(section)
        for label, p_aliases, unit, sort, row_hint in params:
            row = None
            via = ""
            if rng:
                row = _find_param_row(ws, [_norm(a) for a in p_aliases], rng[0], rng[1])
                via = "anchor"
            if row is None and row_hint:
                row = row_hint
                via = "row-hint"

            actual = cum = None
            cell_ref = ""
            if row is not None:
                actual = _clean(ws.cell(row=row, column=month_col).value)
                cum    = _clean(ws.cell(row=row, column=cum_col).value)
                cell_ref = f"{col_letter(month_col)}{row}/{col_letter(cum_col)}{row}"

            status = ("ok" if (actual is not None or cum is not None)
                      else ("no value" if row is not None else "not found"))
            rows.append({
                "plant": PLANT,
                "group_code": GROUP,
                "section": section,
                "parameter": label,
                "unit": unit,
                "sort_order": sort,
                "month": report_month,
                "actual": actual,
                "cum_actual": cum,
                "cell": cell_ref,
                "found_via": via if row is not None else "",
                "status": status,
            })

    return {
        "plant": PLANT,
        "month": report_month,
        "sheet": ws.title,
        "month_col": col_letter(month_col),
        "cum_col": col_letter(cum_col),
        "columns_detected": bool(mc_found and cc_found),
        "shops_found": sorted(shop_rows.keys()),
        "rows": rows,
    }
