import re
import calendar
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
import logging
import os
from datetime import datetime
from typing import Optional

logger = logging.getLogger("excel_extractor")

MONTH_NAMES = {
    "01": "January", "02": "February", "03": "March", "04": "April",
    "05": "May",     "06": "June",     "07": "July",  "08": "August",
    "09": "September","10": "October", "11": "November","12": "December"
}
MONTH_NUMS = {v: k for k, v in MONTH_NAMES.items()}


def clean_val(val) -> Optional[float]:
    if val is None or str(val).strip().lower() in ("nan", "###", "-", "#div/0!", ""):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# DPR Mail production — CUM column is authoritative; M.RATE is a cross-check
# ---------------------------------------------------------------------------
# Both production tables on the DPR sheet carry the plant's own full-month
# figure one column to the RIGHT of the month-to-date cumulative:
#   "MAIN UNITS"        block  — col O = CUM, col P  = M.RATE
#   "SALEABLE STEEL"    block  — col Z = CUM, col AA = M.RATE
# M.RATE is *supposed* to already be the whole-month figure (a no-op copy of
# CUM on the last day, or BSL's own projection mid-month) but BSL's own
# report is known to sometimes get it wrong even at month-end — confirmed
# against real files: e.g. a 31/05/2026 month-end report's "H R SHEET" row
# reads CUM 3872 vs M.RATE 3168, an 18% gap, with no such gap anywhere else
# on that sheet. That is a flaw in BSL's own report, not a legitimate
# rework/downgrade adjustment, so CUM (or, mid-month, CUM projected to the
# full month — cum x days_in_month / report_day) is always what gets saved.
# M.RATE is only read to cross-check that projection and surface an alert
# when the two disagree by more than 10% (see _dpr_value_and_alert) — it is
# never used to pick the saved value.

def _dpr_report_day(o1_raw):
    """(report_day, days_in_month, 'YYYY-MM') from DPR cell O1 — an Excel
    datetime or a DD.MM.YYYY string. (None, None, None) if unreadable."""
    y = mth = day = None
    if isinstance(o1_raw, datetime):
        y, mth, day = o1_raw.year, o1_raw.month, o1_raw.day
    elif o1_raw:
        m = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", str(o1_raw))
        if m:
            day, mth, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if not (y and mth and day):
        return None, None, None
    return day, calendar.monthrange(y, mth)[1], f"{y}-{str(mth).zfill(2)}"


def _dpr_is_mid_month(report_day, days_in_month) -> bool:
    return bool(report_day and days_in_month and 0 < report_day < days_in_month)


def _dpr_cum_cell(mrate_addr: str) -> str:
    """The CUM cell one column left of an M.RATE cell (P->O, AA->Z)."""
    col, row = coordinate_from_string(mrate_addr)
    return f"{get_column_letter(column_index_from_string(col) - 1)}{row}"


def _dpr_expected(cum, report_day, days_in_month, per_day: bool):
    """The full-month figure derived purely from CUM and the report date: a
    per-day average for nos/day items, else the month-to-date total scaled
    to a full month (a no-op at month-end). None if not computable."""
    if cum is None or not report_day:
        return None
    if per_day:
        return cum / report_day
    if 0 < report_day < (days_in_month or 0):
        return cum * days_in_month / report_day
    return cum


_DPR_ERRATIC_LO, _DPR_ERRATIC_HI = 0.9, 1.1  # ±10% — see module note above


def _dpr_value_and_alert(mrate, cum, report_day, days_in_month, per_day=False,
                          label=""):
    """The item's month figure — always derived from CUM, never from
    M.RATE (see module note above) — plus an alert string when M.RATE
    disagrees with that figure by more than 10%, else None.

    Returns (value, basis, alert). basis is 'projected' (mid-month),
    'cum' (month-end/undated), or '' (CUM itself is missing)."""
    mid = bool(report_day and 0 < report_day < (days_in_month or 0))
    exp = _dpr_expected(cum, report_day, days_in_month, per_day)
    if exp is not None:
        value, basis = exp, ("projected" if mid else "cum")
    elif cum is not None:
        value, basis = cum, "cum"
    else:
        value, basis = None, ""

    alert = None
    if mrate not in (None, 0) and exp not in (None, 0):
        ratio = mrate / exp
        if not (_DPR_ERRATIC_LO <= ratio <= _DPR_ERRATIC_HI):
            alert = (
                f"{label}: sheet's M.RATE ({mrate:,.1f}) differs from the "
                f"cumulative-derived figure ({exp:,.1f}, from CUM "
                f"{cum:,.1f}) by {abs(ratio - 1) * 100:.0f}% — the "
                f"cumulative figure was used; M.RATE looks erratic on "
                f"this report."
            )
            logger.warning("BSL DPR: %s", alert)
    return value, basis, alert


def extract_and_save_excel(file_path: str, report_month: str = "", source_file_name: str = "") -> bool:
    """
    Dispatcher: auto-detects BSL file type by sheet name and calls the correct extractor.

    Supported file types:
      • DPR Mail (.xlsx)         — sheet 'DPR'  (month-end daily production report)
      • Final Monthly Report     — (commented out, to be implemented)
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames
        logger.info(f"BSL: loading file. Sheets: {sheet_names}")

        if "DPR" in sheet_names:
            return _extract_dpr_report(wb, source_file_name)

        raise ValueError(
            "Uploaded BSL file does not match any known format. "
            "Expected sheet 'DPR' (DPR Mail month-end report) or "
            "Sheet1 with 'SPECIAL STEEL' title (Corporate Office Special Steel report — "
            "use Extract & Preview workflow for that file)."
        )
    except ValueError as ve:
        logger.error(f"BSL validation error: {ve}")
        raise ve
    except Exception as e:
        logger.error(f"BSL extraction error: {e}")
        return False


# ---------------------------------------------------------------------------
# Extractor 1 — BSL DPR Mail (month-end daily production report, sheet: DPR)
# ---------------------------------------------------------------------------

def _dpr_config():
    """
    Single source of truth for the BSL DPR Mail cell mapping — shared by the
    DB-writing extractor (_extract_dpr_report) and the preview-only extractor
    (_extract_dpr_preview) so they can never drift apart.

    Every cell points at the plant's own M.RATE figure — column P for the
    "MAIN UNITS" block, column AA for the "PRODUCTION (SALEABLE STEEL)"
    block. The month-to-date CUM figure sits one column to the left (O / Z).
    These addresses are only a starting point / last-resort fallback now —
    see _resolve_dpr_cells, which re-locates every row by its own label text
    on each upload (rows here drift between report vintages; see its
    docstring for a confirmed real example).

    Reads excel_cells_config.json (section 'bsl_dpr'); falls back to these
    hardcoded defaults only if that config section is missing.
    """
    import sys
    sys.path.insert(0, os.path.dirname(__file__))
    from cells_loader import get_extractor_config
    cfg = get_extractor_config("bsl_dpr")

    cells = cfg.get("cells", {
        "Oven Pushing (nos/day)": "P6",
        "Total Sinter":        "P7",
        "Hot Metal":           "P8",
        "Pig Iron":            "AA21",
        "SMS-1 CCM-1":         "P10",
        "SMS-2 CCM-1&2":       "P11",
        "Total Crude Steel":   "P12",
        "HSM Total HR Coil":   "P14",
        "HSM HR Coil (Sale)":  "AA6",
        "HSM HR Plate":        "AA7",
        "HR Sheet":            "AA8",
        "CR I/II CR(Coil) Sale": "AA9",
        "CR Sheets":           "AA10",
        "CR III CR(Coil) Sale": "AA15",
        "GPC3":                "AA16",
        "CRSALE":              "AA18",
        "Saleable Steel":      "P31",
        "Saleable Semis":      "AA19",
        "Thick Plate":         "AA20",
    })
    no_convert = set(cfg.get("no_convert", ["Oven Pushing (nos/day)"]))
    derived = cfg.get("derived", [
        {"item": "Finished Steel", "op": "subtract", "a_item": "Saleable Steel", "b_item": "Saleable Semis"},
        {"item": "GP/GC", "op": "add_labels", "labels": ["_GP_SHEET", "_GC_SHEET", "_GP_COIL"]},
        # Replaces the old "CRC&S(1&2)" item, which read the CRM(1+2)
        # mother-mill row (AA14/"CRM(1+2)" label) directly — confirmed to
        # diverge from the sales-side figures below in some months (see
        # page_catwise_saleable.py's note, e.g. Sep'25: 35.7 vs the true
        # 47.4). "CR(1&2) Total Saleable" is unambiguously defined instead,
        # as the sum of its own already-resolved sibling items — see the
        # "sum_items" op below, which runs after GP/GC so it's available.
        {"item": "CR(1&2) Total Saleable", "op": "sum_items",
         "items": ["CR I/II CR(Coil) Sale", "CR Sheets", "GP/GC"]},
        # Replaces the old "CRC(3)" item, which read the CRM-3 mother-mill
        # row (AA17/"CRM-3" label) directly — like CRC&S(1&2) above, this
        # diverged from the sales-side total in most months once the DPR
        # route started supplying it regularly (confirmed: 10 of 17
        # overlapping months disagreed even with CR III CR(Coil) Sale
        # alone, let alone once GPC3 is added). "CR III Total Saleable" is
        # unambiguously defined instead, mirroring CR(1&2) Total Saleable.
        {"item": "CR III Total Saleable", "op": "sum_items",
         "items": ["CR III CR(Coil) Sale", "GPC3"]},
    ])
    return cells, no_convert, derived


# Label anchors for every DPR production row — resistant to rows being
# inserted or removed between report vintages, unlike a fixed cell address.
# Confirmed against real files spanning 2016-2026: a "- CHQ PLATE" sub-line
# (under H R PLATE, "SALEABLE STEEL" table) and a "G P COIL }" line have
# each independently come and gone across eras — e.g. 31/08/2026's report
# has "- CHQ PLATE" but not "G P COIL }", while 31/05/2026's has the
# opposite — shifting every row below the one that appeared/disappeared.
# A fixed cell address silently read the wrong neighbouring row whenever
# this happened (HR Sheet, CR I/II CR(Coil) Sale, CR Sheets and the GP/GC
# sum were all wrong on any file with "- CHQ PLATE" present, extracting
# the row above their real one).
#
# value: (label_col, [substrings]). label_col is 'L' for the "MAIN UNITS"
# table (cum=O, m.rate=P) or 'W' for "SALEABLE STEEL" (cum=Z, m.rate=AA).
# First row in range 6-40 whose label cell contains any substring wins.
_DPR_LABELS = {
    "Oven Pushing (nos/day)": ("L", ["OVENS PUSHED"]),
    "Total Sinter":          ("L", ["G.SINTER"]),
    "Hot Metal":             ("L", ["HOT METAL"]),
    "SMS-1 CCM-1":           ("L", ["SMS NEW-SLAB"]),
    "SMS-2 CCM-1&2":         ("L", ["SMS-2/CCS"]),
    "Total Crude Steel":     ("L", ["CRUDE STEEL"]),
    "HSM Total HR Coil":     ("L", ["HSM H.R.COIL"]),
    "Saleable Steel":        ("L", ["SAL.STEEL(P)"]),
    "HSM HR Coil (Sale)":    ("W", ["H R COIL"]),
    "HSM HR Plate":          ("W", ["H R PLATE"]),
    # "- CHQ PLATE" is a sub-line of H R PLATE, present only on some report
    # vintages (see above) — HSM HR Plate already includes it, so this is
    # purely an additional breakdown, not summed into anything else. Saved
    # as "Checkered plate" (not "CHQ Plate") to match the item_name
    # page_catwise_saleable.py/page_segment_wise.py already display this
    # under and that manual entries already use — confirmed no overlapping
    # month between the two spellings (Checkered plate: May-Jul'26 manual;
    # CHQ Plate: Aug'26, this extractor's first month writing it at all).
    "Checkered plate":       ("W", ["CHQ PLATE"]),
    "HR Sheet":              ("W", ["H R SHEET"]),
    "CR I/II CR(Coil) Sale": ("W", ["C R CL/SL"]),
    "CR Sheets":             ("W", ["C R SHEET"]),
    # Neither CRM-3 (mill 3's own mother-mill row) nor CRM(1+2) (mills
    # 1&2's equivalent) are read here any more — both retired in favour of
    # "CR III Total Saleable" / "CR(1&2) Total Saleable" (see _dpr_config's
    # derived rules), which are unambiguously defined from the sales-side
    # rows below instead.
    "CR III CR(Coil) Sale":  ("W", ["CR COIL-3"]),
    "GPC3":                  ("W", ["GP COIL-3"]),
    "CRSALE":                ("W", ["TOT CR SAL"]),
    "Saleable Semis":        ("W", ["SLAB"]),
    "Thick Plate":           ("W", ["COBB PLT"]),
    "Pig Iron":              ("W", ["PIG IRON"]),
}

# The GP/GC derived sum's three possible component rows — not themselves
# saved as DB items, and (unlike everything in _DPR_LABELS) not always all
# present: "G P COIL }" only exists on some vintages (see _DPR_LABELS'
# docstring), so each is searched independently and summed with whichever
# are found (none found → GP/GC is skipped, same as every other item).
_DPR_GPGC_HELPERS = {
    "_GP_SHEET": ("W", ["G P SHEET"]),
    "_GC_SHEET": ("W", ["G C SHEET"]),
    "_GP_COIL":  ("W", ["G P COIL"]),
}

_DPR_TABLE_COLS = {"L": ("O", "P"), "W": ("Z", "AA")}  # label_col -> (cum_col, mrate_col)


def _find_dpr_label_row(ws, col_letter, substrings, row_range=(6, 40)):
    """First row (1-based) in row_range whose col_letter cell's text
    contains any of substrings (case-insensitive, stripped), else None."""
    col = column_index_from_string(col_letter)
    for r in range(row_range[0], row_range[1] + 1):
        label = str(ws.cell(row=r, column=col).value or "").strip().upper()
        if not label:
            continue
        if any(s in label for s in substrings):
            return r
    return None


def _resolve_dpr_cells(ws, defaults):
    """{item_name: mrate_cell_address} — every row re-located by its own
    label text (see _DPR_LABELS) rather than trusted at a fixed address,
    so an inserted/removed row elsewhere on the sheet can't shift a
    neighbouring item's data into the wrong one. Falls back to `defaults`
    (the configured/hardcoded cell) only for an item _DPR_LABELS doesn't
    cover; an item whose label can't be found at all is dropped rather
    than risk reading the wrong row silently."""
    resolved = dict(defaults)
    for item, (label_col, substrings) in _DPR_LABELS.items():
        row = _find_dpr_label_row(ws, label_col, substrings)
        cum_col, mrate_col = _DPR_TABLE_COLS[label_col]
        if row is not None:
            resolved[item] = f"{mrate_col}{row}"
        else:
            resolved.pop(item, None)
    return resolved


def _resolve_dpr_gpgc_helpers(ws):
    """{helper_name: mrate_cell_address or None} for _DPR_GPGC_HELPERS."""
    out = {}
    for name, (label_col, substrings) in _DPR_GPGC_HELPERS.items():
        row = _find_dpr_label_row(ws, label_col, substrings)
        _, mrate_col = _DPR_TABLE_COLS[label_col]
        out[name] = f"{mrate_col}{row}" if row is not None else None
    return out


def _extract_dpr_report(wb, source_file_name: str) -> bool:
    """
    Extracts cumulative production data from the BSL DPR Mail report (sheet 'DPR').

    Date is auto-detected from cell O1 (stored as a Python datetime by Excel).

    Every item's row is re-located by its own label text (see _DPR_LABELS)
    rather than a fixed cell address — column P/AA (M.RATE) and O/Z (CUM)
    stay fixed per table, but which row they land on shifts between report
    vintages. Value saved is always CUM-derived (see _dpr_value_and_alert);
    M.RATE is cross-checked and logged as a warning when it disagrees by
    more than 10%, never used to pick the value.

    Items — see _DPR_LABELS for the label each is found by:
      Oven Pushing (nos/day), Total Sinter, Hot Metal, SMS-1 CCM-1,
      SMS-2 CCM-1&2, Total Crude Steel, HSM Total HR Coil, Saleable Steel
        — "MAIN UNITS" table (col O = CUM, col P = M.RATE)
      HSM HR Coil (Sale), HSM HR Plate, Checkered plate (a sub-line of HSM
      HR Plate, present only on some report vintages — already included in
      HSM HR Plate, saved separately too when its row exists), HR Sheet,
      CR I/II CR(Coil) Sale, CR Sheets (sales-side, mills 1&2), CR III
      CR(Coil) Sale (sales-side, mill 3), GPC3, CRSALE, Saleable Semis,
      Thick Plate, Pig Iron
        — "SALEABLE STEEL" table (col Z = CUM, col AA = M.RATE)
      Finished Steel        derived: Saleable Steel − Saleable Semis
                             (already includes Thick Plate; Thick Plate is
                             also saved separately)
      GP/GC                  derived: G P SHEET + G C SHEET + G P COIL
                             (whichever of those three rows exist on this
                             report — see _DPR_GPGC_HELPERS)
      CR(1&2) Total Saleable  derived: CR I/II CR(Coil) Sale + CR Sheets
                             + GP/GC — mills 1&2's saleable output only.
                             Replaces the old "CRC&S(1&2)" mother-mill
                             (CRM(1+2)) row, which sometimes diverged from
                             this sum — see the derived-rules comment in
                             _dpr_config.
      CR III Total Saleable   derived: CR III CR(Coil) Sale + GPC3 — mill
                             3's saleable output only. Replaces the old
                             "CRC(3)" mother-mill (CRM-3) row, same
                             rationale as CR(1&2) Total Saleable above.
    """
    import sys
    sys.path.insert(0, os.path.dirname(__file__))
    import db

    ws = wb["DPR"]

    # Auto-detect month from O1 (Excel stores it as a Python datetime object)
    o1_raw = ws["O1"].value
    if isinstance(o1_raw, datetime):
        day   = o1_raw.day
        m_num = str(o1_raw.month).zfill(2)
        year  = str(o1_raw.year)
        db_report_month = f"{year}-{m_num}"
    elif o1_raw:
        # Fallback: try to parse date string formats DD.MM.YYYY or YYYY-MM-DD
        date_match = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", str(o1_raw))
        if date_match:
            d, m_num, year = date_match.groups()
            day = int(d)
            db_report_month = f"{year}-{m_num}"
        else:
            raise ValueError(
                f"Cannot parse date from cell O1: {repr(o1_raw)}. "
                "Expected a date value (DD.MM.YYYY or Excel date)."
            )
    else:
        raise ValueError("Cell O1 is empty — cannot determine report month.")

    # CUM (projected to the full month mid-month) is always what gets saved;
    # a mid-month upload is allowed, with each item's M.RATE cross-checked
    # against that projection and logged as a warning when it's erratic.
    last_day = calendar.monthrange(int(year), int(m_num))[1]
    days_in_month = last_day
    report_day = day
    if _dpr_is_mid_month(report_day, days_in_month):
        logger.warning(
            "BSL DPR: cell O1 date (%02d.%s.%s) is not the last day of %s %s "
            "(last day is %02d) — mid-month DPR upload. Every item is CUM x "
            "%d/%d; M.RATE is cross-checked and logged if erratic.",
            report_day, m_num, year, MONTH_NAMES[m_num], year, last_day,
            days_in_month, report_day,
        )

    logger.info(f"BSL DPR: month auto-detected from O1 → {db_report_month}")

    default_cells, NO_CONVERT, derived_rules = _dpr_config()
    production_cells = _resolve_dpr_cells(ws, default_cells)
    gpgc_helpers = _resolve_dpr_gpgc_helpers(ws)

    def _mrate_and_cum(addr):
        return (clean_val(ws[addr].value),
                clean_val(ws[_dpr_cum_cell(addr)].value))

    conn = db.connect()
    cursor = conn.cursor()
    vals_extracted = 0
    # Final (post-conversion) value already saved for each item this run —
    # lets a later derived rule (e.g. "sum_items") add up already-resolved
    # sibling items directly, instead of re-deriving from raw cells.
    resolved_vals = {}

    def _save(item_name, val, pre_converted=False):
        nonlocal vals_extracted
        if val is not None:
            vals_extracted += 1
            if not pre_converted and item_name not in NO_CONVERT:
                val = round(val / 1000.0, 3)
        resolved_vals[item_name] = val
        cursor.execute("""
            INSERT INTO production_table (report_month, plant_name, item_name, month_actual)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(report_month, plant_name, item_name)
            DO UPDATE SET month_actual = excluded.month_actual
        """, (db_report_month, "BSL", item_name, val))

    for item_name, cell in production_cells.items():
        mrate, cum = _mrate_and_cum(cell)
        val, _basis, _alert = _dpr_value_and_alert(
            mrate, cum, report_day, days_in_month,
            per_day=item_name in NO_CONVERT, label=item_name,
        )
        _save(item_name, val)

    # Derived values, driven by config, referencing already-resolved items
    # (or, for GP/GC, the independently-resolved helper rows) so they share
    # the same label-based row lookup as everything else.
    for d in derived_rules:
        item = d["item"]
        if d["op"] == "subtract":
            a_cell, b_cell = production_cells.get(d["a_item"]), production_cells.get(d["b_item"])
            a_m, a_c = _mrate_and_cum(a_cell) if a_cell else (None, None)
            b_m, b_c = _mrate_and_cum(b_cell) if b_cell else (None, None)
            if a_m is not None and b_m is not None:
                mrate_val = a_m - b_m
                cum_val = (a_c - b_c) if (a_c is not None and b_c is not None) else None
            elif a_m is not None:
                mrate_val, cum_val = a_m, a_c
            else:
                continue
            val, _basis, _alert = _dpr_value_and_alert(mrate_val, cum_val, report_day,
                                                        days_in_month, label=item)
            _save(item, val)
        elif d["op"] in ("add_items", "add_labels"):
            addrs = [production_cells.get(i) for i in d["items"]] if d["op"] == "add_items" \
                else [gpgc_helpers.get(l) for l in d["labels"]]
            addrs = [a for a in addrs if a]
            m_parts = [v for v in (clean_val(ws[c].value) for c in addrs) if v is not None]
            c_parts = [v for v in (clean_val(ws[_dpr_cum_cell(c)].value) for c in addrs) if v is not None]
            if m_parts:
                val, _basis, _alert = _dpr_value_and_alert(
                    sum(m_parts), sum(c_parts) if c_parts else None,
                    report_day, days_in_month, label=item,
                )
                _save(item, val)
        elif d["op"] == "sum_items":
            # Adds already-resolved (post-conversion) sibling values rather
            # than raw cells — works for any mix of plain production_cells
            # items and earlier derived items (e.g. GP/GC) in one sum,
            # without needing its own cell address.
            parts = [resolved_vals.get(i) for i in d["items"]]
            parts = [p for p in parts if p is not None]
            if parts:
                _save(item, round(sum(parts), 3), pre_converted=True)

    if vals_extracted == 0:
        raise ValueError(
            "No numeric data found at the expected cell locations in sheet 'DPR'. "
            "Please verify this is the correct BSL DPR Mail file."
        )

    conn.commit()
    conn.close()

    db.log_extraction(
        plant="BSL",
        report_month=db_report_month,
        file_name=source_file_name,
        sheet_name="DPR",
        source_type="DPR Mail (Month-End)",
        items_extracted=vals_extracted,
    )
    logger.info(f"BSL DPR extraction done: {vals_extracted} values saved for {db_report_month}.")
    return True


# ---------------------------------------------------------------------------
# Extractor 2 — BSL Techno-Economic Parameters (TECHNO <MON><YYYY>.XLS)
# ---------------------------------------------------------------------------
#
# Sheet layout:
#   Sheet1 — COKE AND COAL CHEMICALS, SINTER PLANT + shared plant KPIs
#   Sheet2 — Coke Oven parameters
#   SMS-I  — Steel Melting Shop (BOF / Converter) parameters
#
# Column convention: mon_col = month actual; mon_col+1 = till the month (cumulative)
#
# Parameters: (sheet, row_1based, mon_col, multiplier,
#              group_code, section, parameter, unit)
# ---------------------------------------------------------------------------

try:
    import xlrd as _xlrd
    _XLRD_OK = True
except ImportError:
    _XLRD_OK = False


class _XlsCell:
    __slots__ = ("value",)
    def __init__(self, value):
        self.value = None if value == "" else value


class _XlsSheet:
    def __init__(self, sheet):
        self._s = sheet
        self.max_row    = sheet.nrows
        self.max_column = sheet.ncols

    def cell(self, row: int, col: int) -> "_XlsCell":
        r, c = row - 1, col - 1
        if r < 0 or r >= self._s.nrows or c < 0 or c >= self._s.ncols:
            return _XlsCell(None)
        v = self._s.cell_value(r, c)
        return _XlsCell(v)


class _XlsWb:
    def __init__(self, wb):
        self._wb = wb

    @property
    def sheetnames(self):
        return self._wb.sheet_names()

    def __getitem__(self, name: str) -> "_XlsSheet":
        return _XlsSheet(self._wb.sheet_by_name(name))

    def __contains__(self, name: str) -> bool:
        return name in self._wb.sheet_names()


def _open_wb(file_path: str):
    """Open .xls (xlrd) or .xlsx (openpyxl) and return a unified workbook."""
    if file_path.lower().endswith(".xls"):
        if not _XLRD_OK:
            raise ImportError(
                "xlrd is required for Excel 97-2003 (.xls) files. "
                "Install with: pip install xlrd"
            )
        return _XlsWb(_xlrd.open_workbook(file_path))
    return openpyxl.load_workbook(file_path, data_only=True)


def _cell_float(ws, row: int, col: int) -> Optional[float]:
    """Read a cell and convert to float; return None on blank/error."""
    v = ws.cell(row, col).value
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "-", "—", "nan", "###", "#DIV/0!", "#VALUE!", "#N/A"):
        return None
    try:
        return float(s.replace(",", ""))
    except (ValueError, TypeError):
        return None


# (sheet_name, row, mon_col, multiplier, group_code, section, parameter, unit)
# cum_col is always mon_col + 1 (next column, same row = "till the month")
_TECHNO_PARAM_MAP = [
    # ── Sheet1 ─────────────────────────────────────────────────────────────
    ("Sheet1", 10,  6, 1000.0, "COKE_SINTER", "Energy",                "Sp. Heat Cons.",                    "Kcal/TCO"),
    # Plant-wide "POWER CONSUMPN/T OF SAL.STEEL" (item 3 in Sheet1's own
    # numbered list) — NOT the same as row 24's CRM-3-specific figure a few
    # rows above it (see _CRM_ROWS' own comment on that distinction, from an
    # earlier direct user correction). Section text is "Sp Power
    # Consumption" (not "Specific...") so _to_snake() in
    # bsl_technopara_extractor.py's MAJOR-group branch yields
    # "sp_power_consumption" with no _MAJOR_KEY_NORM override needed.
    ("Sheet1", 26,  6,    1.0, "MAJOR",       "Sp Power Consumption", "BSL",                                "KWH"),
    ("Sheet1", 28,  6,    1.0, "COKE_SINTER", "Energy",                "Specific Energy Consumption",       "G.Cal/TCS"),
    ("Sheet1", 31,  6,    1.0, "COKE_SINTER", "Sinter Plant",          "Machine Productivity",              "T/m²/hr"),
    ("Sheet1", 33,  6,    1.0, "IRON_MAKING", "BF Productivity",       "BSL",                               "T/m³/day"),
    ("Sheet1", 35,  6,    1.0, "IRON_MAKING", "BF Coke Rate",          "BSL",                               "Kg/THM"),
    ("Sheet1", 37,  6,    1.0, "IRON_MAKING", "CDI",                   "BSL",                               "Kg/THM"),
    ("Sheet1", 37,  6,    1.0, "MAJOR",       "CDI Rate",              "BSL",                               "kg/Thm"),
    ("Sheet1", 39,  6,    1.0, "IRON_MAKING", "Fuel Rate",             "BSL",                               "Kg/THM"),
    ("Sheet1", 41,  6,    1.0, "MAJOR",       "Coal to Hot Metal",     "BSL",                               "Ratio"),
    # CRM 1&2 / CRM 3 "Yield of HR Coil", Refractory and Water Consumption
    # used to be hardcoded here at rows 49/51/55/57 — confirmed on Techno
    # March 2024.xls that the rows shift (a duplicated "GROSS METALLIC INPUT"
    # sub-row above pushes everything below it down by 2), which silently
    # mismapped SMS-II's Gross Metallic Input (~1100-1140 Kg/TCS) into
    # "Yield of HR Coil" (a %) for that month, and likewise shifted
    # Refractory/Water. Moved to _extract_sheet1_label_search_rows() below,
    # which searches for each label instead of trusting a fixed row.
    # ── Sheet2 (Coke Ovens), Sheet3 (Sinter Plant), Sheet4, SMS-I/SMS-II ────
    # All extracted by _extract_keyword_block()/_extract_sheet4_iron_making_
    # rows() below instead of hardcoded rows here. Confirmed on real files
    # (e.g. Techno April 2025 vs June 2025) that Sheet2's Ash/VM/Crushing
    # Index/CSR/CRI/M-10 block and Sheet3's Basicity/Fe%/Availability/
    # Utilisation block were off by a fixed number of rows against every
    # file we have — not a layout drift over time, just wrong row numbers —
    # so a bounded keyword search replaces them entirely rather than being
    # patched to a different fixed offset that could drift again.
]

# ---------------------------------------------------------------------------
# FAX GM OPRN sheet — monthly fax to Corp Office OPRN Directorate. Adds a
# handful of parameters not present on Sheet1-4/SMS-I/SMS-II: per-furnace BF
# dimensions (Useful/Working Volume, Not Dry Cast%) and per-furnace CDI, plus
# Tar Injection, LD Slag, CO Gas Yield, Mn in HM, Si/S in HM, Slag Basicity/
# MgO, and SMS Avg Blow/Day, Avg Heat Weight, Gross HM Consumption, Lime/
# Aluminium Consumption, Caster Yield.
#
# This sheet's remaining rows (Fuel Rate, CDI overall, Sinter in Burden, Coke
# Screen Loss, Dry Coal Charge, M-10, Ammonium Sulphate/Crude Tar/Benzol
# Yield, Coke/Flux Crushing, Sp. Heat Consumption, and every SMS-I/SMS-II
# consumption/lining-life/reblown-heat figure) are intentionally NOT mapped
# here even though they're present on the sheet — each already has a more
# direct source elsewhere in this same workbook (Sheet1's fixed rows, or
# Sheet2/Sheet3/Sheet4/SMS-I/SMS-II's keyword-searched blocks — see
# _SHEET2_KEYWORDS/_SHEET3_KEYWORDS/_SHEET4_KEYWORDS/_SMS_KEYWORDS below).
# Duplicating them here would risk two conflicting values for the same
# real-world figure depending on extraction order. Verified row-by-row
# against those other maps/keyword lists before adding anything new — don't
# add a FAX GM OPRN row without checking there first.
#
# Column layout differs from _TECHNO_PARAM_MAP — month/cum columns are not
# adjacent, so cum_col is given explicitly rather than assumed as mon_col+1:
#   rows 12-32   (particulars table 1): mon_col=7, cum_col=9
#   rows 34-42   (per-furnace BF table): mon/cum columns adjacent, per param
#   rows 48-93   (particulars table 2): mon_col=8, cum_col=10
#
# FCE 3 / BF-3 is intentionally excluded: its cells hold obvious placeholder
# values (e.g. repeated "7") rather than real readings, consistent with BF-3
# not being an operating furnace in BSL's current PDF-based BF reporting.
# ---------------------------------------------------------------------------
_FAX_SHEET = "FAX GM OPRN"

_FAX_SINGLE_PARAM_MAP = [
    # row, mon_col, cum_col, mult, group_code,   section,                 parameter,                unit
    (15,  7,  9, 1.0, "COKE_SINTER", "Coke Ovens",            "Coke Oven Gas Yield",    "%"),
    (22,  7,  9, 1.0, "IRON_MAKING", "Nut Coke",              "BSL",                    "Kg/THM"),
    (23,  7,  9, 1.0, "IRON_MAKING", "BF-1",                  "CDI",                    "Kg/THM"),
    (24,  7,  9, 1.0, "IRON_MAKING", "BF-2",                  "CDI",                    "Kg/THM"),
    (26,  7,  9, 1.0, "IRON_MAKING", "BF-4",                  "CDI",                    "Kg/THM"),
    (27,  7,  9, 1.0, "IRON_MAKING", "BF-5",                  "CDI",                    "Kg/THM"),
    (29,  7,  9, 1.0, "IRON_MAKING", "Tar Injection",         "BSL",                    "Kg/THM"),
    (32,  7,  9, 1.0, "SMS",         "LD Slag Cons",          "BSL",                    "Kg/TGS"),
    (53,  8, 10, 1.0, "IRON_MAKING", "Blast Furnaces",        "Manganese in HM",        "%"),
    # BF slag/HM quality params with no source anywhere else in this
    # extractor — confirmed against Sheet1's _TECHNO_PARAM_MAP and Sheet2/3/4's
    # keyword lists (_SHEET2_KEYWORDS/_SHEET3_KEYWORDS/_SHEET4_KEYWORDS) that
    # none of them cover these 4, unlike most of this table's other
    # candidates (Sinter in Burden, Coke Screen Loss, Fuel Rate, CDI overall,
    # and every SMS-I/SMS-II consumption figure are all already read
    # elsewhere from a more direct source — deliberately NOT duplicated here).
    (51,  8, 10, 1.0, "IRON_MAKING", "Blast Furnaces",        "Si in HM",               "%"),
    (52,  8, 10, 1.0, "IRON_MAKING", "Blast Furnaces",        "S in HM",                "%"),
    (55,  8, 10, 1.0, "IRON_MAKING", "Blast Furnaces",        "Slag Basicity",          "%"),
    (56,  8, 10, 1.0, "IRON_MAKING", "Blast Furnaces",        "Slag MgO",               "%"),
    (57,  8, 10, 1.0, "SMS",         "SMS-I",                 "Average Blows Per Day",  "Nos."),
    (58,  8, 10, 1.0, "SMS",         "SMS-II",                "Average Blows Per Day",  "Nos."),
    (59,  8, 10, 1.0, "SMS",         "SMS-I",                 "Average Heat Weight",    "T"),
    (60,  8, 10, 1.0, "SMS",         "SMS-II",                "Average Heat Weight",    "T"),
    (63,  8, 10, 1.0, "SMS",         "SMS-I",                 "Gross HM Consumption",   "Kg/TCS"),
    (64,  8, 10, 1.0, "SMS",         "SMS-II",                "Gross HM Consumption",   "Kg/TCS"),
    (65,  8, 10, 1.0, "SMS",         "Gross HM Consumption",  "BSL",                    "Kg/TCS"),
    (83,  8, 10, 1.0, "SMS",         "SMS-I",                 "Lime Consumption",       "Kg/TCS"),
    (84,  8, 10, 1.0, "SMS",         "SMS-II",                "Lime Consumption",       "Kg/TCS"),
    (85,  8, 10, 1.0, "SMS",         "Lime Consumption",      "BSL",                    "Kg/TCS"),
    (86,  8, 10, 1.0, "SMS",         "SMS-I",                 "Aluminium Consumption",  "Kg/TCS"),
    (87,  8, 10, 1.0, "SMS",         "SMS-II",                "Aluminium Consumption",  "Kg/TCS"),
    (88,  8, 10, 1.0, "SMS",         "Aluminium Consumption", "BSL",                    "Kg/TCS"),
    (92,  8, 10, 1.0, "SMS",         "SMS-I",                 "Caster Yield",           "%"),
    (93,  8, 10, 1.0, "SMS",         "SMS-II",                "Caster Yield",           "%"),
]

# Per-furnace BF dimension/quality table (rows 37-42); month/cum columns are
# adjacent pairs here, one pair per parameter. FCE 3 (row 39) excluded — see
# note above.
#
# NOTE: columns 2-3 and 4-5 are NOT furnace volume dimensions despite the
# sheet's original "Useful Volume"/"Working Volume" column headers — their
# values (~1.6-2.3) are on the T/m³/day scale of BF Productivity, not the
# thousands-of-m³ scale of an actual furnace volume. They are BF Productivity
# computed on a useful/working-volume basis, so they're labelled and unit'd
# accordingly here (see bsl_technopara_extractor.py's _BF_KEY_NORM for how
# the working-volume figure folds onto the shared "bf_productivity" key).
_FAX_BF_FURNACE_ROWS = {37: "BF-1", 38: "BF-2", 40: "BF-4", 41: "BF-5", 42: "BF_Shop"}
_FAX_BF_FURNACE_PARAMS = [
    # mon_col, cum_col, parameter,                    unit
    (2,   3,   "BF Productivity (Useful Volume)",  "T/m³/day"),
    (4,   5,   "BF Productivity (Working Volume)", "T/m³/day"),
    (6,   7,   "Hot Blast Temp",                   "°C"),
    (8,   9,   "Coke Rate",                        "Kg/THM"),
    (10,  11,  "Not Dry Cast",                     "%"),
]

# ---------------------------------------------------------------------------
# Sheet4 iron-making block (Sinter/Pellet in Burden, O2 Enrichment, Coke
# Screening Loss, Slag Rate, BF Gas Yield, CV of BF Gas, Furnace Availability/
# Utilisation). Unlike _TECHNO_PARAM_MAP's fixed row numbers, these params are
# located by searching a bounded row window for each label's keyword in
# column B — confirmed via TECHNO APRIL2026.xls vs TECHNO JUNE 2026.xls that
# a new "PELLET IN BURDEN" line item appears from Jun'26 onward and shifts
# every row below it (O2 Enrichment onward) down by 2, so a fixed row number
# can't handle both the old and new report layouts at once. The labels in
# this block are all distinct (no repeats to disambiguate, unlike RSP's
# furnace blocks), so a plain keyword search is safe here.
# Column F = month actual, Column G = till-the-month cumulative.
# ---------------------------------------------------------------------------
_SHEET4_SHEET     = "Sheet4"
_SHEET4_LABEL_COL = 2   # column B
_SHEET4_MON_COL   = 6   # column F
_SHEET4_CUM_COL   = 7   # column G
_SHEET4_WINDOW    = (28, 48)  # covers both the old and new row positions

# (keyword to match in the column-B label, section, parameter, unit, optional)
# optional=True means the row is genuinely expected to be missing on some
# files (not a sign of a reshuffle) — logged at info level, not a warning.
_SHEET4_KEYWORDS = [
    ("sinter in burden",         "Sinter in Burden", "BSL",                  "%",       False),
    ("pellet in burden",         "Pellet in Burden",  "BSL",                  "%",       True),
    ("oxygen enrichment",        "O2 Enrichment",     "BSL",                  "%",       False),
    ("coke screening loss",      "Coke Screen Loss",  "BSL",                  "%",       False),
    ("slag rate",                "Blast Furnaces",    "Slag Rate",            "Kg/THM",  False),
    ("bf gas yield",             "Blast Furnaces",    "BF Gas Yield",         "Nm³/THM", False),
    ("cv of bf gas",             "Blast Furnaces",    "CV of BF Gas",         "Kcal/Nm³",False),
    ("availability against cal", "Blast Furnaces",    "Furnace Availability", "%",       False),
    ("utilin. against avl",      "Blast Furnaces",    "Furnace Utilization",  "%",       False),
]


def _extract_sheet4_iron_making_rows(wb, db_month: str) -> list:
    """Extract BSL's Sheet4 iron-making params by searching a bounded row
    window for each label's keyword, rather than trusting fixed row numbers —
    row positions shift whenever SAIL inserts a new line item (e.g. "PELLET
    IN BURDEN", added from Jun'26 onward, which pushed every row below it
    down by 2). "Pellet in Burden" itself is only present from Jun'26 onward
    — its absence on older files is expected, not a reshuffle."""
    if _SHEET4_SHEET not in wb.sheetnames:
        logger.info("BSL techno: sheet %r not found — skipping Sheet4 iron-making block", _SHEET4_SHEET)
        return []

    ws = wb[_SHEET4_SHEET]
    lo, hi = _SHEET4_WINDOW
    labels = {
        r: str(ws.cell(r, _SHEET4_LABEL_COL).value or "").strip().lower()
        for r in range(lo, hi + 1)
    }

    out = []
    for i, (keyword, section, param, unit, optional) in enumerate(_SHEET4_KEYWORDS):
        match_row = next((r for r in range(lo, hi + 1) if keyword in labels[r]), None)
        if match_row is None:
            if optional:
                logger.info(
                    "BSL techno: Sheet4 label containing %r not found in "
                    "rows %d-%d — expected on reports before Jun'26.",
                    keyword, lo, hi,
                )
            else:
                logger.warning(
                    "BSL techno: Sheet4 label containing %r not found in rows "
                    "%d-%d — verify _SHEET4_KEYWORDS in excel_extractor_bsl.py "
                    "still matches the report's wording.",
                    keyword, lo, hi,
                )
            continue

        actual_raw = _cell_float(ws, match_row, _SHEET4_MON_COL)
        cum_raw    = _cell_float(ws, match_row, _SHEET4_CUM_COL)
        actual = round(actual_raw, 4) if actual_raw is not None else None
        cum    = round(cum_raw,    4) if cum_raw    is not None else None

        mon_ltr = get_column_letter(_SHEET4_MON_COL)
        cum_ltr = get_column_letter(_SHEET4_CUM_COL)

        out.append({
            "group_code": "IRON_MAKING",
            "section":    section,
            "parameter":  param,
            "unit":       unit,
            "actual":     actual,
            "cum_actual": cum,
            "sort_order": 370 + i * 10,
            "cell":       f"{mon_ltr}{match_row}/{cum_ltr}{match_row}",
            "file_label": labels[match_row],
            "plant":      "BSL",
            "month":      db_month,
            "found_via":  f"Sheet4 label-search (matched {keyword!r} at row {match_row})",
            "status":     "ok" if (actual is not None or cum is not None) else "skip",
        })
    return out


# ---------------------------------------------------------------------------
# Generic bounded-window keyword-search extractor, shared by the Coke Ovens
# (Sheet2), Sinter Plant (Sheet3), and SMS-I/SMS-II blocks below. These used
# to be fixed row numbers in _TECHNO_PARAM_MAP, which turned out to already be
# wrong on every file we have (Sheet2's Ash/VM/Crushing Index/Ash-in-BF-Coke/
# Fixed Carbon/CSR/CRI/M-10 block was off by 1 row; Sheet3's Basicity/Fe%/
# Availability/Utilisation block was off by 2) — the same class of bug fixed
# for Sheet4's iron-making block (_extract_sheet4_iron_making_rows).
# ---------------------------------------------------------------------------
def _extract_keyword_block(wb, sheet_name: str, window: tuple, keywords: list,
                            group_code: str, section: str, db_month: str,
                            label_col: int = 2, mon_col: int = 6, cum_col: int = 7,
                            sort_base: int = 0) -> list:
    """Locate each keyword's row within a bounded row window by searching
    column label_col, instead of trusting a fixed row number.

    keywords: list of (keyword, parameter_label, unit, optional). Matched in
    list order; each row can satisfy only one keyword, so when one label's
    text is a substring of another's (e.g. "Iron Ore" inside "Iron Ore
    Pellet"), list the longer/more specific one first — once a row is
    claimed it's excluded from later keyword searches.
    """
    if sheet_name not in wb.sheetnames:
        logger.info("BSL techno: sheet %r not found — skipping %s block", sheet_name, section)
        return []

    ws = wb[sheet_name]
    lo, hi = window
    # Collapse internal whitespace too — source labels are inconsistently
    # spaced (e.g. "CONV.UTILISATION  ON AVL. Hrs." with a double space,
    # "AVERAGE LINNING LIFE      " with trailing spaces).
    labels = {
        r: re.sub(r'\s+', ' ', str(ws.cell(r, label_col).value or "")).strip().lower()
        for r in range(lo, hi + 1)
    }

    used_rows = set()
    out = []
    for i, (keyword, param, unit, optional) in enumerate(keywords):
        match_row = next(
            (r for r in range(lo, hi + 1) if r not in used_rows and keyword in labels[r]),
            None,
        )
        if match_row is None:
            if optional:
                logger.info(
                    "BSL techno: %s label containing %r not found in rows %d-%d.",
                    section, keyword, lo, hi,
                )
            else:
                logger.warning(
                    "BSL techno: %s label containing %r not found in rows %d-%d — "
                    "verify the keyword still matches the report's wording.",
                    section, keyword, lo, hi,
                )
            continue
        used_rows.add(match_row)

        actual_raw = _cell_float(ws, match_row, mon_col)
        cum_raw    = _cell_float(ws, match_row, cum_col)
        actual = round(actual_raw, 4) if actual_raw is not None else None
        cum    = round(cum_raw,    4) if cum_raw    is not None else None

        mon_ltr = get_column_letter(mon_col)
        cum_ltr = get_column_letter(cum_col)

        out.append({
            "group_code": group_code,
            "section":    section,
            "parameter":  param,
            "unit":       unit,
            "actual":     actual,
            "cum_actual": cum,
            "sort_order": sort_base + i * 10,
            "cell":       f"{mon_ltr}{match_row}/{cum_ltr}{match_row}",
            "file_label": labels[match_row],
            "plant":      "BSL",
            "month":      db_month,
            "found_via":  f"{sheet_name} label-search (matched {keyword!r} at row {match_row})",
            "status":     "ok" if (actual is not None or cum is not None) else "skip",
        })
    return out


# (keyword, parameter, unit, optional) — Sheet2, "Coke Ovens" block.
_SHEET2_WINDOW = (9, 34)
_SHEET2_KEYWORDS = [
    ("dry coal charge",             "Dry Coal Charge per Oven",  "T/oven",   False),
    ("avg. coking time",            "Average Coking Time",       "Hrs",      False),
    ("- gross coke",                "Gross Coke Yield",          "%",        False),
    ("- bf coke",                   "BF Coke Yield",             "%",        False),
    ("ammonium sulphate",           "Ammonium Sulphate",         "Kg/TCO",   False),
    ("crude tar",                   "Crude Tar",                 "Kg/TCO",   False),
    ("crude benzol",                "Crude Benzol",              "Kg/TCO",   False),
    ("coke oven gas",               "Coke Oven Gas",             "Nm³/TCO",  False),
    ("cv of coke oven gas",         "CV of Coke Oven Gas",       "Kcal/Nm³", False),
    ("ash in coal blend",           "Ash in Coal Blend",         "%",        False),
    ("v.m. in coal blend",          "VM in Coal Blend",          "%",        False),
    ("coal crushing index",         "Coal Crushing Index",       "%",        False),
    ("ash in bf coke",              "Ash in BF Coke",            "%",        False),
    ("fixed carbon in b.f. coke",   "Fixed Carbon in BF Coke",   "%",        False),
    ("c.s. r value",                "Coke CSR",                  "%",        False),
    ("c.r.i. value",                "Coke CRI",                  "%",        False),
    ("m - 10 value",                "Coke M-10",                 "%",        False),
]

# Sheet3, "Sinter Plant" block.
_SHEET3_WINDOW = (18, 48)
_SHEET3_KEYWORDS = [
    ("coke crushing index",   "Coke Crushing Index for Sinter", "%",     False),
    ("flux crushing index",   "Flux Crushing Index for Sinter", "%",     False),
    ("sinter return",         "Sinter Return",                  "%",     False),
    ("feo consistency",       "FeO in Sinter",                  "%",     False),
    ("basicity of sinter",    "Basicity of Sinter",             "ratio", False),
    ("fe content in sinter",  "Fe% in Sinter",                  "%",     False),
    ("m/c availability",      "Sinter M/c Availability",         "%",     False),
    ("m/c utilisation",       "Sinter M/c Utilization",          "%",     False),
]

# SMS-I / SMS-II — identical label wording on both sheets, only row positions
# differ, so one keyword list serves both (called once per sheet name).
_SMS_WINDOW = (10, 52)
_SMS_KEYWORDS = [
    ("tap to tap time on avl",   "Tap to Tap Time (Avail. Hrs)",       "Min",     False),
    ("tap to tap time on work",  "Tap to Tap Time (Working Hrs)",      "Min",     False),
    ("average linning life",     "Average Lining Life",                "Heats",   False),
    ("conv.availabilty on cal",  "Converter Availability (Cal. Hr)",   "%",       False),
    ("utilisation on avl",       "Converter Availability (Avail. Hr)", "%",       False),
    ("net hot metal",            "Sp. Hot Metal Cons.",                "Kg/TCS",  False),
    ("scrap + pig iron",         "Sp. Scrap Cons.",                    "Kg/TCS",  False),
    ("iron ore pellet",          "Sp. Pellet Cons.",                   "Kg/TCS",  False),
    ("iron ore",                 "Sp. Iron Ore Cons.",                 "Kg/TCS",  False),
    ("fe - si",                  "Fe-Si Cons.",                        "Kg/TCS",  False),
    ("fe - mn",                  "Fe-Mn Cons.",                        "Kg/TCS",  False),
    ("si - mn",                  "Si-Mn Cons.",                        "Kg/TCS",  False),
    ("refractory",               "Refractory Cons.",                   "Kg/TCS",  False),
    ("oxygen",                   "Oxygen Blow per T Crude",            "Nm³/T CS",False),
    ("heat cons",                "Heat Consumed",                      "Kcal/T CS",False),
    ("power cons",               "Power Consumed",                     "KWH/T CS",False),
    ("reblown heat",             "Reblown Heat",                       "%",       False),
    ("feo in slag",              "FeO in Slag",                        "%",       False),
]


# ---------------------------------------------------------------------------
# HSM (Hot Strip Mill) — Sheet8. Unlike _TECHNO_PARAM_MAP (which reads column
# A purely for a display label and trusts the hardcoded row number for
# meaning), Sheet8's parameter name lives in column B and is used here as the
# source of truth for which HSM parameter each row holds: matched by keyword
# rather than assumed from row position, so if a future month's file reshuffles
# these rows, extraction logs a warning instead of silently mismapping data
# into the wrong parameter (e.g. Utilisation's number landing in Availability).
# Column F = month actual, Column G = till-the-month cumulative.
# ---------------------------------------------------------------------------
_HSM_SHEET     = "Sheet8"
# Row 35 ("ROLLING HOURS/DAY") was previously registered here too, but it
# isn't one of the 4 tracked HSM parameters below at all, and it also matches
# the "rolling" keyword — so it silently raced with row 29 ("SLAB ROLLING
# RATE", the real Rolling Rate figure) for the same param, overwriting
# whichever one got processed last. Removed.
#
# Heat/Power Consumption used to be read from this sheet (rows 17/18) but
# per direct user correction those rows don't hold the right figures — HSM's
# Sp. Heat/Power Consumption are Sheet1 sub-items of BSL's overall Heat/Power
# Consumption tables ("SLAB AT R.H. FURNACE" / "HR COIL + THICK PLATE"), not
# anything on Sheet8. See _CRM_ROWS below (despite the name, it now also
# carries these 2 HSM rows — same fixed-row/label-verified Sheet1 pattern).
_HSM_ROWS      = [29, 37, 41, 43]
_HSM_LABEL_COL = 2   # column B
_HSM_MON_COL   = 6   # column F
_HSM_CUM_COL   = 7   # column G

# (keyword to match in the column-B label, canonical parameter, unit)
# Matched in order — first keyword found in the (lowercased) label wins.
_HSM_KEYWORDS = [
    ("yield",   "Yield",             "%"),
    ("availab", "Availability",      "%"),
    ("utili",   "Utilisation",       "%"),
    ("rolling", "Rolling Rate",      "t/hr"),
]


def _extract_hsm_rows(wb, db_month: str) -> list:
    """Extract BSL HSM techno params from Sheet8, self-checking each row's
    identity against its column-B label rather than trusting row position alone."""
    if _HSM_SHEET not in wb.sheetnames:
        logger.info("BSL techno: sheet %r not found — skipping HSM", _HSM_SHEET)
        return []

    ws = wb[_HSM_SHEET]
    out = []
    for row in _HSM_ROWS:
        label = str(ws.cell(row, _HSM_LABEL_COL).value or "").strip()
        match = next((m for m in _HSM_KEYWORDS if m[0] in label.lower()), None)
        if match is None:
            logger.warning(
                "BSL techno: HSM %s row %d column B label %r didn't match any "
                "known HSM parameter (yield/availability/utilisation/rolling/"
                "heat/power) — verify _HSM_ROWS in excel_extractor_bsl.py "
                "still points at the right rows.",
                _HSM_SHEET, row, label,
            )
            continue
        _, param, unit = match

        actual_raw = _cell_float(ws, row, _HSM_MON_COL)
        cum_raw    = _cell_float(ws, row, _HSM_CUM_COL)
        actual = round(actual_raw, 4) if actual_raw is not None else None
        cum    = round(cum_raw,    4) if cum_raw    is not None else None

        out.append({
            "group_code": "MILL_BSL",
            "section":    "HSM",
            "parameter":  param,
            "unit":       unit,
            "actual":     actual,
            "cum_actual": cum,
            "sort_order": 480 + row,
            "cell":       f"{_HSM_SHEET}!{get_column_letter(_HSM_MON_COL)}{row}/{get_column_letter(_HSM_CUM_COL)}{row}",
            "file_label": label,
            "plant":      "BSL",
            "month":      db_month,
            "found_via":  f"hardcoded {_HSM_SHEET} R{row}C{_HSM_MON_COL} (label-matched {label!r} -> {param})",
            "status":     "ok" if (actual is not None or cum is not None) else "skip",
        })
    return out


# ---------------------------------------------------------------------------
# Cold Rolling Mill complex — CRM 1&2 (Tandem Mill figures, Sheet9) and CRM 3
# (PLTCM/SPM figures, Sheet11; Specific Power Consumption, Sheet1) each report
# a handful of specific named figures directly under the CRM 1&2 / CRM 3 unit
# itself (not as separate per-sub-machine units/rows) — confirmed by direct
# user instruction. Parameter names are prefixed with the sub-machine they
# come from (TM-1, PLTCM, SPM, ...) since several sub-machines' figures now
# share one unit and need distinct keys.
#
# (sheet, row, mon_col, cum_col, unit_name, parameter, unit_str, label_keyword)
# label_keyword is checked against column B for the same self-verification
# HSM uses above — logs a warning (never crashes) if a future file reshuffles
# rows, rather than silently mismapping data into the wrong parameter.
# ---------------------------------------------------------------------------
_CRM_LABEL_COL = 2   # column B — Sheet9/Sheet11/Sheet1 all agree on this

_CRM_ROWS = [
    # (sheet, row, mon_col, cum_col, unit_name, parameter, unit_str, label_keyword, multiplier)
    # CRM 1&2 (Sheet9) — Tandem Mill-I / Tandem Mill-II Utilisation only.
    ("Sheet9", 41, 6, 7, "CRM 1&2", "TM-1 Utilisation", "%", "utilis", 1.0),
    ("Sheet9", 51, 6, 7, "CRM 1&2", "TM-2 Utilisation", "%", "utilis", 1.0),
    # CRM 3 (Sheet11) — PLTCM Yield/Availability/Utilisation, SPM's "Yield
    # of CR Coil from HR Coil" (the material-in/material-out yield, not its
    # "Yield from Annealed Coil" process yield)/Availability/Utilisation.
    ("Sheet11", 30, 5, 6, "CRM 3", "PLTCM Yield",             "%", "yield", 1.0),
    ("Sheet11", 36, 5, 6, "CRM 3", "PLTCM Availability",      "%", "availab", 1.0),
    ("Sheet11", 38, 5, 6, "CRM 3", "PLTCM Utilisation",       "%", "utilis", 1.0),
    ("Sheet11", 44, 5, 6, "CRM 3", "SPM Yield of CR Coil",    "%", "yield of cr coil", 1.0),
    ("Sheet11", 48, 5, 6, "CRM 3", "SPM Availability",        "%", "availab", 1.0),
    ("Sheet11", 50, 5, 6, "CRM 3", "SPM Utilisation",         "%", "utilis", 1.0),
    # CRM 3 — Specific Power Consumption. Row 26 ("POWER CONSUMPN/T OF
    # SAL.STEEL") was used previously, but per direct user correction that's
    # a plant-wide saleable-steel figure, not CRM 3's own — row 24 ("CR
    # SALEABLE (III)"), under the "POWER CONSUMPN. IN KWH/T OF" table, is the
    # actual CRM 3-specific figure.
    ("Sheet1", 24, 6, 7, "CRM 3", "Specific Power Consumption", "kWh/t", "cr saleable (iii)", 1.0),

    # HSM — Sp. Heat/Power Consumption. Previously read from Sheet8 (rows
    # 17/18), which per direct user correction don't hold these figures at
    # all — they're sub-items of Sheet1's plant-wide Heat/Power Consumption
    # tables: "SLAB AT R.H. FURNACE" (Heat) and "HR COIL + THICK PLATE"
    # (Power), both under the "...CONSUMPN. IN .../T OF" header blocks. The
    # sheet's own "G.CAL/T" header for the Heat table is per-mille of what
    # the kcal/t unit needs (per direct user correction) — e.g. a sheet value
    # of 0.412 is 412 kcal/t — so that row alone gets a x1000 multiplier.
    ("Sheet1", 13, 6, 7, "HSM", "Heat Consumption",  "kcal/t", "slab at r.h", 1000.0),
    ("Sheet1", 21, 6, 7, "HSM", "Power Consumption", "kWh/t",  "hr coil + thick plate", 1.0),
]


def _extract_crm_rows(wb, db_month: str) -> list:
    """Extract BSL's Cold Rolling Mill sub-machine techno params from
    Sheet9/Sheet10 (CRM 1&2) and Sheet11 (CRM 3), plus HSM's and CRM 3's
    Heat/Power Consumption figures from Sheet1 (see _CRM_ROWS — despite the
    function name, it's not CRM-only anymore), self-checking each row's
    identity against its column-B label rather than trusting row position
    alone (same approach as _extract_hsm_rows)."""
    out = []
    for i, (sheet, row, mon_col, cum_col, unit, param, unit_str, keyword, mult) in enumerate(_CRM_ROWS):
        if sheet not in wb.sheetnames:
            logger.info("BSL techno: sheet %r not found — skipping %s/%s", sheet, unit, param)
            continue
        ws = wb[sheet]
        label = str(ws.cell(row, _CRM_LABEL_COL).value or "").strip()
        if keyword not in label.lower():
            logger.warning(
                "BSL techno: CRM %s row %d column B label %r didn't contain "
                "expected keyword %r for %s/%s — verify _CRM_ROWS in "
                "excel_extractor_bsl.py still points at the right rows.",
                sheet, row, label, keyword, unit, param,
            )
            continue

        actual_raw = _cell_float(ws, row, mon_col)
        cum_raw    = _cell_float(ws, row, cum_col)
        actual = round(actual_raw * mult, 4) if actual_raw is not None else None
        cum    = round(cum_raw    * mult, 4) if cum_raw    is not None else None

        out.append({
            "group_code": "MILL_BSL",
            "section":    unit,
            "parameter":  param,
            "unit":       unit_str,
            "actual":     actual,
            "cum_actual": cum,
            "sort_order": 700 + i,
            "cell":       f"{sheet}!{get_column_letter(mon_col)}{row}/{get_column_letter(cum_col)}{row}",
            "file_label": label,
            "plant":      "BSL",
            "month":      db_month,
            "found_via":  f"hardcoded {sheet} R{row}C{mon_col} (label-matched {label!r} -> {unit}/{param})",
            "status":     "ok" if (actual is not None or cum is not None) else "skip",
        })
    return out


# ---------------------------------------------------------------------------
# Sheet1 rows whose position drifts between file versions (confirmed on
# Techno March 2024.xls, which has an extra "GROSS METALLIC INPUT" sub-row
# not present in every other file we have, pushing everything below it down
# by 2 rows) — found by searching column B for the label within a window
# instead of trusting a fixed row, so a shifted file skips the row (logged)
# instead of silently mismapping a wrong figure into it.
# (label_keyword, group_code, section, parameter, unit_str, search_start, search_end)
# ---------------------------------------------------------------------------
_SHEET1_LABEL_COL = 2   # column B
_SHEET1_MON_COL   = 6   # column F
_SHEET1_CUM_COL   = 7   # column G

_SHEET1_LABEL_SEARCH_ROWS = [
    ("yield of cr saleable (1", "MILL_BSL",    "CRM 1&2",   "Yield of HR Coil",  "%",       44, 56),
    ("yield of cr saleable (3", "MILL_BSL",    "CRM 3",     "Yield of HR Coil",  "%",       44, 58),
    ("overall refractory",      "SMS",         "Refractory","BSL",              "Kg/TCS",  50, 62),
    ("water consn",             "COKE_SINTER", "Water",     "Water Consumption", "m³/T",    52, 64),
]


def _extract_sheet1_label_search_rows(wb, db_month: str) -> list:
    """Extract the Sheet1 rows in _SHEET1_LABEL_SEARCH_ROWS, self-checking
    each by searching for its label within a window rather than trusting a
    fixed row (same approach as _extract_hsm_rows/_extract_crm_rows)."""
    if "Sheet1" not in wb.sheetnames:
        return []
    ws = wb["Sheet1"]
    out = []
    for i, (keyword, group_code, section, parameter, unit_str, start, end) in enumerate(_SHEET1_LABEL_SEARCH_ROWS):
        found_row, label = None, ""
        for r in range(start, end + 1):
            text = str(ws.cell(r, _SHEET1_LABEL_COL).value or "").strip()
            if keyword in text.lower():
                found_row, label = r, text
                break
        if found_row is None:
            logger.warning(
                "BSL techno: Sheet1 label containing %r not found in rows %d-%d — "
                "verify _SHEET1_LABEL_SEARCH_ROWS in excel_extractor_bsl.py "
                "still matches the report's wording.",
                keyword, start, end,
            )
            continue

        actual_raw = _cell_float(ws, found_row, _SHEET1_MON_COL)
        cum_raw    = _cell_float(ws, found_row, _SHEET1_CUM_COL)
        actual = round(actual_raw, 4) if actual_raw is not None else None
        cum    = round(cum_raw,    4) if cum_raw    is not None else None

        out.append({
            "group_code": group_code,
            "section":    section,
            "parameter":  parameter,
            "unit":       unit_str,
            "actual":     actual,
            "cum_actual": cum,
            "sort_order": 600 + i,
            "cell":       f"Sheet1!{get_column_letter(_SHEET1_MON_COL)}{found_row}/"
                          f"{get_column_letter(_SHEET1_CUM_COL)}{found_row}",
            "file_label": label,
            "plant":      "BSL",
            "month":      db_month,
            "found_via":  f"label search rows {start}-{end} for {keyword!r} -> row {found_row}",
            "status":     "ok" if (actual is not None or cum is not None) else "skip",
        })
    return out


_MONTH_FROM_NAME = {
    "JANUARY": "01", "FEBRUARY": "02", "MARCH": "03", "APRIL": "04",
    "MAY": "05", "JUNE": "06", "JULY": "07", "AUGUST": "08",
    "SEPTEMBER": "09", "OCTOBER": "10", "NOVEMBER": "11", "DECEMBER": "12",
    "JAN": "01", "FEB": "02", "MAR": "03", "APR": "04",
    "JUN": "06", "JUL": "07", "AUG": "08", "SEP": "09",
    "OCT": "10", "NOV": "11", "DEC": "12",
}


def _detect_month_from_filename(file_path: str) -> Optional[str]:
    """Try to extract YYYY-MM from a filename like TECHNO APRIL2026.XLS."""
    import re
    name = os.path.basename(file_path).upper()
    m = re.search(
        r'(JANUARY|FEBRUARY|MARCH|APRIL|MAY|JUNE|JULY|AUGUST|SEPTEMBER|OCTOBER|NOVEMBER|DECEMBER'
        r'|JAN|FEB|MAR|APR|JUN|JUL|AUG|SEP|OCT|NOV|DEC)(\d{4})',
        name
    )
    if m:
        mon_str, yr = m.group(1), m.group(2)
        mon_num = _MONTH_FROM_NAME.get(mon_str)
        if mon_num:
            return f"{yr}-{mon_num}"
    return None


_MONTH_NAME_RE = re.compile(
    r'\b(JANUARY|FEBRUARY|MARCH|APRIL|MAY|JUNE|JULY|AUGUST|SEPTEMBER|OCTOBER|NOVEMBER|DECEMBER)'
    r'[\s\-–]*(\d{4})\b'
)
# BSL BF PDF header: "BF PERFORMANCE & ANALYSIS REPORT For DD/MM/YYYY"
_DATE_DDMMYYYY_RE = re.compile(r'\b\d{2}/(\d{2})/(\d{4})\b')


def _detect_month_from_pdf_text(text: str) -> Optional[str]:
    """Detect YYYY-MM from BSL BF PDF text.

    Primary: DD/MM/YYYY date in the report header line (e.g. 'For 30/04/2026').
    Fallback: month-name + year pattern for any future format variation.
    Searches the first 500 chars first (header), then the full text.
    """
    for chunk in (text[:500], text):
        m = _DATE_DDMMYYYY_RE.search(chunk)
        if m:
            mon, yr = m.group(1), m.group(2)
            if "01" <= mon <= "12":
                return f"{yr}-{mon}"
    for chunk in (text[:500], text):
        m = _MONTH_NAME_RE.search(chunk.upper())
        if m:
            mon_num = _MONTH_FROM_NAME.get(m.group(1))
            if mon_num:
                return f"{m.group(2)}-{mon_num}"
    return None


def _detect_month_from_fax_sheet(fax_ws) -> Optional[str]:
    """Detect what month FAX GM OPRN's data actually holds.

    Primary signal: the actual date value in each particulars table's own
    data-column header (row 11 col G, row 48 col H — the literal headers of
    the columns _FAX_SINGLE_PARAM_MAP/_FAX_BF_FURNACE_PARAMS read "month"
    values from, per their mon_col=7/8). These are structurally guaranteed
    to describe the data since they ARE its column header, not free text
    that can drift out of sync with it.

    Fallback: scan for a month-name + year text pattern (row 9 reads "THE
    REQUIRED INFORMATION FOR <MONTH> <YEAR> IS AS UNDER"), used only if
    neither header cell holds a date. This sheet is a supplementary monthly
    fax to Corporate that isn't always updated when the workbook is reused
    as next month's template — confirmed on a real file where every other
    sheet (Sheet1-4, SMS-I, SMS-II) correctly read "JUNE-2026" but this
    sheet's row 9 title still read "FEBRUARY 2026" while its own data-column
    headers (row 11/48) correctly read May 2026 — i.e. exactly the case the
    header-date check above exists to get right instead of wrongly skipping
    a month's worth of real furnace-wise data (BF-1/2/4/5 Coke Rate, Hot
    Blast Temp, Productivity, CDI, etc. — this sheet's ONLY source for
    per-furnace breakdown) over a stale banner elsewhere on the sheet.
    Returns None if no signal is found at all (never block on an
    undetectable signal)."""
    for row, col in ((11, 7), (48, 8)):
        val = fax_ws.cell(row, col).value
        if isinstance(val, datetime):
            return f"{val.year}-{val.month:02d}"

    for row in range(1, 12):
        for col in range(1, 6):
            try:
                val = str(fax_ws.cell(row, col).value or "").upper()
            except Exception:
                continue
            m = _MONTH_NAME_RE.search(val)
            if m:
                mon_num = _MONTH_FROM_NAME.get(m.group(1))
                if mon_num:
                    return f"{m.group(2)}-{mon_num}"
    return None


def _detect_month_from_sheet1(wb) -> Optional[str]:
    """Scan the first 5 rows of Sheet1 for a month-name + year pattern (Techno Excel header)."""
    try:
        ws = wb["Sheet1"]
    except (KeyError, Exception):
        return None
    for row in range(1, 6):
        for col in range(1, 10):
            try:
                val = str(ws.cell(row, col).value or "").upper()
            except Exception:
                continue
            m = _MONTH_NAME_RE.search(val)
            if m:
                mon_num = _MONTH_FROM_NAME.get(m.group(1))
                if mon_num:
                    return f"{m.group(2)}-{mon_num}"
    return None


def _fmt_month(ym: str) -> str:
    """Format 'YYYY-MM' as 'Month YYYY' for human-readable error messages."""
    try:
        y, mo = ym[:4], ym[5:7]
        return f"{MONTH_NAMES[mo]} {y}"
    except Exception:
        return ym


def _assert_month_match(detected: Optional[str], user_month: str, source: str) -> None:
    """Raise ValueError if detected month does not match user-selected month."""
    if detected and detected != user_month:
        raise ValueError(
            f"Month mismatch: the {source} contains data for "
            f"{_fmt_month(detected)}, but you selected {_fmt_month(user_month)}. "
            f"Please select '{_fmt_month(detected)}' in the month picker, "
            f"or upload the file for {_fmt_month(user_month)}."
        )


# ---------------------------------------------------------------------------
# Extractor 3 — BSL Corporate Office Special Steel Report
# ---------------------------------------------------------------------------
# Sheet: Sheet1
#   Row 1:    Title "SPECIAL STEEL REPORT FOR <MONTH> <YEAR>"
#   Row 2:    Date in cell I2 (col 9) — last day of report month
#   Rows 4-6: Two-row column headers
#   Rows 7+:  Data (product group in col A, grade in col B)
#
# Col  A (1)  Product group  (HR COIL / HR PLATE / HR SHEET / CR COIL/ / SLAB)
# Col  B (2)  Quality / Grade
# Col  D (4)  Actual Up To Last Month  (= last month's monthly actual; for growth ref)
# Col  F (6)  APP FOR THE MONTH  (monthly plan, not stored)
# Col  G (7)  ORDER AVAILABLE TOTAL  (current outstanding orders → order_qty)
# Col  I (9)  Despatch Till Date  (= this month's monthly despatch → actual_despatch)
#
# Note: Col I is the MONTHLY figure for the report month, NOT a running cumulative.
# "Actual Up To Last Month" (D) = last month's monthly actual (for growth comparison).
# ---------------------------------------------------------------------------

_CORP_SS_PRODUCT_MAP = {
    "HR COIL":  "HR COIL",
    "HR PLATE": "HR PLATE",
    "HR SHEET": "HR SHEET",
    "CR COIL/": "CR COIL/SHEET/GP GC",  # spans two col-A cells; "SHEET" continues
    "SLAB":     "SLAB",
}
_CORP_SS_STOP_A  = {"GRND/TOT"}      # grand-total sentinel — stop scan
_CORP_SS_SKIP_A  = {"TOTAL"}          # section-total row — skip, keep product
_CORP_SS_SKIP_B  = {"TOTAL", "GRAND TOTAL"}
_CORP_SS_CONT    = "SHEET"            # continuation of "CR COIL/SHEET/GP GC"


def _find_corp_ss_sheet(wb) -> Optional[str]:
    """Return the sheet name holding a BSL Corporate Office Special Steel
    report, or None. The Corp Office mails this as a single-sheet workbook
    whose tab name varies (e.g. "CORPORATE OFFICE", "Sheet1") — detection
    goes by the "SPECIAL STEEL" title in cell A1, not the tab name."""
    for name in wb.sheetnames:
        try:
            r1 = str(wb[name].cell(1, 1).value or "").upper()
        except Exception:
            continue
        if "SPECIAL STEEL" in r1:
            return name
    return None


def _is_corp_ss_file(wb) -> bool:
    """True if the workbook is a BSL Corporate Office Special Steel report."""
    return _find_corp_ss_sheet(wb) is not None


def _corp_ss_month(ws) -> Optional[str]:
    """Extract 'YYYY-MM' from cell I2 (row 2, col 9)."""
    raw = ws.cell(2, 9).value
    if raw is None:
        return None
    if hasattr(raw, "year") and hasattr(raw, "month"):
        return f"{raw.year}-{str(raw.month).zfill(2)}"
    s = str(raw).strip()
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        return f"{m.group(3)}-{m.group(2).zfill(2)}"
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    return None


def _extract_corp_ss_preview(wb, report_month: str) -> dict:
    """Parse BSL Corporate Office Special Steel Report (Sheet1).

    Col G (ORDER AVAILABLE TOTAL)  = outstanding order quantity → order_qty.
    Col I (Despatch Till Date)     = monthly actual for the report month → actual_despatch.
    Col D (Actual Up To Last Month) = previous month's actual (stored as cply_actual for reference).
    """
    sheet_name = _find_corp_ss_sheet(wb) or "Sheet1"
    ws = wb[sheet_name]
    detected = _corp_ss_month(ws)
    db_month = detected or report_month

    rows = []
    current_product = ""
    sort_order = 0

    for row_num in range(7, ws.max_row + 1):
        a_raw = ws.cell(row_num, 1).value
        b_raw = ws.cell(row_num, 2).value
        d_raw = ws.cell(row_num, 4).value   # last month's actual (growth reference)
        g_raw = ws.cell(row_num, 7).value   # order available total → order_qty
        i_raw = ws.cell(row_num, 9).value   # this month's despatch → actual_despatch

        a_str = str(a_raw).strip() if a_raw is not None else ""
        b_str = str(b_raw).strip() if b_raw is not None else ""

        if a_str in _CORP_SS_STOP_A:
            break

        if a_str:
            if a_str in _CORP_SS_SKIP_A:
                continue  # section TOTAL row — skip, keep current product
            elif a_str in _CORP_SS_PRODUCT_MAP:
                current_product = _CORP_SS_PRODUCT_MAP[a_str]
            elif a_str == _CORP_SS_CONT:
                pass  # "SHEET" continues "CR COIL/SHEET/GP GC"; process col B
            else:
                logger.debug("BSL CorpSS row %d: unknown col-A %r — skipped", row_num, a_str)
                continue

        if not b_str or b_str in _CORP_SS_SKIP_B:
            continue
        if not current_product:
            continue

        d_val = clean_val(d_raw)   # last month's actual (display reference)
        g_val = clean_val(g_raw)   # order available total
        i_val = clean_val(i_raw)   # this month's actual despatch

        sort_order += 1
        has_data = bool(
            (g_val is not None and g_val != 0) or
            (i_val is not None and i_val != 0)
        )

        rows.append({
            "product":         current_product,
            "quality_grade":   b_str,
            "section":         "",
            "sort_order":      sort_order,
            "order_qty":       g_val,    # order available total (col G)
            "actual_despatch": i_val,    # monthly actual despatch (col I)
            "cply_actual":     d_val,    # last month's actual (col D, display only)
            "unit":            "T",
            "cell":            f"R{row_num}:G(order) / I(actual)",
            "status":          "ok" if has_data else "zero",
        })

    ok = sum(1 for r in rows if r["status"] == "ok")
    logger.info("BSL CorpSS: %d ok / %d total rows for %s", ok, len(rows), db_month)

    if ok == 0 and not rows:
        raise ValueError(
            f"No data rows found in sheet {sheet_name!r}. "
            "Verify this is a BSL 'Corporate office Report <MON><YEAR>.xlsx' file."
        )

    note = (f"{ok} grade rows extracted for {db_month}. "
            "Order Qty = ORDER AVAILABLE TOTAL (col G); "
            "Actual = Despatch Till Date (col I, monthly).")

    return {
        "plant":              "BSL",
        "month":              db_month,
        "source_type":        "BSL Corporate Office Special Steel Report",
        "sheets":             sheet_name,
        "workbook_sheets":    list(wb.sheetnames),
        "report_type":        "CORP_SS",
        "detected_month":     detected,
        "production_rows":    [],
        "special_steel_rows": rows,
        "special_steel_note": note,
        "techno_rows":        [],
        "techno_param_rows":  [],
    }


def extract_preview_bf_pdf(file_path: str, report_month: str) -> dict:
    """
    Extract furnace-wise HM production from BSL BF Performance & Analysis
    Report PDF, for production_table (item_name = BF#1/BF#2/BF#3/BF#4/BF#5,
    whichever are actually producing that month — '000 t).

    Parsing is delegated to bsl_mer_parser (shared with the /data-entry/techno
    BF-techno upload, which saves the full 14-parameter set to techno_data —
    production_table's BF-wise weight source stays consistent between the two
    entry points). This path only surfaces production for review/insert here;
    the other techno parameters for this PDF are entered via /data-entry/techno.
    BF_Shop (plant aggregate) is skipped — production_table never carries a
    separate shop-level row (see techno_aggregates.py).
    """
    import pdfplumber
    import sys as _sys

    fname = os.path.basename(file_path)
    logger.info("BSL BF PDF (production, /upload): opening %s for month %s", fname, report_month)

    try:
        with pdfplumber.open(file_path) as pdf:
            full_text = "\n".join(pg.extract_text() or "" for pg in pdf.pages)
    except Exception as exc:
        raise ValueError(f"Cannot open PDF '{fname}': {exc}") from exc

    detected = _detect_month_from_pdf_text(full_text)
    _assert_month_match(detected, report_month, "BSL BF Performance PDF")

    _techno_project_dir = os.path.join(os.path.dirname(__file__), "..", "techno_project")
    if _techno_project_dir not in _sys.path:
        _sys.path.insert(0, _techno_project_dir)
    from bsl_mer_parser import extract_bsl_mer

    records = extract_bsl_mer(full_text, report_month=report_month, filename=fname)

    rows_out = []
    for rec in records:
        unit = rec['unit']
        if not unit.startswith('BF-'):
            continue
        prod = rec['techno_json'].get('month', {}).get('production')
        ok = isinstance(prod, (int, float)) and prod > 0
        rows_out.append({
            # production_table uses "BF#N" (see main.py's normalize_item_name
            # and PRODUCTION_ITEM_ORDER); `unit` stays "BF-N" for the
            # techno-convention labels below (file_label/cell/found_via),
            # matching techno_data's own unit naming.
            'item_name': unit.replace('BF-', 'BF#'),
            'value': round(prod / 1000.0, 3) if ok else None,
            'unit': "'000T",
            'cell': f'BF Performance PDF — {unit} monthly production',
            'file_label': f'{unit} Production',
            'plant': 'BSL', 'month': report_month,
            'found_via': f'BSL {unit} monthly production',
            'status': 'ok' if ok else 'skip',
        })

    ok = sum(1 for r in rows_out if r['status'] == 'ok')
    logger.info("BSL BF PDF (production): %d/%d furnaces ok for %s", ok, len(rows_out), report_month)

    if ok == 0:
        raise ValueError(
            "No furnace-wise HM production extracted. Verify this is a BSL BF "
            "Performance & Analysis Report PDF (BSL_BlastFurnace_DDMMYYYY.pdf)."
        )

    return {
        'source_type':        'BSL BF Performance & Analysis Report (PDF) — furnace-wise production',
        'month':              report_month,
        'detected_month':     detected,
        'plant':              'BSL',
        'workbook_sheets':    ['PDF'],
        'production_rows':    rows_out,
        'techno_rows':        [],
        'techno_param_rows':  [],
        'special_steel_rows': [],
    }


# ---------------------------------------------------------------------------
# Extractor 5 — BSL "Production of Main Products" PPC/Statistical report (PDF)
# ---------------------------------------------------------------------------
# The plant's month-end PDF bundle (e.g. "Rev <Mon><YY> (n).pdf") carries this
# as its page 2 ("PRODUCTION OF MAIN PRODUCTS") and page 3 ("BREAKUP OF PRIME
# & SECONDARY SALEABLE STEEL") — a different report to the BF Performance &
# Analysis PDF above (furnace-wise only) and the DPR Mail Excel (cumulative,
# month-end only). This one is plant-wide production for a single month,
# already finalised (not tentative), covering the same 19 production_table
# items as _dpr_config() (see that function's docstring for the canonical
# item-name set this app already uses for BSL).
#
# Only the MONTH ACTUAL column of each row is used — page 2 also carries
# month plan, cumulative plan/actual and YoY growth%, but no other monthly
# extractor in this app persists plan from a monthly actuals report (ABP
# plan comes from the separate annual Plan upload), so those columns are
# read for validation/preview only, never stored.

_MAIN_PRODUCTS_TITLE_RE = re.compile(r'PRODUCTION\s+OF\s+MAIN\s+PRODUCTS', re.IGNORECASE)
_BREAKUP_TITLE_RE = re.compile(r'BREAKUP\s+OF\s+PRIME', re.IGNORECASE)
_NUM_TOKEN_RE = re.compile(r'^-?[\d,]+\.?\d*$')

# Special Steel & Value Added Items despatch page, elsewhere in the same
# "Rev <Mon><YY> (n).pdf" bundle (page 9 of 23 in the file this was built
# against, titled "PRODUCTION/DESPATCH -- SPECIAL STEEL & VALUE ADDED
# ITEMS") — matched on the distinctive tail of the title only, since the
# leading "PRODUCTION/DESPATCH --" spacing/dash glyph isn't worth pinning
# down across file editions.
_SPECIAL_STEEL_DESPATCH_TITLE_RE = re.compile(
    r'SPECIAL\s+STEEL\s*&\s*VALUE\s+ADDED\s+ITEMS', re.IGNORECASE)


def _extract_special_steel_despatch_rows(pdf) -> tuple:
    """Parse the "PRODUCTION/DESPATCH -- SPECIAL STEEL & VALUE ADDED ITEMS"
    page (searched across the whole document — its position drifts between
    file editions) into special_steel_rows.

    Layout: product-group header rows (SLAB / HR COIL / HR PLATE / HR SHEET
    / CR PRODUCT — bare label, no data), then "- <grade>" rows each with 8
    numeric columns: MONTH (APP, ACTUAL, FULFIL%, PREV.YEAR), CUM (APP,
    ACTUAL, FULFIL%, PREV.YEAR) — a "SUB - TOTAL" row closes each group, and
    "TOTAL SPECIAL STEEL" closes the table.

    Many grade names embed digits/slashes of their own (e.g. "SAE 1006 Si
    Controlled Export", "E-34/E-38/SAPH/BSK 46/E-46", "ISC 370/390/410/440"),
    so splitting label-vs-data by "first numeric token" (as
    _parse_products_line does for the Main Products page) would misfire —
    this page's grade column and its 8 data columns instead sit in fixed,
    well-separated x-position bands, so grade text vs. data is split by each
    word's x0 instead. Only the MONTH-ACTUAL column (the 2nd of the 8 —
    "3rd column" counting the grade name itself as the 1st) is needed, found
    by nearest x-position to that column's own header word rather than by
    a fixed index, so a row missing a leading token can't silently shift
    every value one column over.
    """
    target_page = None
    for pg in pdf.pages:
        text = pg.extract_text() or ""
        if _SPECIAL_STEEL_DESPATCH_TITLE_RE.search(text):
            target_page = pg
            break
    if target_page is None:
        return [], None

    words = target_page.extract_words(use_text_flow=False, keep_blank_chars=False)

    header_actuals = sorted(
        (w for w in words if w["text"] == "ACTUAL" and w["top"] < 120),
        key=lambda w: w["x0"],
    )
    if not header_actuals:
        logger.warning(
            "BSL Special Steel Despatch page %s: 'ACTUAL' column header not "
            "found — skipping despatch extraction.", target_page.page_number)
        return [], target_page.page_number
    month_actual_center = (header_actuals[0]["x0"] + header_actuals[0]["x1"]) / 2
    # Observed column pitch is ~35-40pt on the sample file; half that (with
    # margin) keeps the match from crossing into APP or FULFIL-MENT(%).
    tolerance = 16.0
    label_x_max = 240.0  # every data column starts at x0 >= ~242; grade text never reaches past ~200

    row_groups = []
    cur_top, cur = None, []
    for w in sorted(words, key=lambda w: w["top"]):
        if cur_top is None or abs(w["top"] - cur_top) > 3:
            if cur:
                row_groups.append(sorted(cur, key=lambda ww: ww["x0"]))
            cur, cur_top = [w], w["top"]
        else:
            cur.append(w)
    if cur:
        row_groups.append(sorted(cur, key=lambda ww: ww["x0"]))

    rows_out = []
    current_product = ""
    sort_order = 0
    for row in row_groups:
        if not row or row[0]["top"] < 120:
            continue  # title/header band
        label = " ".join(w["text"] for w in row if w["x0"] < label_x_max).strip()
        data_words = [w for w in row if w["x0"] >= label_x_max]
        if not label:
            continue
        label_upper = label.upper()
        if label_upper.startswith("TOTAL SPECIAL STEEL"):
            break  # grand total — end of table
        if not data_words:
            current_product = label  # section header: SLAB / HR COIL / HR PLATE / ...
            continue
        if label_upper.startswith("SUB - TOTAL") or label_upper.startswith("SUB-TOTAL"):
            continue  # section subtotal — skip row, keep current_product
        if not current_product:
            continue

        grade = label.lstrip("-").strip()
        best = min(data_words, key=lambda w: abs((w["x0"] + w["x1"]) / 2 - month_actual_center))
        if abs((best["x0"] + best["x1"]) / 2 - month_actual_center) > tolerance:
            continue  # no word landed in the ACTUAL column on this row
        val = _num(best["text"])

        sort_order += 1
        rows_out.append({
            "product":         current_product,
            "quality_grade":   grade,
            "section":         "",
            "sort_order":      sort_order,
            "order_qty":       None,
            "actual_despatch": val,
            "unit":            "T",
            "cell":            f"PDF p{target_page.page_number} · MONTH ACTUAL col",
            # Matches the Corp Office extractor's convention (has_data check)
            # of treating a genuine zero the same as "no data" — an
            # all-zero grade row for the month isn't worth persisting.
            "status":          "ok" if val else "zero",
        })

    return rows_out, target_page.page_number


def _parse_products_line(line: str):
    """Split one text line of the products/breakup table into (label, [data
    tokens]). A single leading '-' is a sub-item bullet (e.g. '- HR COIL ...')
    and is stripped before the label starts; a '-' found after the label has
    begun is a genuine "no value" placeholder and marks where data starts."""
    tokens = line.split()
    if not tokens:
        return None, []
    if tokens[0] == '-':
        tokens = tokens[1:]
    i = 0
    while i < len(tokens):
        t = tokens[i]
        if t == '-' or _NUM_TOKEN_RE.match(t):
            break
        i += 1
    label = ' '.join(tokens[:i])
    return label, tokens[i:]


def _num(tok: Optional[str]) -> Optional[float]:
    if tok is None or tok == '-':
        return None
    try:
        return float(tok.replace(',', ''))
    except ValueError:
        return None


def _index_products_lines(text: str) -> dict:
    """{label: [data tokens]} for every parsable line, first occurrence wins
    (labels are unique within each of page 2 / page 3 — verified against a
    real report; see extract_preview_main_products_pdf's mapping comments)."""
    out = {}
    for line in text.split('\n'):
        label, data = _parse_products_line(line)
        if label and data:
            out.setdefault(label, data)
    return out


def extract_preview_main_products_pdf(file_path: str, report_month: str) -> dict:
    """Extract BSL's "Production of Main Products" PDF (PPC/Statistical,
    typically pages 2-3 of the month-end PDF bundle) into the same 19
    production_table items the DPR Mail Excel path uses.

    Mapping (PDF label → item_name), each cross-checked against another PDF
    total so the source row for every item is unambiguous:
      OVEN PUSHED                    → Oven Pushing (nos/day)  [no unit convert]
      GROSS SINTER                   → Total Sinter
      HOT METAL                      → Hot Metal
      PIG IRON                       → Pig Iron
      FS/CC SLAB (SMS NEW)           → SMS-1 CCM-1
      CC SLAB (SMS II)               → SMS-2 CCM-1&2
      CRUDE STEEL                    → Total Crude Steel        (= SMS-1 + SMS-2, verified)
      H R COIL (TOTAL)**             → HSM Total HR Coil
      SALEABLE STEEL section:
        HR COIL                      → HSM HR Coil (Sale)
        HR PLATE                     → HSM HR Plate             (matches page 3 HR PLATE)
        HR SHEET                     → HR Sheet                 (matches page 3 HR SHEET)
        CR COIL III                  → CR III CR(Coil) Sale
        CR COIL III + GP Coil III    → CR III Total Saleable    (mill 3 makes coil only,
                                                                   no separate CR Sheet row)
        CR COIL I & II + CR SHEET
          + GP/GC I & II             → CR(1&2) Total Saleable   (mills 1&2's own equivalent)
        CR COIL I & II               → CR I/II CR(Coil) Sale    (Coil-only, no CR Sheet —
                                                                   pages 17/18 product mix)
        CR SHEET                     → CR Sheets                (Sheet-only, same source
                                                                   row used in CR(1&2) Total
                                                                   Saleable above)
        GP Coil III                  → GPC3
        GP/GC I & II                 → GP/GC
      SAL.CR PRODUCTS                → CRSALE                   (= CR III Total Saleable+
                                                                   CR(1&2) Total Saleable, i.e.
                                                                   all mills combined — verified
                                                                   (50893+21319)+(40548+380)
                                                                   = 113140 on the report this
                                                                   was built against, using the
                                                                   old CRC(3)/CRC&S(1&2)+GP/GC
                                                                   split that CR III/CR(1&2) Total Saleable
                                                                   now bundles into one figure)
      SALEABLE STEEL                 → Saleable Steel
    Page 3 breakup table (month-total = 3rd number of the first PRIME/SEC/
    TOTAL group):
      FIN. STEEL + THICK PLATE       → Finished Steel  (unchanged — Thick
                                                          Plate stays folded
                                                          in here too)
      THICK PLATE                    → Thick Plate       (also saved on its
                                                            own — matches the
                                                            "Thick Plate"/Z20
                                                            item the DPR Mail
                                                            route saves)
      SLAB                           → Saleable Semis
    """
    import pdfplumber

    fname = os.path.basename(file_path)
    logger.info("BSL Main Products PDF: opening %s for month %s", fname, report_month)

    try:
        with pdfplumber.open(file_path) as pdf:
            pages_text = [pg.extract_text() or "" for pg in pdf.pages[:4]]
            ss_rows, ss_page_num = _extract_special_steel_despatch_rows(pdf)
    except Exception as exc:
        raise ValueError(f"Cannot open PDF '{fname}': {exc}") from exc

    ss_ok = sum(1 for r in ss_rows if r["status"] == "ok")
    if ss_page_num is None:
        logger.info("BSL Main Products PDF: no Special Steel & VAI despatch page found in %s", fname)
    else:
        logger.info("BSL Main Products PDF: %d/%d Special Steel despatch grade rows ok "
                    "(page %d) for %s", ss_ok, len(ss_rows), ss_page_num, fname)

    p2_text = next((t for t in pages_text if _MAIN_PRODUCTS_TITLE_RE.search(t)), None)
    if p2_text is None:
        raise ValueError(
            "This doesn't look like a BSL 'Production of Main Products' report — "
            "no page with that title was found in the first 4 pages."
        )
    p3_text = next((t for t in pages_text if _BREAKUP_TITLE_RE.search(t)), "")

    full_text = "\n".join(pages_text)
    detected = _detect_month_from_pdf_text(full_text)
    _assert_month_match(detected, report_month, "BSL Production of Main Products PDF")
    db_month = detected or report_month

    p2 = _index_products_lines(p2_text)
    p3 = _index_products_lines(p3_text)

    rows = []

    def add(item_name, value, cell):
        no_convert = item_name == "Oven Pushing (nos/day)"
        stored = value if (value is None or no_convert) else round(value / 1000.0, 3)
        rows.append({
            "item_name": item_name,
            "value": stored,
            "unit": "nos/d" if no_convert else "'000T",
            "cell": cell or f"{item_name} — not found in PDF",
            "status": "ok" if stored is not None else "no value",
        })

    def month_actual(index, label):
        """2nd token of a PRODUCTS row (APP, ACTUAL, %, ...) = the month's actual."""
        toks = index.get(label)
        if not toks or len(toks) < 2:
            return None, None
        return _num(toks[1]), f"p2 {label!r}"

    def p3_month_total(label):
        """Breakup rows are PRIME SEC TOTAL groups; month total = 3rd token."""
        toks = p3.get(label)
        if not toks or len(toks) < 3:
            return None, None
        return _num(toks[2]), f"p3 {label!r}"

    for pdf_label, item_name in (
        ("OVEN PUSHED", "Oven Pushing (nos/day)"),
        ("GROSS SINTER", "Total Sinter"),
        ("HOT METAL", "Hot Metal"),
        ("PIG IRON", "Pig Iron"),
        ("FS/CC SLAB (SMS NEW)", "SMS-1 CCM-1"),
        ("CC SLAB (SMS II)", "SMS-2 CCM-1&2"),
        ("CRUDE STEEL", "Total Crude Steel"),
        ("H R COIL (TOTAL)**", "HSM Total HR Coil"),
        ("HR COIL", "HSM HR Coil (Sale)"),
        ("HR PLATE", "HSM HR Plate"),
        ("HR SHEET", "HR Sheet"),
        ("GP Coil III", "GPC3"),
        ("GP/GC I & II", "GP/GC"),
        ("SAL.CR PRODUCTS", "CRSALE"),
        ("SALEABLE STEEL", "Saleable Steel"),
    ):
        v, c = month_actual(p2, pdf_label)
        add(item_name, v, c)

    # CR(1&2) Total Saleable = CR Coil I&II (Saleable) + CR Sheet (report
    # carries CR Sheet as one combined figure, not split by mill —
    # attributed entirely to mills 1&2 since mill 3's row is coil-only in
    # this report) + GP/GC (mills 1&2's own GP/GC product, already saved
    # above under its own item_name too). Replaces the old "CRC&S(1&2)"
    # item, which only summed CR Coil I&II + CR Sheet (missing GP/GC) and
    # was confirmed to disagree with the true sales-side total in some
    # months (see page_catwise_saleable.py's note, e.g. Sep'25: 35.7 vs the
    # true 47.4).
    coil12, c1 = month_actual(p2, "CR COIL I & II")
    sheet, c2 = month_actual(p2, "CR SHEET")
    gpgc, c3 = month_actual(p2, "GP/GC I & II")
    cr12_parts = [v for v in (coil12, sheet, gpgc) if v is not None]
    if cr12_parts:
        add("CR(1&2) Total Saleable", sum(cr12_parts), f"{c1} + {c2} + {c3}")
    else:
        add("CR(1&2) Total Saleable", None, None)

    # CR I/II CR(Coil) Sale / CR III CR(Coil) Sale — the Coil-Sale-only
    # figures (pages 17/18 product mix), as distinct from CR(1&2)/CR III
    # Total Saleable which bundle CR Sheet/GP-GC into each mill's figure.
    # "CR COIL III" is already Coil-only in this report (mill 3 makes coil
    # only).
    add("CR I/II CR(Coil) Sale", coil12, c1)
    add("CR Sheets", sheet, c2)
    crc3_val, crc3_cell = month_actual(p2, "CR COIL III")
    add("CR III CR(Coil) Sale", crc3_val, crc3_cell)

    # CR III Total Saleable = CR Coil III (Saleable) + GP Coil III — mill
    # 3's saleable output only (no separate CR Sheet row, mill 3 makes coil
    # only). Replaces the old "CRC(3)" item, which read this same "CR COIL
    # III" figure alone (missing GPC3) and, on the DPR route, sometimes read
    # a different mother-mill row entirely — see excel_extractor_bsl's
    # _dpr_config comment for the confirmed divergence.
    gpc3_val, gpc3_cell = month_actual(p2, "GP Coil III")
    cr3_parts = [v for v in (crc3_val, gpc3_val) if v is not None]
    if cr3_parts:
        add("CR III Total Saleable", sum(cr3_parts), f"{crc3_cell} + {gpc3_cell}")
    else:
        add("CR III Total Saleable", None, None)

    # Finished Steel / Saleable Semis aren't on page 2 — page 3's breakup
    # table carries them. Thick Plate stays folded into Finished Steel (as
    # before) AND is now also saved under its own item_name.
    fin, cf = p3_month_total("FIN. STEEL")
    thick, ct = p3_month_total("THICK PLATE")
    if fin is not None or thick is not None:
        add("Finished Steel", (fin or 0) + (thick or 0), f"{cf} + {ct}")
    else:
        add("Finished Steel", None, None)
    add("Thick Plate", thick, ct)

    slab, cs1 = p3_month_total("SLAB")
    add("Saleable Semis", slab, cs1)

    ok = sum(1 for r in rows if r["status"] == "ok")
    logger.info("BSL Main Products PDF: %d/%d items ok for %s", ok, len(rows), db_month)

    if ok == 0:
        raise ValueError(
            "No production values could be matched in this PDF. Verify this is "
            "the BSL 'Production of Main Products' PPC/Statistical report."
        )

    return {
        "source_type":        "BSL Production of Main Products (PDF)",
        "month":              db_month,
        "detected_month":     detected,
        "plant":              "BSL",
        "workbook_sheets":    ["PDF p2/p3"] + ([f"PDF p{ss_page_num} (Special Steel)"] if ss_page_num else []),
        "production_rows":    rows,
        "techno_rows":        [],
        "techno_param_rows":  [],
        "special_steel_rows": ss_rows,
    }


_SALEABLE_STEEL_FY_TITLE_RE = re.compile(r'Table No\.\s*2\.1', re.I)

_FY_MONTH_ROW_RE = re.compile(
    r"^(APR|MAY|JUN|JULY|JUL|AUG|SEP|OCT|NOV|DEC|JAN|FEB|MAR)('?(\d{2}))?\s+(.+)$"
)
_FY_MONTH_NUM = {
    "APR": 4, "MAY": 5, "JUN": 6, "JUL": 7, "JULY": 7, "AUG": 8, "SEP": 9,
    "OCT": 10, "NOV": 11, "DEC": 12, "JAN": 1, "FEB": 2, "MAR": 3,
}
_FY_TABLE_COLS = [
    "Slab", "ThickPlate", "HRCoil", "HRPlate", "HRSheet",
    "CRM_I_II", "CRM_III", "CRTotal", "CRSheet",
    "GPCRM_I_II", "GPCRM_III", "GPSheet", "GCSheet", "GPGCTotal",
    "CRSaleable", "TotalSaleableSteel",
]
_FY_MONTH_NAMES = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]

# Items this extractor writes. Neither legacy bundled item ("CRC&S(1&2)",
# "CRC(3)") is written — this extractor never wrote them in the first
# place; both are retired everywhere now (see the DPR/Main Products PDF
# extractors' own comments). "CR(1&2) Total Saleable" (CR I/II CR(Coil)
# Sale + CR Sheets + GP/GC) and "CR III Total Saleable" (CR III CR(Coil)
# Sale + GPC3) replace them, computed from components already resolved
# below.
_FY_ITEM_ORDER = [
    "Thick Plate", "HSM HR Coil (Sale)", "HSM HR Plate", "HR Sheet",
    "CR I/II CR(Coil) Sale", "CR Sheets", "CR III CR(Coil) Sale",
    "GP/GC", "GPC3", "CR(1&2) Total Saleable", "CR III Total Saleable",
    "CRSALE", "Saleable Semis", "Saleable Steel", "Finished Steel",
]


def _fy_start_from_month(report_month: str) -> int:
    y, m = int(report_month[:4]), int(report_month[5:7])
    return y if m >= 4 else y - 1


def _parse_saleable_steel_fy_table(full_text: str):
    """Parse the 'SALEABLE STEEL ... Table No. 2.1' PRODUCTION SUMMARY page
    into {month_idx (0=Apr..11=Mar): {col_name: value_in_tonnes}}, plus
    {month_idx: 2-digit-year} for whichever rows carried a "'YY" suffix
    (always APR and JAN in a normal report). Quarter-subtotal rows ("1ST
    QTR."), the annual recap row ("2025-26 ..."), and prior-year recap rows
    are all skipped automatically — their first token never matches a month
    name, so the row regex simply doesn't match those lines."""
    by_month_idx = {}
    seen_years = {}
    for line in full_text.splitlines():
        m = _FY_MONTH_ROW_RE.match(line.strip())
        if not m:
            continue
        mon_key, _, yr_suffix, rest = m.groups()
        mon_num = _FY_MONTH_NUM.get(mon_key)
        if mon_num is None:
            continue
        nums = re.findall(r'-?\d[\d,]*(?:\.\d+)?', rest)
        if len(nums) < len(_FY_TABLE_COLS):
            continue
        vals = [float(n.replace(',', '')) for n in nums[:len(_FY_TABLE_COLS)]]
        month_idx = (mon_num - 4) % 12   # Apr=0 .. Mar=11
        if month_idx in by_month_idx:
            continue   # first occurrence wins (defends against a recap table reusing month labels)
        by_month_idx[month_idx] = dict(zip(_FY_TABLE_COLS, vals))
        if yr_suffix:
            seen_years[month_idx] = int(yr_suffix)
    return by_month_idx, seen_years


def _derive_saleable_steel_fy_items(cols: dict, month_label: str = "") -> dict:
    """PDF column → production_table item_name, in Tonnes → '000T. Confirmed
    against a real BSL FY25-26 Table 2.1 page, cell-by-cell:
      Thick Plate            → Thick Plate
      HR Coil                → HSM HR Coil (Sale)
      HR Plate                → HSM HR Plate
      HR Sheet                → HR Sheet
      CR Coil, CRM-I/II       → CR I/II CR(Coil) Sale
      CR Sheet                → CR Sheets   (only — NOT "New CR Sheet", an
                                              unrelated item left untouched)
      CR Coil, CRM-III        → CR III CR(Coil) Sale
      GP Sheet + GC Sheet     → GP/GC        (the CRM-I/II GP product)
      GP Coil, CRM-III        → GPC3
      CR Saleable             → CRSALE       (both CRM mills' finished
                                              product total — cross-checked
                                              below against its own
                                              components; see the note there)
      Slab                    → Saleable Semis
      Total Saleable Steel    → Saleable Steel
      (computed)              → Finished Steel = Saleable Steel - Saleable Semis

    CRSALE cross-check: the PDF's own "CR Saleable" column always equals
    CR I/II CR(Coil) Sale + CR III CR(Coil) Sale + CR Sheets + New CR Sheet
    (0 — this table carries no separate "new" CR sheet split) + GP/GC + GPC3
    + the always-zero GP Coil CRM-I/II column — i.e. both CRM mills' saleable
    output summed across coil, sheet and GP/GC. Verified against a real
    FY25-26 Table 2.1 page for all 12 months. Logged (not raised) on
    mismatch — a future year's layout drifting shouldn't block extraction,
    but should be visible.
    """
    slab = round(cols["Slab"] / 1000.0, 3)
    total_saleable = round(cols["TotalSaleableSteel"] / 1000.0, 3)
    gp_gc = round((cols["GPSheet"] + cols["GCSheet"]) / 1000.0, 3)
    # CR(1&2) Total Saleable — mills 1&2's own saleable output only (coil +
    # sheet + GP/GC), summed once from raw tonnes rather than adding the
    # three already-rounded '000T figures, to avoid compounding rounding.
    cr12_total_saleable = round(
        (cols["CRM_I_II"] + cols["CRSheet"] + cols["GPSheet"] + cols["GCSheet"]) / 1000.0, 3)
    # CR III Total Saleable — mill 3's own saleable output only (coil +
    # GPC3, no separate CR Sheet row for mill 3).
    cr3_total_saleable = round((cols["CRM_III"] + cols["GPCRM_III"]) / 1000.0, 3)

    expected_crsale = (cols["CRM_I_II"] + cols["CRM_III"] + cols["CRSheet"]
                       + cols["GPCRM_I_II"] + cols["GPCRM_III"] + cols["GPSheet"] + cols["GCSheet"])
    if abs(expected_crsale - cols["CRSaleable"]) > 1.0:   # 1 tonne tolerance
        logger.warning(
            "BSL Saleable Steel FY PDF: CR Saleable cross-check mismatch for %s — "
            "PDF says %.3f T, components (CR I/II + CR III + CR Sheet + GP/GC + GPC3) sum to %.3f T",
            month_label, cols["CRSaleable"], expected_crsale,
        )

    return {
        "Thick Plate": round(cols["ThickPlate"] / 1000.0, 3),
        "HSM HR Coil (Sale)": round(cols["HRCoil"] / 1000.0, 3),
        "HSM HR Plate": round(cols["HRPlate"] / 1000.0, 3),
        "HR Sheet": round(cols["HRSheet"] / 1000.0, 3),
        "CR I/II CR(Coil) Sale": round(cols["CRM_I_II"] / 1000.0, 3),
        "CR Sheets": round(cols["CRSheet"] / 1000.0, 3),
        "CR III CR(Coil) Sale": round(cols["CRM_III"] / 1000.0, 3),
        "GP/GC": gp_gc,
        "GPC3": round(cols["GPCRM_III"] / 1000.0, 3),
        "CR(1&2) Total Saleable": cr12_total_saleable,
        "CR III Total Saleable": cr3_total_saleable,
        "CRSALE": round(cols["CRSaleable"] / 1000.0, 3),
        "Saleable Semis": slab,
        "Saleable Steel": total_saleable,
        "Finished Steel": round(total_saleable - slab, 3),
    }


def extract_preview_saleable_steel_fy_pdf(file_path: str, report_month: str, all_months: bool = False) -> dict:
    """Extract BSL's year-wise 'SALEABLE STEEL ... Table No. 2.1' PRODUCTION
    SUMMARY PDF (Corporate MIS page, one row per month across a full
    financial year). `report_month` anchors which FY to read — any month
    within that FY works, since the whole table is parsed regardless.
    `all_months=True` returns all 12 months' rows (each row carries its own
    `report_month`) for a one-shot FY backfill; otherwise only the single
    selected month's rows are returned, matching every other extractor's
    contract.
    """
    import pdfplumber

    fname = os.path.basename(file_path)
    try:
        with pdfplumber.open(file_path) as pdf:
            n_pages = len(pdf.pages)
            full_text = "\n".join((pg.extract_text() or "") for pg in pdf.pages)
    except Exception as exc:
        raise ValueError(f"Cannot open PDF '{fname}': {exc}") from exc

    by_month_idx, seen_years = _parse_saleable_steel_fy_table(full_text)
    if not by_month_idx:
        raise ValueError(
            "No monthly rows found. Verify this is a BSL 'SALEABLE STEEL ... "
            "Table No. 2.1' PRODUCTION SUMMARY page (12 month rows, Apr-Mar)."
        )

    fy_start = _fy_start_from_month(report_month)
    detected_fy_start = None
    if 0 in seen_years:
        detected_fy_start = 2000 + seen_years[0]
    elif 9 in seen_years:
        detected_fy_start = 2000 + seen_years[9] - 1
    if detected_fy_start is not None and detected_fy_start != fy_start:
        raise ValueError(
            f"This report's own APR/JAN row labels show FY {detected_fy_start}-"
            f"{str(detected_fy_start + 1)[2:]}, but you selected a month in FY "
            f"{fy_start}-{str(fy_start + 1)[2:]}. Pick a month within FY "
            f"{detected_fy_start}-{str(detected_fy_start + 1)[2:]}, or upload the right file."
        )

    def ym_for(idx):
        return f"{fy_start}-{idx + 4:02d}" if idx <= 8 else f"{fy_start + 1}-{idx - 8:02d}"

    months_to_emit = list(range(12)) if all_months else [(int(report_month[5:7]) - 4) % 12]
    cell_tag = f"PDF ({n_pages}p) · Table 2.1 FY{fy_start}-{str(fy_start + 1)[2:]}"

    rows = []
    for idx in months_to_emit:
        ym = ym_for(idx)
        mon_label = _FY_MONTH_NAMES[idx]
        cols = by_month_idx.get(idx)
        if cols is None:
            for item_name in _FY_ITEM_ORDER:
                rows.append({
                    "item_name": f"(not found) {item_name}", "value": None, "unit": "'000T",
                    "cell": f"{cell_tag} · {mon_label} row not found", "report_month": ym,
                    "pdf_label": f"({mon_label} row not found)", "status": "unmapped",
                })
            continue
        derived = _derive_saleable_steel_fy_items(cols, month_label=f"{mon_label} (FY{fy_start})")
        for item_name in _FY_ITEM_ORDER:
            rows.append({
                "item_name": item_name, "value": derived[item_name], "unit": "'000T",
                "cell": f"{cell_tag} · {mon_label}", "report_month": ym,
                "pdf_label": mon_label, "status": "ok",
            })

    ok = sum(1 for r in rows if r["status"] == "ok")
    logger.info("BSL Saleable Steel FY PDF: %d/%d rows ok (FY%s, all_months=%s)",
                ok, len(rows), fy_start, all_months)

    return {
        "plant": "BSL",
        "month": report_month,
        "detected_month": None,
        "source_type": "BSL Saleable Steel — Table 2.1 (Year-wise PRODUCTION SUMMARY)",
        "sheets": f"PDF ({n_pages}p) — Table 2.1",
        "workbook_sheets": [f"PDF ({n_pages} pages) — Table 2.1"],
        "report_type": "Table 2.1 FY Summary",
        "production_rows": rows,
        "special_steel_rows": [],
        "special_steel_note": "",
        "techno_rows": [],
        "techno_param_rows": [],
    }


# ---------------------------------------------------------------------------
# Annual Book — PRODUCTION SUMMARY (Table 2.1) + FURNACEWISE PRODUCTION
# (Table 7.3), one row per FY month like the Saleable Steel FY PDF above.
# ---------------------------------------------------------------------------

# Table 7.3's own number (unlike Table 2.1, not reused by any other BSL PDF
# this app already supports) — checked before the generic Table 2.1 probe so
# this file (whose page 1 is ALSO titled "... Table No. 2.1 (Contd.)", a
# different sub-table from the Saleable Steel one) isn't misrouted there.
_ANNUAL_BOOK_TITLE_RE = re.compile(r'Table No\.\s*7\.3', re.I)

# Same month-row shape as the Saleable Steel FY table, but this source
# sometimes trails the abbreviation with a period ("JUN.", "JUL.") and
# sometimes doesn't ("APR'25", "MAY") — \.? makes that optional.
_ANNUAL_BOOK_MONTH_ROW_RE = re.compile(
    r"^(APR|MAY|JUN|JULY|JUL|AUG|SEP|OCT|NOV|DEC|JAN|FEB|MAR)\.?('?(\d{2}))?\s+(.+)$"
)

# PRODUCTION SUMMARY (page 1) column order, left to right after PERIOD.
_ANNUAL_BOOK_SUMMARY_COLS = [
    "OvenPushingAvgDay", "OvenPushingTotal", "GrossCoke", "BFCoke", "Sinter",
    "HotMetal", "PigIron", "SteelToFoundry", "CastSlabSMSNew", "CastSlabSMSII",
    "CrudeSteel",
]
# col_name → (item_name, convert_to_000T). item_name=None means "not written"
# (OvenPushingTotal isn't one of the requested items; Avg./day is, under the
# same item_name/no-convert convention the DPR/Main-Products routes use).
# Cast Slab (SMS New)/(SMS II) reuse the exact item names the Main Products
# PDF route already writes for the same two figures (SMS-1 CCM-1 /
# SMS-2 CCM-1&2) — verified: Total Crude Steel = SMS-1 CCM-1 + SMS-2 CCM-1&2
# on every sampled month, same identity the Main Products route documents.
_ANNUAL_BOOK_SUMMARY_ITEMS = {
    "OvenPushingAvgDay": ("Oven Pushing (nos/day)", False),
    "OvenPushingTotal":  (None, False),
    "GrossCoke":         ("Gross Coke", True),
    "BFCoke":            ("BF Coke", True),
    "Sinter":            ("Total Sinter", True),
    "HotMetal":          ("Hot Metal", True),
    "PigIron":           ("Pig Iron", True),
    # "Steel to Foundry" on this PDF is the same figure page5_6.py /
    # page17_concast.py / page_prod_by_process.py already know as
    # "SMS-1 Ingot" (the legacy RSP/BSL ingot-route item, backfilled from
    # Report_format/RSPBSL_INGOT.xlsx — see scripts/backfill_rspbsl_ingot.py)
    # — confirmed identical on every one of the 12 months where both this
    # PDF and that backfill cover the same month. Written under the older,
    # more widely-used name so both sources land on one row instead of two.
    "SteelToFoundry":    ("SMS-1 Ingot", True),
    "CastSlabSMSNew":    ("SMS-1 CCM-1", True),
    "CastSlabSMSII":     ("SMS-2 CCM-1&2", True),
    "CrudeSteel":        ("Total Crude Steel", True),
}

# FURNACEWISE PRODUCTION: each furnace contributes 4 columns (Basic Gr., Fdy.
# Gr., Off Gr., Total) — only the Total column (furnace-wise Hot Metal) is
# wanted, at item_name "BF#<n>" (production_table's convention — see
# main.py's normalize_item_name/PRODUCTION_ITEM_ORDER and
# extract_preview_bf_pdf above, which now matches it too; plant_name stays
# 'BSL', furnace number becomes the item_name).
# Page 2 has furnaces 1-3, page 3 has furnaces 4-5 + a trailing Granulated
# Slag column (not extracted — no item_name convention for it yet).
_ANNUAL_BOOK_FURNACE_PAGE2 = ["BF#1", "BF#2", "BF#3"]
_ANNUAL_BOOK_FURNACE_PAGE3 = ["BF#4", "BF#5"]


def _parse_annual_book_month_rows(text: str, n_cols: int):
    """{month_idx (0=Apr..11=Mar): [n_cols floats]} from one page's text.
    Quarter-subtotal ("1ST QTR."), annual recap ("2025-26 ..."), and
    prior-year recap rows are skipped automatically — their first token
    never matches a month abbreviation, so the row regex just doesn't match."""
    by_month_idx = {}
    for line in text.splitlines():
        m = _ANNUAL_BOOK_MONTH_ROW_RE.match(line.strip())
        if not m:
            continue
        mon_key, _, _, rest = m.groups()
        mon_num = _FY_MONTH_NUM.get(mon_key)
        if mon_num is None:
            continue
        nums = re.findall(r'-?\d[\d,]*(?:\.\d+)?', rest)
        if len(nums) < n_cols:
            continue
        vals = [float(n.replace(',', '')) for n in nums[:n_cols]]
        month_idx = (mon_num - 4) % 12   # Apr=0 .. Mar=11
        if month_idx in by_month_idx:
            continue   # first occurrence wins (defends against a recap table reusing month labels)
        by_month_idx[month_idx] = vals
    return by_month_idx


def extract_preview_annual_book_pdf(file_path: str, report_month: str, all_months: bool = False) -> dict:
    """Extract BSL's Annual Book 3-page extract: PRODUCTION SUMMARY (Table
    2.1 — Oven Pushing, Gross/BF Coke, Sinter, Hot Metal, Pig Iron, Steel to
    Foundry, Cast Slab SMS New/II, Crude Steel) and FURNACEWISE PRODUCTION
    (Table 7.3, split across pages 2-3 — Hot Metal per furnace's own Total
    column). `report_month` anchors which FY to read — any month within
    that FY works, since the whole table is parsed regardless. `all_months=
    True` returns every month in the FY (each row carries its own
    `report_month`) for a one-shot FY backfill, same contract as
    extract_preview_saleable_steel_fy_pdf."""
    import pdfplumber

    fname = os.path.basename(file_path)
    try:
        with pdfplumber.open(file_path) as pdf:
            n_pages = len(pdf.pages)
            pages_text = [pg.extract_text() or "" for pg in pdf.pages]
    except Exception as exc:
        raise ValueError(f"Cannot open PDF '{fname}': {exc}") from exc

    summary_text = pages_text[0] if n_pages >= 1 else ""
    furnace_text_23 = "\n".join(pages_text[1:3]) if n_pages >= 2 else ""

    by_month_summary = _parse_annual_book_month_rows(summary_text, len(_ANNUAL_BOOK_SUMMARY_COLS))
    by_month_f23  = _parse_annual_book_month_rows(pages_text[1], 12) if n_pages >= 2 else {}
    by_month_f45  = _parse_annual_book_month_rows(pages_text[2], 9)  if n_pages >= 3 else {}

    if not by_month_summary:
        raise ValueError(
            "No monthly rows found on page 1. Verify this is the BSL Annual "
            "Book 3-page extract (PRODUCTION SUMMARY — MAIN PRODUCTS, Table "
            "No. 2.1, followed by FURNACEWISE PRODUCTION, Table No. 7.3)."
        )

    fy_start = _fy_start_from_month(report_month)

    def ym_for(idx):
        return f"{fy_start}-{idx + 4:02d}" if idx <= 8 else f"{fy_start + 1}-{idx - 8:02d}"

    months_to_emit = list(range(12)) if all_months else [(int(report_month[5:7]) - 4) % 12]
    cell_tag = f"PDF ({n_pages}p) · FY{fy_start}-{str(fy_start + 1)[2:]}"

    rows = []
    for idx in months_to_emit:
        ym = ym_for(idx)
        mon_label = _FY_MONTH_NAMES[idx]

        cols = by_month_summary.get(idx)
        if cols is None:
            for col_name in _ANNUAL_BOOK_SUMMARY_COLS:
                item_name, _ = _ANNUAL_BOOK_SUMMARY_ITEMS[col_name]
                if item_name is None:
                    continue
                rows.append({
                    "item_name": f"(not found) {item_name}", "value": None, "unit": "'000T",
                    "cell": f"{cell_tag} · {mon_label} row not found", "report_month": ym,
                    "pdf_label": f"({mon_label} row not found)", "status": "unmapped",
                })
        else:
            for col_name, val in zip(_ANNUAL_BOOK_SUMMARY_COLS, cols):
                item_name, convert = _ANNUAL_BOOK_SUMMARY_ITEMS[col_name]
                if item_name is None:
                    continue
                out_val = round(val / 1000.0, 3) if convert else val
                rows.append({
                    "item_name": item_name, "value": out_val,
                    "unit": "nos/day" if not convert else "'000T",
                    "cell": f"{cell_tag} · {mon_label}", "report_month": ym,
                    "pdf_label": col_name, "status": "ok",
                })

        for furnace_names, by_month_f, total_offset, n_cols in (
            (_ANNUAL_BOOK_FURNACE_PAGE2, by_month_f23, 3, 4),
            (_ANNUAL_BOOK_FURNACE_PAGE3, by_month_f45, 3, 4),
        ):
            fvals = by_month_f.get(idx)
            for i, item_name in enumerate(furnace_names):
                total = fvals[i * n_cols + total_offset] if fvals else None
                ok = isinstance(total, (int, float)) and total > 0
                rows.append({
                    "item_name": item_name if ok else f"(no value) {item_name}",
                    "value": round(total / 1000.0, 3) if ok else None,
                    "unit": "'000T",
                    "cell": f"{cell_tag} · {mon_label} · {item_name} Total col",
                    "report_month": ym, "pdf_label": f"{item_name} Total",
                    "status": "ok" if ok else "no value",
                })

    ok = sum(1 for r in rows if r["status"] == "ok")
    logger.info("BSL Annual Book PDF: %d/%d rows ok (FY%s, all_months=%s)",
                ok, len(rows), fy_start, all_months)

    return {
        "plant": "BSL",
        "month": report_month,
        "detected_month": None,
        "source_type": "BSL Annual Book (Production Summary + Furnacewise Production)",
        "sheets": f"PDF ({n_pages}p) — Table 2.1 + Table 7.3",
        "workbook_sheets": [f"PDF ({n_pages} pages) — Table 2.1 + Table 7.3"],
        "report_type": "Annual Book FY Summary",
        "production_rows": rows,
        "special_steel_rows": [],
        "special_steel_note": "",
        "techno_rows": [],
        "techno_param_rows": [],
    }


def _extract_delhi_report_stock(wb, db_month: str) -> list:
    """Extract next-month opening stock from 'Delhi Report' sheet.

    Column M = closing stock as on last day of db_month = opening of next month.
    Cell mapping (Tonnes → stored as '000T, 3 d.p.):
      M67 → SLABS INPROCESS
      M66 → FINISHED STEEL
      M69 → PIG IRON
    """
    dr_name = next((s for s in wb.sheetnames if s.strip().lower() == "delhi report"), None)
    if dr_name is None:
        return []

    ws = wb[dr_name]
    y, m = int(db_month[:4]), int(db_month[5:7])
    next_month = f"{y+1 if m == 12 else y}-{1 if m == 12 else m+1:02d}"

    def _t(addr):
        v = clean_val(ws[addr].value)
        return round(v / 1000, 3) if v is not None else None

    def _row(item_type, stock_type, value, formula):
        return {
            "plant": "BSL", "item_type": item_type, "stock_type": stock_type,
            "stock_month": next_month, "value": value, "formula": formula,
            "status": "ok" if value is not None else "skip",
        }

    return [
        _row("SLABS",         "INPROCESS", _t("M67"), "Delhi Report!M67"),
        _row("FINISHED STEEL", "",          _t("M66"), "Delhi Report!M66"),
        _row("PIG IRON",      "",          _t("M69"), "Delhi Report!M69"),
    ]


def _calculate_burden_percentages(rows_out: list, report_month: str):
    """
    Calculate Sinter % in Burden and Pellet % in Burden from consumption values.
    Delegates to shared utility function.
    """
    import sys
    sys.path.insert(0, os.path.dirname(__file__))
    from techno_calc_utils import calculate_burden_percentages as calc_burden
    calc_burden(rows_out, report_month, plant_name="BSL")


def _extract_dpr_preview(wb, report_month: str) -> dict:
    """Preview BSL DPR Mail report (sheet 'DPR') — no DB writes."""
    ws = wb["DPR"]

    # Auto-detect month + report day from O1 (Excel datetime or DD.MM.YYYY string)
    o1_raw = ws["O1"].value
    report_day, days_in_month, db_month = _dpr_report_day(o1_raw)
    db_month = db_month or report_month
    if not db_month:
        raise ValueError("Cannot determine report month: cell O1 is empty and no month supplied.")

    if db_month != report_month and report_month and len(report_month) >= 7:
        raise ValueError(
            f"Month mismatch: the DPR file's date cell (O1) shows "
            f"{_fmt_month(db_month)}, but you selected {_fmt_month(report_month)}. "
            f"Please select '{_fmt_month(db_month)}' in the month picker, "
            f"or upload the DPR file for {_fmt_month(report_month)}."
        )

    logger.info("BSL DPR preview: month from O1 → %s", db_month)

    mid_month = _dpr_is_mid_month(report_day, days_in_month)
    if mid_month:
        logger.warning(
            "BSL DPR preview: mid-month file (day %d of %d) — every item is "
            "CUM x %d/%d; M.RATE is cross-checked and flagged if erratic.",
            report_day, days_in_month, days_in_month, report_day,
        )

    def _mrate_and_cum(addr):
        return (clean_val(ws[addr].value),
                clean_val(ws[_dpr_cum_cell(addr)].value))

    default_cells, NO_CONVERT, derived_rules = _dpr_config()
    CELL_MAP = _resolve_dpr_cells(ws, default_cells)
    gpgc_helpers = _resolve_dpr_gpgc_helpers(ws)

    rows = []
    alerts = []
    for item_name, addr in CELL_MAP.items():
        per_day = item_name in NO_CONVERT
        mrate, cum = _mrate_and_cum(addr)
        raw, basis, alert = _dpr_value_and_alert(mrate, cum, report_day, days_in_month,
                                                  per_day=per_day, label=item_name)
        if alert:
            alerts.append(alert)
        cum_addr = _dpr_cum_cell(addr)
        cell = f"DPR!{cum_addr}" + (f" x{days_in_month}/{report_day}" if basis == "projected" else "")
        if alert:
            cell += f" (M.RATE {addr} erratic)"
        if raw is not None:
            if per_day:
                stored, unit = raw, "nos/d"
            else:
                stored, unit = round(raw / 1000.0, 3), "'000T"
            rows.append({"item_name": item_name, "value": stored, "unit": unit,
                         "cell": cell, "pdf_label": addr,
                         "basis": basis, "status": "ok"})
        else:
            unit = "nos/d" if per_day else "'000T"
            rows.append({"item_name": item_name, "value": None, "unit": unit,
                         "cell": cell, "pdf_label": addr,
                         "basis": basis, "status": "skip"})

    # Derived values, referencing already-resolved items (or, for GP/GC, the
    # independently-resolved helper rows — see _DPR_GPGC_HELPERS) so they
    # share the same label-based row lookup as everything else.
    for d in derived_rules:
        item = d["item"]
        if d["op"] == "subtract":
            a_cell, b_cell = CELL_MAP.get(d["a_item"]), CELL_MAP.get(d["b_item"])
            a_m, a_c = _mrate_and_cum(a_cell) if a_cell else (None, None)
            b_m, b_c = _mrate_and_cum(b_cell) if b_cell else (None, None)
            if a_m is not None and b_m is not None:
                mrate_val = a_m - b_m
                cum_val = (a_c - b_c) if (a_c is not None and b_c is not None) else None
                lbl = f"{a_cell}-{b_cell}"
            elif a_m is not None:
                mrate_val, cum_val, lbl = a_m, a_c, a_cell
            else:
                continue
            val, basis, alert = _dpr_value_and_alert(mrate_val, cum_val, report_day,
                                                      days_in_month, label=item)
            if alert:
                alerts.append(alert)
            if val is not None:
                rows.append({"item_name": item, "value": round(val / 1000.0, 3),
                             "unit": "'000T", "cell": f"DPR!{lbl} (computed)",
                             "pdf_label": lbl, "basis": basis, "status": "ok"})
        elif d["op"] in ("add_items", "add_labels"):
            addrs = [CELL_MAP.get(i) for i in d["items"]] if d["op"] == "add_items" \
                else [gpgc_helpers.get(l) for l in d["labels"]]
            addrs = [a for a in addrs if a]
            m_parts = [v for v in (clean_val(ws[c].value) for c in addrs) if v is not None]
            c_parts = [v for v in (clean_val(ws[_dpr_cum_cell(c)].value) for c in addrs) if v is not None]
            if m_parts:
                val, basis, alert = _dpr_value_and_alert(
                    sum(m_parts), sum(c_parts) if c_parts else None,
                    report_day, days_in_month, label=item,
                )
                if alert:
                    alerts.append(alert)
                if val is not None:
                    rows.append({"item_name": item, "value": round(val / 1000.0, 3),
                                 "unit": "'000T", "cell": f"DPR!{'+'.join(addrs)} (computed)",
                                 "pdf_label": "+".join(addrs), "basis": basis, "status": "ok"})

    ok = sum(1 for r in rows if r["status"] == "ok")
    logger.info("BSL DPR preview: %d/%d ok for %s", ok, len(rows), db_month)

    if ok == 0:
        raise ValueError(
            "No values found at expected cell locations in DPR sheet. "
            "Verify this is a BSL DPR Mail file (sheet 'DPR', date in O1)."
        )

    stock_rows = _extract_delhi_report_stock(wb, db_month)
    stock_ok = sum(1 for r in stock_rows if r["status"] == "ok")
    logger.info("BSL DPR stock (Delhi Report): %d/%d ok for next_month of %s",
                stock_ok, len(stock_rows), db_month)

    return {
        "source_type":        "BSL DPR Mail (Month-End)",
        "month":              db_month,
        "plant":              "BSL",
        "workbook_sheets":    list(wb.sheetnames),
        "production_rows":    rows,
        "techno_rows":        [],
        "techno_param_rows":  [],
        "special_steel_rows": [],
        "stock_rows":         stock_rows,
        "morning_report_day":    report_day,
        "morning_days_in_month": days_in_month,
        "morning_is_month_end":  (not report_day) or report_day >= (days_in_month or 0),
        "mrate_alerts":       alerts,
    }


def extract_preview(file_path: str, report_month: str, all_months: bool = False) -> dict:
    """
    Extract BSL data — auto-detects file type:

    • BF Performance & Analysis Report (.pdf)
      → production_rows populated with BF-wise HM production (BF-1/2/4/5).
        Full techno parameter set for this PDF is entered via /data-entry/techno.

    • Production of Main Products (.pdf) — plant month-end PDF bundle,
      typically pages 2-3 ("PRODUCTION OF MAIN PRODUCTS" / "BREAKUP OF
      PRIME & SECONDARY SALEABLE STEEL")
      → production_rows populated with the same 19 production items as the
        DPR Mail path, read straight from the finalised monthly report
        rather than a cumulative month-end DPR cell dump.
      → special_steel_rows also populated, from this same bundle's
        "PRODUCTION/DESPATCH -- SPECIAL STEEL & VALUE ADDED ITEMS" page
        (position varies by edition — found by title search, not a fixed
        page index); monthly actual = that page's MONTH-ACTUAL column.

    • Saleable Steel — Table No. 2.1 (.pdf) — Corporate MIS year-wise
      PRODUCTION SUMMARY page, one row per FY month (Apr-Mar) plus
      quarter/prior-year recap rows (skipped). `all_months=True` returns
      every month in the FY anchored by `report_month` instead of just the
      one selected — see extract_preview_saleable_steel_fy_pdf().

    • Annual Book 3-page extract (.pdf) — PRODUCTION SUMMARY (Table No. 2.1,
      MAIN PRODUCTS: Oven Pushing, Gross/BF Coke, Sinter, Hot Metal, Pig
      Iron, SMS-1 Ingot, Cast Slab SMS New/II, Crude Steel) +
      FURNACEWISE PRODUCTION (Table No. 7.3, pages 2-3: Hot Metal per
      furnace's own Total column, item_name BF#1/2/3/4/5). Same year-wide,
      `all_months`-capable shape as Saleable Steel above — see
      extract_preview_annual_book_pdf().

    • DPR Mail Month-End Report (.xlsx, sheet 'DPR')
      → production_rows populated with ~19 production items.

    • Corporate Office Special Steel Report (.xlsx, Sheet1 with "SPECIAL STEEL" title)
      → special_steel_rows populated; monthly actual = Despatch Till Date − Actual Up To Last Month.

    • Techno-Economic Parameters (TECHNO <MON><YYYY>.XLS / .XLSX)
      → techno_param_rows populated from Sheet1/Sheet2/SMS-I/SMS-II.

    Returns a preview dict compatible with /api/confirm-extraction.
    """
    if file_path.lower().endswith('.pdf'):
        import pdfplumber as _plumb_probe
        try:
            with _plumb_probe.open(file_path) as _pdf:
                _probe_text = "\n".join((pg.extract_text() or "") for pg in _pdf.pages[:4])
        except Exception as exc:
            raise ValueError(f"Cannot open PDF: {exc}") from exc
        if _ANNUAL_BOOK_TITLE_RE.search(_probe_text):
            # Checked before the Table 2.1 probe below — this file's own
            # page 1 is ALSO titled "... Table No. 2.1 (Contd.)" (a
            # different sub-table, MAIN PRODUCTS not SALEABLE STEEL), so it
            # would otherwise be misrouted to the Saleable Steel parser.
            return extract_preview_annual_book_pdf(file_path, report_month, all_months=all_months)
        if _SALEABLE_STEEL_FY_TITLE_RE.search(_probe_text):
            return extract_preview_saleable_steel_fy_pdf(file_path, report_month, all_months=all_months)
        if _MAIN_PRODUCTS_TITLE_RE.search(_probe_text):
            return extract_preview_main_products_pdf(file_path, report_month)
        return extract_preview_bf_pdf(file_path, report_month)

    wb = _open_wb(file_path)

    # ── Route 1: DPR Mail Month-End Report ───────────────────────────────
    if "DPR" in wb.sheetnames:
        return _extract_dpr_preview(wb, report_month)

    # ── Route 2: Corporate Office Special Steel Report ────────────────────
    if _is_corp_ss_file(wb):
        return _extract_corp_ss_preview(wb, report_month)

    # ── Determine report month ────────────────────────────────────────────
    detected_month = _detect_month_from_sheet1(wb) or _detect_month_from_filename(file_path)
    db_month = report_month
    if not db_month or len(db_month) < 7:
        db_month = detected_month or ""
    if not db_month:
        raise ValueError(
            "Cannot determine report month. Set the month in the selector or "
            "name the file as TECHNO <MONTH><YYYY>.XLS (e.g. TECHNO APRIL2026.XLS)."
        )
    _assert_month_match(detected_month, db_month, "BSL Techno Excel")

    # ── Check which sheets are present ───────────────────────────────────
    sheets_present = list(wb.sheetnames)

    rows_out = []
    for sort_idx, (sheet_name, row, mon_col, mult,
                   group, section, param, unit) in enumerate(_TECHNO_PARAM_MAP):

        cum_col = mon_col + 1  # till the month is always the next column

        try:
            ws = wb[sheet_name]
        except (KeyError, Exception):
            logger.warning("BSL techno: sheet %r not found — skipping %s", sheet_name, param)
            continue

        actual_raw  = _cell_float(ws, row, mon_col)
        cum_raw     = _cell_float(ws, row, cum_col)

        actual  = round(actual_raw * mult, 4) if actual_raw is not None else None
        cum     = round(cum_raw   * mult, 4) if cum_raw    is not None else None

        mon_ltr = get_column_letter(mon_col)
        cum_ltr = get_column_letter(cum_col)
        cell_ref = f"{mon_ltr}{row}/{cum_ltr}{row}"
        if mult != 1.0:
            cell_ref += f" ×{int(mult)}"

        # Row label from column A of the same sheet for display
        try:
            file_label = str(ws.cell(row, 1).value or "").strip()
        except Exception:
            file_label = ""

        status = "ok" if (actual is not None or cum is not None) else "skip"
        rows_out.append({
            "group_code": group,
            "section":    section,
            "parameter":  param,
            "unit":       unit,
            "actual":     actual,
            "cum_actual": cum,
            "sort_order": sort_idx * 10,
            "cell":       cell_ref,
            "file_label": file_label,
            "plant":      "BSL",
            "month":      db_month,
            "found_via":  f"hardcoded {sheet_name} R{row}C{mon_col}",
            "status":     status,
        })

    # ── FAX GM OPRN sheet — extra parameters not in Sheet1-4/SMS-I/SMS-II ──
    fax_month_stale = False
    if _FAX_SHEET in wb.sheetnames:
        fax_ws = wb[_FAX_SHEET]
        fax_month = _detect_month_from_fax_sheet(fax_ws)
        fax_month_stale = bool(fax_month and fax_month != db_month)
        if fax_month_stale:
            # Skip only this sheet's rows, not the whole extraction — the
            # other sheets (Sheet1-4/SMS-I/SMS-II) already passed their own
            # month check above and remain correct for db_month. Blocking
            # the entire upload over one stale supplementary sheet would
            # throw away otherwise-correct data.
            logger.warning(
                "BSL techno: %r sheet reports %s but upload is for %s — "
                "skipping its rows (incl. the ONLY source of per-furnace "
                "BF-1/2/4/5 Coke Rate/Hot Blast Temp/Productivity/CDI) "
                "rather than attaching stale data to the wrong month.",
                _FAX_SHEET, _fmt_month(fax_month), _fmt_month(db_month))

    if _FAX_SHEET in wb.sheetnames and not fax_month_stale:
        fax_ws = wb[_FAX_SHEET]
        fax_idx = len(rows_out)

        def _fax_row(row, mon_col, cum_col, mult, group, section, param, unit, idx):
            actual_raw = _cell_float(fax_ws, row, mon_col)
            cum_raw    = _cell_float(fax_ws, row, cum_col)
            actual = round(actual_raw * mult, 4) if actual_raw is not None else None
            cum    = round(cum_raw   * mult, 4) if cum_raw    is not None else None
            mon_ltr = get_column_letter(mon_col)
            cum_ltr = get_column_letter(cum_col)
            cell_ref = f"{mon_ltr}{row}/{cum_ltr}{row}"
            try:
                file_label = str(fax_ws.cell(row, 1).value or "").strip()
            except Exception:
                file_label = ""
            status = "ok" if (actual is not None or cum is not None) else "skip"
            return {
                "group_code": group,
                "section":    section,
                "parameter":  param,
                "unit":       unit,
                "actual":     actual,
                "cum_actual": cum,
                "sort_order": idx * 10,
                "cell":       cell_ref,
                "file_label": file_label,
                "plant":      "BSL",
                "month":      db_month,
                "found_via":  f"hardcoded {_FAX_SHEET} R{row}C{mon_col}",
                "status":     status,
            }

        for row, mon_col, cum_col, mult, group, section, param, unit in _FAX_SINGLE_PARAM_MAP:
            rows_out.append(_fax_row(row, mon_col, cum_col, mult, group, section, param, unit, fax_idx))
            fax_idx += 1

        for row, unit_name in _FAX_BF_FURNACE_ROWS.items():
            for mon_col, cum_col, param, unit in _FAX_BF_FURNACE_PARAMS:
                rows_out.append(_fax_row(row, mon_col, cum_col, 1.0, "IRON_MAKING", unit_name, param, unit, fax_idx))
                fax_idx += 1
    elif _FAX_SHEET not in wb.sheetnames:
        logger.info("BSL techno: sheet %r not found — skipping FAX GM OPRN extras", _FAX_SHEET)

    # ── Sheet2 — Coke Ovens (Ash/VM/Crushing Index/CSR/CRI/M-10, etc.) ──────
    rows_out.extend(_extract_keyword_block(
        wb, "Sheet2", _SHEET2_WINDOW, _SHEET2_KEYWORDS,
        "COKE_SINTER", "Coke Ovens", db_month, sort_base=400))

    # ── Sheet3 — Sinter Plant (Basicity/Fe%/Availability/Utilisation, etc.) ─
    rows_out.extend(_extract_keyword_block(
        wb, "Sheet3", _SHEET3_WINDOW, _SHEET3_KEYWORDS,
        "COKE_SINTER", "Sinter Plant", db_month, sort_base=600))

    # ── SMS-I / SMS-II ───────────────────────────────────────────────────────
    rows_out.extend(_extract_keyword_block(
        wb, "SMS-I", _SMS_WINDOW, _SMS_KEYWORDS,
        "SMS", "SMS-I", db_month, sort_base=700))
    rows_out.extend(_extract_keyword_block(
        wb, "SMS-II", _SMS_WINDOW, _SMS_KEYWORDS,
        "SMS", "SMS-II", db_month, sort_base=750))

    # ── Sheet4 — iron-making block (Sinter/Pellet, O2, Coke Screen Loss, BF) ─
    rows_out.extend(_extract_sheet4_iron_making_rows(wb, db_month))

    # ── Sheet8 — HSM (Hot Strip Mill) ──────────────────────────────────────
    rows_out.extend(_extract_hsm_rows(wb, db_month))

    # ── Sheet9/Sheet10/Sheet11 — Cold Rolling Mill sub-machines ─────────────
    rows_out.extend(_extract_crm_rows(wb, db_month))

    # ── Sheet1 rows whose position drifts between file versions ────────────
    rows_out.extend(_extract_sheet1_label_search_rows(wb, db_month))

    ok = sum(1 for r in rows_out if r["status"] == "ok")
    logger.info("BSL techno preview: %d/%d rows ok for %s", ok, len(rows_out), db_month)

    if ok == 0:
        raise ValueError(
            "No numeric values found at any of the expected cell locations. "
            "Verify this is a BSL TECHNO <MON><YYYY>.XLS file with Sheet1, Sheet2, and SMS-I present."
        )

    # Calculate Sinter % in Burden and Pellet % in Burden from raw material consumption
    _calculate_burden_percentages(rows_out, db_month)

    return {
        "source_type":       "BSL Techno-Economic Parameters",
        "month":             db_month,
        "detected_month":    detected_month,
        "plant":             "BSL",
        "workbook_sheets":   sheets_present,
        "production_rows":   [],
        "techno_rows":       [],
        "techno_param_rows": rows_out,
        "special_steel_rows": [],
    }


