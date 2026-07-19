"""RSP Excel extractor — production, techno-table, and mill/sinter techno params
in a single pass with preview support.

Supports these file types (a single workbook may contain either or both of the
first two — see _find_report_sheets):
  • page-9 production table   — sheet name canon-matches "page9"
  • page-1-8 techno table     — sheet name canon-matches "^page1[89]" (RSP
    renames this sheet almost every month: "page-1-8", "PAGE-1-8 & 11,12",
    "PAGE1-8 &11-12", "PAGE-1-9", ...); only feeds preview/display — its
    legacy DB target (techno_table) no longer exists, so it is not persisted
    by extract_and_save_excel. techno_data (read by /data-entry/techno) is
    populated by the separate TechnoExtractor upload flow.
  • Daily Morning Report      — sheet starting with 'RSP Morning Report Data for-'

Both page-9 and page-1-8 column detection are position-tolerant:
  • page-9's month columns are stable across years (COL_MAP_P9, fixed letters).
  • page-1-8 prepends one more legacy fiscal-year column every year, shifting
    the current month's column annually — its month/Cum columns are located
    dynamically per file via rsp_row_scan.find_month_cum_columns instead.
  • Parameter rows are located by scanning anchor text in the first LABEL_COLS
    columns; hardcoded row_hints serve as fallback only.
"""
import re
import sys
import calendar
import datetime
import openpyxl
import logging
import sqlite3
import os
from typing import Optional

from extraction_utils import calculate_tmi_consumption

_TP_DIR = os.path.join(os.path.dirname(__file__), "..", "techno_project")
if _TP_DIR not in sys.path:
    sys.path.insert(0, _TP_DIR)
from rsp_row_scan import find_month_cum_columns, P18_NAME_RE  # noqa: E402  (path set above)

logger = logging.getLogger("excel_extractor")

DB_PATH = os.path.join(os.path.dirname(__file__), "..", "mis_reports.db")

MONTH_NAMES = {
    "01": "January", "02": "February", "03": "March", "04": "April",
    "05": "May",     "06": "June",     "07": "July",  "08": "August",
    "09": "September","10": "October", "11": "November","12": "December",
}
MONTH_NUMS = {v: k for k, v in MONTH_NAMES.items()}


def clean_val(val) -> Optional[float]:
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in ("", "nan", "-", "###", "#div/0!", "#value!", "cr", "na"):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _norm(s) -> str:
    return re.sub(r'[^a-z0-9]+', ' ', str(s).lower()).strip()


def _canon_sheet(s: str) -> str:
    return re.sub(r'[^a-z0-9]', '', str(s).lower())


# The production sheet's name drifts over time the same way P18_NAME_RE
# already tolerates page-1-8 becoming page-1-9 - RSP renamed "Page-9" to
# "Page-10"/"Pg-10" partway through 2026 (confirmed: TECHNOPARA JUNE-2026's
# real sheet list is ['PAGE-1-9', 'Pg-10', 'PAGE-11', 'PAGE-12'], no "page9"
# match at all), which silently broke extraction since it was matched by an
# exact canon-name set. Matched by a loose name pattern instead, and
# confirmed by a "PRODUCTION" title in the sheet's own first few rows so an
# unrelated sheet that happens to canon-match the same pattern (e.g.
# "PAGE-11"/"PAGE-12", which hold Despatch/External-Receipt data in real
# files) is never mistaken for it.
_P9_NAME_RE = re.compile(r'^(page|pag|pg)\d{1,2}$')


def _looks_like_p9_sheet(ws) -> bool:
    # Title wording varies by file edition ("PRODUCTION PERFORMANCE" vs the
    # abbreviated "PRODN. PERFORMANCE" seen in some real files) - "prod"
    # matches both without also matching the Despatch/External-Receipt
    # titles on neighboring same-name-pattern sheets.
    for r in range(1, 4):
        for c in range(1, 5):
            v = ws.cell(row=r, column=c).value
            if v and "prod" in str(v).lower():
                return True
    return False


def _find_report_sheets(wb, sheet_names):
    p9 = next(
        (s for s in sheet_names
         if _P9_NAME_RE.match(_canon_sheet(s))
         and not P18_NAME_RE.match(_canon_sheet(s))
         and _looks_like_p9_sheet(wb[s])),
        None,
    )
    p18 = next((s for s in sheet_names if P18_NAME_RE.match(_canon_sheet(s))), None)
    return p9, p18


# ---------------------------------------------------------------------------
# Month/year validation — the page-9 sheet has no single "this is the report
# month" cell (it's a "growing FY table": all 12 months' columns exist at
# once), so the upload flow previously trusted the user's month-picker
# selection blindly, with no way to catch an accidental wrong-file upload.
# Detected instead from two independent signals already present in the
# sheet's own content.
# ---------------------------------------------------------------------------

_P9_MONTH_DETECT_ROW_LABEL = "total crude steel"


def _detect_p9_report_month(ws) -> Optional[str]:
    """Best-effort month detection ('01'..'12'): the report month is whichever
    month's column is the LAST one with a non-zero value in a reliably-always
    -populated row ("Total Crude Steel"). Returns None if undetectable (never
    guessed wrong in 33 real files spanning 2023-2026 during testing)."""
    target_row = None
    for r in range(1, ws.max_row + 1):
        if _P9_MONTH_DETECT_ROW_LABEL in _norm(ws.cell(r, 1).value or ""):
            target_row = r
            break
    if target_row is None:
        return None

    order = ["04", "05", "06", "07", "08", "09", "10", "11", "12", "01", "02", "03"]
    last_month = None
    for m in order:
        col = COL_MAP_P9.get(m)
        if not col:
            continue
        v = clean_val(ws.cell(target_row, openpyxl.utils.column_index_from_string(col)).value)
        if v:
            last_month = m
    return last_month


def _detect_p9_fy_start_year(ws) -> Optional[int]:
    """FY start year (e.g. 2026 for "2026-27") from the sheet's own title row,
    e.g. "PRODUCTION PERFORMANCE - 2026-27" / "PRODN. PERFORMANCE - 2026-27"."""
    for r in range(1, 4):
        v = str(ws.cell(r, 1).value or "")
        if "prod" not in v.lower():
            continue
        m = re.search(r'(\d{4})\s*-\s*\d{2,4}', v)
        if m:
            return int(m.group(1))
    return None


def _assert_p9_month_year_match(ws, report_month: str) -> None:
    """Raise ValueError if the uploaded file's own content disagrees with the
    user-selected month/year - never blocks upload just because a signal
    couldn't be detected at all, only when detection and selection actively
    disagree."""
    db_month, month_num = _parse_report_month(report_month)
    year = int(db_month[:4])

    detected_month = _detect_p9_report_month(ws)
    if detected_month and detected_month != month_num:
        raise ValueError(
            f"Month mismatch: the uploaded file's production data (sheet {ws.title!r}) "
            f"appears to be for {MONTH_NAMES.get(detected_month, detected_month)}, but you "
            f"selected {MONTH_NAMES.get(month_num, month_num)} {year}. Please select the "
            f"matching month, or upload the file for {MONTH_NAMES.get(month_num, month_num)}."
        )

    fy_start_year = _detect_p9_fy_start_year(ws)
    if fy_start_year:
        expected_fy_start = year if int(month_num) >= 4 else year - 1
        if fy_start_year != expected_fy_start:
            raise ValueError(
                f"Year mismatch: the uploaded file is for FY {fy_start_year}-"
                f"{str(fy_start_year + 1)[2:]}, but your selected "
                f"{MONTH_NAMES.get(month_num, month_num)} {year} implies FY "
                f"{expected_fy_start}-{str(expected_fy_start + 1)[2:]}. Please verify the "
                f"uploaded file matches the selected period."
            )


def _parse_report_month(report_month: str):
    """Returns (db_month 'YYYY-MM', month_num_str '01'..'12')."""
    if len(report_month) == 7 and report_month[4] == "-":
        return report_month, report_month[5:7]
    parts = report_month.split()
    m_name, y_str = parts[0], parts[1]
    m_num = MONTH_NUMS.get(m_name, "11")
    return f"{y_str}-{m_num}", m_num


# ---------------------------------------------------------------------------
# Production / techno-table cell maps
# ---------------------------------------------------------------------------

COL_MAP_P9 = {
    "04": "B", "05": "C", "06": "D", "07": "F", "08": "G", "09": "H",
    "10": "J", "11": "K", "12": "L", "01": "N", "02": "O", "03": "P",
}
# PAGE-1-8's month/Cum columns are NOT a fixed map (see find_month_cum_columns):
# the sheet prepends one more legacy fiscal-year column every year, shifting
# every month's column by one letter annually.

NO_CONVERT = {"Oven Pushing (nos/day)", "COB#6", "COB#1-5"}


# ---------------------------------------------------------------------------
# Production table config  (page-9 sheet)
# ---------------------------------------------------------------------------
# Schema: (item_name, [label_aliases], row_hint)
# row_hint is used only when label scanning finds nothing.
# Aliases below cover every RSP page-9 label edition confirmed across the
# real sample corpus (2023-24 through 2026-27 FY): the original aliases
# ("cob 6", "sp 1", "hot metal", "pig iron", ...) never actually matched ANY
# real file's labels — RSP's actual wording has always been "Battery-6
# (Nos./Day)", "Sinter - I", "Tota Hot Metal" (a standing typo), "Cold Pig",
# etc. Every item that lacked a working alias silently fell through to its
# (also stale) row_hint, which — since the row layout has also shifted
# between editions — frequently landed on a COMPLETELY DIFFERENT item's row,
# corrupting values across the board (confirmed on the Jun'26 file: e.g.
# "SMS-2 CCM-4" and "Total Crude Steel" both silently resolved to the same
# row and the same wrong-for-one-of-them value). New aliases below were
# verified against real files via cross-footing (e.g. Sinter Total = Sinter
# I + II + III; Hot Metal = BF-1 + BF-5; Total Crude Steel = SMS-I Total +
# SMS-II Total), not guessed.
P9_ITEMS = [
    ("COB#6",                ["cob no 6", "cob 6", "cob#6", "battery 6"], 6),
    ("COB#1-5",              ["cob no 1 5", "cob 1 5", "oven pushed nos day"], 7),
    ("Oven Pushing (nos/day)",  ["oven pushing", "equivalent oven pushed"], 8),
    ("SP-1",                 ["sp 1", "sinter plant 1", "sinter i"],       9),
    ("SP-2",                 ["sp 2", "sinter plant 2", "sinter ii"],     10),
    ("SP-3",                 ["sp 3", "sinter plant 3", "sinter iii"],    11),
    ("Total Sinter",         ["total sinter", "sinter total"],            12),
    ("BF#1",                 ["bf 1", "bf1"],                             13),
    ("BF#5",                 ["bf 5", "bf5"],                             14),
    # "Tota Hot Metal" is a standing typo in RSP's own template (missing "l") —
    # present in every recent file, not a one-off mistake to work around.
    ("Hot Metal",            ["hot metal", "tota hot metal"],             15),
    # "Cold Pig" is standard steel-industry terminology for Pig Iron that has
    # been cast and cooled (vs. Hot Metal going straight to steelmaking).
    ("Pig Iron",             ["pig iron", "cold pig"],                    16),
    ("SMS-1 CCM-1",          ["sms 1 ccm 1", "ccm i slabs"],              19),
    ("SMS-2 CCM-1&2",        ["sms 2 ccm 1 2", "sms 2 ccm 1", "ccm ii slabs"], 20),
    ("SMS-2 CCM-3",          ["sms 2 ccm 3", "sms 2 caster iii"],         21),
    ("SMS-2 CCM-4",          ["sms 2 ccm 4", "sms 2 caster 4", "sms 2 caster iv"], 22),
    ("Total Crude Steel",    ["total crude steel"],                        24),
    ("HSM-2 Total HR Coil",  ["hsm 2 total hr coil", "total hr coil", "hsm 2 total hr coils"], 26),
    ("HSM-2 HR Coil (Sale)", ["hr coil sale", "hsm 2 hr coil sale", "hsm 2 h r c s"], 27),
    # No confident match found in the current template for a distinct
    # "HSM-2 HR Plate" row — "SSL - HR Plates" (Silicon Steel Mill Line) is a
    # different product line, not HSM-2's own. Left unresolved (no row_hint)
    # rather than guessed, so it's correctly absent instead of silently wrong.
    ("HSM-2 HR Plate",       [],                                          None),
    # 2023-24 era files use "Plates (PM)" / "New Plate Mill" instead of
    # "OPM Plates" / "NPM Plates" — confirmed via plant_registry.py's RSP
    # mill units, which register both "PM" and "NPM" as distinct mills, and
    # "New Plate Mill" is literally the expansion of "NPM".
    ("OPM Plate",            ["opm plate", "opm plates", "plates pm"],    29),
    ("NPM Plate",            ["npm plate", "npm plates", "new plate mill"], 30),
    ("CRNO Coils",           ["crno coils", "crno"],                      31),
    ("ERW Pipes",            ["erw pipes", "erw pipe"],                   32),
    ("SW Pipes",             ["sw pipes", "sw pipe"],                     33),
    ("Saleable Steel",       ["saleable steel"],                           34),
    # "Finished Steel" has never had its own row in any real file — the
    # original code pointed both items at the same row_hint (34), meaning
    # RSP's report doesn't distinguish them. Resolved as an explicit mirror
    # of Saleable Steel in _build_p9_cells instead of guessing a row here.
    ("Finished Steel",       [],                                          None),
]

# ---------------------------------------------------------------------------
# Techno-table config  (page-1-8 sheet)
# ---------------------------------------------------------------------------
# Schema: (param_name, [label_aliases], row_hint)
# More-specific entries (CDI BF-1/BF-5) are listed BEFORE the generic "CDI"
# so that exact-first scanning assigns them before "CDI" can claim their rows.
P18_TECHNO_ITEMS = [
    # Battery 1-5 entries listed FIRST so label scan claims their rows before Battery-6
    ("COB#1-5 Hard Coke Yield%",  ["h coke yield 25mm dry coal", "h coke yield 25mm"],       9),
    ("COB#1-5 Dry Coal Charge",   ["dry coal charge t oven", "dry coal charge"],             11),
    # Battery 6
    ("COB-6 Dry Coal Charge per Oven",
                                  ["cob 6 dry coal charge", "dry coal charge per oven"],    17),
    ("COB#6 Coke yield%",         ["h coke yield 25mm dry coal", "cob no 6 coke yield"],    21),
    ("Coke oven tar yield",        ["coke oven tar yield", "tar yield"],                     27),
    ("Coke oven Ammonia Sulphate yield",
                                  ["coke oven ammonia sulphate", "ammonia sulphate yield"],  28),
    ("Coke Screen Loss",           ["coke screen loss"],                                      31),
    ("SP-1 Sinter Productivity",   ["sp 1 sinter productivity"],                             38),
    ("SP-2 Sinter Productivity",   ["sp 2 sinter productivity"],                             58),
    ("SP-3 Sinter Productivity",   ["sp 3 sinter productivity"],                             81),
    # Furnace-wise productivity BEFORE generic so exact-scan claims them first
    ("BF Productivity BF-1",       ["productivity bf 1"],                                     97),
    ("BF Productivity BF-4",       ["productivity bf 4"],                                     98),
    ("BF Productivity BF-5",       ["productivity bf 5"],                                     99),
    # "Productivity - (Shop)" normalises to "productivity shop"
    ("BF Productivity",            ["productivity shop", "bf productivity"],                100),
    # Furnace-wise coke rate entries BEFORE generic so exact-scan claims them first
    ("Coke Rate BF-1",             ["coke rate bf 1", "coke rate bf#1"],                      101),
    ("Coke Rate BF-4",             ["coke rate bf 4", "coke rate bf#4"],                      102),
    ("Coke Rate BF-5",             ["coke rate bf 5", "coke rate bf#5"],                      103),
    ("Coke Rate",                  ["coke rate"],                                            104),
    ("CDI BF-1",                   ["cdi bf 1", "pci bf 1"],                                105),
    ("CDI BF-4",                   ["cdi bf 4", "pci bf 4"],                                  106),
    ("CDI BF-5",                   ["cdi bf 5", "pci bf 5"],                                107),
    ("CDI",                        ["cdi", "pci"],                                           108),
    ("Coal to Hot metal ratio",    ["coal to hot metal ratio", "coal to hm ratio","Coal to H M Ratio **"],          113),
    ("Nut Coke BF-1",              ["nut coke bf 1", "nut coke bf#1", "n c bf 1"],            109),
    ("Nut Coke BF-4",              ["nut coke bf 4", "nut coke bf#4", "n c bf 4"],            110),
    ("Nut Coke BF-5",              ["nut coke bf 5", "nut coke bf#5", "n c bf 5"],            111),
    ("Nut Coke Rate",              ["nut coke rate", "nut coke"],                            112),
    ("Hot Blast Temp BF-1",        ["hot blast temp bf 1", "hot blast temperature bf 1",
                                    "hbt bf 1"],                                               138),
    ("Hot Blast Temp BF-4",        ["hot blast temp bf 4", "hot blast temperature bf 4",
                                    "hbt bf 4"],                                               139),
    ("Hot Blast Temp BF-5",        ["hot blast temp bf 5", "hot blast temperature bf 5",
                                    "hbt bf 5"],                                               140),
    ("Hot Blast Temp",             ["hot blast temp", "hot blast temperature", "hbt"],         141),
    ("Si% in HM BF-1",             ["si% in hm bf 1", "si in hm bf 1", "silicon in hm bf 1"],143),
    ("Si% in HM BF-4",             ["si% in hm bf 4", "si in hm bf 4", "silicon in hm bf 4"],144),
    ("Si% in HM BF-5",             ["si% in hm bf 5", "si in hm bf 5", "silicon in hm bf 5"],145),
    ("Si% in HM",                  ["si% in hm", "si in hm", "silicon% in hm",
                                    "silicon in hm"],                                          146),
    ("O2 Enrichment BF-1",         ["o2 enrichment bf 1", "oxygen enrichment bf 1"],           133),
    ("O2 Enrichment BF-4",         ["o2 enrichment bf 4", "oxygen enrichment bf 4"],           134),
    ("O2 Enrichment BF-5",         ["o2 enrichment bf 5", "oxygen enrichment bf 5"],           135),
    ("O2 Enrichment",              ["o2 enrichment", "oxygen enrichment", "o 2 enrichment"],   136),
    ("Sinter% in Burden",          ["sinter in burden", "sinter% in burden"],               124),
    ("Pellet% in Burden",          ["pellet in burden", "pellet% in burden"],               125),
    ("Fuel Rate",                  ["fuel rate"],                                            157),
    ("SMS-1 HM consumption per ton of crude steel",
                                  ["sms 1 hm consumption per ton", "sms 1 hm cons per ton"],164),
    ("SMS-1 Scrap consumption per ton of crude steel",
                                  ["sms 1 scrap consumption per ton", "sms 1 scrap cons"],  165),

    ("SMS-1 TMI consumption per ton of crude steel",
                                  ["sms 1 TMI consumption per ton", "sms 1 TMI cons"],  163),               
    ("SMS-1 Avg heat wt",          ["sms 1 avg heat wt"],                                   175),
    ("SMS-1 Avg Blows per day",    ["blow sms1", "blow sms 1", "sms 1 avg blows per day"],  184),
    ("SMS-2 HM consumption per ton of crude steel",
                                  ["sms 2 hm consumption per ton", "sms 2 hm cons per ton"],191),
    ("SMS-2 Scrap consumption per ton of crude steel",
                                  ["sms 2 scrap consumption per ton", "sms 2 scrap cons"],  192),
    ("SMS-2 TMI consumption per ton of crude steel",
                                  ["sms 2 tmi consumption per ton", "sms 2 tmi cons"],  190),
    
    
    ("SMS-2 Avg heat wt",          ["sms 2 avg heat wt"],                                   204),
    ("SMS-1 lining life",          ["sms 1 lining life", "avg lining life nos"],           177),
    ("SMS-2 lining life",          ["sms 2 lining life"],                                   206),
    ("SMS-2 Avg Blows per day",    ["blow sms2", "blow sms 2", "sms 2 avg blows per day"],  214),
    ("Oven heat Consumption per ton of Dry coke Input",
                                  ["oven heat consumption", "oven heat cons per ton"],       305),
    ("Energy consumption",         ["energy consumption"],                                   340),
]


TECHNO_UNIT_MAP = {
    "Coal to Hot metal ratio": "--",
    "Coke Rate": "Kg/THM",
    "Coke Rate BF-1": "Kg/THM", "Coke Rate BF-4": "Kg/THM", "Coke Rate BF-5": "Kg/THM",
    "Nut Coke Rate": "Kg/THM",
    "Nut Coke BF-1": "Kg/THM", "Nut Coke BF-4": "Kg/THM",
    "CDI": "Kg/THM",
    "CDI BF-1": "Kg/THM", "CDI BF-4": "Kg/THM", "CDI BF-5": "Kg/THM",
    "Fuel Rate": "Kg/THM",
    "BF Productivity": "T/m³/day",
    "BF Productivity BF-1": "T/m³/day", "BF Productivity BF-4": "T/m³/day", "BF Productivity BF-5": "T/m³/day",
    "Sinter% in Burden": "%", "Pellet% in Burden": "%",
    "Si% in HM": "%", "Si% in HM BF-1": "%", "Si% in HM BF-4": "%", "Si% in HM BF-5": "%",
    "O2 Enrichment": "%", "O2 Enrichment BF-1": "%", "O2 Enrichment BF-4": "%", "O2 Enrichment BF-5": "%",
    "Hot Blast Temp": "°C",
    "Hot Blast Temp BF-1": "°C", "Hot Blast Temp BF-4": "°C", "Hot Blast Temp BF-5": "°C",
    "Hot Blast Volume": "Nm³/min",
    "Hot Blast Volume BF-1": "Nm³/min", "Hot Blast Volume BF-4": "Nm³/min", "Hot Blast Volume BF-5": "Nm³/min",
    "Energy consumption": "Gcal/tcs",
    "SMS-1 HM consumption per ton of crude steel": "kg/tcs",
    "SMS-1 Scrap consumption per ton of crude steel": "kg/tcs",
    "SMS-1 TMI consumption per ton of crude steel": "kg/tcs",
    "SMS-2 HM consumption per ton of crude steel": "kg/tcs",
    "SMS-2 Scrap consumption per ton of crude steel": "kg/tcs",
    "SMS-2 TMI consumption per ton of crude steel": "kg/tcs",
    "SMS-1 Oxygen Consumption": "Nm³/T", "SMS-2 Oxygen Consumption": "Nm³/T",
    "SMS-1 Converter Yield": "%",        "SMS-2 Converter Yield": "%",
    "SMS-1 Caster Yield": "%",           "SMS-2 Caster Yield": "%",
    "SMS-1 Avg heat wt": "T/Blow",        "SMS-2 Avg heat wt": "T/Blow",
    "SMS-1 lining life": "Nos/Lining",   "SMS-2 lining life": "Nos/Lining",
    "SMS-1 Avg Blows per day": "nos/day","SMS-2 Avg Blows per day": "nos/day",
    "COB#1-5 Hard Coke Yield%": "% dry coal",
    "COB#1-5 Dry Coal Charge":  "T/Oven",
}

# Parameters whose workbook value is total blows/month and must be divided
# by days-in-month to yield the average blows per day.
_BLOWS_DAILY_AVG = {"SMS-1 Avg Blows per day", "SMS-2 Avg Blows per day"}

# ---------------------------------------------------------------------------
# RSP Special Steel Excel — product group name normalisation
# ---------------------------------------------------------------------------
# Maps the exact string found in col A of the Excel to the canonical product
# name used as `product` in special_steel_orders (matches _gen_rsp in page_special_steel.py).

_RSP_SS_PRODUCT_MAP = {
    "PM PLATES":               "PM PLATES",
    "New PM PLATES":           "New PM PLATES",
    "New PM PLATES ":          "New PM PLATES",          # trailing space in workbook
    "HR PLATES -NDL":          "HR PLATES -NDL",
    "HR COILS (SALE) - HSM-1": "HR COILS (SALE) - HSM-1",
    "HR PLATES  SSL":          "HR PLATES SSL",           # double-space → single
    "HR PLATES SSL":           "HR PLATES SSL",
    "HR COILS (SALE) -HSM-2":  "HR COILS (SALE) -HSM-2",
    "Pipes, CRNO":             "Pipes, CRNO",
    "SPP":                     "SPP",
    "Value Added SLABs":       "Value Added SLABs",
}


# Expected workbook-label fragments (normalized) for each techno parameter.
# Used in preview to flag rows where the resolved cell does not match expectations.
# Derived from P18_TECHNO_ITEMS aliases; anchor-based params listed explicitly.
_LABEL_CHECKS: dict = {name: aliases for name, aliases, _ in P18_TECHNO_ITEMS}
_LABEL_CHECKS.update({
    "SMS-1 Avg heat wt":        ["blow weight"],
    "SMS-2 Avg heat wt":        ["blow weight"],
    "SMS-1 lining life":        ["avg lining life"],
    "SMS-2 lining life":        ["avg lining life"],
    "SMS-1 Oxygen Consumption": ["oxygen consmn ld"],
    "SMS-2 Oxygen Consumption": ["oxygen cons bof ii"],
    "SMS-1 Converter Yield":    ["converter yield"],
    "SMS-2 Converter Yield":    ["converter yield"],
    "SMS-1 Caster Yield":       ["caster yield"],
    "SMS-2 Caster Yield":       ["caster yield"],
    "SMS-1 Avg Blows per day":  ["blow sms1", "blow sms 1"],
    "SMS-2 Avg Blows per day":  ["blow sms2", "blow sms 2"],
    # COB-6 workbook labels are identical to Battery-1-5 rows; accept them directly
    "COB-6 Dry Coal Charge per Oven":   ["dry coal charge"],
    "COB#6 Coke yield%":                ["h coke yield 25mm"],
})


def _ytd_days(year_i: int, month_i: int) -> int:
    """Total calendar days from April 1 of the financial year to end of report month.
    Financial year starts in April (RSP convention).
    """
    fy_year = year_i if month_i >= 4 else year_i - 1
    total, y, m = 0, fy_year, 4
    while True:
        total += calendar.monthrange(y, m)[1]
        if y == year_i and m == month_i:
            break
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return total

# ---------------------------------------------------------------------------
# Label-scanning helpers for dynamic cell resolution
# ---------------------------------------------------------------------------

def _scan_rows_for_items(ws, items, max_scan: int) -> dict:
    """Return {item_name: row_number} by scanning ws labels.

    Strategy (in priority order):
      1. Exact normalized-label match against each alias.
      2. Word-boundary prefix match (alias must end at a space or row-label end).
      3. Row hint from the config (fallback when sheet rows shift).

    More-specific aliases should be listed first in *items* so that exact
    matching assigns them before a shorter generic alias can claim their row.
    """
    # Single pass: build {normalized_label: first_row} for every non-empty row.
    label_rows: dict = {}
    for r in range(1, max_scan + 1):
        t = _label_at(ws, r)
        if t and t not in label_rows:
            label_rows[t] = r

    row_map: dict = {}
    claimed: set  = set()

    # Phase 1 — exact alias match (highest confidence)
    for name, aliases, _ in items:
        for a in aliases:
            na = _norm(a)
            if na in label_rows and label_rows[na] not in claimed:
                row_map[name] = label_rows[na]
                claimed.add(label_rows[na])
                break

    # Phase 2 — word-boundary prefix match for still-unresolved items
    for name, aliases, _ in items:
        if name in row_map:
            continue
        for a in aliases:
            na = _norm(a)
            for t, r in label_rows.items():
                if r in claimed:
                    continue
                if t.startswith(na) and (len(t) == len(na) or not t[len(na)].isalnum()):
                    row_map[name] = r
                    claimed.add(r)
                    break
            if name in row_map:
                break

    # Phase 3 — row hint fallback. Must still respect `claimed`: a stale hint
    # landing on a row another item already resolved via real label matching
    # would silently duplicate that item's value under the wrong name (this
    # is exactly how the Jun'26 RSP file's SMS-2 CCM-4/Total Crude Steel
    # collision happened — an unresolved item's stale hint overwrote a row a
    # different item had already correctly claimed).
    for name, _, hint in items:
        if name not in row_map and hint and hint not in claimed:
            row_map[name] = hint
            claimed.add(hint)

    return row_map


def _row_near_anchor(ws, anchor_aliases: list, offset: int,
                     max_scan: int, fallback: int) -> int:
    """Return anchor_row + offset for the first row matching any anchor alias (prefix match).
    offset=0  → anchor row itself (e.g. O2 consumption)
    offset=-1 → row immediately before (e.g. Avg. Lining Life)
    offset=+N → N rows after the anchor (e.g. Converter/Caster Yield)
    Returns fallback if anchor is not found.
    """
    norm_aliases = [_norm(a) for a in anchor_aliases]
    for r in range(1, max_scan + 1):
        t = _label_at(ws, r)
        for na in norm_aliases:
            if t == na or (t.startswith(na) and (len(t) == len(na) or not t[len(na)].isalnum())):
                return max(1, r + offset)
    return fallback


def _build_p9_cells(ws, col: str) -> dict:
    """Dynamically resolve page-9 production rows; return {item_name: cell_ref}."""
    row_map = _scan_rows_for_items(ws, P9_ITEMS, max_scan=60)
    # "Finished Steel" has no row of its own in any real file — mirror
    # Saleable Steel's resolved row rather than guessing one (see P9_ITEMS).
    if "Finished Steel" not in row_map and "Saleable Steel" in row_map:
        row_map["Finished Steel"] = row_map["Saleable Steel"]
    return {name: f"{col}{row}" for name, row in row_map.items()}


def _build_p18_cells(ws, col: str, cum_col: str) -> dict:
    """Dynamically resolve page-1-8 techno rows; return {param: (month_cell, cum_cell)}."""
    row_map = _scan_rows_for_items(ws, P18_TECHNO_ITEMS, max_scan=400)
    # Several SMS params share labels across SMS-1 and SMS-2 sections — label scanning
    # finds only the first occurrence. All are located by offset from the section-unique
    # O2 anchor (LD for SMS-1, BOF II for SMS-2):
    #   -3 → Blow Weight (Avg heat wt),  -1 → Avg. Lining Life
    #    0 → O2 consumption,  +3 → Converter Yield,  +4 → Caster Yield
    for sms, o2_anchor, hw_fb, ll_fb, o2_fb, cv_fb, cs_fb in [
        ("SMS-1", ["oxygen consmn ld"],   175, 177, 178, 181, 182),
        ("SMS-2", ["oxygen cons bof ii"], 204, 206, 207, 210, 211),
    ]:
        row_map[sms + " Avg heat wt"]        = _row_near_anchor(ws, o2_anchor, -3, 400, hw_fb)
        row_map[sms + " lining life"]        = _row_near_anchor(ws, o2_anchor, -1, 400, ll_fb)
        row_map[sms + " Oxygen Consumption"] = _row_near_anchor(ws, o2_anchor,  0, 400, o2_fb)
        row_map[sms + " Converter Yield"]    = _row_near_anchor(ws, o2_anchor, +3, 400, cv_fb)
        row_map[sms + " Caster Yield"]       = _row_near_anchor(ws, o2_anchor, +4, 400, cs_fb)
    return {
        name: (f"{col}{row}", f"{cum_col}{row}")
        for name, row in row_map.items()
    }


# ---------------------------------------------------------------------------
# Morning-report cells (fixed mapping — columns vary too much for label scan)
# ---------------------------------------------------------------------------
MORNING_CELLS = {
    "COB#1-5":             "F10",
    "COB#6":               "F11",
    "Oven Pushing (nos/day)": "F12",
    "SP-1":                "E41",
    "SP-2":                "E42",
    "SP-3":                "E43",
    "Total Sinter":        "E44",
    "BF#1":                "K50",
    "BF#5":                "K52",
    "Hot Metal":           "K53",
    "Pig Iron":            "E296",
    "SMS-1 CCM-1":         "E92",
    "SMS-2 CCM-1&2":       ("X69", "X74"),
    "SMS-2 CCM-3":         "L99",
    "SMS-2 CCM-4":         "X79",
    "Total Crude Steel":   "F94",
    "HSM-2 Total HR Coil": "AB209",
    "HSM-2 HR Coil (Sale)":"Z263",
    "HSM-2 HR Plate":      "AB210",
    "OPM Plate":           "F204",
    "NPM Plate":           "F215",
    "CRNO Coils":          "E267",
    "ERW Pipes":           "E265",
    "SW Pipes":            "E266",
    "Saleable Steel":      "E268",
    "Finished Steel":      "E268",
}


# ---------------------------------------------------------------------------
# Mill techno-parameter configuration  (MILL_RSP group)
# ---------------------------------------------------------------------------
# Schema: (section_name, [shop_anchors], [(label, [aliases], unit, sort, row_hint)])
#
# • shop_anchors: normalized strings that identify the section header row.
# • aliases: startswith-matched against the normalized label of each row.
# • row_hint: sample-mapping row number used only when anchor search fails.
#   Specific Heat / Specific Power live in a HEAT & POWER block far below the
#   mill section, so they always rely on row_hints.

MAX_SCAN_ROWS = 600
LABEL_COLS    = 14

SHOPS = [
    ("Plate Mill", ["plate mill"], [
        ("Yield - Primes",      ["yield primes", "yield prime"],
                                                          "%",       1, 221),
        ("Yield - Total",       ["yield total"],          "%",       2, 222),
        ("Average Slab Weight", ["average slab weight", "avg slab weight"],
                                                          "Tons",    3, 223),
        ("Mill Availability",   ["mill availability"],    "%",       4, 224),
        # file label: "Util. on Avail. Time"
        ("Mill Utilisation",    ["util on avail", "mill utilisation", "mill utilization"],
                                                          "%Avl.",   5, 225),
        ("Rolling Rate",        ["rolling rate"],         "T/Hr.",   6, 226),
        ("Specific Heat",       ["specific heat", "sp heat cons"],
                                                          "M.Cal/T", 7, 310),
        ("Specific Power",      ["specific power", "sp power cons"],
                                                          "Kwh/T",   8, 328),
    ]),
    ("HSM-II", ["hsm ii", "hsm 2", "hot strip mill ii", "hot strip mill 2"], [
        ("H R Coil Yield",         ["h r coil yield", "hr coil yield"],
                                                          "%",      10, 253),
        ("Average Slab Weight",    ["average slab weight"],
                                                          "Tons",   11, 254),
        ("Mill Availability",      ["mill availability"], "%",      12, 255),
        # file label: "Utilisation on Avail. Hrs"
        ("Mill Utilisation",       ["utilisation on avail", "utili on avail",
                                    "mill utilisation", "mill utilization"],
                                                          "%Avl.",  13, 256),
        ("Rolling Rate -HR Coils", ["rolling rate hr coils", "rolling rate"],
                                                          "T/Hr.",  14, 257),
        ("Specific Heat",          ["specific heat"],     "M.Cal/T",15, 313),
        ("Specific Power",         ["specific power"],    "Kwh/T",  16, 331),
    ]),
    ("New Plate Mill", ["new plate mill", "npm"], [
        ("Yield - Primes",      ["yield primes", "yield prime"],
                                                          "%",      20, 234),
        ("Yield - Total",       ["yield total"],          "%",      21, 235),
        ("Average Slab Weight", ["average slab weight"],  "Tons",   22, 236),
        ("Mill Availability",   ["mill availability"],    "%",      23, 237),
        # file label: "Util. on Avail. Time"
        ("Mill Utilisation",    ["util on avail", "mill utilisation", "mill utilization"],
                                                          "%Avl.",  24, 238),
        ("Rolling Rate",        ["rolling rate"],         "T/Hr.",  25, 239),
        ("Specific Heat",       ["specific heat"],        "M.Cal/T",26, 311),
        ("Specific Power",      ["specific power"],       "Kwh/T",  27, 329),
    ]),
    ("SSM", ["ssm", "silicon steel mill"], [
        ("Yield from HRC-CRNO",   ["yield from hrc crno", "yield from hrc"],
                                                          "%",     30, 286),
        ("Acid Cons. in AP Line", ["acid cons in ap line", "acid cons"],
                                                          "kg/T",  31, 287),
        # file label: "REV. MILL Availability On Cal Hrs"
        ("Mill Availability",     ["rev mill availability", "mill availability on cal",
                                   "mill availability"],  "%",     32, 288),
        # file label: "REV.MILL Utilisation.on Avail Hrs"
        ("Mill Utilisation",      ["rev mill utilisation", "rev mill utilization",
                                   "mill utilisation on avail", "mill utilisation"],
                                                          "%Avl.", 33, 289),
        ("Rolling Rate",          ["rolling rate rev mill", "rolling rate"],
                                                          "T/Hr.", 34, 290),
    ]),
    ("ERW Pipe Plant", ["erw pipe plant", "erw pipe"], [
        ("Yield from HR Coils", ["yield from hr coils"],  "%",     40, 261),
        ("Mill Availability",   ["mill availability"],     "%",     41, 263),
        # file label: "Utili.on Available Time"
        ("Mill Utilisation",    ["utili on available", "utili on avail",
                                 "mill utilisation", "mill utilization"],
                                                          "%Avl.", 42, 264),
        ("Rolling Rate",        ["rolling rate"],          "T/Hr.", 43, 265),
    ]),
    ("SW Pipe Plant", ["sw pipe plant", "sw pipe", "spiral weld pipe"], [
        ("Yield from HR Coils", ["yield from hr coils"],  "%",     50, 268),
        ("Mill Availability",   ["mill availability"],     "%",     51, 270),
        # file label: "Utili. on Avail. Time"
        ("Mill Utilisation",    ["mill utilization"],
                                                          "%Avl.", 52, 271),
        ("Rolling Rate",        ["rolling rate"],          "T/Hr.", 53, 272),
    ]),
]


# ---------------------------------------------------------------------------
# Sinter productivity configuration  (COKE_SINTER group)
# ---------------------------------------------------------------------------
# Schema: (section_name, [anchor_aliases], unit, sort_order, prod_row_hint)
#
# Anchor matching uses EXACT match only to avoid 'Sinter Plant I' ambiguously
# matching 'Sinter Plant II' / 'Sinter Plant III'.

SINTER_SECTIONS = [
    ("RSP SP-1", ["sinter plant i",   "sinter plant 1"],  "T/m2/hr", 33, 38),
    ("RSP SP-2", ["sinter plant ii",  "sinter plant 2"],  "T/m2/hr", 34, 58),
    ("RSP SP-3", ["sinter plant iii", "sinter plant 3"],  "T/m2/hr", 35, 81),
]
_SINTER_GROUP      = "COKE_SINTER"
_SINTER_PARAM      = "Sinter Productivity"
_SINTER_P_ALIASES  = ["specific productivity", "sinter productivity", "productivity"]


# ---------------------------------------------------------------------------
# Mill techno helper functions
# ---------------------------------------------------------------------------

def _label_at(ws, r: int) -> str:
    """Concatenated normalized text of the first LABEL_COLS cells of a row."""
    parts = []
    for c in range(1, LABEL_COLS + 1):
        v = ws.cell(row=r, column=c).value
        if v is not None and not isinstance(v, (int, float, datetime.datetime)):
            parts.append(str(v))
    return _norm(" ".join(parts))


def _find_shop_rows(ws, max_row: int) -> dict:
    """Returns {section_name: anchor_row} for SHOPS using fuzzy label match."""
    anchors = {}
    for r in range(1, max_row + 1):
        t = _label_at(ws, r)
        if not t:
            continue
        for section, aliases, _ in SHOPS:
            if section in anchors:
                continue
            max_alias = max(aliases, key=len)
            if any(t == a or t.startswith(a + " ") or (" " + a) in (" " + t)
                   for a in aliases):
                if len(t) <= len(max_alias) + 25:
                    anchors[section] = r
    return anchors


def _find_sinter_anchors(ws, max_row: int) -> dict:
    """Returns {section_name: anchor_row} for SINTER_SECTIONS.

    Uses startswith + word-boundary guard to distinguish 'sinter plant i'
    from 'sinter plant ii' / 'sinter plant iii' even when the row label is
    long (the sinter header row often has historical year columns appended).
    """
    anchors = {}
    for r in range(1, max_row + 1):
        t = _label_at(ws, r)
        if not t:
            continue
        for section, aliases, *_ in SINTER_SECTIONS:
            if section in anchors:
                continue
            for a in aliases:
                if t.startswith(a):
                    # word-boundary: next char must not be alphanumeric
                    nxt = len(a)
                    if nxt >= len(t) or not t[nxt].isalnum():
                        anchors[section] = r
                        break
    return anchors


def _find_param_row(ws, aliases: list, start: int, end: int) -> Optional[int]:
    """Find first row in [start, end] whose normalized label startswith any alias."""
    for r in range(start, end + 1):
        t = _label_at(ws, r)
        if not t:
            continue
        for a in aliases:
            if t == a or t.startswith(a):
                return r
    return None


def _pick_mill_sheet(wb):
    """Return the worksheet with the most SHOPS section anchors."""
    best, best_hits = wb.worksheets[0], -1
    for ws in wb.worksheets:
        hits = len(_find_shop_rows(ws, min(ws.max_row, MAX_SCAN_ROWS)))
        if hits > best_hits:
            best, best_hits = ws, hits
    return best


# ---------------------------------------------------------------------------
# Unified techno-param extraction  (mill + sinter, no DB writes)
# ---------------------------------------------------------------------------

def extract_techno_params(wb, report_month: str) -> dict:
    """Extract MILL_RSP + COKE_SINTER rows from an already-open workbook.
    Returns a preview dict; caller decides whether to persist.

    Column detection scans the sheet's own header row for the month/Cum tokens
    (see rsp_row_scan.find_month_cum_columns) rather than assuming a fixed
    month->column-letter map — the techno sheet prepends one more legacy
    fiscal-year column every year, so a fixed map goes stale annually.
    """
    db_report_month, month_num = _parse_report_month(report_month)

    ws      = _pick_mill_sheet(wb)
    month_col, cum_col = find_month_cum_columns(ws, month_num)
    if month_col is None:
        raise ValueError(f"Cannot locate month '{month_num}' column header on sheet {ws.title!r}.")
    max_row = min(ws.max_row, MAX_SCAN_ROWS)

    def col_letter_of(c):
        return openpyxl.utils.get_column_letter(c)

    # ── mill shops ────────────────────────────────────────────────────────────
    shop_rows = _find_shop_rows(ws, max_row)
    ordered   = sorted(shop_rows.items(), key=lambda kv: kv[1])
    ranges    = {}
    for i, (sec, r) in enumerate(ordered):
        end = ordered[i + 1][1] - 1 if i + 1 < len(ordered) else max_row
        ranges[sec] = (r, end)

    rows = []
    for section, _aliases, params in SHOPS:
        rng = ranges.get(section)
        for label, p_aliases, unit, sort, row_hint in params:
            row  = None
            via  = ""
            norm_aliases = [_norm(a) for a in p_aliases]
            if rng:
                row = _find_param_row(ws, norm_aliases, rng[0], rng[1])
                if row is not None:
                    via = "anchor"
            if row is None and row_hint:
                row = row_hint
                via = "row-hint"

            actual = cum = None
            cell_ref = ""
            if row is not None:
                actual   = clean_val(ws.cell(row=row, column=month_col).value)
                cum      = clean_val(ws.cell(row=row, column=cum_col).value)
                cell_ref = f"{col_letter_of(month_col)}{row}/{col_letter_of(cum_col)}{row}"

            status = ("ok"        if (actual is not None or cum is not None)
                      else "no value" if row is not None
                      else "not found")
            rows.append({
                "plant":      "RSP",
                "group_code": "MILL_RSP",
                "section":    section,
                "parameter":  label,
                "unit":       unit,
                "sort_order": sort,
                "month":      db_report_month,
                "actual":     actual,
                "cum_actual": cum,
                "cell":       cell_ref,
                "found_via":  via if row is not None else "",
                "status":     status,
            })

    # ── sinter plants ─────────────────────────────────────────────────────────
    sinter_anchors = _find_sinter_anchors(ws, max_row)
    all_anchors    = sorted(sinter_anchors.values()) + [max_row + 1]

    for section, _aliases, unit, sort, hint in SINTER_SECTIONS:
        anchor = sinter_anchors.get(section)
        row    = None
        via    = ""
        if anchor is not None:
            # range: this anchor to the next anchor (or max_row)
            next_start = next((v for v in all_anchors if v > anchor), max_row + 1)
            row = _find_param_row(ws, [_norm(a) for a in _SINTER_P_ALIASES],
                                  anchor, next_start - 1)
            if row is not None:
                via = "anchor"
        if row is None and hint:
            row = hint
            via = "row-hint"

        actual = cum = None
        cell_ref = ""
        if row is not None:
            actual   = clean_val(ws.cell(row=row, column=month_col).value)
            cum      = clean_val(ws.cell(row=row, column=cum_col).value)
            cell_ref = f"{col_letter_of(month_col)}{row}/{col_letter_of(cum_col)}{row}"

        status = ("ok"        if (actual is not None or cum is not None)
                  else "no value" if row is not None
                  else "not found")
        rows.append({
            "plant":      "RSP",
            "group_code": _SINTER_GROUP,
            "section":    _SINTER_PARAM,
            "parameter":  section,
            "unit":       unit,
            "sort_order": sort,
            "month":      db_report_month,
            "actual":     actual,
            "cum_actual": cum,
            "cell":       cell_ref,
            "found_via":  via if row is not None else "",
            "status":     status,
        })

    return {
        "plant":      "RSP",
        "month":      db_report_month,
        "sheet":      ws.title,
        "month_col":  col_letter_of(month_col),
        "cum_col":    col_letter_of(cum_col),
        "shops_found": sorted(shop_rows.keys()),
        "sinter_found": sorted(sinter_anchors.keys()),
        "rows":       rows,
    }


# ---------------------------------------------------------------------------
# Preview extraction (no DB writes)
# ---------------------------------------------------------------------------

def _preview_production_from_cells(ws, cells):
    rows = []
    for item, spec in cells.items():
        if isinstance(spec, tuple):
            parts    = [clean_val(ws[c].value) for c in spec]
            val      = None if all(p is None for p in parts) else sum(p for p in parts if p is not None)
            cell_ref = "+".join(spec)
        else:
            val      = clean_val(ws[spec].value)
            cell_ref = spec
        if val is not None and item not in NO_CONVERT:
            val = round(val / 1000.0, 3)
        rows.append({
            "item_name": item,
            "value":     val,
            "cell":      cell_ref,
            "unit":      "nos/d" if item in NO_CONVERT else "'000T",
            "status":    "ok" if val is not None else "no value",
        })
    return rows


def _preview_techno_from_cells(ws, cells, days_in_month=None, ytd_days=None):
    rows = []
    techno_values_with_cells = {}

    # First pass: extract all values with cell references
    for param, (m_cell, y_cell) in cells.items():
        mv = clean_val(ws[m_cell].value)
        yv = clean_val(ws[y_cell].value)
        if param in _BLOWS_DAILY_AVG:
            if days_in_month and mv is not None:
                mv = round(mv / days_in_month, 1)
            if ytd_days and yv is not None:
                yv = round(yv / ytd_days, 1)
        techno_values_with_cells[param] = (mv, yv, m_cell, y_cell)

    # Extract just the values for TMI calculation
    techno_values = {k: (v[0], v[1]) for k, v in techno_values_with_cells.items()}

    # Calculate TMI as HM Consumption + Scrap Consumption for all SMS units
    techno_values = calculate_tmi_consumption(techno_values)

    # Rebuild with cell references
    for param, (mv, yv) in techno_values.items():
        if param in techno_values_with_cells:
            _, _, m_cell, y_cell = techno_values_with_cells[param]
        else:
            # New calculated TMI parameter - use a dummy cell ref
            m_cell = y_cell = "CALC"
        techno_values_with_cells[param] = (mv, yv, m_cell, y_cell)

    # Second pass: build rows (including calculated TMI)
    for param, (mv, yv, m_cell, y_cell) in techno_values_with_cells.items():
        # Mapping correctness check: read the workbook label at the resolved row
        row_num   = openpyxl.utils.coordinate_to_tuple(m_cell)[0]
        row_label = str(ws.cell(row=row_num, column=1).value or "").strip()
        norm_lbl  = _norm(row_label)
        expected  = _LABEL_CHECKS.get(param, [])
        mapping_ok = any(
            norm_lbl == _norm(a) or
            (norm_lbl.startswith(_norm(a)) and
             (len(norm_lbl) == len(_norm(a)) or not norm_lbl[len(_norm(a))].isalnum()))
            for a in expected
        ) if expected else bool(row_label)

        rows.append({
            "parameter":    param,
            "unit":         TECHNO_UNIT_MAP.get(param, ""),
            "month_actual": mv,
            "ytd_actual":   yv,
            "cell":         f"{m_cell}/{y_cell}",
            "row_label":    row_label,
            "mapping_ok":   mapping_ok,
            "status":       "ok" if (mv is not None or yv is not None) else "no value",
        })
    return rows


# ---------------------------------------------------------------------------
# RSP Special Steel Excel extraction
# ---------------------------------------------------------------------------

def _is_rsp_ss_file(wb) -> bool:
    """True when the workbook is the RSP Special Steel report.
    Signature: the FIRST sheet's R1C1 starts with 'Special Steel'. Extra
    unrelated sheets (e.g. old scratch tabs left over in a compilation
    workbook — one real sample has 10) don't disqualify it; only the
    first sheet's own content does."""
    ws = wb.worksheets[0]
    r1c1 = str(ws.cell(1, 1).value or "").strip().lower()
    return r1c1.startswith("special steel")


_MON_ABBR_TO_NUM = {name[:3]: num for num, name in MONTH_NAMES.items()}


def _parse_rsp_ss_header_month(ws) -> Optional[str]:
    """Read the report-month label from R4C2 (e.g. "Jan'26") → '2026-01'.

    The header spells the month as a 3-letter abbreviation ("Aug", "Jan"),
    not the full name — MONTH_NUMS is keyed by full names ("August"), so
    looking it up there always missed and this returned None on every real
    file, silently disabling month detection (and with it, the mismatch
    check) entirely.
    """
    raw = str(ws.cell(4, 2).value or "").strip().strip("'\"")
    m = re.match(r"([A-Za-z]+)'(\d{2})", raw)
    if not m:
        return None
    mon_name, yr2 = m.group(1).capitalize(), int(m.group(2))
    yr_full = 2000 + yr2 if yr2 < 50 else 1900 + yr2
    mon_num = _MON_ABBR_TO_NUM.get(mon_name[:3])
    if not mon_num:
        return None
    return f"{yr_full}-{mon_num}"


def _fmt_ym(ym: str) -> str:
    """'2026-01' -> 'January 2026'"""
    try:
        y, mo = ym[:4], ym[5:7]
        return f"{MONTH_NAMES.get(mo, mo)} {y}"
    except Exception:
        return ym


def _extract_rsp_ss_preview(wb, report_month: str) -> dict:
    """Parse the RSP Special Steel Excel workbook.

    Column layout (row 4-5 header):
      C1  = grade / product label
      C3  = monthly Orders        → order_qty saved to DB
      C4  = monthly Despatch      → actual_despatch saved to DB
      C7  = YTD Orders            (cross-check only, not saved)
      C8  = YTD Despatch          (cross-check only, not saved)
      C10 = CPLY YTD Despatch     (cross-check only, not saved)

    Returns the standard preview dict with special_steel_rows.
    """
    ws = wb.worksheets[0]
    detected_month = _parse_rsp_ss_header_month(ws)
    if detected_month and detected_month != report_month:
        raise ValueError(
            f"Month mismatch: this RSP Special Steel report's own header (cell B4) shows "
            f"{_fmt_ym(detected_month)}, but you selected {_fmt_ym(report_month)}. "
            f"Please select '{_fmt_ym(detected_month)}' in the month picker, "
            f"or upload the report for {_fmt_ym(report_month)}."
        )
    db_month = detected_month or report_month

    rows = []
    cur_product = ""
    sort = 0

    for r in range(6, ws.max_row + 1):
        c1_raw = ws.cell(r, 1).value
        if c1_raw is None:
            continue
        label = str(c1_raw).strip()
        if not label:
            continue

        # Footer sentinel rows — stop scanning
        if (label.startswith("GRAND TOTAL") or
                label.startswith("Total Saleable") or
                label.startswith("% Special")):
            break

        # Product group header
        canonical = _RSP_SS_PRODUCT_MAP.get(label) or _RSP_SS_PRODUCT_MAP.get(label.rstrip())
        if canonical:
            cur_product = canonical
            continue

        c3  = clean_val(ws.cell(r, 3).value)    # monthly order
        c4  = clean_val(ws.cell(r, 4).value)    # monthly despatch
        c7  = clean_val(ws.cell(r, 7).value)    # YTD order (cross-check)
        c8  = clean_val(ws.cell(r, 8).value)    # YTD despatch (cross-check)
        c10 = clean_val(ws.cell(r, 10).value)   # CPLY YTD (cross-check)

        # Skip rows where every value is None or zero (sub-section headings, blank lines)
        if not any(v for v in (c3, c4, c7, c8, c10)):
            continue

        is_total = label.upper().startswith("TOTAL")
        sort += 1

        # CRNO grade belongs to "Pipes, CRNO" regardless of where it sits in the Excel layout
        row_product = cur_product
        if not is_total and label.upper() == "CRNO":
            row_product = "Pipes, CRNO"

        rows.append({
            "product":         "" if is_total else row_product,
            "quality_grade":   label,
            "section":         "",
            "sort_order":      sort,
            "order_qty":       c3,
            "prodn":           None,
            "actual_despatch": c4,
            "cum_order_qty":   c7,    # preview cross-check only
            "cum_actual":      c8,    # preview cross-check only
            "cply_ytd":        c10,   # preview cross-check only
            "unit":            "T",
            "cell":            f"R{r}C3/C4",
            "status":          "total" if is_total else "ok",
        })

    return {
        "plant":             "RSP",
        "month":             db_month,
        "detected_month":    detected_month,
        "source_type":       "Special Steel Report",
        "sheets":            ws.title,
        "workbook_sheets":   wb.sheetnames,
        "production_rows":   [],
        "techno_rows":       [],
        "techno_param_rows": [],
        "special_steel_rows": rows,
    }


def _save_rsp_special_steel(wb, report_month: str, source_file_name: str) -> bool:
    """Persist RSP special steel rows to special_steel_orders."""
    import db as _db
    result  = _extract_rsp_ss_preview(wb, report_month)
    db_month = result["month"]
    saved = 0
    _ss_rows = [r for r in result["special_steel_rows"]
                if r["status"] == "ok" and (r.get("order_qty") is not None or r.get("actual_despatch") is not None)]
    if _ss_rows:
        _db.clear_special_steel_orders(db_month, "RSP")
    for r in _ss_rows:
        _db.save_special_steel_entry(
            db_month, "RSP",
            r["product"], r["quality_grade"],
            r["sort_order"],
            r.get("order_qty"), r.get("actual_despatch"),
            section="",
        )
        saved += 1
    if saved:
        _db.log_extraction(
            plant="RSP", report_month=db_month,
            file_name=source_file_name, sheet_name=result["sheets"],
            source_type="Special Steel Report",
            items_extracted=saved)
        logger.info(f"RSP Special Steel: {saved} rows saved for {db_month}.")
    return saved > 0


def _extract_morning_stock(ws, db_report_month: str) -> list:
    """Extract opening stock from RSP Daily Morning Report worksheet.

    Two months of data per run:
      this_month = stock as on 1st of db_report_month  (C301, J268)
      next_month = stock as on 1st of next month        (D301, K268, L296)

    Cell mapping (raw Tonnes → stored as '000T, 3 d.p.):
      C301 → SLABS INPROCESS, this_month
      D301 → SLABS INPROCESS, next_month
      J268 → FINISHED STEEL,  this_month
      K268 → FINISHED STEEL,  next_month
      L296 → PIG IRON,        next_month only
    """
    y, m = int(db_report_month[:4]), int(db_report_month[5:7])
    this_month = db_report_month
    next_month = f"{y+1 if m == 12 else y}-{1 if m == 12 else m+1:02d}"

    def _t(v):
        f = clean_val(v)
        return round(f / 1000, 3) if f is not None else None

    def _row(item_type, stock_type, stock_month, value, formula):
        return {
            "plant": "RSP", "item_type": item_type, "stock_type": stock_type,
            "stock_month": stock_month, "value": value, "formula": formula,
            "status": "ok" if value is not None else "skip",
        }

    return [
        _row("SLABS",         "INPROCESS", this_month, _t(ws["C301"].value), "C301 (opening this month)"),
        _row("SLABS",         "INPROCESS", next_month, _t(ws["D301"].value), "D301 (opening next month)"),
        _row("FINISHED STEEL", "",         this_month, _t(ws["J268"].value), "J268 (opening this month)"),
        _row("FINISHED STEEL", "",         next_month, _t(ws["K268"].value), "K268 (opening next month)"),
        _row("PIG IRON",      "",          next_month, _t(ws["L296"].value), "L296 (opening next month)"),
    ]


def extract_preview(file_path: str, report_month: str) -> dict:
    """Unified RSP preview: production + techno_table + mill/sinter techno params.
    Auto-detects file type. No database writes."""
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Special Steel Report takes priority — single sheet, R1C1 starts with "Special Steel"
    if _is_rsp_ss_file(wb):
        return _extract_rsp_ss_preview(wb, report_month)

    sheet_names = wb.sheetnames

    production_rows, techno_rows, stock_rows = [], [], []
    source_type, sheets_used = "Techno Parameters File", ""
    db_report_month, month_num = _parse_report_month(report_month)

    p9_sheet, p18_sheet = _find_report_sheets(wb, sheet_names)
    if p9_sheet or p18_sheet:
        # Production (p9) and techno-table (p18) sheets are detected and processed
        # independently — a workbook where only one of the two names/sheets is
        # found (common: RSP's p18-style sheet name varies far more than p9's)
        # should still yield whichever half is actually present, rather than
        # silently discarding both.
        source_type = "Final Monthly Report"
        sheets_used = ", ".join(s for s in (p9_sheet, p18_sheet) if s)
        year_i, month_i = map(int, db_report_month.split('-'))
        days_in_month   = calendar.monthrange(year_i, month_i)[1]
        ytd_days_val    = _ytd_days(year_i, month_i)

        if p9_sheet:
            col_p9 = COL_MAP_P9.get(month_num)
            if not col_p9:
                raise ValueError(f"Month column mapping not found for month '{month_num}'.")
            ws_p9 = wb[p9_sheet]
            _assert_p9_month_year_match(ws_p9, report_month)
            production_rows = _preview_production_from_cells(ws_p9, _build_p9_cells(ws_p9, col_p9))

        if p18_sheet:
            ws_p18 = wb[p18_sheet]
            month_col_idx, cum_col_idx = find_month_cum_columns(ws_p18, month_num)
            if month_col_idx is not None and cum_col_idx is not None:
                col_p18 = openpyxl.utils.get_column_letter(month_col_idx)
                cum_p18 = openpyxl.utils.get_column_letter(cum_col_idx)
                techno_rows = _preview_techno_from_cells(
                    ws_p18, _build_p18_cells(ws_p18, col_p18, cum_p18),
                    days_in_month, ytd_days_val)
            else:
                logger.warning(
                    f"RSP page-1-8 techno: could not locate month '{month_num}'/Cum "
                    f"header columns on sheet {ws_p18.title!r} — skipping techno_rows.")
    else:
        morning_sheet = next(
            (s for s in sheet_names
             if s.strip().lower().startswith("rsp morning report data for-")), None)
        if morning_sheet:
            source_type = "Daily Morning Report"
            sheets_used = morning_sheet
            ws = wb[morning_sheet]
            a2_raw = ws["A2"].value or ""
            dm = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", str(a2_raw))
            if dm:
                _d, m_num2, year = dm.groups()
                db_report_month = f"{year}-{m_num2}"
            production_rows = _preview_production_from_cells(ws, MORNING_CELLS)
            stock_rows = _extract_morning_stock(ws, db_report_month)

    techno_param_rows, mill_meta = [], {}
    if source_type != "Daily Morning Report":
        try:
            t = extract_techno_params(wb, db_report_month)
            if t.get("shops_found") or t.get("sinter_found"):
                techno_param_rows = t["rows"]
                mill_meta = {
                    "mill_sheet":        t["sheet"],
                    "month_col":         t["month_col"],
                    "cum_col":           t["cum_col"],
                    "shops_found":       t["shops_found"],
                    "sinter_found":      t.get("sinter_found", []),
                }
        except Exception as e:
            logger.warning(f"RSP mill/sinter techno scan skipped: {e}")

    if not production_rows and not techno_rows and not techno_param_rows and not stock_rows:
        raise ValueError(
            "No extractable data found. Expected an RSP Final Monthly Report, "
            "Daily Morning Report, or a techno-parameters workbook with mill sections.")

    return {
        "plant":              "RSP",
        "month":              db_report_month,
        "source_type":        source_type,
        "sheets":             sheets_used,
        "workbook_sheets":    sheet_names,
        "production_rows":    production_rows,
        "techno_rows":        techno_rows,
        "techno_param_rows":  techno_param_rows,
        "stock_rows":         stock_rows,
        **mill_meta,
    }


# ---------------------------------------------------------------------------
# Save dispatcher
# ---------------------------------------------------------------------------

def extract_and_save_excel(file_path: str, report_month: str,
                           source_file_name: str = "") -> bool:
    """Auto-detects RSP file type and writes all extracted data to the DB."""
    try:
        wb          = openpyxl.load_workbook(file_path, data_only=True)

        if _is_rsp_ss_file(wb):
            return _save_rsp_special_steel(wb, report_month, source_file_name)

        sheet_names = wb.sheetnames

        p9_sheet, p18_sheet = _find_report_sheets(wb, sheet_names)
        if p9_sheet or p18_sheet:
            # Production (p9) and page-1-8 techno-table data are extracted and
            # persisted independently — a workbook where only one of the two
            # sheets is found still gets whichever half is actually present.
            # (Page-1-8's techno values feed only the legacy techno_table,
            # which no longer exists as a DB table — RSP's techno_data table,
            # read by /data-entry/techno, is populated exclusively via the
            # dedicated TechnoExtractor upload flow, not from here.)
            return _extract_monthly_report(wb, report_month, source_file_name,
                                           p9_sheet, p18_sheet)

        morning_sheet = next(
            (s for s in sheet_names
             if s.strip().lower().startswith("rsp morning report data for-")), None)
        if morning_sheet:
            return _extract_morning_report(wb, morning_sheet, source_file_name)

        raise ValueError(
            "Uploaded RSP file does not match any known format. "
            "Expected a sheet named 'page-9' (production) and/or a 'page-1-8'-style "
            "sheet (techno table), or 'RSP Morning Report Data for-...' "
            "(Daily Morning Report). For a standalone RSP Technopara file, use the "
            "techno-parameters upload flow instead.")
    except ValueError as ve:
        logger.error(f"RSP validation error: {ve}")
        raise ve
    except Exception as e:
        logger.error(f"RSP extraction error: {e}")
        return False


# ---------------------------------------------------------------------------
# Extractor 1 — Final Monthly Report
# ---------------------------------------------------------------------------

def _extract_monthly_report(wb, report_month: str, source_file_name: str,
                             p9_name: Optional[str] = "page-9",
                             p18_name: Optional[str] = None) -> bool:
    """Persist production_table data from the page-9 sheet.

    p18_name (the page-1-8 techno-table sheet) is accepted for backward
    compatibility but is not persisted here: its target table (techno_table)
    no longer exists in the schema — RSP's techno_data table (read by
    /data-entry/techno) is populated exclusively via the dedicated
    TechnoExtractor upload flow. Page-1-8 values remain available read-only
    through extract_preview()/_build_p18_cells() for that upload flow's own use.
    """
    if not p9_name:
        raise ValueError("No page-9 production sheet found — nothing to save.")

    import db as _db

    db_report_month, month_num = _parse_report_month(report_month)
    col_p9 = COL_MAP_P9.get(month_num)
    if not col_p9:
        raise ValueError(f"Month column mapping not found for month '{month_num}'.")

    sheet_p9 = wb[p9_name]
    _assert_p9_month_year_match(sheet_p9, report_month)

    conn   = db.connect()
    cursor = conn.cursor()
    vals_extracted = 0

    for item_name, cell in _build_p9_cells(sheet_p9, col_p9).items():
        val = clean_val(sheet_p9[cell].value)
        if val is not None:
            vals_extracted += 1
            if item_name not in NO_CONVERT:
                val = round(val / 1000.0, 3)
        cursor.execute("""
            INSERT INTO production_table (report_month, plant_name, item_name, month_actual)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(report_month, plant_name, item_name)
            DO UPDATE SET month_actual = excluded.month_actual
        """, (db_report_month, "RSP", item_name, val))

    if vals_extracted == 0:
        conn.close()
        raise ValueError("No numeric data found in the RSP monthly report sheets.")

    conn.commit()
    conn.close()

    _db.log_extraction(
        plant="RSP", report_month=db_report_month,
        file_name=source_file_name,
        sheet_name=", ".join(s for s in (p9_name, p18_name) if s),
        source_type="Final Monthly Report",
        items_extracted=vals_extracted)
    logger.info(f"RSP Monthly Report: {vals_extracted} values for {db_report_month}.")
    return True


# ---------------------------------------------------------------------------
# Extractor 2 — Daily Morning Report
# ---------------------------------------------------------------------------

def _extract_morning_report(wb, sheet_name: str, source_file_name: str) -> bool:
    import db as _db

    ws = wb[sheet_name]
    a2_raw     = ws["A2"].value or ""
    date_match = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", str(a2_raw))
    if not date_match:
        raise ValueError(
            f"Cannot parse date from cell A2: {repr(a2_raw)}. "
            "Expected format DD.MM.YYYY.")
    _day, m_num, year = date_match.groups()
    db_report_month = f"{year}-{m_num}"
    logger.info(f"RSP Morning Report: month auto-detected → {db_report_month}")

    conn   = db.connect()
    cursor = conn.cursor()
    vals_extracted = 0

    for item_name, cell_spec in MORNING_CELLS.items():
        if isinstance(cell_spec, tuple):
            parts = [clean_val(ws[c].value) for c in cell_spec]
            val   = sum(p for p in parts if p is not None) or None
            if all(p is None for p in parts):
                val = None
        else:
            val = clean_val(ws[cell_spec].value)
        if val is not None:
            vals_extracted += 1
            if item_name not in NO_CONVERT:
                val = round(val / 1000.0, 3)
        cursor.execute("""
            INSERT INTO production_table (report_month, plant_name, item_name, month_actual)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(report_month, plant_name, item_name)
            DO UPDATE SET month_actual = excluded.month_actual
        """, (db_report_month, "RSP", item_name, val))

    if vals_extracted == 0:
        conn.close()
        raise ValueError(
            "No numeric data found at expected cell locations in the Morning Report.")

    conn.commit()
    conn.close()

    _db.log_extraction(
        plant="RSP", report_month=db_report_month,
        file_name=source_file_name, sheet_name=sheet_name,
        source_type="Daily Morning Report",
        items_extracted=vals_extracted)
    logger.info(f"RSP Morning Report: {vals_extracted} values for {db_report_month}.")
    return True
