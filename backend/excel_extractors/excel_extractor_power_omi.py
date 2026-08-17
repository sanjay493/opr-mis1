"""Power-OIS monthly report extractor — "MONTHLY SUMMARY OF POWER DATA"
workbook (Report_format/Power-OIS/*.xlsx), one sheet per FY (e.g. "2026-27"),
one 13-row block per plant (12 month rows Apr->Mar + a "Cum." row) for BSP,
DSP, RSP, BSL, ISP, SSP, VISP, CFP, and a SAIL total block.

Column layout (confirmed stable across FY2022-23, FY2024-25, and FY2026-27
samples via ws.merged_cells.ranges + direct cell inspection — only whether a
cell HAS data varies between vintages, never its position):
    A       plant name (only populated on each block's first row)
    B       per-row date stamp (dropped — not meaningful to store)
    C-G     PLAN generation: Own(OLD), New P&BS, JV-PP-II, Drawal JV PP3, Total (MW)
    H-L     ACTUAL generation: same 5 sub-items (MW)
    M-U     grid/consumption: Wheeling from PX, Purchase from PX, Renewable+
            GDAM, Drawal from Grid, Export to Grid, Total Power Consump,
            Decarbon Nos, Decarbon Hrs, Specific Power Cons (KWH/TSS) — label
            wording drifted ("NIL GRID POWER"->"RE Grid Drawal"/"Decarbon")
            but column order never has.
    V       last-year date stamp (dropped)
    W-      last-year comparison: Own CPP, JV CPP, Drawal PP3, Total Gen,
            Total Power Consump — located by header-TEXT search (not a fixed
            column), see _find_last_year_cols.

Row layout is NOT positionally stable: some plants carry an extra plant-
name-only header row before April in some vintages (BSL, ISP), some have a
stray blank row before "Cum." (SSP) — see _find_month_rows, which locates
each month row by matching column B's own date value instead of counting a
fixed offset from the plant-name cell.

Each monthly upload restates the WHOLE FY-to-date grid (verified: April's
PLAN values are byte-identical between a April-only file and a later
April-July cumulative file for the same FY), so extraction is naturally
idempotent — every upload's records simply upsert on (report_month,
plant_name, item_name), later uploads overwriting earlier ones for the same
month with no special "first upload of the FY" handling needed.
"""
import datetime
from typing import Dict, List, Optional

import openpyxl

_PLANTS = ["BSP", "DSP", "RSP", "BSL", "ISP", "SSP", "VISP", "CFP", "SAIL"]

# 1-based column indices for the current-FY metrics — see module docstring.
_FIXED_COLS: Dict[str, int] = {
    "plan_own": 3, "plan_new_pbs": 4, "plan_jv_pp2": 5, "plan_drawal_jv_pp3": 6, "plan_total": 7,
    "actual_own": 8, "actual_new_pbs": 9, "actual_jv_pp2": 10, "actual_drawal_jv_pp3": 11, "actual_total": 12,
    "wheeling_px": 13, "purchase_px": 14, "renewable_gdam": 15, "drawal_grid": 16, "export_grid": 17,
    "total_power_consump": 18, "decarbon_nos": 19, "decarbon_hrs": 20, "specific_power_cons": 21,
}

# Header-text needles (all must appear, any order, in the same cell) for the
# 4 labeled last-year sub-columns — searched across a bounded column window
# so a future re-wording (already seen: spacing/newline differences between
# vintages) doesn't silently break extraction the way a fixed offset would.
_LAST_YEAR_LABELS: Dict[str, List[str]] = {
    "last_year_own_cpp": ["OWN", "CPP"],
    "last_year_jv_cpp": ["JV", "CPP"],
    "last_year_drawal_pp3": ["DRAWAL", "PP3"],
    "last_year_total_gen": ["TOTAL", "GEN"],
}


def _find_last_year_cols(ws) -> Dict[str, int]:
    """The 5th last-year column ("Total Power Consump", last year) has no
    row5 sub-label of its own in any sample vintage (it's the trailing,
    unlabeled column of the wider "ACTUAL LAST YEAR" merge) — derived as
    one column past the last of the 4 labeled ones instead, which is
    correct in every sample regardless of where that group starts."""
    found: Dict[str, int] = {}
    for r in (4, 5, 6):
        for c in range(20, 34):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            up = v.upper().replace("\n", " ")
            for key, needles in _LAST_YEAR_LABELS.items():
                if key in found:
                    continue
                if all(n in up for n in needles):
                    found[key] = c
    if found:
        found["last_year_total_power_consump"] = max(found.values()) + 1
    return found


def _find_month_rows(ws, plant_row: int, max_scan: int = 16):
    """From plant_row (column A == the plant's own code), scan forward for
    (YYYY-MM -> row) by column B's date value, plus the 'Cum.' row —
    never a fixed row offset, see module docstring. A stray duplicate
    plant-name row or blank separator row (both seen in real files) is
    simply skipped: this only looks at column B, never column A, while
    scanning."""
    month_rows: Dict[str, int] = {}
    cum_row: Optional[int] = None
    r = plant_row
    end_row = plant_row + max_scan
    while r < end_row:
        b = ws.cell(row=r, column=2).value
        if isinstance(b, datetime.datetime):
            month_rows[f"{b.year}-{b.month:02d}"] = r
        elif isinstance(b, str) and b.strip().upper().startswith("CUM"):
            cum_row = r
            break
        r += 1
    return month_rows, cum_row


def _clean(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "-", "—", "#DIV/0!", "#VALUE!", "#N/A", "#REF!"):
        return None
    try:
        return float(s.replace(",", ""))
    except (ValueError, TypeError):
        return None


def extract_power_omi(file_path: str) -> dict:
    """Parse a Power-OIS workbook — no DB writes, safe to call from a
    preview endpoint. Returns:
        {"records": [{"report_month","plant_name","item_name","value"}, ...],
         "months": ["2026-04", ...],
         "plants_found": [...],
         "warnings": [...]}
    Raises ValueError if nothing recognizable was found at all.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    warnings: List[str] = []
    ly_cols = _find_last_year_cols(ws)
    if len(ly_cols) < 5:
        warnings.append("Could not locate all 5 'last year' comparison columns.")

    all_cols = dict(_FIXED_COLS)
    all_cols.update(ly_cols)

    records: List[dict] = []
    months_seen = set()
    plants_found = []

    r = 7
    for plant in _PLANTS:
        prow = None
        while r <= ws.max_row:
            a = ws.cell(row=r, column=1).value
            if isinstance(a, str) and a.strip() == plant:
                prow = r
                break
            r += 1
        if prow is None:
            warnings.append(f"Plant block '{plant}' not found.")
            continue
        plants_found.append(plant)

        month_rows, cum_row = _find_month_rows(ws, prow)
        if len(month_rows) < 12:
            warnings.append(f"{plant}: only found {len(month_rows)}/12 month rows.")
        months_seen.update(month_rows)

        for month, row_idx in month_rows.items():
            for item, col in all_cols.items():
                val = _clean(ws.cell(row=row_idx, column=col).value)
                if val is not None:
                    records.append({
                        "report_month": month, "plant_name": plant,
                        "item_name": item, "value": val,
                    })

        if cum_row is not None and month_rows:
            # NOT max(month_rows) — every month row exists year-round since
            # PLAN is pre-filled for the whole FY even for months that
            # haven't happened yet (see module docstring), so that would
            # always resolve to next March regardless of how far the file
            # has actually been reported. The Cum row instead belongs to
            # the latest month that actually HAS an actual_total figure.
            reported_months = [
                mo for mo, row_idx in month_rows.items()
                if ws.cell(row=row_idx, column=_FIXED_COLS["actual_total"]).value is not None
            ]
            latest_month = max(reported_months) if reported_months else max(month_rows)
            for item, col in all_cols.items():
                val = _clean(ws.cell(row=cum_row, column=col).value)
                if val is not None:
                    records.append({
                        "report_month": latest_month, "plant_name": plant,
                        "item_name": f"{item}_cum", "value": val,
                    })

        r = (cum_row + 1) if cum_row is not None else (prow + 14)

    if not records:
        raise ValueError(
            "No data extracted — file may not match the expected Power-OIS "
            "layout (single sheet, plant blocks in column A: "
            + ", ".join(_PLANTS) + ")."
        )

    return {
        "records": records,
        "months": sorted(months_seen),
        "plants_found": plants_found,
        "warnings": warnings,
    }
