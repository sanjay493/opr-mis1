"""BSP (Bhilai Steel Plant) Excel extractor — unified module.

Handles five distinct BSP file types, auto-detected:

  1. BSP PPC MIS daily report (.xls, sheet "S1")
       → extract_and_save_excel() — direct DB write, no preview
  2. BSP MIS-2 month-end report (.xls/.xlsx, row 2 = "BSP MIS-2") — furnace-
     wise Hot Metal production (tentative), column D "CUM" on a month-end-
     dated report is the for-the-month figure
       → _extract_mis2_furnace_preview()
  3. BSP 3-page Techno parameters (.xlsx, sheet "Sheet1", month in R3C1)
       → _extract_techno_3page_preview()
  4. BSP OISCO Techno parameters (.xlsx, R3C3 contains "TECHNO ECONOMIC PARAMETERS")
       → _extract_oisco_preview()
  5. BSP Special Steel report (.xlsx, sheet "CORP", R3C1 = "BHILAI STEEL PLANT")
       → _extract_bsp_ss_preview()

Public API:
  extract_and_save_excel(file_path, report_month, source_file_name)
  extract_preview(file_path, report_month)  ← unified, auto-detects type (2-5)
"""

import logging
import os
import re
import sqlite3
from typing import Optional, List, Dict, Any

import openpyxl
from openpyxl.utils import get_column_letter

from extraction_utils import calculate_tmi_consumption

try:
    import xlrd
    _XLRD_AVAILABLE = True
except ImportError:
    _XLRD_AVAILABLE = False

logger = logging.getLogger("excel_extractor")

DB_PATH = os.path.join(os.path.dirname(__file__), "..", "mis_reports.db")


# ---------------------------------------------------------------------------
# Excel 97-2003 (.xls) compatibility shim
# Makes an xlrd workbook/sheet look like openpyxl (1-based cell access).
# ---------------------------------------------------------------------------

class _XlsCell:
    __slots__ = ("value",)
    def __init__(self, value):
        self.value = None if value == "" else value


class _XlsSheet:
    def __init__(self, sheet):
        self._s = sheet
        self.max_column = sheet.ncols
        self.max_row = sheet.nrows

    def cell(self, row: int, col: int) -> _XlsCell:
        r, c = row - 1, col - 1
        if r < 0 or r >= self._s.nrows or c < 0 or c >= self._s.ncols:
            return _XlsCell(None)
        return _XlsCell(self._s.cell_value(r, c))

    @property
    def title(self):
        return self._s.name


class _XlsWorkbook:
    def __init__(self, wb):
        self._wb = wb
        self.worksheets = [_XlsSheet(wb.sheets()[i]) for i in range(wb.nsheets)]

    @property
    def sheetnames(self):
        return self._wb.sheet_names()

    def __getitem__(self, name: str) -> _XlsSheet:
        return _XlsSheet(self._wb.sheet_by_name(name))

    @property
    def active(self) -> _XlsSheet:
        return _XlsSheet(self._wb.sheets()[0])


def _open_workbook(file_path: str):
    """Open .xlsx (openpyxl) or .xls (xlrd) and return a unified interface."""
    if file_path.lower().endswith(".xls"):
        if not _XLRD_AVAILABLE:
            raise ImportError("xlrd is required for .xls files: pip install xlrd")
        return _XlsWorkbook(xlrd.open_workbook(file_path))
    return openpyxl.load_workbook(file_path, data_only=True)


# ---------------------------------------------------------------------------
# Shared month constants
# ---------------------------------------------------------------------------

_MONTH_NAME_TO_NUM: Dict[str, str] = {
    "JAN": "01", "FEB": "02", "MAR": "03", "APR": "04",
    "MAY": "05", "JUN": "06", "JUL": "07", "AUG": "08",
    "SEP": "09", "OCT": "10", "NOV": "11", "DEC": "12",
    "JANUARY": "01", "FEBRUARY": "02", "MARCH": "03", "APRIL": "04",
    "JUNE": "06", "JULY": "07", "AUGUST": "08",
    "SEPTEMBER": "09", "OCTOBER": "10", "NOVEMBER": "11", "DECEMBER": "12",
}

_MONTH_NUM_TO_HDR: Dict[str, str] = {
    "04": "APR", "05": "MAY", "06": "JUN", "07": "JUL",
    "08": "AUG", "09": "SEP", "10": "OCT", "11": "NOV",
    "12": "DEC", "01": "JAN", "02": "FEB", "03": "MAR",
}

# report month-number (MM) → 1-based Excel column index (for 3-page-Tech)
_MONTH_NUM_TO_COL: Dict[str, int] = {
    "04": 4, "05": 5, "06": 6, "07": 7, "08": 8, "09": 9,
    "10": 10, "11": 11, "12": 12, "01": 13, "02": 14, "03": 15,
}

_MONTH_FULL: Dict[str, str] = {
    "01": "January", "02": "February", "03": "March", "04": "April",
    "05": "May", "06": "June", "07": "July", "08": "August",
    "09": "September", "10": "October", "11": "November", "12": "December",
}


# ---------------------------------------------------------------------------
# Shared clean helpers
# ---------------------------------------------------------------------------

def _clean(v) -> Optional[float]:
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "-", "—", "nan", "###", "#DIV/0!", "#VALUE!", "#N/A", "#REF!"):
        return None
    try:
        return float(s.replace(",", ""))
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Section 1 — BSP PPC MIS daily report (.xls, sheet "S1")
# ---------------------------------------------------------------------------

MONTH_NAMES = {
    1: "January", 2: "February", 3: "March", 4: "April",
    5: "May", 6: "June", 7: "July", 8: "August",
    9: "September", 10: "October", 11: "November", 12: "December"
}


def _ppc_mis_config() -> Dict[str, tuple]:
    """
    Single source of truth for the BSP PPC MIS cell mapping — shared by the
    DB-writing extractor (extract_and_save_excel) and the preview-only
    extractor (_extract_ppc_mis_preview) so they can never drift apart.

    Reads excel_cells_config.json (section 'bsp_ppc_mis'); falls back to these
    hardcoded (row_0based, col_0based, divide_by_1000) tuples only if that
    config section is missing.
    """
    import sys
    sys.path.insert(0, os.path.dirname(__file__))
    from cells_loader import get_extractor_config
    cfg = get_extractor_config("bsp_ppc_mis")

    default_rc = {
        "COB#1-8":             (3,  5,  False),
        "Oven Pushing (nos/day)": (5,  5,  False),
        "SP-2":                (7,  5,  True),
        "SP-3":                (8,  5,  True),
        "Total Sinter":        (9,  5,  True),
        "BF#1-7":              (10, 5,  True),
        "BF#8":                (11, 5,  True),
        "Hot Metal":           (12, 5,  True),
        "SMS-2":               (13, 5,  True),
        "SMS-3":               (15, 5,  True),
        "Total Crude Steel":   (16, 5,  True),
        "RSM_RAIL":            (17, 5,  True),
        "URM_RAIL":            (21, 5,  True),
        "MM":                  (23, 5,  True),
        "WIRERODS":            (24, 5,  True),
        "BARS&RODMILL":        (25, 5,  True),
        "PLATEMILL":           (26, 5,  True),
        "Finished Steel":      (27, 5,  True),
        "Saleable Semis":      (34, 5,  True),
        "Saleable Steel":      (35, 5,  True),
        "RSMPRIME":            (38, 5,  True),
        "URMPRIME":            (38, 9,  True),
        "Pig Iron":            (62, 13, True),
    }
    # Config stores [row, col, divide] lists; convert to tuples matching code expectation
    cfg_rc = cfg.get("cells_rc", {})
    return {
        k: tuple(v) if isinstance(v, list) else v
        for k, v in cfg_rc.items()
    } if cfg_rc else default_rc


def _detect_ppc_date(ws):
    """Scan row 0 for a plausible Excel date serial.

    The date lived at N1 (col 13) from ~2021-Mar onward, but older (~2012-2020)
    reports have it one column to the left (M1, col 12) instead — everything
    below it shifted too (see _PPC_STABLE_LABELS / _PPC_GUARDED_ITEMS), so a
    fixed column here isn't safe either. Bounded plausibility window (roughly
    1982-2036) avoids false positives from stray numeric cells nearby.
    """
    for c in range(9, 16):
        v = ws.cell_value(0, c)
        if isinstance(v, float) and 25000 < v < 55000:
            return v
    return None


# Items whose column-A label text has been verified to identify the same row
# reliably across every report vintage seen (~2012-2026), despite surrounding
# rows being inserted/removed/renamed between eras. Matched by substring
# (case-insensitive) against the *whole sheet*, first match wins.
_PPC_STABLE_LABELS = {
    "Oven Pushing(nos/d)": (False, ["EQV. PUSHING"]),
    "SP-2":                (True,  ["SP-2"]),
    "SP-3":                (True,  ["SP-3"]),
    "Total Sinter":        (True,  ["TOTAL SINTER"]),
    "BF#1-7":              (True,  ["BF-1 TO 7"]),
    "BF#8":                (True,  ["BF-8"]),
    "MM":                  (True,  ["MERCH"]),
    "WIRERODS":            (True,  ["WIRE RODS"]),
    "BARS&RODMILL":        (True,  ["BRM"]),
    "PLATEMILL":           (True,  ["PLATES"]),
    "Finished Steel":      (True,  ["FINISHED STEEL"]),
    "Saleable Semis":      (True,  ["TOTAL SEMIS"]),
    "Saleable Steel":      (True,  ["SAL. STEEL", "SAL STEEL", "SAL. STEEL PROD"]),
}

# Items whose section genuinely restructured across eras (rows added/removed,
# e.g. "Total Hot Metal" didn't exist before ~2020; SMS-1 disappeared after
# ~2021) — not safe to search for by label across the whole sheet the way
# _PPC_STABLE_LABELS is, since old and new eras mean different things by
# similar text. Instead keep the current config's fixed cell, but only trust
# it if the label actually sitting there (or a nearby anchor cell, for the
# two sub-tables that use a different column) still matches — otherwise the
# item is skipped for that file rather than risk silently-wrong data.
# value: (guard_row, guard_col, required_substring)
_PPC_GUARDED_ITEMS = {
    "Hot Metal":         (12, 0, "HOT METAL"),
    "SMS-2":             (13, 0, "SMS-2"),
    "SMS-3":             (15, 0, "SMS-3"),
    "Total Crude Steel": (16, 0, "CR STEEL"),
    "RSM_RAIL":          (17, 0, "RSM"),
    "URM_RAIL":          (21, 0, "URM"),
    "COB#1-8":           (3,  0, "BATT"),
    "RSMPRIME":          (37, 2, "RSM"),
    "URMPRIME":          (37, 7, "URM"),
    "Pig Iron":          (60, 12, "PIG IRON"),
}


def _find_ppc_label_row(ws, substrings, max_row=45):
    """First row (0-based) whose column-A text contains any of substrings."""
    for r in range(0, max_row):
        label = str(ws.cell_value(r, 0) or "").strip().upper()
        if not label:
            continue
        if any(s in label for s in substrings):
            return r
    return None


def extract_and_save_excel(file_path: str, report_month: str = None,
                           source_file_name: str = "") -> bool:
    """Extract BSP production actuals from the daily PPC MIS .xls report (sheet S1).

    Month auto-detected from the date cell (see _detect_ppc_date — its column
    has drifted between report vintages). report_month used only as fallback.
    Units: raw Tonnes → stored as '000 T (divide by 1000). Coke items as-is (nos/day).
    """
    try:
        wb = xlrd.open_workbook(file_path)
        if "S1" not in wb.sheet_names():
            raise ValueError(
                "Uploaded BSP file is missing required sheet 'S1'. "
                "Please upload the correct BSP PPC MIS daily report."
            )

        ws = wb.sheet_by_name("S1")

        date_serial = _detect_ppc_date(ws)
        if date_serial is not None:
            y, m, *_ = xlrd.xldate_as_tuple(date_serial, wb.datemode)
            db_report_month = f"{y}-{m:02d}"
            logger.info(f"BSP PPC MIS: month auto-detected → {db_report_month}")
        elif report_month:
            db_report_month = report_month
        else:
            raise ValueError(
                "Cannot determine report month: no plausible date found and "
                "no report_month was provided."
            )

        production_cells = dict(_ppc_mis_config())

        # Stable items: search for the row by label instead of trusting the
        # config's row number, which is only valid for the current-era layout.
        for item_name, (convert, substrings) in _PPC_STABLE_LABELS.items():
            row = _find_ppc_label_row(ws, substrings)
            default_col = production_cells.get(item_name, (None, 5, None))[1]
            if row is not None:
                production_cells[item_name] = (row, default_col, convert)
            else:
                production_cells.pop(item_name, None)

        # Restructured items: keep the config's row/col, but only if the
        # label actually there (or its section anchor) still matches —
        # otherwise this file's layout doesn't match what that row assumes,
        # so skip rather than risk reading the wrong quantity.
        for item_name, (grow, gcol, needle) in _PPC_GUARDED_ITEMS.items():
            if item_name not in production_cells:
                continue
            guard_label = str(ws.cell_value(grow, gcol) or "").strip().upper()
            if needle not in guard_label:
                production_cells.pop(item_name, None)

        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        vals_extracted = 0

        for item_name, (row_idx, col_idx, do_convert) in production_cells.items():
            raw = ws.cell_value(row_idx, col_idx)
            val = _clean(raw)
            if val is not None:
                if do_convert:
                    val = round(val / 1000.0, 3)
                vals_extracted += 1

            cursor.execute("""
                INSERT INTO production_table (report_month, plant_name, item_name, month_actual)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(report_month, plant_name, item_name)
                DO UPDATE SET month_actual = excluded.month_actual
            """, (db_report_month, "BSP", item_name, val))

        if vals_extracted == 0:
            raise ValueError(
                "No numeric data found at expected cell locations in sheet S1. "
                "Please verify this is the correct BSP PPC MIS file."
            )

        conn.commit()
        conn.close()

        import sys
        sys.path.insert(0, os.path.dirname(__file__))
        import db as _db
        _db.log_extraction(
            plant="BSP", report_month=db_report_month,
            file_name=source_file_name, sheet_name="S1",
            source_type="Daily PPC MIS Report",
            items_extracted=vals_extracted,
        )
        logger.info(f"BSP PPC MIS extraction done: {vals_extracted} values for {db_report_month}.")
        return True

    except ValueError:
        raise
    except Exception as e:
        logger.error(f"BSP PPC MIS extraction error: {e}")
        return False


# ---------------------------------------------------------------------------
# Section 2 — BSP 3-page Techno parameters (.xlsx, sheet "Sheet1")
#
# Sheet layout:
#   Row  3, Col A : Report month name, e.g. "MAY"
#   Row  4, Cols D-O : Column month headers APR … MAR
#   Row  4, Col P : "ACTUAL" / cumulative header
#   Col  B (2)    : Unit
#   Col  D–O      : Monthly data (APR–MAR)
#   Col  P (16)   : Cumulative "till month" value — ALWAYS
# ---------------------------------------------------------------------------

_TECHNO_3PAGE_CUM_FALLBACK = 16
_TECHNO_3PAGE_CUM_HEADERS = {
    "ACTUAL", "CUM", "CUMULATIVE", "CUM.", "TILL DATE", "TILL MONTH", "APR-ACTUAL",
}

PARAM_MAP_3PAGE: List[tuple] = [
    # (row_1based, group_code, section, parameter_label, unit)
    # ── Coke & By-products (rows 37-40) ─────────────────────────────────────
    (37, "COKE_SINTER", "Coke Yield",           "BF Coke",                   "%"),
    (38, "COKE_SINTER", "Coke Yield",           "Crude Tar",                 "%"),
    (39, "COKE_SINTER", "Coke Yield",           "Ammonium Sulphate",         "%"),
    (40, "COKE_SINTER", "Coke Yield",           "Crude Benzol",              "%"),
    # ── Sinter Plant 2 (rows 45-49) ─────────────────────────────────────────
    (45, "COKE_SINTER", "Sinter Plant SP-2",    "Machine Availability",      "%"),
    (46, "COKE_SINTER", "Sinter Plant SP-2",    "Machine Utilisation",       "%"),
    (48, "COKE_SINTER", "Sinter Plant SP-2",    "Productivity",              "T/m²/hr"),
    (49, "COKE_SINTER", "Sinter Plant SP-2",    "Basicity",                  ""),
    # ── Sinter Plant 3 (rows 51-55) ─────────────────────────────────────────
    (51, "COKE_SINTER", "Sinter Plant SP-3",    "Machine Availability",      "%"),
    (52, "COKE_SINTER", "Sinter Plant SP-3",    "Machine Utilisation",       "%"),
    (54, "COKE_SINTER", "Sinter Plant SP-3",    "Productivity",              "T/m²/hr"),
    (55, "COKE_SINTER", "Sinter Plant SP-3",    "Basicity",                  ""),
    # ── Blast Furnaces (rows 58-70) ─────────────────────────────────────────
    (58, "IRON_MAKING", "Sinter in Burden",       "BSP",           "%"),
    (59, "IRON_MAKING", "BF Coke Rate",           "BSP",           "Kg/THM"),
    (60, "IRON_MAKING", "Coke Screen Loss",        "BSP",           "25mm%"),
    (68, "IRON_MAKING", "CDI",                    "BSP",           "Kg/THM"),
    (69, "IRON_MAKING", "Slag Rate",              "BSP",           "Kg/THM"),
    (70, "IRON_MAKING", "Nut Coke Rate",           "BSP",           "Kg/THM"),
    # ── BF Productivity (rows 65-67) ────────────────────────────────────────
    (65, "IRON_MAKING", "BF Productivity",        "BSP BF-7",                 "T/m³/day"),
    (66, "IRON_MAKING", "BF Productivity",        "BSP BF-8",                 "T/m³/day"),
    (67, "IRON_MAKING", "BF Productivity",        "BSP",           "T/m³/day"),
    # ── SMS-II Consumption (rows 76-78) ─────────────────────────────────────
    (76, "SMS", "SMS-II Consumption",            "Hot Metal",                "Kg/TCS"),
    (77, "SMS", "SMS-II Consumption",            "Scrap",                    "Kg/TCS"),
    (78, "SMS", "SMS-II Consumption",            "Total Metallic Charge",    "Kg/TCS"),
    # ── SMS-III Consumption (rows 90-92) ────────────────────────────────────
    (90, "SMS", "SMS-III Consumption",           "Hot Metal",                "Kg/TCS"),
    (91, "SMS", "SMS-III Consumption",           "Scrap",                    "Kg/TCS"),
    (92, "SMS", "SMS-III Consumption",           "Total Metallic Charge",    "Kg/TCS"),
    # ── Rail & Structural Mill — RSM ─────────────────────────────────────────
    ( 97, "MILL_BSP", "Rail & Structural Mill",  "Yield",                    "%"),
    ( 98, "MILL_BSP", "Rail & Structural Mill",  "Rolling Rate",             "T/Hr"),
    ( 99, "MILL_BSP", "Rail & Structural Mill",  "Mill Availability",        "%"),
    (100, "MILL_BSP", "Rail & Structural Mill",  "Mill Utilisation",         "%"),
    (137, "MILL_BSP", "Rail & Structural Mill",  "Heat Consumption",         "103Kcal/T"),
    (153, "MILL_BSP", "Rail & Structural Mill",  "Power Consumption",        "Kwh/T"),
    # ── Universal Rail Mill — URM ─────────────────────────────────────────────
    (103, "MILL_BSP", "Universal Rail Mill",     "Yield",                    "%"),
    (104, "MILL_BSP", "Universal Rail Mill",     "Rolling Rate",             "T/Hr"),
    (105, "MILL_BSP", "Universal Rail Mill",     "Mill Availability",        "%"),
    (106, "MILL_BSP", "Universal Rail Mill",     "Mill Utilisation",         "%"),
    (138, "MILL_BSP", "Universal Rail Mill",     "Heat Consumption",         "103Kcal/T"),
    (154, "MILL_BSP", "Universal Rail Mill",     "Power Consumption",        "Kwh/T"),
    # ── Merchant Mill — MM ───────────────────────────────────────────────────
    (109, "MILL_BSP", "Merchant Mill",           "Yield",                    "%"),
    (110, "MILL_BSP", "Merchant Mill",           "Rolling Rate",             "T/Hr"),
    (111, "MILL_BSP", "Merchant Mill",           "Mill Availability",        "%"),
    (112, "MILL_BSP", "Merchant Mill",           "Mill Utilisation",         "%"),
    (139, "MILL_BSP", "Merchant Mill",           "Heat Consumption",         "103Kcal/T"),
    (155, "MILL_BSP", "Merchant Mill",           "Power Consumption",        "Kwh/T"),
    # ── Wire Rod Mill — WRM ──────────────────────────────────────────────────
    (115, "MILL_BSP", "Wire Rod Mill",           "Yield",                    "%"),
    (116, "MILL_BSP", "Wire Rod Mill",           "Rolling Rate",             "T/Hr"),
    (117, "MILL_BSP", "Wire Rod Mill",           "Mill Availability",        "%"),
    (118, "MILL_BSP", "Wire Rod Mill",           "Mill Utilisation",         "%"),
    (140, "MILL_BSP", "Wire Rod Mill",           "Heat Consumption",         "103Kcal/T"),
    (156, "MILL_BSP", "Wire Rod Mill",           "Power Consumption",        "Kwh/T"),
    # ── Bar & Rod Mill — BRM ─────────────────────────────────────────────────
    (121, "MILL_BSP", "Bar & Rod Mill",          "Yield",                    "%"),
    (122, "MILL_BSP", "Bar & Rod Mill",          "Rolling Rate",             "T/Hr"),
    (123, "MILL_BSP", "Bar & Rod Mill",          "Mill Availability",        "%"),
    (124, "MILL_BSP", "Bar & Rod Mill",          "Mill Utilisation",         "%"),
    # ── Plate Mill ───────────────────────────────────────────────────────────
    (127, "MILL_BSP", "Plate Mill",              "Yield",                    "%"),
    (128, "MILL_BSP", "Plate Mill",              "Rolling Rate",             "T/Hr"),
    (130, "MILL_BSP", "Plate Mill",              "Mill Availability",        "%"),
    (131, "MILL_BSP", "Plate Mill",              "Mill Utilisation",         "%"),
    (141, "MILL_BSP", "Plate Mill",              "Heat Consumption",         "103Kcal/T"),
    (157, "MILL_BSP", "Plate Mill",              "Power Consumption",        "Kwh/T"),
    # ── Energy ───────────────────────────────────────────────────────────────
    (161, "MILL_BSP", "Energy",                  "Sp. Energy Consumption",   "G.Cal/T"),
]


# ---------------------------------------------------------------------------
# Section 3 — BSP OISCO Techno parameters
#
# Sheet layout (single sheet "Sheet1"):
#   Row  3, Col C : Title "TECHNO ECONOMIC PARAMETERS_<MON>'YY"
#   Row  6, Cols E+ : Rolling month headers (APR | MAY | CUM)
#   Col  C (3)    : Parameter label
#   Col  D (4)    : Unit
#   Data col      : Column in row 6 matching user-selected month abbreviation
#   CUM  col      : Always one column to the right of the data column
# ---------------------------------------------------------------------------

_OISCO_HEADER_ROW = 6
_OISCO_LABEL_COL  = 3
_OISCO_UNIT_COL   = 4

PARAM_MAP_OISCO: List[tuple] = [
    # (row_1based, group_code, section, label, unit)
    # ── Blast Furnace Operations (rows 9-20) ─────────────────────────────────
    ( 9, "IRON_MAKING", "CDI",                 "BSP",          "Kg/THM"),
    (11, "IRON_MAKING", "CDI",                 "BSP BF-4",                "Kg/THM"),
    (12, "IRON_MAKING", "CDI",                 "BSP BF-5",                "Kg/THM"),
    (13, "IRON_MAKING", "CDI",                 "BSP BF-6",                "Kg/THM"),
    (14, "IRON_MAKING", "CDI",                 "BSP BF-7",                "Kg/THM"),
    (15, "IRON_MAKING", "CDI",                 "BSP BF-8",                "Kg/THM"),
    (16, "IRON_MAKING", "Fuel Rate",           "BSP",          "Kg/THM"),
    (17, "IRON_MAKING", "Pellet in Burden",    "BSP",          "%"),
    (18, "IRON_MAKING", "LD Slag Usage",       "BSP",          "Kg/THM"),
    (19, "IRON_MAKING", "Not Dry Cast",        "BSP",          "%"),
    (20, "MAJOR",       "Coal to Hot Metal",   "BSP",                     "Ratio"),
    # ── SMS-2 Operations (rows 22-36) ───────────────────────────────────────
    (22, "SMS", "BSP SMS-2", "Converter Availability",      "% ICH"),
    (23, "SMS", "BSP SMS-2", "Converter Utilisation",       "% Avail hr"),
    (24, "SMS", "BSP SMS-2", "Tap to Tap Time",             "Min"),
    (25, "SMS", "BSP SMS-2", "Average Blows/Day",           "Heats/Day"),
    (26, "SMS", "BSP SMS-2", "Average Heat Weight",         "T"),
    (27, "SMS", "BSP SMS-2", "Avg. Lining Life",            "Heats"),
    (28, "SMS", "BSP SMS-2", "Fe-Mn Consumption",           "Kg/TCS"),
    (29, "SMS", "BSP SMS-2", "Fe-Si Consumption",           "Kg/TCS"),
    (30, "SMS", "BSP SMS-2", "Si-Mn Consumption",           "Kg/TCS"),
    (31, "SMS", "BSP SMS-2", "Oxygen Consumption",          "Nm³/TCS"),
    (32, "SMS", "BSP SMS-2", "Alumina Consumption",         "Kg/TCS"),
    (35, "SMS", "BSP SMS-2", "LD Gas Recovery",             "m³/T"),
    (36, "SMS", "BSP SMS-2", "DS Heats",                    "Nos."),
    # ── SMS-3 Operations (rows 38-51) ──────────────────────────────────────
    (38, "SMS", "BSP SMS-3", "Converter Availability",     "% ICH"),
    (39, "SMS", "BSP SMS-3", "Converter Utilisation",      "% Avail hr"),
    (40, "SMS", "BSP SMS-3", "Tap to Tap Time",            "Min"),
    (41, "SMS", "BSP SMS-3", "Average Blows/Day",          "Heats/Day"),
    (42, "SMS", "BSP SMS-3", "Average Heat Weight",        "T"),
    (43, "SMS", "BSP SMS-3", "Fe-Mn Consumption",          "Kg/TCS"),
    (44, "SMS", "BSP SMS-3", "Fe-Si Consumption",          "Kg/TCS"),
    (45, "SMS", "BSP SMS-3", "Si-Mn Consumption",          "Kg/TCS"),
    (46, "SMS", "BSP SMS-3", "Oxygen Consumption",         "Nm³/TCS"),
    (47, "SMS", "BSP SMS-3", "Alumina Consumption",        "Kg/TCS"),
    (50, "SMS", "BSP SMS-3", "LD Gas Recovery",            "m³/T"),
    (51, "SMS", "BSP SMS-3", "DS Heats",                   "Nos."),
    # ── Utilities (row 54) ───────────────────────────────────────────────────
    (54, "GENERAL",      "Utilities", "Sp. Water Consumption",             "m³/TCS"),
]


# ---------------------------------------------------------------------------
# Section 4 — BSP Special Steel report (.xlsx, sheet "CORP")
#
# Sheet layout:
#   R2C9  : Report date "DD.MM.YY" (e.g. "31.03.26") — month extracted from here
#   R3C1  : "BHILAI STEEL PLANT"
#   R4-R5 : Column headers
#   Col A (1): Products (product group for first row of each group)
#   Col B (2): Quality/Grade
#   Col C (3): ABP Annual
#   Col D (4): ABP Monthly (current month)
#   Col E (5): Orders Available Total
#   Col F (6): Orders Available Effective  → saved as order_qty
#   Col G (7): Loading (Till Date)         → saved as actual_despatch
#   Col H (8): Loadable in stock
#   Col I (9): Semis/Fin in process
#   Data rows from R6 onwards
# ---------------------------------------------------------------------------

# Map from Col A label (uppercase) → canonical product group name stored in DB
_BSP_SS_GROUP_MAP: Dict[str, Optional[str]] = {
    "SEMIS":       "Semis",
    "WIRE RODS":   "Wire Rods",
    "MERCHANT":    "Merchant Products",  # first row of a 2-row merged label
    "PRODUCTS":    None,                 # continuation of "Merchant Products" — skip group change
    "BRM PRODUCT": "BRM Product",
    "RAILS":       "Rails",
    "PLATES":      "Plates",
}

# Labels in Col A that signal a total/subtotal row (prefix match)
_BSP_SS_TOTAL_PREFIXES = ("TOTAL ", "TOTAL\n")


def _is_bsp_ss_file(wb) -> bool:
    """True if workbook is BSP Special Steel: sheet 'CORP' with BHILAI STEEL PLANT in R3C1."""
    if "CORP" not in wb.sheetnames:
        return False
    ws = wb["CORP"]
    return "BHILAI STEEL PLANT" in str(ws.cell(3, 1).value or "").upper()


def _is_oisco_file(wb) -> bool:
    """True if R3C3 contains 'TECHNO ECONOMIC PARAMETERS' (OISCO format)."""
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    return "TECHNO ECONOMIC PARAMETERS" in str(ws.cell(3, 3).value or "").upper()


def _parse_bsp_ss_month(ws) -> Optional[str]:
    """Parse report month from R2C9 date string like '31.03.26' → '2026-03'."""
    raw = str(ws.cell(2, 9).value or "").strip()
    # Format: DD.MM.YY
    m = re.match(r"(\d{1,2})\.(\d{2})\.(\d{2})$", raw)
    if m:
        month_num = m.group(2)
        yr2 = int(m.group(3))
        yr_full = 2000 + yr2 if yr2 < 50 else 1900 + yr2
        return f"{yr_full}-{month_num}"
    return None


def _extract_bsp_ss_preview(wb, report_month: str) -> dict:
    """Parse BSP Special Steel Excel (CORP sheet).

    Returns the standard preview dict with special_steel_rows.
    order_qty = Col F (Orders Available Effective)
    actual_despatch = Col G (Loading Till Date)
    """
    ws = wb["CORP"]
    db_month = _parse_bsp_ss_month(ws) or report_month

    rows: List[Dict[str, Any]] = []
    cur_product = ""
    sort = 0

    for r in range(6, ws.max_row + 1):
        col_a_val = ws.cell(r, 1).value
        col_b_val = ws.cell(r, 2).value

        if col_a_val is None and col_b_val is None:
            continue

        label_a = str(col_a_val).strip() if col_a_val is not None else ""
        label_b = str(col_b_val).strip() if col_b_val is not None else ""
        label_a_up = label_a.upper()

        # Grand total sentinel — include then stop
        if "TOTAL SPECIAL STEEL" in label_a_up:
            cf = _clean(ws.cell(r, 6).value)
            cg = _clean(ws.cell(r, 7).value)
            sort += 1
            rows.append({
                "product": "", "quality_grade": label_a, "section": "",
                "sort_order": sort,
                "order_qty": cf, "prodn": None, "actual_despatch": cg,
                "abp_month": _clean(ws.cell(r, 4).value),
                "unit": "T", "cell": f"R{r}C6/C7",
                "status": "total",
            })
            break

        # Intermediate total / subtotal row
        is_total = any(label_a_up.startswith(p) for p in _BSP_SS_TOTAL_PREFIXES)

        # Update current product group
        canonical = _BSP_SS_GROUP_MAP.get(label_a_up)
        if canonical is not None:
            cur_product = canonical
        # If canonical is explicitly None it's a continuation row — keep cur_product

        # Determine grade label
        if is_total:
            grade = label_a
        elif label_b:
            grade = label_b
        elif label_a and label_a_up not in _BSP_SS_GROUP_MAP:
            grade = label_a
        else:
            continue  # pure group-header row with no grade

        # Extract values
        cf   = _clean(ws.cell(r, 6).value)   # effective orders → order_qty
        cg   = _clean(ws.cell(r, 7).value)   # loading till date → actual_despatch
        c_d  = _clean(ws.cell(r, 4).value)   # ABP monthly (cross-check)
        c_e  = _clean(ws.cell(r, 5).value)   # total orders (cross-check)

        # Skip structurally-empty rows (but keep total rows even if zero)
        if not is_total and not any(v for v in (cf, cg, c_d, c_e)):
            continue

        sort += 1
        rows.append({
            "product":         "" if is_total else cur_product,
            "quality_grade":   grade,
            "section":         "",
            "sort_order":      sort,
            "order_qty":       cf,
            "prodn":           None,
            "actual_despatch": cg,
            "abp_month":       c_d,    # preview cross-check only
            "unit":            "T",
            "cell":            f"R{r}C6/C7",
            "status":          "total" if is_total else "ok",
        })

    ok_count = sum(1 for r in rows if r["status"] == "ok")
    logger.info(
        "BSP Special Steel preview: %d ok rows + %d total rows for %s",
        ok_count, len(rows) - ok_count, db_month,
    )

    return {
        "plant":              "BSP",
        "month":              db_month,
        "source_type":        "BSP Special Steel Report",
        "sheets":             "CORP",
        "workbook_sheets":    wb.sheetnames,
        "production_rows":    [],
        "techno_rows":        [],
        "techno_param_rows":  [],
        "special_steel_rows": rows,
    }


# ---------------------------------------------------------------------------
# Section 2 helpers — 3-page-Tech
# ---------------------------------------------------------------------------

def _detect_3page_cum_col(ws) -> int:
    """Scan row 4 for a known cumulative header; return its 1-based column."""
    for c in range(1, 30):
        v = str(ws.cell(4, c).value or "").strip().upper()
        if v in _TECHNO_3PAGE_CUM_HEADERS:
            return c
    return _TECHNO_3PAGE_CUM_FALLBACK


def _detect_3page_month_from_file(ws) -> Optional[str]:
    """Read month name from row 3, col A. Returns 'MM' string or None."""
    raw = ws.cell(3, 1).value
    if not raw:
        return None
    key = str(raw).strip().upper()
    return _MONTH_NAME_TO_NUM.get(key) or _MONTH_NAME_TO_NUM.get(key[:3])


def _extract_techno_3page_preview(wb, file_path: str, report_month: str) -> dict:
    """Extract BSP techno params from 3-page-Tech.xlsx or S2 sheet of PPC MIS."""
    if "Sheet1" in wb.sheetnames:
        ws = wb["Sheet1"]
    elif "S2" in wb.sheetnames:
        ws = wb["S2"]
    else:
        ws = wb.active

    m_num: Optional[str] = None
    y_str: Optional[str] = None

    if report_month and len(report_month) >= 7 and report_month[4] == "-":
        y_str = report_month[:4]
        m_num = report_month[5:7]
    else:
        m_num = _detect_3page_month_from_file(ws)
        import datetime
        y_str = str(datetime.date.today().year)

    if not m_num or m_num not in _MONTH_NUM_TO_COL:
        raise ValueError(
            f"Cannot determine a valid report month (got m_num={m_num!r}). "
            "Please provide a valid month in YYYY-MM format."
        )

    db_month    = f"{y_str}-{m_num}"
    month_col   = _MONTH_NUM_TO_COL[m_num]
    month_label = _MONTH_FULL.get(m_num, m_num)

    expected_hdr = _MONTH_NUM_TO_HDR.get(m_num, "")
    actual_hdr   = str(ws.cell(4, month_col).value or "").strip().upper()
    header_ok    = actual_hdr == expected_hdr
    if not header_ok:
        logger.warning(
            "BSP 3-page-Tech: expected row-4 header %r in col %s but found %r. Proceeding.",
            expected_hdr, get_column_letter(month_col), actual_hdr,
        )

    cum_col     = _detect_3page_cum_col(ws)
    cum_col_ltr = get_column_letter(cum_col)
    mon_col_ltr = get_column_letter(month_col)

    rows_out: List[Dict[str, Any]] = []
    for sort_idx, (row_1b, group, section, param, unit) in enumerate(PARAM_MAP_3PAGE):
        actual_val = _clean(ws.cell(row_1b, month_col).value)
        cum_val    = _clean(ws.cell(row_1b, cum_col).value)
        file_label = str(ws.cell(row_1b, 1).value or "").strip()
        cell_ref   = f"{mon_col_ltr}{row_1b}/{cum_col_ltr}{row_1b}"
        status     = "ok" if (actual_val is not None or cum_val is not None) else "skip"

        rows_out.append({
            "group_code": group, "section": section, "parameter": param,
            "unit": unit, "actual": actual_val, "cum_actual": cum_val,
            "sort_order": sort_idx * 10, "cell": cell_ref,
            "file_label": file_label, "plant": "BSP", "month": db_month,
            "found_via": f"hardcoded R{row_1b}", "status": status,
        })

    # Calculate TMI as HM Consumption + Scrap Consumption
    rows_out = calculate_tmi_consumption(rows_out)

    # Calculate Sinter % in Burden and Pellet % in Burden
    from techno_calc_utils import calculate_burden_percentages
    calculate_burden_percentages(rows_out, db_month, plant_name="BSP")

    ok_count = sum(1 for r in rows_out if r["status"] == "ok")
    logger.info(
        "BSP 3-page-Tech preview: %d/%d rows ok for %s (col %s, cum col %s)",
        ok_count, len(rows_out), db_month, mon_col_ltr, cum_col_ltr,
    )

    return {
        "source_type":        "BSP-3-page-Tech.xlsx",
        "month":              db_month,
        "plant":              "BSP",
        "workbook_sheets":    wb.sheetnames,
        "month_col_letter":   mon_col_ltr,
        "cum_col_letter":     cum_col_ltr,
        "header_verified":    header_ok,
        "file_month":         month_label,
        "production_rows":    [],
        "techno_rows":        [],
        "techno_param_rows":  rows_out,
        "special_steel_rows": [],
    }


# ---------------------------------------------------------------------------
# Section 3 helpers — OISCO
# ---------------------------------------------------------------------------

def _detect_oisco_month_from_title(ws) -> Optional[str]:
    """Parse month abbreviation from title cell R3C3, e.g. '...PARAMETERS_MAY'25'."""
    raw = str(ws.cell(3, 3).value or "").upper()
    m = re.search(r"_([A-Z]{3})'?\d{2}", raw)
    if m:
        return _MONTH_NAME_TO_NUM.get(m.group(1))
    for abbr, num in _MONTH_NAME_TO_NUM.items():
        if len(abbr) == 3 and abbr in raw:
            return num
    return None


def _detect_oisco_columns(ws, m_num: str):
    """Scan row 6 for month abbreviation and CUM header; return (data_col, cum_col) 1-based.

    The OISCO layout is: ... | APR | MAY | ... | CUM. |
    For earlier months (e.g. April), unused month columns exist between the
    data column and the CUM. column, so cum_col = data_col + 1 would land on an
    empty month column, not on CUM.  We therefore scan explicitly for CUM.
    """
    expected = _MONTH_NUM_TO_HDR.get(m_num, "")
    data_col = None
    cum_col  = None
    for c in range(1, ws.max_column + 2):
        v = str(ws.cell(_OISCO_HEADER_ROW, c).value or "").strip().upper()
        if v == expected:
            data_col = c
        if v.startswith("CUM"):          # "CUM." / "CUMULATIVE"
            cum_col = c
    if data_col is None:
        found = [
            str(ws.cell(_OISCO_HEADER_ROW, c).value or "").strip()
            for c in range(1, ws.max_column + 1)
            if ws.cell(_OISCO_HEADER_ROW, c).value
        ]
        raise ValueError(
            f"Month header '{expected}' not found in row {_OISCO_HEADER_ROW}. "
            f"Found: {found}. Ensure the file is for the selected month."
        )
    if cum_col is None:
        cum_col = data_col + 1           # fallback: adjacent column
    return data_col, cum_col


def _extract_oisco_preview(wb, report_month: str) -> dict:
    """Extract BSP OISCO techno params from OISCO_<Mon>'YY.xlsx."""
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active

    m_num: Optional[str] = None
    y_str: Optional[str] = None

    if report_month and len(report_month) >= 7 and report_month[4] == "-":
        y_str = report_month[:4]
        m_num = report_month[5:7]
    else:
        m_num = _detect_oisco_month_from_title(ws)
        import datetime
        y_str = str(datetime.date.today().year)

    if not m_num or m_num not in _MONTH_NUM_TO_HDR:
        raise ValueError(
            f"Cannot determine report month (got {m_num!r}). "
            "Provide a valid YYYY-MM month."
        )

    db_month   = f"{y_str}-{m_num}"
    month_full = _MONTH_FULL.get(m_num, m_num)

    data_col, cum_col = _detect_oisco_columns(ws, m_num)
    data_ltr = get_column_letter(data_col)
    cum_ltr  = get_column_letter(cum_col)

    file_m_num = _detect_oisco_month_from_title(ws)
    header_ok  = (file_m_num == m_num)

    # Check if extracting for a month prior to the file's report month
    extracting_prior_month = False
    if file_m_num and m_num and file_m_num != m_num:
        file_m_int = int(file_m_num)
        ext_m_int = int(m_num)
        # If extraction month < file month, it's a prior month (same FY)
        # Cumulative in the file (column Q) is till report month, not extraction month
        extracting_prior_month = ext_m_int < file_m_int  # e.g., extracting 02 from 03 file

        if extracting_prior_month:
            logger.warning(
                "OISCO: Extracting data for month %r from file with report month %r. "
                "Cumulative data will be skipped (cumulative in file is till report month %r, not %r).",
                _MONTH_FULL.get(m_num, m_num),
                _MONTH_FULL.get(file_m_num, file_m_num),
                _MONTH_FULL.get(file_m_num, file_m_num),
                _MONTH_FULL.get(m_num, m_num),
            )
        else:
            logger.warning(
                "OISCO: file title month %r does not match user-selected %r. Proceeding.",
                file_m_num, m_num,
            )

    rows_out: List[Dict[str, Any]] = []
    for sort_idx, (row_1b, group, section, param, unit) in enumerate(PARAM_MAP_OISCO):
        actual_val = _clean(ws.cell(row_1b, data_col).value)
        # Skip cumulative when extracting for a prior month (cumulative in file is till report month)
        cum_val    = None if extracting_prior_month else _clean(ws.cell(row_1b, cum_col).value)
        file_label = str(ws.cell(row_1b, _OISCO_LABEL_COL).value or "").strip().replace("\n", " ")
        cell_ref   = f"{data_ltr}{row_1b}/{cum_ltr}{row_1b}"
        status     = "ok" if (actual_val is not None or cum_val is not None) else "skip"

        rows_out.append({
            "group_code": group, "section": section, "parameter": param,
            "unit": unit, "actual": actual_val, "cum_actual": cum_val,
            "sort_order": sort_idx * 10, "cell": cell_ref,
            "file_label": file_label, "plant": "BSP", "month": db_month,
            "found_via": f"hardcoded R{row_1b}", "status": status,
        })

    ok_count = sum(1 for r in rows_out if r["status"] == "ok")
    logger.info(
        "BSP OISCO preview: %d/%d rows ok for %s (col %s, cum col %s)",
        ok_count, len(rows_out), db_month, data_ltr, cum_ltr,
    )

    return {
        "source_type":        "BSP-OISCO-Techno.xlsx",
        "month":              db_month,
        "plant":              "BSP",
        "workbook_sheets":    wb.sheetnames,
        "month_col_letter":   data_ltr,
        "cum_col_letter":     cum_ltr,
        "header_verified":    header_ok,
        "file_month":         month_full,
        "production_rows":    [],
        "techno_rows":        [],
        "techno_param_rows":  rows_out,
        "special_steel_rows": [],
    }


# ---------------------------------------------------------------------------
# Section 1b — BSP PPC MIS preview (S1 production + S2 techno)
# ---------------------------------------------------------------------------

def _extract_ppc_mis_preview(file_path: str, report_month: str) -> dict:
    """Preview BSP PPC MIS monthly report (S1 sheet): production + opening stock.

    Supports both .xls (xlrd) and .xlsx (openpyxl) formats.
    """
    import datetime as _dt

    is_xls = file_path.lower().endswith(".xls")

    if is_xls:
        wb_raw = xlrd.open_workbook(file_path)
        ws_s1  = wb_raw.sheet_by_name("S1")
        wb_sheets = wb_raw.sheet_names()

        n1_raw = ws_s1.cell_value(0, 13)
        if n1_raw and isinstance(n1_raw, float) and n1_raw > 0:
            y, m, *_ = xlrd.xldate_as_tuple(n1_raw, wb_raw.datemode)
            db_month = f"{y}-{m:02d}"
        elif report_month:
            db_month = report_month
        else:
            db_month = "unknown"

        def _cv(r0, c0):        # 0-based
            return ws_s1.cell_value(r0, c0)

    else:
        wb_raw = openpyxl.load_workbook(file_path, data_only=True)
        ws_s1  = wb_raw["S1"]
        wb_sheets = wb_raw.sheetnames

        n1_raw = ws_s1.cell(row=1, column=14).value
        if isinstance(n1_raw, (_dt.datetime, _dt.date)):
            db_month = f"{n1_raw.year}-{n1_raw.month:02d}"
        elif report_month:
            db_month = report_month
        else:
            db_month = "unknown"

        def _cv(r0, c0):        # 0-based → openpyxl 1-based
            v = ws_s1.cell(row=r0 + 1, column=c0 + 1).value
            return None if v == "" else v

    month_mismatch = bool(
        report_month and db_month not in ("unknown", report_month)
    )
    if month_mismatch:
        logger.warning(
            "BSP PPC MIS: file month %s ≠ selected month %s — file month will be used",
            db_month, report_month,
        )
    else:
        logger.info("BSP PPC MIS preview: month → %s", db_month)

    production_rows = []
    for item_name, (row_0, col_0, do_convert) in _ppc_mis_config().items():
        raw = _cv(row_0, col_0)
        val = _clean(raw)
        if val is not None and do_convert:
            val = round(val / 1000.0, 3)
        unit = "nos/d" if not do_convert else "'000T"
        cell_ref = f"{get_column_letter(col_0 + 1)}{row_0 + 1}"
        production_rows.append({
            "item_name": item_name,
            "value": val,
            "unit": unit,
            "cell": f"S1!{cell_ref}",
            "pdf_label": cell_ref,
            "status": "ok" if val is not None else "skip",
        })

    ok_prod = sum(1 for r in production_rows if r["status"] == "ok")
    logger.info("BSP PPC MIS preview: %d/%d production rows ok for %s", ok_prod, len(production_rows), db_month)

    # ── Opening stock extraction (sheet S1, closing stock = next month) ───────
    def _next_month_str(m):
        y, mo = int(m[:4]), int(m[5:7])
        return f"{y+1 if mo == 12 else y}-{1 if mo == 12 else mo+1:02d}"

    stock_month = _next_month_str(db_month) if db_month != "unknown" else "unknown"

    N47 = _clean(_cv(46, 13))   # Slabs saleable stock (FOR SALE)
    K47 = _clean(_cv(46, 10))   # Slabs gross stock (INPROCESS + FOR SALE)
    N50 = _clean(_cv(49, 13))   # Total semis FOR SALE (Slabs + BB)
    K50 = _clean(_cv(49, 10))   # Total semis gross
    N58 = _clean(_cv(57, 13))   # Finished Steel stock
    O63 = _clean(_cv(62, 14))   # Pig Iron stock

    def _t(v):
        return round(v / 1000, 3) if v is not None else None

    sl_b = _t(N47)
    sl_a = _t(K47 - N47)              if K47 is not None and N47 is not None else None
    bb_b = _t(N50 - N47)              if N50 is not None and N47 is not None else None
    bb_a = _t(K50 - N50 - K47 + N47) if all(x is not None for x in [K50, N50, K47, N47]) else None

    stock_rows = [
        {"plant": "BSP", "item_type": "SLABS",          "stock_type": "FOR SALE",  "stock_month": stock_month, "value": sl_b, "formula": "N47",            "status": "ok" if sl_b is not None else "skip"},
        {"plant": "BSP", "item_type": "SLABS",          "stock_type": "INPROCESS", "stock_month": stock_month, "value": sl_a, "formula": "K47-N47",        "status": "ok" if sl_a is not None else "skip"},
        {"plant": "BSP", "item_type": "BLOOM/BILLETS",  "stock_type": "FOR SALE",  "stock_month": stock_month, "value": bb_b, "formula": "N50-N47",        "status": "ok" if bb_b is not None else "skip"},
        {"plant": "BSP", "item_type": "BLOOM/BILLETS",  "stock_type": "INPROCESS", "stock_month": stock_month, "value": bb_a, "formula": "K50-N50-K47+N47","status": "ok" if bb_a is not None else "skip"},
        {"plant": "BSP", "item_type": "FINISHED STEEL", "stock_type": "",          "stock_month": stock_month, "value": _t(N58), "formula": "N58",         "status": "ok" if N58  is not None else "skip"},
        {"plant": "BSP", "item_type": "PIG IRON",       "stock_type": "",          "stock_month": stock_month, "value": _t(O63), "formula": "O63",         "status": "ok" if O63  is not None else "skip"},
    ]
    ok_stock = sum(1 for r in stock_rows if r["status"] == "ok")
    logger.info("BSP PPC MIS preview: %d stock rows ok for stock_month %s", ok_stock, stock_month)

    return {
        "source_type":        "BSP PPC MIS Monthly Report",
        "month":              db_month,
        "plant":              "BSP",
        "workbook_sheets":    wb_sheets,
        "month_mismatch":     month_mismatch,
        "selected_month":     report_month or "",
        "production_rows":    production_rows,
        "techno_rows":        [],
        "techno_param_rows":  [],
        "special_steel_rows": [],
        "stock_rows":         stock_rows,
    }


# ---------------------------------------------------------------------------
# Section 5 — BSP MIS-2 month-end report: furnace-wise Hot Metal production
# (tentative). On a month-end-dated report the column-D "CUM" figure is the
# for-the-month production. Feeds production_table item_names "BF-1".."BF-8"
# (via the shared /api/confirm-extraction normalize_item_name step, which
# rewrites "BF-" → "BF#") so techno_cumulative's per-furnace weighting can
# read real furnace-wise weights instead of falling back to the techno page's
# own 'production' field. BF-8 intentionally targets the SAME item_name
# ("BF#8") already written by the PPC MIS upload — the preview screen shows
# the current DB value alongside this tentative one so the user can decide
# whether to let it stand until the PPC MIS upload replaces it. The shop
# total ('BF 1-8(TOTAL)') is NOT extracted here: BF_Shop's own weighting
# already reads the existing 'Hot Metal' production_table item.
# ---------------------------------------------------------------------------

_MIS2_FURNACE_LABELS = ["BF-1", "BF-4", "BF-5", "BF-6", "BF-7", "BF-8"]
_MIS2_PRODUCTION_COL = 4   # column D — CUM

# Merchant Mill's two product groups — always printed as their own rows
# ("MM Rods & Bar" / "MM Strls.", column D = CUM) directly below the furnace
# block. "MM" (the mill total already fed by the PPC MIS extractor above) is
# re-derived here as their sum on every MIS-2 upload: verified against real
# data that MM Rods & Bar + MM Strls always equals the existing MM total
# exactly, so this is a strictly more granular replacement, not a new figure.
_MIS2_MM_LABELS = {
    "MM RODS & BAR": "TMT BARS(MM)",
    "MM RODS":       "TMT BARS(MM)",   # older (~2015-2019) label, no "& Bar"
    "MM STRLS.":     "LT STRS(MM)",
}

# Two more figures unique to this report (not derivable from anything the
# PPC MIS extractor already covers — verified against real data that
# "Pig Iron Prodn." and "BF 1-8(TOTAL)" on this same sheet exactly match the
# existing "Pig Iron"/"Hot Metal" items, so those two are redundant and
# deliberately NOT added here).
_MIS2_EXTRA_LABELS = {
    "COB #11":                 "COB#11",
    "SP - III (M/C-2 PRODN.)":  "SP-3 M/C-2",
}


def _is_mis2_file(wb) -> bool:
    """True if row 2 (any of the first few cells) reads 'BSP MIS-2'."""
    ws = wb.worksheets[0]
    for c in range(1, 8):
        v = ws.cell(2, c).value
        if isinstance(v, str) and v.strip().upper() == "BSP MIS-2":
            return True
    return False


def _parse_mis2_date(ws) -> Optional[tuple]:
    """Row 2: 'Date:' | day | MONTH | year (e.g. 'Date:' 31 'MAY' 2026) → (y, m, d)."""
    for c in range(1, 10):
        v = ws.cell(2, c).value
        if isinstance(v, str) and v.strip().rstrip(":").upper() == "DATE":
            rest = [ws.cell(2, c2).value for c2 in range(c + 1, c + 5)]
            rest = [x for x in rest if x not in ("", None)][:3]
            try:
                day = int(float(rest[0]))
                mon = _MONTH_NAME_TO_NUM[str(rest[1]).strip().upper()[:3]]
                year = int(float(rest[2]))
                return year, int(mon), day
            except (IndexError, KeyError, ValueError, TypeError):
                return None
    return None


def _extract_mis2_furnace_preview(wb, report_month: str) -> dict:
    """Preview BSP MIS-2 furnace-wise Hot Metal production (column D, CUM)."""
    ws = wb.worksheets[0]

    parsed = _parse_mis2_date(ws)
    if not parsed:
        raise ValueError(
            "Cannot read the report date from the MIS-2 header (row 2 should "
            "contain 'Date:' day MONTH year, e.g. Date: 31 MAY 2026)."
        )
    y, m, d = parsed
    db_month = f"{y}-{m:02d}"
    month_mismatch = bool(report_month and db_month != report_month)
    if month_mismatch:
        logger.warning(
            "BSP MIS-2: file month %s != selected month %s — file month will be used",
            db_month, report_month,
        )

    # item_name uses the "BF#N" spelling (not "BF-N") to match the existing
    # BSP production_table convention directly — main.py's normalize_item_name
    # only runs at confirm/insert time, not during preview, so emitting "BF-8"
    # here would fail to match the existing "BF#8" row's db_value on the
    # preview screen even though it would merge correctly on confirm.
    production_rows: List[Dict[str, Any]] = []
    found = set()
    mm_values: Dict[str, Optional[float]] = {}
    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(r, 2).value or "").strip().upper()   # column B
        if label in _MIS2_FURNACE_LABELS and label not in found:
            found.add(label)
            val = _clean(ws.cell(r, _MIS2_PRODUCTION_COL).value)
            val_000t = round(val / 1000.0, 3) if val is not None else None
            production_rows.append({
                "item_name": label.replace("BF-", "BF#"),
                "value":     val_000t,
                "unit":      "'000T",
                "cell":      f"{ws.title}!D{r}",
                "pdf_label": label,
                "status":    "ok" if val_000t is not None else "skip",
            })
        elif label in _MIS2_MM_LABELS and _MIS2_MM_LABELS[label] not in mm_values:
            # Two label spellings ("MM Rods & Bar" vs. the older "MM Rods")
            # map to the same item — dedupe on the target item, not the label,
            # so a file using the older wording doesn't also get a spurious
            # "skip" row for the newer label further down.
            item_name = _MIS2_MM_LABELS[label]
            val = _clean(ws.cell(r, _MIS2_PRODUCTION_COL).value)
            val_000t = round(val / 1000.0, 3) if val is not None else None
            mm_values[item_name] = val_000t
            production_rows.append({
                "item_name": item_name,
                "value":     val_000t,
                "unit":      "'000T",
                "cell":      f"{ws.title}!D{r}",
                "pdf_label": label,
                "status":    "ok" if val_000t is not None else "skip",
            })
        elif label in _MIS2_EXTRA_LABELS and label not in found:
            found.add(label)
            item_name = _MIS2_EXTRA_LABELS[label]
            val = _clean(ws.cell(r, _MIS2_PRODUCTION_COL).value)
            val_000t = round(val / 1000.0, 3) if val is not None else None
            production_rows.append({
                "item_name": item_name,
                "value":     val_000t,
                "unit":      "'000T",
                "cell":      f"{ws.title}!D{r}",
                "pdf_label": label,
                "status":    "ok" if val_000t is not None else "skip",
            })
    for label in _MIS2_FURNACE_LABELS:
        if label not in found:
            production_rows.append({
                "item_name": label.replace("BF-", "BF#"), "value": None, "unit": "'000T",
                "cell": "", "pdf_label": label, "status": "skip",
            })
    for item_name in dict.fromkeys(_MIS2_MM_LABELS.values()):
        if item_name not in mm_values:
            production_rows.append({
                "item_name": item_name, "value": None, "unit": "'000T",
                "cell": "", "pdf_label": item_name, "status": "skip",
            })
    for label, item_name in _MIS2_EXTRA_LABELS.items():
        if label not in found:
            production_rows.append({
                "item_name": item_name, "value": None, "unit": "'000T",
                "cell": "", "pdf_label": label, "status": "skip",
            })

    # "MM" (mill total) re-derived as the sum of the two product groups above
    # — only when both are present, so a partial/garbled read doesn't silently
    # understate the total. Compare against the number of unique *target*
    # items, not label count — "MM Rods"/"MM Rods & Bar" are two label
    # spellings for the same one target ("TMT BARS(MM)"), so len() of the
    # raw label dict overcounts and would never be satisfied.
    _mm_target_count = len(set(_MIS2_MM_LABELS.values()))
    if len(mm_values) == _mm_target_count and all(v is not None for v in mm_values.values()):
        production_rows.append({
            "item_name": "MM",
            "value":     round(sum(mm_values.values()), 3),
            "unit":      "'000T",
            "cell":      "derived: TMT BARS(MM) + LT STRS(MM)",
            "pdf_label": "MM (derived)",
            "status":    "ok",
        })

    ok = sum(1 for r in production_rows if r["status"] == "ok")
    logger.info("BSP MIS-2 furnace preview: %d/%d furnace rows ok for %s",
                ok, len(production_rows), db_month)

    return {
        "source_type":        "BSP MIS-2 Month-End (Furnace-wise Hot Metal, tentative)",
        "month":              db_month,
        "plant":              "BSP",
        "workbook_sheets":    wb.sheetnames,
        "month_mismatch":     month_mismatch,
        "selected_month":     report_month or "",
        "production_rows":    production_rows,
        "techno_rows":        [],
        "techno_param_rows":  [],
        "special_steel_rows": [],
        "stock_rows":         [],
    }


# ---------------------------------------------------------------------------
# Unified preview entry point (auto-detects file type)
# ---------------------------------------------------------------------------

def extract_preview(file_path: str, report_month: str) -> dict:
    """Unified BSP preview — auto-detects file type. No DB writes.

    Detection priority:
      1. BSP PPC MIS       → sheet 'S1' present (.xls monthly report)
      2. BSP MIS-2         → row 2 reads 'BSP MIS-2' (furnace-wise Hot Metal, tentative)
      3. BSP Special Steel → sheet 'CORP' with 'BHILAI STEEL PLANT' in R3C1
      4. OISCO Techno      → R3C3 contains 'TECHNO ECONOMIC PARAMETERS'
      5. BSP 3-page-Tech   → default (Sheet1 with month name in R3C1)
    """
    wb = _open_workbook(file_path)

    if "S1" in wb.sheetnames:
        logger.info("BSP: detected PPC MIS file (sheet S1) — production + stock")
        return _extract_ppc_mis_preview(file_path, report_month)

    if _is_mis2_file(wb):
        logger.info("BSP: detected MIS-2 file — furnace-wise Hot Metal production (tentative)")
        return _extract_mis2_furnace_preview(wb, report_month)

    if _is_bsp_ss_file(wb):
        logger.info("BSP: detected Special Steel file (sheet CORP)")
        return _extract_bsp_ss_preview(wb, report_month)

    if _is_oisco_file(wb):
        logger.info("BSP: detected OISCO Techno file")
        return _extract_oisco_preview(wb, report_month)

    logger.info("BSP: detected 3-page-Tech file (default)")
    return _extract_techno_3page_preview(wb, file_path, report_month)
