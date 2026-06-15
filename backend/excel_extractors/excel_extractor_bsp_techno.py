"""BSP Techno-Economic Parameters extractor — BSP-3-page-Tech.xlsx

Sheet layout (single sheet "Sheet1"):
  Row  3, Col A : Report month name, e.g. "MAY"  (auto-detection)
  Row  4, Cols D-O : Column month headers APR … MAR
  Row  4, Col P : "ACTUAL" header
  Col  B (2)    : Unit for each parameter row
  Col  D  (4)  = APR    Col  E  (5)  = MAY    Col  F  (6)  = JUN
  Col  G  (7)  = JUL    Col  H  (8)  = AUG    Col  I  (9)  = SEP
  Col  J  (10) = OCT    Col  K  (11) = NOV    Col  L  (12) = DEC
  Col  M  (13) = JAN    Col  N  (14) = FEB    Col  O  (15) = MAR
  Col  P  (16) = Cumulative "till month" (Apr → report month) — ALWAYS

Parameter rows are hardcoded (validated per user mapping):
  Coke & By-products    → COKE_SINTER / Coke Yield
  Sinter Plants 2 & 3   → COKE_SINTER / Sinter Plant SP-2, Sinter Plant SP-3
  Blast Furnaces        → IRON_MAKING / Blast Furnaces, BF Productivity (BSP)
  SMS-2, SMS-3          → BOF / SMS-II Consumption, SMS-III Consumption
  RSM, URM, MM, WRM, BRM, Plate Mill → MILL_BSP / {mill name}
  Energy                → MILL_BSP / Energy
"""

import logging
import os
from typing import Optional, List, Dict, Any
import openpyxl
from openpyxl.utils import get_column_letter

try:
    import xlrd
    _XLRD_AVAILABLE = True
except ImportError:
    _XLRD_AVAILABLE = False

logger = logging.getLogger("excel_extractor")


# ---------------------------------------------------------------------------
# Excel 97-2003 (.xls) compatibility layer
# ---------------------------------------------------------------------------

class _XlsCell:
    __slots__ = ("value",)
    def __init__(self, value):
        self.value = None if value == "" else value


class _XlsSheet:
    """Makes an xlrd sheet look like an openpyxl worksheet (1-based cell access)."""
    def __init__(self, sheet):
        self._s = sheet
        self.max_column = sheet.ncols

    def cell(self, row: int, col: int) -> _XlsCell:
        r, c = row - 1, col - 1
        if r < 0 or r >= self._s.nrows or c < 0 or c >= self._s.ncols:
            return _XlsCell(None)
        return _XlsCell(self._s.cell_value(r, c))


class _XlsWorkbook:
    """Makes an xlrd workbook look like an openpyxl workbook."""
    def __init__(self, wb):
        self._wb = wb

    @property
    def sheetnames(self):
        return self._wb.sheet_names()

    def __getitem__(self, name: str) -> _XlsSheet:
        return _XlsSheet(self._wb.sheet_by_name(name))

    @property
    def active(self) -> _XlsSheet:
        return _XlsSheet(self._wb.sheets()[0])


def _open_workbook(file_path: str):
    """Open .xlsx (openpyxl) or .xls (xlrd) and return a unified workbook interface."""
    if file_path.lower().endswith(".xls"):
        if not _XLRD_AVAILABLE:
            raise ImportError(
                "xlrd is required for Excel 97-2003 (.xls) files. "
                "Install it with: pip install xlrd"
            )
        return _XlsWorkbook(xlrd.open_workbook(file_path))
    return openpyxl.load_workbook(file_path, data_only=True)

# ---------------------------------------------------------------------------
# Month / column constants
# ---------------------------------------------------------------------------
_MONTH_NAME_TO_NUM = {
    "JAN": "01", "FEB": "02", "MAR": "03", "APR": "04",
    "MAY": "05", "JUN": "06", "JUL": "07", "AUG": "08",
    "SEP": "09", "OCT": "10", "NOV": "11", "DEC": "12",
    "JANUARY": "01", "FEBRUARY": "02", "MARCH": "03", "APRIL": "04",
    "MAY": "05", "JUNE": "06", "JULY": "07", "AUGUST": "08",
    "SEPTEMBER": "09", "OCTOBER": "10", "NOVEMBER": "11", "DECEMBER": "12",
}

# report month-number (MM) → 1-based Excel column index
_MONTH_NUM_TO_COL: Dict[str, int] = {
    "04": 4, "05": 5, "06": 6, "07": 7, "08": 8, "09": 9,
    "10": 10, "11": 11, "12": 12, "01": 13, "02": 14, "03": 15,
}

# Expected header text in row 4 for each month column
_MONTH_NUM_TO_HDR: Dict[str, str] = {
    "04": "APR", "05": "MAY", "06": "JUN", "07": "JUL",
    "08": "AUG", "09": "SEP", "10": "OCT", "11": "NOV",
    "12": "DEC", "01": "JAN", "02": "FEB", "03": "MAR",
}

_MONTH_FULL: Dict[str, str] = {
    "01": "January", "02": "February", "03": "March", "04": "April",
    "05": "May", "06": "June", "07": "July", "08": "August",
    "09": "September", "10": "October", "11": "November", "12": "December",
}

_CUM_COL_FALLBACK = 16   # column P: used when header scan finds nothing
_CUM_HEADERS = {"ACTUAL", "CUM", "CUMULATIVE", "CUM.", "TILL DATE", "TILL MONTH", "APR-ACTUAL"}


# ---------------------------------------------------------------------------
# Parameter map
# (row_1based, group_code, section, parameter_label, unit)
# sort_order = index × 10 (assigned dynamically in extract_preview)
# ---------------------------------------------------------------------------
PARAM_MAP: List[tuple] = [
    # ── Coke & By-products (rows 37-40) ──────────────────────────────────
    (37, "COKE_SINTER", "Coke Yield",           "BF Coke",                   "%"),
    (38, "COKE_SINTER", "Coke Yield",           "Crude Tar",                 "%"),
    (39, "COKE_SINTER", "Coke Yield",           "Ammonium Sulphate",         "%"),
    (40, "COKE_SINTER", "Coke Yield",           "Crude Benzol",              "%"),
    # ── Sinter Plant 2 (rows 45-49) ──────────────────────────────────────
    (45, "COKE_SINTER", "Sinter Plant SP-2",    "Machine Availability",      "%"),
    (46, "COKE_SINTER", "Sinter Plant SP-2",    "Machine Utilisation",       "%"),
    (48, "COKE_SINTER", "Sinter Plant SP-2",    "Productivity",              "T/m2/hr"),
    (49, "COKE_SINTER", "Sinter Plant SP-2",    "Basicity",                  ""),
    # ── Sinter Plant 3 (rows 51-55) ──────────────────────────────────────
    (51, "COKE_SINTER", "Sinter Plant SP-3",    "Machine Availability",      "%"),
    (52, "COKE_SINTER", "Sinter Plant SP-3",    "Machine Utilisation",       "%"),
    (54, "COKE_SINTER", "Sinter Plant SP-3",    "Productivity",              "T/m2/hr"),
    (55, "COKE_SINTER", "Sinter Plant SP-3",    "Basicity",                  ""),
    # ── Blast Furnaces — burden & rates (rows 58-60, 68-70) ──────────────
    (58, "IRON_MAKING", "Blast Furnaces",        "Sinter in Burden (BF 1-8)","%" ),
    (59, "IRON_MAKING", "Blast Furnaces",        "Coke Rate (BF 1-8)",       "Kg./THM"),
    (60, "IRON_MAKING", "Blast Furnaces",        "Coke Screening (BF 1-8)",  "25mm%"),
    (68, "IRON_MAKING", "Blast Furnaces",        "CDI Rate",                 "Kg./THM"),
    (69, "IRON_MAKING", "Blast Furnaces",        "Slag Rate (1-8)",          "Kg./THM"),
    (70, "IRON_MAKING", "Blast Furnaces",        "Sp. Nut Coke Consumption", "Kg./THM"),
    # ── BF Productivity (rows 65-67) ────────────────────────────────────
    (65, "IRON_MAKING", "BF Productivity (BSP)", "BF-7 (2104 Cu.M)",         "T/m3/day"),
    (66, "IRON_MAKING", "BF Productivity (BSP)", "BF-8 (3445 Cu.M)",         "T/m3/day"),
    (67, "IRON_MAKING", "BF Productivity (BSP)", "Overall (BF 1-8)",         "T/m3/day"),
    # ── SMS-II Consumption (rows 76-78) ─────────────────────────────────
    (76, "BOF", "SMS-II Consumption",            "Hot Metal",                "Kg/T CS"),
    (77, "BOF", "SMS-II Consumption",            "Scrap",                    "Kg/T CS"),
    (78, "BOF", "SMS-II Consumption",            "Total Metallic Charge",    "Kg/T CS"),
    # ── SMS-III Consumption (rows 90-92) ────────────────────────────────
    (90, "BOF", "SMS-III Consumption",           "Hot Metal",                "Kg/T CS"),
    (91, "BOF", "SMS-III Consumption",           "Scrap",                    "Kg/T CS"),
    (92, "BOF", "SMS-III Consumption",           "Total Metallic Charge",    "Kg/T CS"),
    # ── Rail & Structural Mill — RSM (rows 97-100, 137, 153) ─────────────
    ( 97, "MILL_BSP", "Rail & Structural Mill",  "Yield",                    "%"),
    ( 98, "MILL_BSP", "Rail & Structural Mill",  "Rolling Rate",             "T/Hr."),
    ( 99, "MILL_BSP", "Rail & Structural Mill",  "Mill Availability",        "%"),
    (100, "MILL_BSP", "Rail & Structural Mill",  "Mill Utilisation",         "%"),
    (137, "MILL_BSP", "Rail & Structural Mill",  "Heat Consumption",         "103Kcal/T"),
    (153, "MILL_BSP", "Rail & Structural Mill",  "Power Consumption",        "Kwh/T"),
    # ── Universal Rail Mill — URM (rows 103-106, 138, 154) ───────────────
    (103, "MILL_BSP", "Universal Rail Mill",     "Yield",                    "%"),
    (104, "MILL_BSP", "Universal Rail Mill",     "Rolling Rate",             "T/Hr."),
    (105, "MILL_BSP", "Universal Rail Mill",     "Mill Availability",        "%"),
    (106, "MILL_BSP", "Universal Rail Mill",     "Mill Utilisation",         "%"),
    (138, "MILL_BSP", "Universal Rail Mill",     "Heat Consumption",         "103Kcal/T"),
    (154, "MILL_BSP", "Universal Rail Mill",     "Power Consumption",        "Kwh/T"),
    # ── Merchant Mill — MM (rows 109-112, 139, 155) ──────────────────────
    (109, "MILL_BSP", "Merchant Mill",           "Yield",                    "%"),
    (110, "MILL_BSP", "Merchant Mill",           "Rolling Rate",             "T/Hr."),
    (111, "MILL_BSP", "Merchant Mill",           "Mill Availability",        "%"),
    (112, "MILL_BSP", "Merchant Mill",           "Mill Utilisation",         "%"),
    (139, "MILL_BSP", "Merchant Mill",           "Heat Consumption",         "103Kcal/T"),
    (155, "MILL_BSP", "Merchant Mill",           "Power Consumption",        "Kwh/T"),
    # ── Wire Rod Mill — WRM (rows 115-118, 140, 156) ─────────────────────
    (115, "MILL_BSP", "Wire Rod Mill",           "Yield",                    "%"),
    (116, "MILL_BSP", "Wire Rod Mill",           "Rolling Rate",             "T/Hr."),
    (117, "MILL_BSP", "Wire Rod Mill",           "Mill Availability",        "%"),
    (118, "MILL_BSP", "Wire Rod Mill",           "Mill Utilisation",         "%"),
    (140, "MILL_BSP", "Wire Rod Mill",           "Heat Consumption",         "103Kcal/T"),
    (156, "MILL_BSP", "Wire Rod Mill",           "Power Consumption",        "Kwh/T"),
    # ── Bar & Rod Mill — BRM (rows 121-124, no heat/power rows) ──────────
    (121, "MILL_BSP", "Bar & Rod Mill",          "Yield",                    "%"),
    (122, "MILL_BSP", "Bar & Rod Mill",          "Rolling Rate",             "T/Hr."),
    (123, "MILL_BSP", "Bar & Rod Mill",          "Mill Availability",        "%"),
    (124, "MILL_BSP", "Bar & Rod Mill",          "Mill Utilisation",         "%"),
    # ── Plate Mill (rows 127-128, 130-131, 141, 157) ─────────────────────
    (127, "MILL_BSP", "Plate Mill",              "Yield",                    "%"),
    (128, "MILL_BSP", "Plate Mill",              "Rolling Rate",             "T/Hr."),
    (130, "MILL_BSP", "Plate Mill",              "Mill Availability",        "%"),
    (131, "MILL_BSP", "Plate Mill",              "Mill Utilisation",         "%"),
    (141, "MILL_BSP", "Plate Mill",              "Heat Consumption",         "103Kcal/T"),
    (157, "MILL_BSP", "Plate Mill",              "Power Consumption",        "Kwh/T"),
    # ── Energy (row 161) ─────────────────────────────────────────────────
    (161, "MILL_BSP", "Energy",                  "Sp. Energy Consumption",   "G.Cal/T"),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def clean_val(v) -> Optional[float]:
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "-", "—", "nan", "###", "#DIV/0!", "#VALUE!", "#N/A"):
        return None
    try:
        return float(s.replace(",", ""))
    except (ValueError, TypeError):
        return None


def _detect_cum_col(ws) -> int:
    """Scan row 4 for a known cumulative header; return its 1-based column index."""
    for c in range(1, 30):
        cell_val = ws.cell(4, c).value
        if cell_val is None:
            continue
        v = str(cell_val).strip().upper()
        if v in _CUM_HEADERS:
            logger.debug("BSP techno: CUM column detected at col %d (header %r)", c, v)
            return c
    logger.warning(
        "BSP techno: CUM column header not found in row 4; falling back to col %d",
        _CUM_COL_FALLBACK,
    )
    return _CUM_COL_FALLBACK


def _detect_month_from_file(ws) -> Optional[str]:
    """Read month name from row 3, col A. Returns 'MM' string or None."""
    raw = ws.cell(3, 1).value
    if not raw:
        return None
    key = str(raw).strip().upper()
    # try full name first, then 3-char abbrev
    m = _MONTH_NAME_TO_NUM.get(key) or _MONTH_NAME_TO_NUM.get(key[:3])
    return m


def _open_sheet(file_path: str):
    """Open workbook (.xlsx or .xls) and return the data sheet."""
    wb = _open_workbook(file_path)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    return wb, ws


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def extract_preview(file_path: str, report_month: str) -> Dict[str, Any]:
    """
    Extract BSP techno parameters from BSP-3-page-Tech.xlsx.

    Args:
        file_path:    Path to the uploaded .xlsx file.
        report_month: 'YYYY-MM' string selected by user.

    Returns:
        Preview dict compatible with /api/confirm-extraction:
          production_rows   = []  (file has no production data)
          techno_rows       = []  (legacy table not used for BSP techno)
          techno_param_rows = [{group_code, section, parameter, unit,
                                actual, cum_actual, sort_order, cell,
                                file_label, plant, month, found_via, status}]
          special_steel_rows = []
    """
    wb, ws = _open_sheet(file_path)

    # ── Determine report month ────────────────────────────────────────────
    m_num: Optional[str] = None
    y_str: Optional[str] = None

    if report_month and len(report_month) >= 7 and report_month[4] == "-":
        y_str = report_month[:4]
        m_num = report_month[5:7]
    else:
        # fall back to file auto-detection
        m_num = _detect_month_from_file(ws)
        import datetime
        y_str = str(datetime.date.today().year)

    if not m_num or m_num not in _MONTH_NUM_TO_COL:
        raise ValueError(
            f"Cannot determine a valid report month (got m_num={m_num!r}). "
            "Please provide a valid month in YYYY-MM format."
        )

    db_month    = f"{y_str}-{m_num}"
    month_col   = _MONTH_NUM_TO_COL[m_num]
    month_label = _MONTH_FULL.get(m_num, m_num)

    # ── Verify column header in row 4 ────────────────────────────────────
    expected_hdr = _MONTH_NUM_TO_HDR.get(m_num, "")
    actual_hdr   = str(ws.cell(4, month_col).value or "").strip().upper()
    header_ok    = actual_hdr == expected_hdr
    if not header_ok:
        logger.warning(
            "BSP techno: expected row-4 header %r in col %s but found %r. "
            "Proceeding anyway — verify the file is for the correct FY.",
            expected_hdr, get_column_letter(month_col), actual_hdr,
        )

    # ── Locate cumulative column dynamically ─────────────────────────────
    cum_col     = _detect_cum_col(ws)
    cum_col_ltr = get_column_letter(cum_col)

    # ── Extract rows ──────────────────────────────────────────────────────
    mon_col_ltr = get_column_letter(month_col)
    rows_out: List[Dict[str, Any]] = []

    for sort_idx, (row_1b, group, section, param, unit) in enumerate(PARAM_MAP):
        actual_val  = clean_val(ws.cell(row_1b, month_col).value)
        cum_val     = clean_val(ws.cell(row_1b, cum_col).value)
        file_label  = str(ws.cell(row_1b, 1).value or "").strip()
        cell_ref    = f"{mon_col_ltr}{row_1b}/{cum_col_ltr}{row_1b}"

        # Row 40 (Crude Benzol) often has "-" — still include with None values
        status = "ok" if (actual_val is not None or cum_val is not None) else "skip"

        rows_out.append({
            "group_code": group,
            "section":    section,
            "parameter":  param,
            "unit":       unit,
            "actual":     actual_val,
            "cum_actual": cum_val,
            "sort_order": sort_idx * 10,
            "cell":       cell_ref,
            "file_label": file_label,
            "plant":      "BSP",
            "month":      db_month,
            "found_via":  f"hardcoded R{row_1b}",
            "status":     status,
        })

    ok_count = sum(1 for r in rows_out if r["status"] == "ok")
    logger.info(
        "BSP techno preview: %d/%d rows ok for %s (col %s, cum col %s)",
        ok_count, len(rows_out), db_month, mon_col_ltr, cum_col_ltr,
    )

    return {
        "source_type":       "BSP-3-page-Tech.xlsx",
        "month":             db_month,
        "plant":             "BSP",
        "workbook_sheets":   wb.sheetnames,
        "month_col_letter":  mon_col_ltr,
        "cum_col_letter":    cum_col_ltr,
        "header_verified":   header_ok,
        "file_month":        month_label,
        "production_rows":   [],
        "techno_rows":       [],
        "techno_param_rows": rows_out,
        "special_steel_rows": [],
    }
