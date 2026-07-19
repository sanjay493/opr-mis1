import re
import openpyxl
from openpyxl.utils import get_column_letter
import logging
import sqlite3
import os
from datetime import datetime
from typing import Optional

logger = logging.getLogger("excel_extractor")

DB_PATH = os.path.join(os.path.dirname(__file__), "..", "mis_reports.db")

MONTH_NAMES = {
    "01": "January", "02": "February", "03": "March", "04": "April",
    "05": "May",     "06": "June",     "07": "July",  "08": "August",
    "09": "September","10": "October", "11": "November","12": "December"
}
MONTHS_MAP = {v: k for k, v in MONTH_NAMES.items()}


def clean_val(val) -> Optional[float]:
    if val is None or str(val).strip().lower() in ("nan", "###", "-", "#div/0!", ""):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _parse_report_month(report_month: str):
    """Returns (db_report_month_yyyymm, month_num_str) from 'YYYY-MM' or legacy 'Month Year'."""
    if len(report_month) == 7 and report_month[4] == "-":
        return report_month, report_month[5:7]
    parts = report_month.split()
    m_name, y_str = parts[0], parts[1]
    m_num = MONTHS_MAP.get(m_name, "01")
    return f"{y_str}-{m_num}", m_num


def extract_and_save_excel(file_path: str, report_month: str = "", source_file_name: str = "") -> bool:
    """
    Dispatcher: auto-detects ISP file type by sheet name.

    Supported file types:
      • Morning Report (.xlsx)       — sheet 'DAILYREPORT1'. Month auto-detected from K5.
      • Final Monthly Report (.xlsx) — sheet 'Maj Production Summ'. Month set manually.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames
        logger.info(f"ISP: loading file. Sheets: {sheet_names}")

        if "DAILYREPORT1" in sheet_names:
            return _extract_morning_report(wb, source_file_name)

        if "Maj Production Summ" in sheet_names:
            return _extract_monthly_report(wb, report_month, source_file_name)

        raise ValueError(
            "Uploaded ISP file does not match any known format. "
            "Expected sheet 'DAILYREPORT1' (Morning Report) or "
            "'Maj Production Summ' (Final Monthly Report)."
        )
    except ValueError as ve:
        logger.error(f"ISP validation error: {ve}")
        raise
    except Exception as e:
        logger.error(f"ISP extraction error: {e}")
        return False


# ---------------------------------------------------------------------------
# Extractor 1 — ISP Morning Report (month-end daily, sheet: DAILYREPORT1)
# ---------------------------------------------------------------------------

def _morning_report_config():
    """
    Single source of truth for the ISP Morning Report cell mapping — shared by
    the DB-writing extractor (_extract_morning_report) and the preview-only
    extractor (_preview_morning_report_rows) so they can never drift apart.

    Reads excel_cells_config.json (section 'isp_morning'); falls back to these
    hardcoded defaults only if that config section is missing.
    """
    import sys
    sys.path.insert(0, os.path.dirname(__file__))
    from cells_loader import get_extractor_config
    cfg = get_extractor_config("isp_morning")

    no_convert = set(cfg.get("no_convert", ["COB#10", "COB#11", "Oven Pushing (nos/day)"]))
    cells = cfg.get("cells", {
        "COB#10":               "E9",
        "COB#11":               "E10",
        "Oven Pushing (nos/day)":  "E11",
        "SP M/c-1":             "E12",
        "SP M/c-2":             "E13",
        "Total Sinter":         "E14",
        "Hot Metal":            "E16",
        "SMS CCM-1&2":              "E30",
        "SMS CCM-3":                "E31",
        "Total Crude Steel":    "E32",
        "WRMILL":               "E33",
        "BARMILL":              "E34",
        "USMILL":               "E35",
        "Finished Steel":       "E36",
        "Saleable 150 Billets": "E37",
        "200 Blooms":           "E38",
        "Saleable Steel":       "E39",
    })
    derived = cfg.get("derived", [
        {"item": "Pig Iron",       "op": "add", "cells": ["E18", "E20"]},
        {"item": "Saleable Semis", "op": "add", "cells": ["E37", "E38"]},
    ])
    return cells, no_convert, derived


def _extract_morning_report(wb, source_file_name: str) -> bool:
    """
    Extracts cumulative monthly production from ISP Morning Report (sheet 'DAILYREPORT1').

    Date is auto-detected from K5 (Python datetime object). Column E = Monthly Rate.

    Cell map — sheet 'DAILYREPORT1':
      E9:       COB#10              — monthly avg, no unit conversion
      E10:      COB#11              — monthly avg, no unit conversion
      E11:      Oven Pushing (nos/day) — monthly avg nos/day, no unit conversion
      E12:      SP M/c-1            — tonnes → /1000
      E13:      SP M/c-2            — tonnes → /1000
      E14:      Total Sinter        — tonnes → /1000
      E16:      Hot Metal           — tonnes → /1000
      E18+E20:  Pig Iron            — derived sum, tonnes → /1000
      E30:      CCM-1&2             — tonnes → /1000
      E31:      CCM-3               — tonnes → /1000
      E32:      Total Crude Steel   — tonnes → /1000
      E33:      WRMILL              — tonnes → /1000
      E34:      BARMILL             — tonnes → /1000
      E35:      USMILL              — tonnes → /1000
      E36:      Finished Steel      — tonnes → /1000
      E37:      Saleable 150 Billets — tonnes → /1000
      E38:      200 Blooms          — tonnes → /1000
      E37+E38:  Saleable Semis      — derived sum, tonnes → /1000
      E39:      Saleable Steel      — tonnes → /1000
    """
    import sys
    sys.path.insert(0, os.path.dirname(__file__))
    import db

    ws = wb["DAILYREPORT1"]

    # Date from K5 (Excel stores it as a Python datetime)
    k5_raw = ws["K5"].value
    if isinstance(k5_raw, datetime):
        m_num = str(k5_raw.month).zfill(2)
        year  = str(k5_raw.year)
        db_report_month = f"{year}-{m_num}"
    elif k5_raw:
        date_match = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', str(k5_raw))
        if date_match:
            _d, m_num, year = date_match.groups()
            db_report_month = f"{year}-{m_num}"
        else:
            raise ValueError(
                f"Cannot parse date from K5: {repr(k5_raw)}. "
                "Expected a datetime value or DD.MM.YYYY string."
            )
    else:
        raise ValueError("Cell K5 is empty — cannot determine report month.")

    logger.info(f"ISP Morning Report: month auto-detected from K5 → {db_report_month}")

    production_cells, NO_CONVERT, derived_rules = _morning_report_config()

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    vals_extracted = 0

    def _save(item_name, val):
        nonlocal vals_extracted
        if val is not None:
            vals_extracted += 1
            if item_name not in NO_CONVERT:
                val = round(val / 1000.0, 3)
        cursor.execute("""
            INSERT INTO production_table (report_month, plant_name, item_name, month_actual)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(report_month, plant_name, item_name)
            DO UPDATE SET month_actual = excluded.month_actual
        """, (db_report_month, "ISP", item_name, val))

    for item_name, cell in production_cells.items():
        _save(item_name, clean_val(ws[cell].value))

    # Derived values driven by config
    for d in derived_rules:
        item = d["item"]
        if d["op"] == "subtract":
            a_val = clean_val(ws[d["a"]].value)
            b_val = clean_val(ws[d["b"]].value)
            if a_val is not None and b_val is not None:
                _save(item, a_val - b_val)
            elif a_val is not None:
                _save(item, a_val)
        elif d["op"] == "add":
            parts = [clean_val(ws[c].value) for c in d.get("cells", [])]
            valid  = [v for v in parts if v is not None]
            if valid:
                _save(item, sum(valid))

    if vals_extracted == 0:
        raise ValueError(
            "No numeric data found at the expected cell locations in sheet 'DAILYREPORT1'. "
            "Please verify this is the correct ISP Morning Report file."
        )

    conn.commit()
    conn.close()

    db.log_extraction(
        plant="ISP",
        report_month=db_report_month,
        file_name=source_file_name,
        sheet_name="DAILYREPORT1",
        source_type="Daily Morning Report",
        items_extracted=vals_extracted,
    )
    logger.info(f"ISP Morning Report extraction done: {vals_extracted} values saved for {db_report_month}.")
    return True


# ---------------------------------------------------------------------------
# Extractor 2 — ISP Final Monthly Report (sheet: Maj Production Summ)
# ---------------------------------------------------------------------------

def _extract_monthly_report(wb, report_month: str, source_file_name: str) -> bool:
    """
    Extracts production data from ISP Final Monthly Report (sheet 'Maj Production Summ').
    Month must be provided via report_month ('Month Year' or 'YYYY-MM').

    Cell map: items are in fixed rows; month selects the column via col_map.
    """
    import sys
    sys.path.insert(0, os.path.dirname(__file__))
    import db

    if not report_month:
        raise ValueError(
            "report_month is required for ISP Final Monthly Report. "
            "Set the month selector before uploading."
        )

    db_report_month, month_num = _parse_report_month(report_month)

    col_map = {
        "04": "F",  "05": "H",  "06": "L",  "07": "P",
        "08": "T",  "09": "X",  "10": "AD", "11": "AH",
        "12": "AL", "01": "AR", "02": "AV", "03": "AZ",
    }
    col = col_map.get(month_num)
    if not col:
        raise ValueError(f"Month column mapping not found for month code '{month_num}'.")

    ws = wb["Maj Production Summ"]

    NO_CONVERT = {"COB#10", "COB#11", "Oven Pushing (nos/day)"}

    # item_name → row number (column is dynamic per month via col_map)
    production_cells = {
        "COB#10":                6,
        "COB#11":                7,
        "Oven Pushing (nos/day)":   8,
        "Total Sinter":          16,
        "Hot Metal":             17,
        "Pig Iron":              26,
        "SMS CCM-1&2":               19,
        "SMS CCM-3":                 20,
        "Total Crude Steel":     18,
        "WRMILL":                30,
        "BARMILL":               31,
        "USMILL":                32,
        "Finished Steel":        33,
        "Saleable 150 Billets":  34,
        "200 Blooms":            35,
        "Saleable Semis":        36,
        "Saleable Steel":        37,
    }

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    vals_extracted = 0

    for item_name, row_num in production_cells.items():
        cell_ref = f"{col}{row_num}"
        val = clean_val(ws[cell_ref].value)
        if val is not None:
            vals_extracted += 1
            if item_name not in NO_CONVERT:
                val = round(val / 1000.0, 3)
        cursor.execute("""
            INSERT INTO production_table (report_month, plant_name, item_name, month_actual)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(report_month, plant_name, item_name)
            DO UPDATE SET month_actual = excluded.month_actual
        """, (db_report_month, "ISP", item_name, val))

    # Machine-wise sinter production ("SP M/c-1"/"SP M/c-2") lives on the
    # separate 'SINTER' tab under "MACHINE WISE SINTER PRODUCTION", not on
    # 'Maj Production Summ' (which only has the "Total Sinter" figure). Its
    # row position has shifted between report template revisions (R61 in
    # Apr'25, R63 May'25-Feb'26, R65 from Mar'26) so locate it by label
    # instead of a fixed row, same as the ISP techno extractor does for the
    # 'COKE OVENS' sheet's shifting rows (see _find_label_rows below).
    if "SINTER" in wb.sheetnames:
        ws_sinter = wb["SINTER"]
        mc_rows = _find_label_rows(ws_sinter, "M/C #", count=2)
        for item_name, row_num in zip(("SP M/c-1", "SP M/c-2"), mc_rows):
            cell_ref = f"{col}{row_num}"
            val = clean_val(ws_sinter[cell_ref].value)
            if val is not None:
                vals_extracted += 1
                val = round(val / 1000.0, 3)
            cursor.execute("""
                INSERT INTO production_table (report_month, plant_name, item_name, month_actual)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(report_month, plant_name, item_name)
                DO UPDATE SET month_actual = excluded.month_actual
            """, (db_report_month, "ISP", item_name, val))

    if vals_extracted == 0:
        raise ValueError(
            "No numeric data could be extracted from 'Maj Production Summ'. "
            "Please verify the contents of the Excel file."
        )

    conn.commit()
    conn.close()

    db.log_extraction(
        plant="ISP",
        report_month=db_report_month,
        file_name=source_file_name,
        sheet_name="Maj Production Summ",
        source_type="Final Monthly Report",
        items_extracted=vals_extracted,
    )
    logger.info(f"ISP Monthly Report extraction done: {vals_extracted} values saved for {db_report_month}.")
    return True


# ---------------------------------------------------------------------------
# Preview functions (no DB writes)
# ---------------------------------------------------------------------------

def _preview_morning_report_rows(wb):
    """Returns (production_rows, db_report_month) for a Morning Report workbook."""
    ws = wb["DAILYREPORT1"]

    k5_raw = ws["K5"].value
    if isinstance(k5_raw, datetime):
        m_num = str(k5_raw.month).zfill(2)
        year  = str(k5_raw.year)
        db_report_month = f"{year}-{m_num}"
    elif k5_raw:
        date_match = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', str(k5_raw))
        if date_match:
            _, m_num, year = date_match.groups()
            db_report_month = f"{year}-{m_num}"
        else:
            raise ValueError(f"Cannot parse date from K5: {repr(k5_raw)}.")
    else:
        raise ValueError("Cell K5 is empty — cannot determine report month.")

    production_cells, NO_CONVERT, derived_rules = _morning_report_config()

    rows = []
    for item_name, cell in production_cells.items():
        val = clean_val(ws[cell].value)
        if val is not None and item_name not in NO_CONVERT:
            val = round(val / 1000.0, 3)
        rows.append({
            "item_name": item_name,
            "value":     val,
            "cell":      cell,
            "unit":      "nos/d" if item_name in NO_CONVERT else "'000T",
            "status":    "ok" if val is not None else "no value",
        })

    # Derived values driven by the same config used by the DB-writing extractor.
    for d in derived_rules:
        item = d["item"]
        if d["op"] == "add":
            addrs = d.get("cells", [])
            parts = [clean_val(ws[c].value) for c in addrs]
            valid = [v for v in parts if v is not None]
            cell_label = "+".join(addrs)
            if valid:
                rows.append({"item_name": item, "value": round(sum(valid) / 1000.0, 3),
                             "cell": cell_label, "unit": "'000T", "status": "ok"})
            else:
                rows.append({"item_name": item, "value": None,
                             "cell": cell_label, "unit": "'000T", "status": "no value"})
        elif d["op"] == "subtract":
            a_val = clean_val(ws[d["a"]].value)
            b_val = clean_val(ws[d["b"]].value)
            cell_label = f"{d['a']}-{d['b']}"
            if a_val is not None and b_val is not None:
                rows.append({"item_name": item, "value": round((a_val - b_val) / 1000.0, 3),
                             "cell": cell_label, "unit": "'000T", "status": "ok"})
            elif a_val is not None:
                rows.append({"item_name": item, "value": round(a_val / 1000.0, 3),
                             "cell": cell_label, "unit": "'000T", "status": "ok"})
            else:
                rows.append({"item_name": item, "value": None,
                             "cell": cell_label, "unit": "'000T", "status": "no value"})

    return rows, db_report_month


def _preview_monthly_report_rows(wb, report_month: str):
    """Returns (production_rows, db_report_month) for a Final Monthly Report workbook."""
    if not report_month:
        raise ValueError(
            "report_month is required for ISP Final Monthly Report. "
            "Set the month selector before uploading."
        )

    db_report_month, month_num = _parse_report_month(report_month)

    col_map = {
        "04": "F",  "05": "H",  "06": "L",  "07": "P",
        "08": "T",  "09": "X",  "10": "AD", "11": "AH",
        "12": "AL", "01": "AR", "02": "AV", "03": "AZ",
    }
    col = col_map.get(month_num)
    if not col:
        raise ValueError(f"Month column mapping not found for month code '{month_num}'.")

    ws = wb["Maj Production Summ"]
    NO_CONVERT = {"COB#10", "COB#11", "Oven Pushing (nos/day)"}

    production_cells = {
        "COB#10":               6,
        "COB#11":               7,
        "Oven Pushing (nos/day)":  8,
        "Total Sinter":         16,
        "Hot Metal":            17,
        "Pig Iron":             26,
        "SMS CCM-1&2":              19,
        "SMS CCM-3":                20,
        "Total Crude Steel":    18,
        "WRMILL":               30,
        "BARMILL":              31,
        "USMILL":               32,
        "Finished Steel":       33,
        "Saleable 150 Billets": 34,
        "200 Blooms":           35,
        "Saleable Semis":       36,
        "Saleable Steel":       37,
    }

    rows = []
    for item_name, row_num in production_cells.items():
        cell_ref = f"{col}{row_num}"
        val = clean_val(ws[cell_ref].value)
        if val is not None and item_name not in NO_CONVERT:
            val = round(val / 1000.0, 3)
        rows.append({
            "item_name": item_name,
            "value":     val,
            "cell":      cell_ref,
            "unit":      "nos/d" if item_name in NO_CONVERT else "'000T",
            "status":    "ok" if val is not None else "no value",
        })

    # See _extract_monthly_report for why this is looked up by label on the
    # separate 'SINTER' tab instead of a fixed row on 'Maj Production Summ'.
    if "SINTER" in wb.sheetnames:
        ws_sinter = wb["SINTER"]
        mc_rows = _find_label_rows(ws_sinter, "M/C #", count=2)
        for item_name, row_num in zip(("SP M/c-1", "SP M/c-2"), mc_rows):
            cell_ref = f"{col}{row_num}"
            val = clean_val(ws_sinter[cell_ref].value)
            if val is not None:
                val = round(val / 1000.0, 3)
            rows.append({
                "item_name": item_name,
                "value":     val,
                "cell":      f"SINTER!{cell_ref}",
                "unit":      "'000T",
                "status":    "ok" if val is not None else "no value",
            })

    return rows, db_report_month


# ---------------------------------------------------------------------------
# Summarized Monthly Report helpers (no DB writes)
# ---------------------------------------------------------------------------

_MONTH_ABBR3 = {
    "01": "Jan", "02": "Feb", "03": "Mar", "04": "Apr",
    "05": "May", "06": "Jun", "07": "Jul", "08": "Aug",
    "09": "Sep", "10": "Oct", "11": "Nov", "12": "Dec",
}


# YTD column offset from monthly ACT column — applies to ALL ISP summarized monthly sheets.
# All sheets (B-FCE, SINTER, SMS, WRM, BM, USM, COKE OVENS) share the same column layout:
# each month has ABP+ACT, followed by QTR/half-year/annual cumulative columns at irregular spots.
# Offset = (ytd_act_col_index) - (monthly_act_col_index).
_ISP_SUMM_CUM_OFFSET = {
    "04": 0,   # Apr: no separate YTD col — use monthly value
    "05": 2,   # May: 2M cumulative
    "06": 2,   # Jun: QTR-1 cumulative
    "07": 2,   # Jul: 4M cumulative
    "08": 2,   # Aug: 5M cumulative
    "09": 4,   # Sep: H1 cumulative (skip QTR-2 ACT at offset +2)
    "10": 2,   # Oct: 7M cumulative
    "11": 2,   # Nov: 8M cumulative
    "12": 4,   # Dec: 9M cumulative (skip QTR-3 ACT at offset +2)
    "01": 2,   # Jan: 10M cumulative
    "02": 2,   # Feb: 11M cumulative
    "03": 6,   # Mar: Full-year cumulative (skip QTR-4 at +2, H2 at +4)
}


def _find_label_rows(ws, label: str, count: int = 2, label_col: int = 2, max_row: int = 260):
    """Scan label_col for cells whose text contains `label` (case-insensitive).
    Returns a list of up to `count` 1-based row numbers in order of appearance.
    Falls back to an empty list if not found.
    """
    found = []
    lc = label.lower()
    for r in range(1, max_row + 1):
        if lc in str(ws.cell(row=r, column=label_col).value or "").lower():
            found.append(r)
            if len(found) == count:
                break
    return found


def _tco(section: str, parameter: str, unit: str, sort_order: int,
         actual, cum_actual, cell: str, found_via: str, plant: str, month: str):
    """Build a single techno_param_row dict for COKE_SINTER group."""
    status = "ok" if actual is not None else "no value"
    return {
        "group_code":  "COKE_SINTER",
        "section":     section,
        "parameter":   parameter,
        "unit":        unit,
        "sort_order":  sort_order,
        "actual":      round(float(actual), 4) if actual is not None else None,
        "cum_actual":  round(float(cum_actual), 4) if cum_actual is not None else None,
        "status":      status,
        "plant":       plant,
        "month":       month,
        "cell":        cell,
        "found_via":   found_via,
    }


def _preview_coke_ovens(wb, db_report_month: str, month_num: str, year_2d: str) -> list:
    """Extract ISP COKE_SINTER techno_param_rows from the 'COKE OVENS' sheet.

    Returns a list of dicts in techno_param_rows format.
    """
    ws = wb["COKE OVENS"]
    ac = _find_act_col(ws, month_num, year_2d, 3, 4)
    if ac is None:
        return []

    offset = _ISP_SUMM_CUM_OFFSET.get(month_num, 2)
    # For April (offset=0) the YTD col IS the monthly col.
    cum_ac = ac if offset == 0 else ac + offset
    col = get_column_letter(ac)

    def _v(row, c):
        return clean_val(ws.cell(row=row, column=c).value)

    def _div(a, b):
        return round(a / b, 4) if (a is not None and b) else None

    rows = []

    # ── BF Coke Yield (overall ISP, row 37) ──────────────────────────────────
    v37  = _v(37, ac)
    cv37 = _v(37, cum_ac)
    rows.append(_tco("BF Coke Yield", "ISP", "%", 70,
                     v37, cv37, f"COKE OVENS!{col}37", "R37", "ISP", db_report_month))

    # ── Dry Coal Charge per Oven (Mt ÷ Ovens pushed = T/oven) ─────────────────
    mt10  = _v(17, ac);  ov10  = _v(8, ac)
    mt11  = _v(18, ac);  ov11  = _v(9, ac)
    cmt10 = _v(17, cum_ac); cov10 = _v(8, cum_ac)
    cmt11 = _v(18, cum_ac); cov11 = _v(9, cum_ac)
    dcc10  = _div(mt10,  ov10);  cdcc10 = _div(cmt10, cov10)
    dcc11  = _div(mt11,  ov11);  cdcc11 = _div(cmt11, cov11)
    rows.append(_tco("Dry Coal Charge/Oven", "ISP COB#10", "Tonnes", 440,
                     dcc10, cdcc10,
                     f"COKE OVENS!{col}17÷{col}8", "R17÷R8", "ISP", db_report_month))
    rows.append(_tco("Dry Coal Charge/Oven", "ISP COB#11", "Tonnes", 445,
                     dcc11, cdcc11,
                     f"COKE OVENS!{col}18÷{col}9", "R18÷R9", "ISP", db_report_month))

    # ── Sp. Heat Cons. — locate rows by label (row shifts across months) ─────────
    sh_rows = _find_label_rows(ws, "Sp Heat Cons", count=2)
    sh_r10  = sh_rows[0] if len(sh_rows) > 0 else 169   # fallback R169
    sh_r11  = sh_rows[1] if len(sh_rows) > 1 else 195   # fallback R195
    def _kc(x): return round(x * 1000, 4) if x is not None else None
    rows.append(_tco("Sp. Heat Cons.", "ISP COB#10", "000 K.Cal/TDC", 170,
                     _kc(_v(sh_r10, ac)), _kc(_v(sh_r10, cum_ac)),
                     f"COKE OVENS!{col}{sh_r10}×1000", f"R{sh_r10}×1000", "ISP", db_report_month))
    rows.append(_tco("Sp. Heat Cons.", "ISP COB#11", "000 K.Cal/TDC", 175,
                     _kc(_v(sh_r11, ac)), _kc(_v(sh_r11, cum_ac)),
                     f"COKE OVENS!{col}{sh_r11}×1000", f"R{sh_r11}×1000", "ISP", db_report_month))

    # ── Coke Oven Gas Yield — locate rows by label (row shifts across months) ───
    cog_rows = _find_label_rows(ws, "C.O.Gas Yield", count=2)
    cog_r10  = cog_rows[0] if len(cog_rows) > 0 else 164  # fallback R164
    cog_r11  = cog_rows[1] if len(cog_rows) > 1 else 189  # fallback R189
    cog10  = _v(cog_r10, ac);  cog11  = _v(cog_r11, ac)
    ccog10 = _v(cog_r10, cum_ac); ccog11 = _v(cog_r11, cum_ac)
    rows.append(_tco("Coke Oven Gas Yield", "ISP COB#10", "M3/TDC", 540,
                     cog10, ccog10,
                     f"COKE OVENS!{col}{cog_r10}", f"R{cog_r10}", "ISP", db_report_month))
    rows.append(_tco("Coke Oven Gas Yield", "ISP COB#11", "M3/TDC", 545,
                     cog11, ccog11,
                     f"COKE OVENS!{col}{cog_r11}", f"R{cog_r11}", "ISP", db_report_month))

    # ── Coal Tar Yield (Kg/Tdcc) ──────────────────────────────────────────────
    ct10  = _v(79, ac);  ct11  = _v(80, ac)
    cct10 = _v(79, cum_ac); cct11 = _v(80, cum_ac)
    rows.append(_tco("Coal Tar Yield", "ISP COB#10", "kg/TDC", 270,
                     ct10, cct10,
                     f"COKE OVENS!{col}79", "R79", "ISP", db_report_month))
    rows.append(_tco("Coal Tar Yield", "ISP COB#11", "kg/TDC", 275,
                     ct11, cct11,
                     f"COKE OVENS!{col}80", "R80", "ISP", db_report_month))

    # ── Amm. Sulphate Yield (Kg/Tdcc) ────────────────────────────────────────
    am10  = _v(82, ac);  am11  = _v(83, ac)
    cam10 = _v(82, cum_ac); cam11 = _v(83, cum_ac)
    rows.append(_tco("Amm. Sulphate Yld", "ISP COB#10", "kg/TDC", 740,
                     am10, cam10,
                     f"COKE OVENS!{col}82", "R82", "ISP", db_report_month))
    rows.append(_tco("Amm. Sulphate Yld", "ISP COB#11", "kg/TDC", 745,
                     am11, cam11,
                     f"COKE OVENS!{col}83", "R83", "ISP", db_report_month))

    return rows


def _find_act_col(ws, month_num: str, year_2d: str, header_row: int, subheader_row: int):
    """Scan header_row for the month label; return 1-based ACT column index or None.

    Matches cells that contain the 3-letter month abbreviation AND the 2-digit year
    suffix (e.g. "'25"), then verifies the next column in subheader_row is "ACT".
    Handles variations like "Jul'25" vs "July'25".
    """
    abbr = _MONTH_ABBR3[month_num].lower()
    year_suffix = f"'{year_2d}"
    for col in range(1, 120):
        raw = ws.cell(row=header_row, column=col).value
        if raw is None:
            continue
        s = str(raw).strip()
        if abbr in s.lower() and year_suffix in s:
            act_check = str(ws.cell(row=subheader_row, column=col + 1).value or "").strip()
            if act_check == "ACT":
                return col + 1
    return None



def _mill_params(ws, ac, cum_ac, col, section,
                 yield_label, avail_off, util_off, spow_off, rrate_off,
                 heat_off=None, total_off=None, max_row=220):
    """Extract rolling-mill techno params by anchoring on the yield-label row
    and using fixed relative offsets (consistent across all monthly files).

    Offsets verified across Apr25 / May25 / May26 / Jun26 / Mar26:
      WRM:  avail=+5, util=+7, spow=+12, rrate=+17  (no heat)
      BM:   total=-3, avail=+9, util=+11, heat=+14, spow=+16, rrate=+21
      USM:  total=-3, avail=+9, util=+11, heat=+14, spow=+16, rrate=+21
    """
    found = _find_label_rows(ws, yield_label, count=1, label_col=2, max_row=max_row)
    if not found:
        return []
    yr = found[0]

    def _v(r): return clean_val(ws.cell(row=r, column=ac).value)
    def _ytd(r): return clean_val(ws.cell(row=r, column=cum_ac).value) if cum_ac else None
    def _lbl(r): return str(ws.cell(row=r, column=2).value or '').strip()

    rows = []
    if total_off is not None:
        r = yr + total_off
        rows.append(_tr(section, "Total Production", "T",
                        _v(r), f"{section}!{col}{r}", _lbl(r), _ytd(r)))
    rows.append(_tr(section, "Mill Yield",        "%",
                    _v(yr), f"{section}!{col}{yr}", _lbl(yr), _ytd(yr)))
    for off, name, unit in [(avail_off, "Mill Availability", "%"),
                             (util_off,  "Mill Utilisation",  "%")]:
        r = yr + off
        rows.append(_tr(section, name, unit, _v(r), f"{section}!{col}{r}", _lbl(r), _ytd(r)))
    if heat_off is not None:
        r = yr + heat_off
        kc = lambda x: round(x * 1000, 4) if x is not None else None
        rows.append(_tr(section, "Sp Heat Consumption", "1000 KCal/T",
                        kc(_v(r)), f"{section}!{col}{r}×1000", _lbl(r), kc(_ytd(r))))
    for off, name, unit in [(spow_off,  "Sp Power Cons", "KWh/T"),
                             (rrate_off, "Rolling Rate",  "T/Hr")]:
        r = yr + off
        rows.append(_tr(section, name, unit, _v(r), f"{section}!{col}{r}", _lbl(r), _ytd(r)))
    return rows


def _tr(section: str, name: str, unit: str, val, cell_ref: str, row_label: str = "", ytd_val=None):
    """Build a single techno_row dict."""
    return {
        "parameter":    f"{section} {name}",
        "unit":         unit,
        "month_actual": round(float(val), 4) if val is not None else None,
        "ytd_actual":   round(float(ytd_val), 4) if ytd_val is not None else None,
        "cell":         cell_ref,
        "row_label":    row_label or name,
        "mapping_ok":   True,
        "status":       "ok" if val is not None else "no value",
    }


def _preview_maj_techno_summ(wb, db_report_month: str, month_num: str,
                              year_2d: str, cum_offset: int) -> list:
    """Extract ISP MAJOR techno params from 'Maj Techno Summ' sheet.

    Sheet layout: row 3 = month headers ("Apr'25"), row 4 = ABP/ACT labels.
    Writes to MAJOR group at source_priority=5 (authoritative plant summary).
    """
    ws = wb["Maj Techno Summ"]
    ac = _find_act_col(ws, month_num, year_2d, 3, 4)
    if not ac:
        logger.warning("ISP Maj Techno Summ: ACT column not found for %s-%s", month_num, year_2d)
        return []

    cum_ac = ac + cum_offset if cum_offset else None
    col = get_column_letter(ac)

    # (search_label, MAJOR_section, unit, sort_order)
    _MAP = [
        ("BF Coke Rate",            "Coke Rate",                        "Kg/THM",    11),
        ("Nut Coke rate",           "Nut Coke Consumption",             "Kg/THM",    31),
        ("CDI rate",                "CDI Rate",                         "Kg/THM",    41),
        ("Total Fuel",              "Fuel Rate",                        "Kg/THM",    51),
        ("BF Productivity",         "BF Productivity (Working Volume)", "T/m³/day",  81),
        ("Gross Coal to Hot Metal", "Coal to Hot Metal",                "Ratio",     91),
        ("Sp Energy Consumption",   "Specific Energy Consumption",      "G.Cal/TCS", 121),
    ]

    rows = []
    for label, section, unit, so in _MAP:
        found = _find_label_rows(ws, label, count=1, label_col=2, max_row=80)
        if not found and label == "Sp Energy Consumption":
            # 2016-18 vintage files truncate this label to "Overall Sp Energy
            # Consumpt" — retry with the shortened form.
            found = _find_label_rows(ws, "Sp Energy Consumpt", count=1, label_col=2, max_row=80)
        if not found:
            rows.append({
                "group_code": "MAJOR", "section": section, "parameter": "ISP",
                "unit": unit, "sort_order": so, "actual": None, "cum_actual": None,
                "cell": f"Maj Techno Summ!?",
                "found_via": f"ISP Maj Techno Summ {label}",
                "status": "no value", "plant": "ISP", "month": db_report_month,
                "source_priority": 5,
            })
            continue
        r = found[0]
        v = clean_val(ws.cell(row=r, column=ac).value)
        y = clean_val(ws.cell(row=r, column=cum_ac).value) if cum_ac else None
        rows.append({
            "group_code":      "MAJOR",
            "section":         section,
            "parameter":       "ISP",
            "unit":            unit,
            "sort_order":      so,
            "actual":          v,
            "cum_actual":      y,
            "cell":            f"Maj Techno Summ!{col}{r}",
            "found_via":       f"ISP Maj Techno Summ {label}",
            "status":          "ok" if v is not None else "no value",
            "plant":           "ISP",
            "month":           db_report_month,
            "source_priority": 5,
        })
    return rows


def _summ_header_months(ws, header_row: int) -> list:
    """Ordered list of 'YYYY-MM' month columns present on a summarized-sheet
    header row (labels like "Apr'17", "Jul'25"). Cumulative columns ("2M",
    "Qrt-1", "H-1") don't match and are skipped."""
    out = []
    for c in range(1, 130):
        raw = ws.cell(row=header_row, column=c).value
        if raw is None:
            continue
        m = re.match(r"^\s*([A-Za-z]{3,9})\s*'\s*(\d{2})\s*$", str(raw).strip())
        if not m:
            continue
        ab = m.group(1)[:3].title()
        for num, abbr in _MONTH_ABBR3.items():
            if abbr == ab:
                ym = f"20{m.group(2)}-{num}"
                if ym not in out:
                    out.append(ym)
                break
    return out


# ---------------------------------------------------------------------------
# Old-vintage (2016-18 "Summarised monthly statistical") extraction.
# Same ABP/ACT column layout as modern files, but different sheet names
# (BF/CO in 2016-17), all headers on rows 3/4, and different row anchors.
# Everything below is label-anchored — no bare fixed-row fallbacks, so a row
# that isn't found is simply skipped instead of silently reading garbage.
# ---------------------------------------------------------------------------

def _old_bf_rows(ws, ac, cum_ac, col, techno_rows, techno_param_rows):
    """BF sheet, old vintage: rates under 'NET RAW MATERIALS CONSUMPTION RATE
    (DRY BASIS)' (+1 Coke, +2 Nut Coke, +3 CDI), productivity/labels scanned."""
    _CROSS_BF = {
        "PCI Rate":          ("CDI",              "Kg/THM",   17),
        "Coke Rate":         ("BF Coke Rate",     "Kg/THM",   45),
        "Nut Coke Rate":     ("Nut Coke Rate",    "Kg/THM",   55),
        "Productivity":      ("BF Productivity",  "T/m³/day", 65),
        "Sinter in Burden":  ("Sinter in Burden", "%",       115),
        "Blast Temperature": ("HBT",              "°C",      100),
        "Si in HM":          ("Si in HM",         "%",        75),
        "S in HM":           ("S in HM",          "%",        85),
    }

    def _v3(r):
        v = clean_val(ws.cell(row=r, column=ac).value)
        y = clean_val(ws.cell(row=r, column=cum_ac).value) if cum_ac else None
        l = str(ws.cell(row=r, column=2).value or '').strip()
        return v, y, l

    def _emit(name, cross, unit, r):
        v, y, l = _v3(r)
        techno_rows.append(_tr("BF", name, unit, v, f"BF!{col}{r}", l, y))
        if cross in _CROSS_BF:
            sec, u, so = _CROSS_BF[cross]
            techno_param_rows.append({
                "group_code": "IRON_MAKING", "section": sec,
                "parameter": "ISP", "unit": u,
                "sort_order": so, "actual": v, "cum_actual": y,
                "cell": f"BF!{col}{r}", "found_via": f"ISP BF (old) {name}",
                "status": "ok" if v is not None else "no value",
            })

    anchor = _find_label_rows(ws, "NET RAW MATERIALS CONSUMPTION RATE", count=1, max_row=60)
    if anchor:
        a = anchor[0]
        _emit("Coke Rate",     "Coke Rate",     "Kg/THM", a + 1)
        _emit("Nut Coke Rate", "Nut Coke Rate", "Kg/THM", a + 2)
        _emit("PCI Rate",      "PCI Rate",      "Kg/THM", a + 3)

    # Label-scanned rows — emitted only when the label exists (old files have
    # no Pellet or Oxygen-enrichment rows, so those are simply absent here).
    for label, name, cross, unit in [
        ("BF Productivity",      "Productivity",      "Productivity",      "T/M3/Day"),
        ("Sinter in Fe bearing", "Sinter% in Burden", "Sinter in Burden",  "%"),
        ("Blast Temperature",    "Hot Blast Temp",    "Blast Temperature", "°C"),
        ("Slag Rate",            "Slag Rate",         None,                "Kg/T"),
        ("Si in HM",             "Si in HM",          "Si in HM",          "%"),
        ("S in HM",              "S in HM",           "S in HM",           "%"),
        ("Sp Power Cons",        "Sp Power Cons",     None,                "KWh/T"),
    ]:
        found = _find_label_rows(ws, label, count=1)
        if found:
            _emit(name, cross, unit, found[0])


def _old_sms_rows(ws, ac, cum_ac, col, techno_rows, techno_param_rows):
    """SMS sheet, old vintage — all rows located by label. Old files carry
    Steel Scrap and Iron Scrap as separate rate rows; their sum is the
    'Scrap Consumption' figure modern files report as one row."""
    def _val(r, c):
        return clean_val(ws.cell(row=r, column=c).value)

    def _lbl(r):
        return str(ws.cell(row=r, column=2).value or '').strip()

    def _find1(label):
        f = _find_label_rows(ws, label, count=1, max_row=130)
        return f[0] if f else None

    days     = _val(2, ac) or 30
    cum_days = _val(2, cum_ac) if cum_ac else None

    r_heats = _find1("Total heats")
    blows_day = cum_blows = None
    if r_heats:
        heats = _val(r_heats, ac)
        blows_day = round(heats / days, 2) if heats is not None else None
        cum_heats = _val(r_heats, cum_ac) if cum_ac else None
        cum_blows = round(cum_heats / cum_days, 2) if (cum_heats is not None and cum_days) else None
        techno_rows.append(_tr("SMS", "Blows per Day", "Nos/Day", blows_day,
                               f"SMS!{col}{r_heats}÷{int(days)}days", _lbl(r_heats), cum_blows))

    r_acw = _find1("Avg. Cast Weight")
    acw = acw_c = None
    if r_acw:
        acw   = _val(r_acw, ac)
        acw_c = _val(r_acw, cum_ac) if cum_ac else None
        techno_rows.append(_tr("SMS", "Avg Cast Wt", "Mt/Cast", acw,
                               f"SMS!{col}{r_acw}", _lbl(r_acw), acw_c))

    r_hm = _find1("Hot Metal Consumption rate")
    if r_hm:
        techno_rows.append(_tr("SMS", "HM Consumption", "Kg/TCS", _val(r_hm, ac),
                               f"SMS!{col}{r_hm}", _lbl(r_hm),
                               _val(r_hm, cum_ac) if cum_ac else None))

    r_ss = _find1("Steel Scrap Consumption rate")
    r_is = _find1("Iron Scrap Consumption rate")
    if r_ss:
        sv  = _val(r_ss, ac)
        iv  = _val(r_is, ac) if r_is else None
        scv = (sv or 0) + (iv or 0) if (sv is not None or iv is not None) else None
        scy = None
        if cum_ac:
            svc = _val(r_ss, cum_ac)
            ivc = _val(r_is, cum_ac) if r_is else None
            scy = (svc or 0) + (ivc or 0) if (svc is not None or ivc is not None) else None
        cell = f"SMS!{col}{r_ss}" + (f"+{col}{r_is}" if r_is else "")
        techno_rows.append(_tr("SMS", "Scrap Consumption", "Kg/TCS", scv,
                               cell, "Steel+Iron Scrap Consumption rate", scy))

    r_tmi = _find1("TMI")
    if r_tmi:
        techno_rows.append(_tr("SMS", "TMI", "Kg/TCS", _val(r_tmi, ac),
                               f"SMS!{col}{r_tmi}", _lbl(r_tmi),
                               _val(r_tmi, cum_ac) if cum_ac else None))

    r_o2 = _find1("Sp Oxygen Consumption")
    o2 = o2c = None
    if r_o2:
        o2  = _val(r_o2, ac)
        o2c = _val(r_o2, cum_ac) if cum_ac else None
        techno_rows.append(_tr("SMS", "Sp O2 Consumption", "Nm³/T", o2,
                               f"SMS!{col}{r_o2}", _lbl(r_o2), o2c))

    for section, unit, so, val, cum in [
        ("Average Blows (Per Day)", "Nos.",    8,  blows_day, cum_blows),
        ("Average Heat Weight",     "T",       15, acw,       acw_c),
        ("Oxygen Blowing",          "Nm³/TCS", 34, o2,        o2c),
    ]:
        techno_param_rows.append({
            "group_code": "SMS", "section": section,
            "parameter": "ISP SMS", "unit": unit, "sort_order": so,
            "actual": val, "cum_actual": cum,
            "cell": f"SMS (old vintage, label-scanned)",
            "found_via": f"ISP SMS {section}",
            "status": "ok" if val is not None else "no value",
        })


def _old_coke_rows(ws, sheet_label, db_report_month, month_num, ac, cum_ac, col) -> list:
    """COKE OVENS / CO sheet, old vintage.

    Ovens-pushed rows here are per-DAY averages (row 2 holds calendar days),
    and dry-coal-charge rows are Old Batts (COB#8+10) / New Batt (COB#11)
    monthly totals — COB#8 was idle through 2016-18, so Old Batts ≈ COB#10.
    """
    def _v(r, c):
        return clean_val(ws.cell(row=r, column=c).value)

    def _find1(label, max_row=260):
        f = _find_label_rows(ws, label, count=1, max_row=max_row)
        return f[0] if f else None

    def _div(a, b):
        return round(a / b, 4) if (a is not None and b) else None

    rows = []
    days     = _v(2, ac)
    cum_days = _v(2, cum_ac) if cum_ac else None

    # BF Coke Yield: '+1 under the OVERALL ... YIELD header
    ya = _find1("OVERALL COKE & COKE FRACTION YIELD")
    if ya:
        rows.append(_tco("BF Coke Yield", "ISP", "%", 70,
                         _v(ya + 1, ac), _v(ya + 1, cum_ac) if cum_ac else None,
                         f"{sheet_label}!{col}{ya+1}", f"R{ya+1}", "ISP", db_report_month))

    # Dry Coal Charge/Oven = monthly charge ÷ (ovens-per-day × days)
    r_ov10 = _find1("Ovens Pushed (COB#10")
    r_ov11 = _find1("Ovens Pushed (COB#11")
    r_ch_old = _find1("Dry Coal Charge (Old Batts)")
    r_ch_new = _find1("Dry Coal Charge (New Batt)")
    for cob, r_ov, r_ch, so in [("COB#10", r_ov10, r_ch_old, 440),
                                ("COB#11", r_ov11, r_ch_new, 445)]:
        if not (r_ov and r_ch):
            continue
        ov_pd  = _v(r_ov, ac)
        charge = _v(r_ch, ac)
        dcc = _div(charge, ov_pd * days) if (ov_pd and days) else None
        cdcc = None
        if cum_ac and cum_days:
            ov_pd_c  = _v(r_ov, cum_ac)
            charge_c = _v(r_ch, cum_ac)
            cdcc = _div(charge_c, ov_pd_c * cum_days) if ov_pd_c else None
        rows.append(_tco("Dry Coal Charge/Oven", f"ISP {cob}", "Tonnes", so,
                         dcc, cdcc,
                         f"{sheet_label}!{col}{r_ch}÷({col}{r_ov}×days)",
                         f"R{r_ch}÷(R{r_ov}×d)", "ISP", db_report_month))

    # Sp Heat Cons ×1000 and C.O.Gas Yield — two occurrences (COB#8/10, COB#11)
    sh = _find_label_rows(ws, "Sp Heat Cons", count=2, max_row=260)
    def _kc(x): return round(x * 1000, 4) if x is not None else None
    for i, (cob, so) in enumerate([("COB#10", 170), ("COB#11", 175)]):
        if i < len(sh):
            r = sh[i]
            rows.append(_tco("Sp. Heat Cons.", f"ISP {cob}", "000 K.Cal/TDC", so,
                             _kc(_v(r, ac)), _kc(_v(r, cum_ac)) if cum_ac else None,
                             f"{sheet_label}!{col}{r}×1000", f"R{r}×1000", "ISP", db_report_month))

    cog = _find_label_rows(ws, "C.O.Gas Yield", count=2, max_row=260)
    for i, (cob, so) in enumerate([("COB#10", 540), ("COB#11", 545)]):
        if i < len(cog):
            r = cog[i]
            rows.append(_tco("Coke Oven Gas Yield", f"ISP {cob}", "M3/TDC", so,
                             _v(r, ac), _v(r, cum_ac) if cum_ac else None,
                             f"{sheet_label}!{col}{r}", f"R{r}", "ISP", db_report_month))

    # By-product yields under 'Yield of Coal Chemicals': +1/+2 = Coal Tar
    # (Existing→COB#8/10 batteries, New→COB#11), +4/+5 = Amm. Sulphate
    yc = _find1("Yield of Coal Chemicals")
    if yc:
        for off, sec, cob, so in [(1, "Coal Tar Yield", "COB#10", 270),
                                  (2, "Coal Tar Yield", "COB#11", 275),
                                  (4, "Amm. Sulphate Yld", "COB#10", 740),
                                  (5, "Amm. Sulphate Yld", "COB#11", 745)]:
            r = yc + off
            rows.append(_tco(sec, f"ISP {cob}", "kg/TDC", so,
                             _v(r, ac), _v(r, cum_ac) if cum_ac else None,
                             f"{sheet_label}!{col}{r}", f"R{r}", "ISP", db_report_month))
    return rows


def _old_prod_summ_rows(wb, prod_sheet, month_num, year_2d) -> list:
    """Production from the old-vintage 'Maj Prod Summ' sheet (identical row
    layout in the 2016-17 and 2017-18 files). Each fixed row is verified
    against its expected label text before use — a mismatch skips the row."""
    ws = wb[prod_sheet]
    ac = _find_act_col(ws, month_num, year_2d, 3, 4)
    if ac is None:
        return []
    col = get_column_letter(ac)

    NO_CONVERT = {"COB#10", "COB#11", "Oven Pushing (nos/day)"}
    # (item_name, row, expected label substring)
    _ROWS = [
        ("COB#10",                  7,  "COB#10"),
        ("COB#11",                  8,  "COB#11"),
        ("Oven Pushing (nos/day)",  9,  "Ovens Pushed"),
        ("Total Sinter",            14, "SINTER"),
        ("Hot Metal",               15, "HOT METAL"),
        ("Total Crude Steel",       16, "CRUDE STEEL"),
        ("200 Blooms",              27, "Bloom Caster"),
        ("Saleable 150 Billets",    28, "Billet Caster"),
        ("WRMILL",                  29, "WR Mill"),
        ("BARMILL",                 30, "Bar Mill"),
        ("USMILL",                  31, "US Mill"),
        ("Saleable Steel",          32, "TOTAL"),
        ("Saleable Semis",          34, "Semis"),
    ]
    rows = []
    by_item = {}
    for item, r, expect in _ROWS:
        lbl = str(ws.cell(row=r, column=2).value or "").strip()
        if expect.lower() not in lbl.lower():
            continue
        val = clean_val(ws.cell(row=r, column=ac).value)
        if val is not None and item not in NO_CONVERT:
            val = round(val / 1000.0, 3)
        by_item[item] = val
        rows.append({
            "item_name": item, "value": val, "cell": f"{prod_sheet}!{col}{r}",
            "unit": "nos/d" if item in NO_CONVERT else "'000T",
            "status": "ok" if val is not None else "no value",
        })

    # Pig Iron: TOTAL row of the 'Cold Pig & Pig Equivalent' block
    cp = _find_label_rows(ws, "Cold Pig & Pig Equivalent", count=1, max_row=40)
    if cp:
        for r in range(cp[0] + 1, cp[0] + 8):
            if str(ws.cell(row=r, column=2).value or "").strip().upper() == "TOTAL":
                val = clean_val(ws.cell(row=r, column=ac).value)
                rows.append({
                    "item_name": "Pig Iron",
                    "value": round(val / 1000.0, 3) if val is not None else None,
                    "cell": f"{prod_sheet}!{col}{r}", "unit": "'000T",
                    "status": "ok" if val is not None else "no value",
                })
                break

    # Finished Steel = Saleable Steel − Saleable Semis
    ss, sem = by_item.get("Saleable Steel"), by_item.get("Saleable Semis")
    if ss is not None and sem is not None:
        rows.append({
            "item_name": "Finished Steel", "value": round(ss - sem, 3),
            "cell": f"{prod_sheet}!{col}32-{col}34", "unit": "'000T", "status": "ok",
        })

    # SP M/c-1 / M/c-2 from the SINTER sheet
    if "SINTER" in wb.sheetnames:
        ws_s = wb["SINTER"]
        ac_s = _find_act_col(ws_s, month_num, year_2d, 3, 4)
        if ac_s:
            col_s = get_column_letter(ac_s)
            mc_rows = _find_label_rows(ws_s, "M/C #", count=2)
            for item_name, r in zip(("SP M/c-1", "SP M/c-2"), mc_rows):
                val = clean_val(ws_s.cell(row=r, column=ac_s).value)
                rows.append({
                    "item_name": item_name,
                    "value": round(val / 1000.0, 3) if val is not None else None,
                    "cell": f"SINTER!{col_s}{r}", "unit": "'000T",
                    "status": "ok" if val is not None else "no value",
                })
    return rows


def _preview_summarized_monthly(wb, report_month: str, sheet_names: list) -> dict:
    """Extract ISP techno parameters from the Summarized Monthly Report.

    Sheet groups and their header rows:
      B-FCE / SINTER / SMS  — month-label row 3 or 4, ABP/ACT row 4 or 5 (varies)
      WRM / BM              — month-label row 3, ABP/ACT row 4
      USM                   — month-label row 3, ABP/ACT row 4  (col shift: Apr ABP=F)

    Column for a given month is detected dynamically so this works regardless of
    how many cumulative columns are interspersed between monthly columns.
    """
    if not report_month:
        raise ValueError(
            "report_month is required for ISP Summarized Monthly Report. "
            "Set the month selector before uploading."
        )

    db_report_month, month_num = _parse_report_month(report_month)
    year_2d = db_report_month[2:4]          # "2025-04" → "25"

    # Cumulative column offset — same for all ISP summarized monthly sheets.
    _co = _ISP_SUMM_CUM_OFFSET.get(month_num, 2)
    def _cum(ac): return ac if _co == 0 else ac + _co

    # ── Month-scope validation against the file's own header ─────────────────
    # A summarized workbook carries one column per elapsed month of ITS fiscal
    # year, so any earlier month of the same FY is extractable from a later
    # (e.g. up-to-March) file. A month outside the file's FY is a hard error —
    # never silently extract mismatched-FY data.
    bf_sheet = "B-FCE" if "B-FCE" in sheet_names else ("BF" if "BF" in sheet_names else None)
    file_months, bf_hr = [], None
    if bf_sheet:
        ws_bf = wb[bf_sheet]
        for hr in (4, 3):
            file_months = _summ_header_months(ws_bf, hr)
            if file_months:
                bf_hr = hr
                break
    # Old vintage (2016-18 files): month headers on ROW 3 of the BF sheet
    # (modern B-FCE files put them on row 4), plus the old rate-block label.
    # Both conditions — the label alone also exists on modern B-FCE sheets.
    old_vintage = (
        bf_hr == 3
        and bf_sheet is not None
        and bool(_find_label_rows(wb[bf_sheet], "NET RAW MATERIALS CONSUMPTION RATE",
                                  count=1, max_row=60))
    )

    month_mismatch = None
    if file_months:
        if db_report_month not in file_months:
            raise ValueError(
                f"Month mismatch: this file covers {file_months[0]} to {file_months[-1]} "
                f"only — selected month {db_report_month} belongs to a different FY. "
                f"Nothing was extracted; select a month within this file's FY or upload "
                f"the correct FY's report."
            )
        if db_report_month != file_months[-1]:
            month_mismatch = {
                "selected_month": db_report_month,
                "actual_month":   file_months[-1],
                "warning_only":   True,
                "message": (
                    f"This file's latest month is {file_months[-1]}; extracting the "
                    f"earlier month {db_report_month} from its {db_report_month} column "
                    f"(same FY, so the data is authoritative)."
                ),
            }

    techno_rows = []
    techno_param_rows = []

    # ── Old vintage (2016-18 'Summarised monthly statistical' files) ─────────
    if old_vintage:
        ws = wb[bf_sheet]
        ac = _find_act_col(ws, month_num, year_2d, 3, 4)
        if ac:
            _old_bf_rows(ws, ac, _cum(ac), get_column_letter(ac),
                         techno_rows, techno_param_rows)

        if "SINTER" in sheet_names:
            ws_s = wb["SINTER"]
            ac_s = _find_act_col(ws_s, month_num, year_2d, 3, 4)
            if ac_s:
                col_s = get_column_letter(ac_s)
                found = _find_label_rows(ws_s, "Specific Productivity", count=1)
                if found:
                    r = found[0]
                    techno_rows.append(_tr(
                        "Sinter", "Sp Productivity", "T/M2/Utlz Hr",
                        clean_val(ws_s.cell(row=r, column=ac_s).value),
                        f"SINTER!{col_s}{r}",
                        str(ws_s.cell(row=r, column=2).value or "").strip(),
                        clean_val(ws_s.cell(row=r, column=_cum(ac_s)).value)))

        if "SMS" in sheet_names:
            ws_m = wb["SMS"]
            ac_m = _find_act_col(ws_m, month_num, year_2d, 3, 4)
            if ac_m:
                _old_sms_rows(ws_m, ac_m, _cum(ac_m), get_column_letter(ac_m),
                              techno_rows, techno_param_rows)

        # Mills — same anchors as modern but old-vintage offsets
        # (rows verified identical across both 2016-17 and 2017-18 files)
        for sn, anchor, kw in [("WRM", "WR MILL YIELD",  dict(avail_off=5, util_off=7, spow_off=12, rrate_off=16)),
                               ("BM",  "BAR MILL YIELD", dict(avail_off=5, util_off=7, heat_off=10, spow_off=12, rrate_off=16, total_off=-1)),
                               ("USM", "US MILL YIELD",  dict(avail_off=5, util_off=7, heat_off=10, spow_off=12, rrate_off=16, total_off=-1))]:
            if sn in sheet_names:
                ws_x = wb[sn]
                ac_x = _find_act_col(ws_x, month_num, year_2d, 3, 4)
                if ac_x:
                    techno_rows += _mill_params(ws_x, ac_x, _cum(ac_x),
                                                get_column_letter(ac_x), sn, anchor, **kw)

        ok_count = sum(1 for r in techno_rows if r["status"] == "ok")
        if ok_count == 0:
            raise ValueError(
                f"No data found for month '{report_month}' in this Summarised "
                f"monthly statistical file (old format). Verify the month falls "
                f"within the file's FY."
            )

        coke_sheet = "COKE OVENS" if "COKE OVENS" in sheet_names else ("CO" if "CO" in sheet_names else None)
        if coke_sheet:
            try:
                ws_c = wb[coke_sheet]
                ac_c = _find_act_col(ws_c, month_num, year_2d, 3, 4)
                if ac_c:
                    techno_param_rows += _old_coke_rows(
                        ws_c, coke_sheet, db_report_month, month_num,
                        ac_c, _cum(ac_c), get_column_letter(ac_c))
            except Exception as e:
                logger.warning(f"ISP (old vintage): {coke_sheet} extraction failed: {e}")

        if "Maj Techno Summ" in sheet_names:
            try:
                techno_param_rows += _preview_maj_techno_summ(
                    wb, db_report_month, month_num, year_2d, _co)
            except Exception as e:
                logger.warning(f"ISP (old vintage): Maj Techno Summ extraction failed: {e}")

        prod_rows = []
        if "Maj Prod Summ" in sheet_names:
            try:
                prod_rows = _old_prod_summ_rows(wb, "Maj Prod Summ", month_num, year_2d)
            except Exception as e:
                logger.warning(f"ISP (old vintage): Maj Prod Summ extraction failed: {e}")

        logger.info(f"ISP Summarised (old vintage): {ok_count} techno, "
                    f"{len(prod_rows)} production values for {db_report_month}.")
        return {
            "plant":             "ISP",
            "month":             db_report_month,
            "source_type":       "Summarised Monthly Statistical (2016-18 format)",
            "sheets":            f"{bf_sheet}, SINTER, SMS, WRM, BM, USM"
                                 + (f", {coke_sheet}" if coke_sheet else "")
                                 + (", Maj Techno Summ" if "Maj Techno Summ" in sheet_names else "")
                                 + (", Maj Prod Summ" if prod_rows else ""),
            "workbook_sheets":   sheet_names,
            "production_rows":   prod_rows,
            "techno_rows":       techno_rows,
            "techno_param_rows": techno_param_rows,
            "month_mismatch":    month_mismatch,
        }

    # ------------------------------------------------------------------
    # B-FCE   (header row 4, ABP/ACT sub-header row 5, days in row 2)
    # Rows 39/40/41/49/91/92 are stable across all files.
    # Sinter%, Pellet%, Blast Temp, Slag Rate, Si, S shift by a row in some
    # files — use label scanning for those.
    # ------------------------------------------------------------------
    if "B-FCE" in sheet_names:
        ws = wb["B-FCE"]
        ac = _find_act_col(ws, month_num, year_2d, 4, 5)
        if ac:
            cum_ac = _cum(ac)
            col = get_column_letter(ac)

            def _bfv(r):
                v = clean_val(ws.cell(row=r, column=ac).value)
                y = clean_val(ws.cell(row=r, column=cum_ac).value) if cum_ac else None
                l = str(ws.cell(row=r, column=2).value or '').strip()
                return v, y, l

            # ISP has only one active BF (BF-5), so furnace-level == shop-level.
            # _CROSS_BF maps ISP BF param name → (cross-plant section, unit, sort_order)
            _CROSS_BF = {
                "PCI Rate":          ("CDI",              "Kg/THM",   17),
                "Coke Rate":         ("BF Coke Rate",     "Kg/THM",   45),
                "Nut Coke Rate":     ("Nut Coke Rate",    "Kg/THM",   55),
                "Productivity":      ("BF Productivity",  "T/m³/day", 65),
                "Sinter in Burden":  ("Sinter in Burden", "%",       115),
                "Pellet in Burden":  ("Pellet in Burden", "%",       125),
                "Blast Temperature": ("HBT",              "°C",      100),
                "Si in HM":          ("Si in HM",         "%",        75),
                "S in HM":           ("S in HM",          "%",        85),
                "Oxygen enrichment": ("O2 Enrichment",    "%",       145),
            }

            def _isp_bf_param(isp_name, cross_name, cell_ref, v, y):
                if cross_name in _CROSS_BF:
                    sec, u, so = _CROSS_BF[cross_name]
                    techno_param_rows.append({
                        "group_code": "IRON_MAKING", "section": sec,
                        "parameter": "ISP", "unit": u,
                        "sort_order": so, "actual": v, "cum_actual": y,
                        "cell": cell_ref, "found_via": f"ISP B-FCE {isp_name}",
                        "status": "ok" if v is not None else "no value",
                    })

            # Stable rows
            for row, name, cross, unit in [
                (39, "Coke Rate",    "Coke Rate",    "Kg/THM"),
                (40, "Nut Coke Rate","Nut Coke Rate","Kg/THM"),
                (41, "PCI Rate",     "PCI Rate",     "Kg/THM"),
                (49, "Productivity", "Productivity", "T/M3/Day"),
            ]:
                v, y, l = _bfv(row)
                techno_rows.append(_tr("BF", name, unit, v, f"B-FCE!{col}{row}", l, y))
                _isp_bf_param(name, cross, f"B-FCE!{col}{row}", v, y)

            # Label-scanned rows (row numbers shift across files)
            def _bf_labeled(label, name, cross, unit, fallback):
                found = _find_label_rows(ws, label, count=1)
                r = found[0] if found else fallback
                v, y, l = _bfv(r)
                techno_rows.append(_tr("BF", name, unit, v, f"B-FCE!{col}{r}", l, y))
                _isp_bf_param(label, cross, f"B-FCE!{col}{r}", v, y)

            _bf_labeled("Sinter in Fe bearing", "Sinter% in Burden",  "Sinter in Burden",  "%",    51)
            _bf_labeled("Pellet in Fe bearing",  "Pellet% in Burden",  "Pellet in Burden",  "%",    52)
            _bf_labeled("Blast Temperature",      "Hot Blast Temp",    "HBT",               "°C",   80)
            _bf_labeled("Slag Rate",              "Slag Rate",          None,               "Kg/T", 81)
            _bf_labeled("Si in HM",               "Si in HM",          "Si in HM",          "%",    82)
            _bf_labeled("S in HM",                "S in HM",           "S in HM",           "%",    83)
            _bf_labeled("Sp Power Cons",          "Sp Power Cons",      None,               "KWh/T",91)
            _bf_labeled("Oxygen enrichment",      "O2 Enrichment",      None,               "%",    92)

    # ------------------------------------------------------------------
    # SINTER  (header row 3, ABP/ACT row 4)
    # "Specific Productivity" label — row shifts (77/79/81) across files
    # ------------------------------------------------------------------
    if "SINTER" in sheet_names:
        ws = wb["SINTER"]
        ac = _find_act_col(ws, month_num, year_2d, 3, 4)
        if ac:
            cum_ac = _cum(ac)
            col = get_column_letter(ac)
            found = _find_label_rows(ws, "Specific Productivity", count=1)
            r = found[0] if found else 81
            val = clean_val(ws.cell(row=r, column=ac).value)
            ytd_val = clean_val(ws.cell(row=r, column=cum_ac).value) if cum_ac else None
            lbl = str(ws.cell(row=r, column=2).value or "").strip()
            techno_rows.append(_tr("Sinter", "Sp Productivity", "T/M2/Utlz Hr", val, f"SINTER!{col}{r}", lbl, ytd_val))

    # ------------------------------------------------------------------
    # SMS / BOF-CCP  (header row 3, ABP/ACT row 4, days in row 2)
    # Row 10 — Total heats → Blows/day = value ÷ calendar days
    # ------------------------------------------------------------------
    if "SMS" in sheet_names:
        ws = wb["SMS"]
        ac = _find_act_col(ws, month_num, year_2d, 3, 4)
        if ac:
            cum_ac = _cum(ac)
            col = get_column_letter(ac)
            days = clean_val(ws.cell(row=2, column=ac).value) or 30
            heats = clean_val(ws.cell(row=10, column=ac).value)
            blows_day = round(heats / days, 2) if heats is not None else None
            cum_days = clean_val(ws.cell(row=2, column=cum_ac).value) if cum_ac else None
            cum_heats = clean_val(ws.cell(row=10, column=cum_ac).value) if cum_ac else None
            cum_blows = round(cum_heats / cum_days, 2) if (cum_heats is not None and cum_days) else None
            lbl10 = str(ws.cell(row=10, column=2).value or "").strip()
            techno_rows.append(_tr("SMS", "Blows per Day", "Nos/Day", blows_day,
                                   f"SMS!{col}10÷{int(days)}days", lbl10, cum_blows))
            # Write BOF cross-plant params to techno_param_rows so page 30 can display them
            avg_heat_val = clean_val(ws.cell(row=52, column=ac).value)
            avg_heat_cum = clean_val(ws.cell(row=52, column=cum_ac).value) if cum_ac else None
            o2_val  = clean_val(ws.cell(row=96, column=ac).value)
            o2_cum  = clean_val(ws.cell(row=96, column=cum_ac).value) if cum_ac else None
            for section, unit, val, cum in [
                ("Average Blows (Per Day)", "Nos.", blows_day,   cum_blows),
                ("Average Heat Weight",     "T",    avg_heat_val, avg_heat_cum),
                ("Oxygen Blowing",          "Nm³/TCS", o2_val,  o2_cum),
            ]:
                techno_param_rows.append({
                    "group_code": "SMS", "section": section,
                    "parameter": "ISP SMS", "unit": unit,
                    "sort_order": {"Average Blows (Per Day)": 8,
                                   "Average Heat Weight": 15,
                                   "Oxygen Blowing": 34}[section],
                    "actual": val, "cum_actual": cum,
                    "cell": f"SMS!{col}10 or R52/R96", "found_via": f"ISP SMS {section}",
                    "status": "ok" if val is not None else "no value",
                })
            sms_params = [
                (52,  "Avg Cast Wt",           "Mt/Cast"),
                (89,  "HM Consumption",        "Kg/TCS"),
                (90,  "Scrap Consumption",     "Kg/TCS"),
                (91,  "TMI",                   "Kg/TCS"),
                (96,  "Sp O2 Consumption",     "Nm³/T"),
            ]
            for row, name, unit in sms_params:
                val = clean_val(ws.cell(row=row, column=ac).value)
                ytd_val = clean_val(ws.cell(row=row, column=cum_ac).value) if cum_ac else None
                lbl = str(ws.cell(row=row, column=2).value or "").strip()
                techno_rows.append(_tr("SMS", name, unit, val, f"SMS!{col}{row}", lbl, ytd_val))

    # ------------------------------------------------------------------
    # WRM — Wire & Rod Mill  (header row 3, ABP/ACT row 4)
    # Anchor: "WR MILL YIELD" row.  Offsets verified across all 5 files:
    #   +5=% Mill Availability, +7=% Mill Utilisation,
    #   +12=Sp Power Consumption, +17=Rolling Rate on yield Wt
    # WRM has no Sp Heat row.
    # ------------------------------------------------------------------
    if "WRM" in sheet_names:
        ws = wb["WRM"]
        ac = _find_act_col(ws, month_num, year_2d, 3, 4)
        if ac:
            cum_ac = _cum(ac)
            col = get_column_letter(ac)
            techno_rows += _mill_params(ws, ac, cum_ac, col, "WRM",
                                        "WR MILL YIELD",
                                        avail_off=5, util_off=7,
                                        spow_off=12, rrate_off=17)

    # ------------------------------------------------------------------
    # BM — Bar Mill  (header row 3, ABP/ACT row 4)
    # Anchor: "BAR MILL YIELD" row.  Offsets verified across all 5 files:
    #   -3=Total Production, +9=% Mill Availability, +11=% Mill Utilisation,
    #   +14=Gross Heat Consumption (×1000→1000KCal/T),
    #   +16=Sp Power Consumption, +21=Rolling Rate on yield Wt
    # ------------------------------------------------------------------
    if "BM" in sheet_names:
        ws = wb["BM"]
        ac = _find_act_col(ws, month_num, year_2d, 3, 4)
        if ac:
            cum_ac = _cum(ac)
            col = get_column_letter(ac)
            techno_rows += _mill_params(ws, ac, cum_ac, col, "BM",
                                        "BAR MILL YIELD",
                                        avail_off=9, util_off=11,
                                        heat_off=14, spow_off=16, rrate_off=21,
                                        total_off=-3)

    # ------------------------------------------------------------------
    # USM — Universal Section Mill  (header row 3, ABP/ACT row 4)
    # Anchor: "US MILL YIELD" row (position varies greatly between old/new
    # files).  Same offsets as BM once the anchor is found.
    # ------------------------------------------------------------------
    if "USM" in sheet_names:
        ws = wb["USM"]
        ac = _find_act_col(ws, month_num, year_2d, 3, 4)
        if ac:
            cum_ac = _cum(ac)
            col = get_column_letter(ac)
            techno_rows += _mill_params(ws, ac, cum_ac, col, "USM",
                                        "US MILL YIELD",
                                        avail_off=9, util_off=11,
                                        heat_off=14, spow_off=16, rrate_off=21,
                                        total_off=-3)

    ok_count = sum(1 for r in techno_rows if r["status"] == "ok")
    if ok_count == 0:
        raise ValueError(
            f"No data found for month '{report_month}' in the ISP Summarized Monthly Report. "
            "Verify the month selector matches the reporting period of this file."
        )

    # Extract COKE OVENS techno_param_rows when sheet is present
    if "COKE OVENS" in sheet_names:
        try:
            techno_param_rows += _preview_coke_ovens(wb, db_report_month, month_num, year_2d)
            cp_ok = sum(1 for r in techno_param_rows if r["status"] == "ok")
            logger.info(f"ISP COKE OVENS: {cp_ok} param rows extracted for {db_report_month}.")
        except Exception as e:
            logger.warning(f"ISP: COKE OVENS extraction failed: {e}")

    # Extract Maj Techno Summ → MAJOR group params (authoritative plant-level summary)
    if "Maj Techno Summ" in sheet_names:
        try:
            mts_rows = _preview_maj_techno_summ(wb, db_report_month, month_num, year_2d, _co)
            techno_param_rows += mts_rows
            mts_ok = sum(1 for r in mts_rows if r["status"] == "ok")
            logger.info(f"ISP Maj Techno Summ: {mts_ok} MAJOR param rows extracted for {db_report_month}.")
        except Exception as e:
            logger.warning(f"ISP: Maj Techno Summ extraction failed: {e}")

    # Also extract production from 'Maj Production Summ' when present in the same workbook
    prod_rows = []
    if "Maj Production Summ" in sheet_names:
        try:
            prod_rows, _ = _preview_monthly_report_rows(wb, report_month)
        except Exception as e:
            logger.warning(f"ISP: could not extract production from 'Maj Production Summ': {e}")

    sheets_str = "B-FCE, SINTER, SMS, WRM, BM, USM"
    if "COKE OVENS" in sheet_names:
        sheets_str = "COKE OVENS, " + sheets_str
    if "Maj Techno Summ" in sheet_names:
        sheets_str = "Maj Techno Summ, " + sheets_str
    if prod_rows:
        sheets_str += ", Maj Production Summ"

    logger.info(f"ISP Summarized Monthly Report preview: {ok_count} techno, {len(prod_rows)} production values for {db_report_month}.")
    result = {
        "plant":             "ISP",
        "month":             db_report_month,
        "source_type":       "Summarized Monthly Report",
        "sheets":            sheets_str,
        "workbook_sheets":   sheet_names,
        "production_rows":   prod_rows,
        "techno_rows":       techno_rows,
        "techno_param_rows": techno_param_rows,
    }
    # Key present only when extracting an earlier same-FY month, so output for
    # the normal latest-month case stays byte-identical (golden-tested).
    if month_mismatch:
        result["month_mismatch"] = month_mismatch
    return result


# ---------------------------------------------------------------------------
# Preview dispatcher
# ---------------------------------------------------------------------------

def _extract_morning_stock(ws, report_month: str) -> list:
    """Extract opening stock from ISP DAILYREPORT1 sheet.

    Row 101 column headers: C='6 AM TODAY', D='01.MM.YY' (next month official), E='01.prev.YY'
    K5 = 1st of the month AFTER the report period (file generated on 1st of next month).

    Cell mapping (raw Tonnes → '000T, 3 d.p.):
      C108 → BLOOM/BILLETS FOR SALE,  next_month
      D108 → BLOOM/BILLETS FOR SALE,  report_month
      K102-C108 → BLOOM/BILLETS INPROCESS, next_month
      C105 → FINISHED STEEL, next_month
      D105 → FINISHED STEEL, report_month
      C110 → PIG IRON,       next_month
      D110 → PIG IRON,       report_month
    """
    y, m = int(report_month[:4]), int(report_month[5:7])
    next_month = f"{y+1 if m == 12 else y}-{1 if m == 12 else m+1:02d}"

    def _t(v):
        f = clean_val(v)
        return round(f / 1000, 3) if f is not None else None

    def _row(item_type, stock_type, stock_month, value, formula):
        return {
            "plant": "ISP", "item_type": item_type, "stock_type": stock_type,
            "stock_month": stock_month, "value": value, "formula": formula,
            "status": "ok" if value is not None else "skip",
        }

    c108 = clean_val(ws["C108"].value)
    k102 = clean_val(ws["K102"].value)
    bb_inp = round((k102 - c108) / 1000, 3) if (k102 is not None and c108 is not None) else None

    return [
        _row("BLOOM/BILLETS", "FOR SALE",  next_month,   _t(ws["C108"].value), "C108"),
        _row("BLOOM/BILLETS", "FOR SALE",  report_month, _t(ws["D108"].value), "D108"),
        _row("BLOOM/BILLETS", "INPROCESS", next_month,   bb_inp,               "K102-C108"),
        _row("FINISHED STEEL", "",         next_month,   _t(ws["C105"].value), "C105"),
        _row("FINISHED STEEL", "",         report_month, _t(ws["D105"].value), "D105"),
        _row("PIG IRON",      "",          next_month,   _t(ws["C110"].value), "C110"),
        _row("PIG IRON",      "",          report_month, _t(ws["D110"].value), "D110"),
    ]


def _fmt_ym(ym: str) -> str:
    """'2026-04' → 'April 2026'."""
    mn = {"01":"January","02":"February","03":"March","04":"April","05":"May","06":"June",
          "07":"July","08":"August","09":"September","10":"October","11":"November","12":"December"}
    return f"{mn.get(ym[5:7], ym[5:7])} {ym[:4]}"


def extract_preview(file_path: str, report_month: str) -> dict:
    """ISP preview: production or techno rows depending on file type. No DB writes.

    Detects file type by sheet names:
      DAILYREPORT1       → Morning Report (production_rows + stock_rows)
      Maj Production Summ → Final Monthly Report (production_rows)
      B-FCE              → Summarized Monthly Report (techno_rows)
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet_names = wb.sheetnames
    logger.info(f"ISP preview: loading file. Sheets: {sheet_names}")

    if "DAILYREPORT1" in sheet_names:
        ws = wb["DAILYREPORT1"]

        # ── File validation: K5 must be 1st of next_month(report_month) ──────
        k5_raw = ws["K5"].value
        if isinstance(k5_raw, datetime):
            k5_month = f"{k5_raw.year}-{k5_raw.month:02d}"
        elif k5_raw:
            dm = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", str(k5_raw))
            k5_month = f"{dm.group(3)}-{dm.group(2)}" if dm else None
        else:
            k5_month = None

        if report_month and k5_month:
            if k5_month != report_month:
                raise ValueError(
                    f"Month mismatch: file date (K5 = {k5_month}) means this is a "
                    f"{_fmt_ym(k5_month)} month-end report, but you selected {_fmt_ym(report_month)}. "
                    f"Please select '{_fmt_ym(k5_month)}' or upload the correct month-end file."
                )

        rows, db_report_month = _preview_morning_report_rows(wb)
        stock_rows = _extract_morning_stock(ws, report_month or db_report_month)

        ok_s = sum(1 for r in stock_rows if r["status"] == "ok")
        logger.info("ISP morning stock: %d/%d ok for %s", ok_s, len(stock_rows), report_month)

        return {
            "plant":             "ISP",
            "month":             db_report_month,
            "source_type":       "Daily Morning Report",
            "sheets":            "DAILYREPORT1",
            "workbook_sheets":   sheet_names,
            "production_rows":   rows,
            "techno_rows":       [],
            "techno_param_rows": [],
            "stock_rows":        stock_rows,
        }

    # Check B-FCE before Maj Production Summ — the Summarized Monthly Report
    # contains BOTH sheets; the standalone Final Monthly Report has only the
    # latter. The 2016-17 vintage names its BF sheet just "BF" (and coke ovens
    # "CO") — require Maj Techno Summ alongside so a bare "BF" sheet in some
    # unrelated workbook can't false-positive.
    if "B-FCE" in sheet_names or ("BF" in sheet_names and "Maj Techno Summ" in sheet_names):
        return _preview_summarized_monthly(wb, report_month, sheet_names)

    if "Maj Production Summ" in sheet_names:
        rows, db_report_month = _preview_monthly_report_rows(wb, report_month)
        return {
            "plant":             "ISP",
            "month":             db_report_month,
            "source_type":       "Final Monthly Report",
            "sheets":            "Maj Production Summ",
            "workbook_sheets":   sheet_names,
            "production_rows":   rows,
            "techno_rows":       [],
            "techno_param_rows": [],
        }

    raise ValueError(
        "Uploaded ISP file does not match any known format. "
        "Expected sheet 'DAILYREPORT1' (Morning Report), "
        "'Maj Production Summ' (Final Monthly Report), or "
        "'B-FCE' (Summarized Monthly Report)."
    )

