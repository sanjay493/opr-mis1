"""
Cost Trend Excel extractor — reads one "ELHM CS SS <...>.xlsx" elementwise
cost-of-production workbook (kept in Report_format/Cost/, one per month plus
one cumulative "APRIL-<month>" file per month after April) and pulls out the
per-plant VARIABLE and FIXED cost (Rs/T) from each sheet's "TOTAL COST" row,
for HM/CS/SS and all 6 plant blocks (BSP/DSP/RSP/BSL/ISP/SAIL).

Feeds db.cost_trend_monthly — see page_cost_trend.py and frontend/src/app/
data-entry/cost-trend for how this data is used/entered manually; this
extractor is the automated alternative for months a source workbook exists
for. TOTAL COST itself is never extracted/stored — it stays computed
(VARIABLE + FIXED) in page_cost_trend.py, same as the manual-entry path.

Workbook layout (consistent across HM/CS/SS sheets and both file kinds,
confirmed against Report_format/Cost/*.xlsx):
  - Row 3 holds a cell reading "<Month>'<YYYY>" for a single month's figures,
    or "UPTO <Month>'<YYYY>" for a cumulative (till-month) file — this is
    the authoritative source of report_month/kind, not the filename.
  - Row 6 holds plant labels (BSP, DSP, RSP, BSL, ISP, "SAIL (5 ISP'S)"),
    one per 5-column block (PRICE, USAGE, VARIABLE, FIXED, COST) starting
    at column C.
  - The row whose column B reads "TOTAL COST (...)" holds the final blended
    VARIABLE/FIXED Rs/T figures extracted here. SAIL's block is its own
    reported blended rate — confirmed NOT a sum of the other 5 blocks — so
    it's extracted the same way as any other plant, not derived.
"""
import re

import openpyxl

PLANT_ORDER = ["BSP", "DSP", "RSP", "BSL", "ISP", "SAIL"]
_BLOCK_START_COLS = [3, 8, 13, 18, 23, 28]  # C, H, M, R, W, AB (1-indexed)
_VARIABLE_OFFSET = 2  # PRICE(+0), USAGE(+1), VARIABLE(+2), FIXED(+3), COST(+4)
_FIXED_OFFSET = 3

_MONTH_ABBR = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}
_HEADER_RE = re.compile(r"([A-Z]{3,})'?\s*(\d{4})")


def _parse_header(text: str):
    """"<MONTH>'<YYYY>" or "UPTO <MONTH>'<YYYY>" -> (report_month 'YYYY-MM',
    is_till_month). Returns (None, None) if text doesn't match."""
    if not text:
        return None, None
    t = str(text).strip().upper().replace("’", "'")
    is_till = t.startswith("UPTO")
    m = _HEADER_RE.search(t)
    if not m:
        return None, None
    mon = _MONTH_ABBR.get(m.group(1)[:3])
    if not mon:
        return None, None
    return f"{int(m.group(2))}-{mon:02d}", is_till


def _find_month_cell_text(ws):
    """Row 3 holds the header; scan its first columns for the month text
    rather than trusting one fixed column (observed to vary by sheet)."""
    for c in range(1, 20):
        v = ws.cell(row=3, column=c).value
        if v and re.search(r"[A-Za-z]{3,}'?\s*\d{4}", str(v)):
            return str(v)
    return None


def _find_total_cost_row(ws):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if v and "TOTAL COST" in str(v).upper():
            return r
    return None


def _validate_plant_label(ws, block_start_col: int, expected: str) -> bool:
    """expected's label isn't always at the same offset within its 5-column
    block (SAIL's sits on the block's first column, everyone else's on the
    second) — scan the whole block rather than assuming one offset."""
    for c in range(block_start_col, block_start_col + 5):
        v = ws.cell(row=6, column=c).value
        if v and str(v).strip().upper().startswith(expected):
            return True
    return False


def _num(v):
    return round(v, 3) if isinstance(v, (int, float)) else None


def extract_cost_trend_workbook(file_path) -> dict:
    """-> {"report_month": "YYYY-MM", "is_till_month": bool,
           "products": {"HM": {"BSP": {"variable":.., "fixed":..}, ...,
                                "SAIL": {...}},
                        "CS": {...}, "SS": {...}}}
    Raises ValueError on anything that doesn't match the expected layout —
    this feeds financial figures, so a column/row mismatch must fail loudly
    rather than silently extracting the wrong cell."""
    wb = openpyxl.load_workbook(file_path, data_only=True)

    report_month = None
    is_till_month = None
    products = {}

    for sheet_name in wb.sheetnames:
        key = sheet_name.strip().upper()
        if key not in ("HM", "CS", "SS"):
            continue
        ws = wb[sheet_name]

        header_text = _find_month_cell_text(ws)
        rm, till = _parse_header(header_text)
        if rm is None:
            raise ValueError(
                f"Sheet '{sheet_name}': could not parse a report month from row 3 (found: {header_text!r})"
            )
        if report_month is None:
            report_month, is_till_month = rm, till
        elif (rm, till) != (report_month, is_till_month):
            raise ValueError(
                f"Sheet '{sheet_name}' header ({rm}, till={till}) disagrees with an earlier "
                f"sheet in this workbook ({report_month}, till={is_till_month})"
            )

        total_row = _find_total_cost_row(ws)
        if total_row is None:
            raise ValueError(f"Sheet '{sheet_name}': no 'TOTAL COST' row found in column B")

        plants = {}
        for plant, block_start in zip(PLANT_ORDER, _BLOCK_START_COLS):
            if not _validate_plant_label(ws, block_start, plant):
                raise ValueError(
                    f"Sheet '{sheet_name}': expected plant '{plant}' in the column block "
                    f"starting at column {block_start}, but its row-6 label doesn't match"
                )
            plants[plant] = {
                "variable": _num(ws.cell(row=total_row, column=block_start + _VARIABLE_OFFSET).value),
                "fixed": _num(ws.cell(row=total_row, column=block_start + _FIXED_OFFSET).value),
            }
        products[key] = plants

    if not products:
        raise ValueError("No HM/CS/SS sheet found in this workbook")

    return {"report_month": report_month, "is_till_month": is_till_month, "products": products}
