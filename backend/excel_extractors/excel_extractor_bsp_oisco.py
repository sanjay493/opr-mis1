"""BSP OISCO Techno-Economic Parameters extractor — OISCO_<Mon>'YY.xlsx

Sheet layout (single sheet "Sheet1"):
  Row  3, Col C : Title contains report month, e.g. "TECHNO ECONOMIC PARAMETERS_MAY'25"
  Row  6, Cols E+ : Rolling month headers, e.g. APR | MAY | CUM.
  Col  C (3)    : Parameter label
  Col  D (4)    : Unit
  Data col      : Column in row 6 matching user-selected month abbreviation
  CUM  col      : Always one column to the right of the data column

Coverage (35 parameters):
  BF Operations  → IRON_MAKING / BF Operations (BSP)
  SMS-II Ops     → BOF / SMS-II Operations (BSP)
  SMS-III Ops    → BOF / SMS-III Operations (BSP)
  Utilities      → IRON_MAKING / Utilities (BSP)
"""

import logging
import re
from typing import Optional, Dict, Any, List
import openpyxl
from openpyxl.utils import get_column_letter

logger = logging.getLogger("excel_extractor")

# ---------------------------------------------------------------------------
# Month constants
# ---------------------------------------------------------------------------
_MONTH_NAME_TO_NUM = {
    "JAN": "01", "FEB": "02", "MAR": "03", "APR": "04",
    "MAY": "05", "JUN": "06", "JUL": "07", "AUG": "08",
    "SEP": "09", "OCT": "10", "NOV": "11", "DEC": "12",
}
_MONTH_NUM_TO_HDR = {
    "04": "APR", "05": "MAY", "06": "JUN", "07": "JUL",
    "08": "AUG", "09": "SEP", "10": "OCT", "11": "NOV",
    "12": "DEC", "01": "JAN", "02": "FEB", "03": "MAR",
}
_MONTH_FULL = {
    "01": "January", "02": "February", "03": "March", "04": "April",
    "05": "May", "06": "June", "07": "July", "08": "August",
    "09": "September", "10": "October", "11": "November", "12": "December",
}

HEADER_ROW = 6   # row containing month abbreviations
LABEL_COL  = 3   # col C: parameter label
UNIT_COL   = 4   # col D: unit


# ---------------------------------------------------------------------------
# Parameter map (row_1based, group_code, section, label, unit)
# sort_order = index × 10
# ---------------------------------------------------------------------------
PARAM_MAP: List[tuple] = [
    # ── Blast Furnace Operations (rows 9-20) ─────────────────────────────
    ( 9, "IRON_MAKING", "BF Operations (BSP)", "CDI Rate (Shop 1-8)",     "Kg/THM"),
    (11, "IRON_MAKING", "BF Operations (BSP)", "CDI BF-4",                "Kg/THM"),
    (12, "IRON_MAKING", "BF Operations (BSP)", "CDI BF-5",                "Kg/THM"),
    (13, "IRON_MAKING", "BF Operations (BSP)", "CDI BF-6",                "Kg/THM"),
    (14, "IRON_MAKING", "BF Operations (BSP)", "CDI BF-7",                "Kg/THM"),
    (15, "IRON_MAKING", "BF Operations (BSP)", "CDI BF-8",                "Kg/THM"),
    (16, "IRON_MAKING", "BF Operations (BSP)", "Fuel Rate (Shop 1-8)",    "Kg/THM"),
    (17, "IRON_MAKING", "BF Operations (BSP)", "Pellet % in Burden",      "%"),
    (18, "IRON_MAKING", "BF Operations (BSP)", "LD Slag Usage (Shop 1-8)","Kg/THM"),
    (19, "IRON_MAKING", "BF Operations (BSP)", "Not Dry Cast (Shop 1-8)", "%"),
    (20, "IRON_MAKING", "BF Operations (BSP)", "Coal to HM Ratio",        "Ratio"),
    # ── SMS-II Operations (rows 22-36) ──────────────────────────────────
    (22, "BOF", "SMS-II Operations (BSP)", "Converter Availability",      "% ICH"),
    (23, "BOF", "SMS-II Operations (BSP)", "Converter Utilisation",       "% Avail hr"),
    (24, "BOF", "SMS-II Operations (BSP)", "Tap to Tap Time",             "Minutes"),
    (25, "BOF", "SMS-II Operations (BSP)", "Average Blows/Day",           "Heats/Day"),
    (26, "BOF", "SMS-II Operations (BSP)", "Average Heat Weight",         "Tonnes"),
    (27, "BOF", "SMS-II Operations (BSP)", "Avg. Lining Life",            "Heats"),
    (28, "BOF", "SMS-II Operations (BSP)", "Fe-Mn Consumption",           "Kg/TCS"),
    (29, "BOF", "SMS-II Operations (BSP)", "Fe-Si Consumption",           "Kg/TCS"),
    (30, "BOF", "SMS-II Operations (BSP)", "Si-Mn Consumption",           "Kg/TCS"),
    (31, "BOF", "SMS-II Operations (BSP)", "Oxygen Consumption",          "NM3/TCS"),
    (32, "BOF", "SMS-II Operations (BSP)", "Alumina Consumption",         "Kg/TCS"),
    (35, "BOF", "SMS-II Operations (BSP)", "LD Gas Recovery",             "CuM/T"),
    (36, "BOF", "SMS-II Operations (BSP)", "DS Heats",                    "Nos"),
    # ── SMS-III Operations (rows 38-51) ─────────────────────────────────
    (38, "BOF", "SMS-III Operations (BSP)", "Converter Availability",     "% ICH"),
    (39, "BOF", "SMS-III Operations (BSP)", "Converter Utilisation",      "% Avail hr"),
    (40, "BOF", "SMS-III Operations (BSP)", "Tap to Tap Time",            "Minutes"),
    (41, "BOF", "SMS-III Operations (BSP)", "Average Blows/Day",          "Heats/Day"),
    (42, "BOF", "SMS-III Operations (BSP)", "Average Heat Weight",        "Tonnes"),
    (43, "BOF", "SMS-III Operations (BSP)", "Fe-Mn Consumption",          "Kg/TCS"),
    (44, "BOF", "SMS-III Operations (BSP)", "Fe-Si Consumption",          "Kg/TCS"),
    (45, "BOF", "SMS-III Operations (BSP)", "Si-Mn Consumption",          "Kg/TCS"),
    (46, "BOF", "SMS-III Operations (BSP)", "Oxygen Consumption",         "NM3/TCS"),
    (47, "BOF", "SMS-III Operations (BSP)", "Alumina Consumption",        "Kg/TCS"),
    (50, "BOF", "SMS-III Operations (BSP)", "LD Gas Recovery",            "CuM/T"),
    (51, "BOF", "SMS-III Operations (BSP)", "DS Heats",                   "Nos"),
    # ── Utilities (row 54) ───────────────────────────────────────────────
    (54, "IRON_MAKING", "Utilities (BSP)", "Sp. Water Consumption",       "CuM/TCS"),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _clean(v) -> Optional[float]:
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "-", "—", "nan", "###", "#DIV/0!", "#VALUE!", "#N/A"):
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _detect_month_from_title(ws) -> Optional[str]:
    """Parse month abbreviation from title cell R3/C3, e.g. '...PARAMETERS_MAY'25'."""
    raw = str(ws.cell(3, 3).value or "").upper()
    # Match 3-letter month abbreviation at or near the end
    m = re.search(r"_([A-Z]{3})'?\d{2}", raw)
    if m:
        abbr = m.group(1)
        return _MONTH_NAME_TO_NUM.get(abbr)
    # Fallback: look for any 3-letter month name
    for abbr, num in _MONTH_NAME_TO_NUM.items():
        if abbr in raw:
            return num
    return None


def _detect_columns(ws, m_num: str):
    """Scan row 6 for the matching month abbreviation.
    Returns (data_col, cum_col) as 1-based column indices.
    CUM is always the column immediately after the data column.
    """
    expected = _MONTH_NUM_TO_HDR.get(m_num, "")
    for c in range(1, ws.max_column + 2):
        v = str(ws.cell(HEADER_ROW, c).value or "").strip().upper()
        if v == expected:
            return c, c + 1
    # Build helpful error
    found_hdrs = [
        str(ws.cell(HEADER_ROW, c).value or "").strip()
        for c in range(1, ws.max_column + 1)
        if ws.cell(HEADER_ROW, c).value
    ]
    raise ValueError(
        f"Month header '{expected}' not found in row {HEADER_ROW}. "
        f"Found: {found_hdrs}. Ensure the file is for the selected month."
    )


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def extract_preview(file_path: str, report_month: str) -> Dict[str, Any]:
    """
    Extract BSP OISCO techno parameters from OISCO_<Mon>'YY.xlsx.

    Args:
        file_path:    Path to the uploaded .xlsx file.
        report_month: 'YYYY-MM' string selected by user.

    Returns:
        Preview dict compatible with /api/confirm-extraction, with
        techno_param_rows containing all 35 mapped parameters.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active

    # ── Determine report month ────────────────────────────────────────────
    m_num: Optional[str] = None
    y_str: Optional[str] = None

    if report_month and len(report_month) >= 7 and report_month[4] == "-":
        y_str = report_month[:4]
        m_num = report_month[5:7]
    else:
        m_num = _detect_month_from_title(ws)
        import datetime
        y_str = str(datetime.date.today().year)

    if not m_num or m_num not in _MONTH_NUM_TO_HDR:
        raise ValueError(
            f"Cannot determine report month (got {m_num!r}). "
            "Provide a valid YYYY-MM month."
        )

    db_month   = f"{y_str}-{m_num}"
    month_full = _MONTH_FULL.get(m_num, m_num)

    # ── Locate data column and CUM column from row 6 ─────────────────────
    data_col, cum_col = _detect_columns(ws, m_num)
    data_ltr = get_column_letter(data_col)
    cum_ltr  = get_column_letter(cum_col)

    # Verify file month matches user selection
    file_m_num = _detect_month_from_title(ws)
    header_ok = (file_m_num == m_num)
    if not header_ok:
        logger.warning(
            "OISCO: file title month %r does not match user-selected month %r. Proceeding.",
            file_m_num, m_num,
        )

    # ── Extract rows ──────────────────────────────────────────────────────
    rows_out: List[Dict[str, Any]] = []

    for sort_idx, (row_1b, group, section, param, unit) in enumerate(PARAM_MAP):
        actual_val = _clean(ws.cell(row_1b, data_col).value)
        cum_val    = _clean(ws.cell(row_1b, cum_col).value)
        file_label = str(ws.cell(row_1b, LABEL_COL).value or "").strip().replace("\n", " ")
        cell_ref   = f"{data_ltr}{row_1b}/{cum_ltr}{row_1b}"
        status     = "ok" if (actual_val is not None or cum_val is not None) else "skip"

        rows_out.append({
            "group_code": group,
            "section":    section,
            "parameter":  param,
            "unit":       unit,
            "actual":     actual_val,
            "cum_actual": cum_val,
            "sort_order": sort_idx * 10,
            "cell":       cell_ref,
            "file_label": file_label,
            "plant":      "BSP",
            "month":      db_month,
            "found_via":  f"hardcoded R{row_1b}",
            "status":     status,
        })

    ok_count = sum(1 for r in rows_out if r["status"] == "ok")
    logger.info(
        "BSP OISCO preview: %d/%d rows ok for %s (col %s, cum col %s)",
        ok_count, len(rows_out), db_month, data_ltr, cum_ltr,
    )

    return {
        "source_type":        "BSP-OISCO-Techno.xlsx",
        "month":              db_month,
        "plant":              "BSP",
        "workbook_sheets":    wb.sheetnames,
        "month_col_letter":   data_ltr,
        "cum_col_letter":     cum_ltr,
        "header_verified":    header_ok,
        "file_month":         month_full,
        "production_rows":    [],
        "techno_rows":        [],
        "techno_param_rows":  rows_out,
        "special_steel_rows": [],
    }
