"""
API endpoints for BSP Techno data (BSP-3-page-Tech.xlsx, OISCO Excel and the
BSP Flash monthly PDF).

All file types extract data into the same techno_data table with plant='BSP'.
Multiple files can be uploaded for the same month — parameters are merged
(merge_upsert_techno_data) so one file doesn't wipe the other's data.

Endpoints (prefix /api/bsp-techno):
  POST /preview/techno    — preview from BSP-3-page-Tech.xlsx (no DB write)
  POST /preview/oisco     — preview from OISCO Excel (no DB write)
  POST /preview/flash-pdf — preview from BSP flash-<mon>YY.pdf (no DB write)
  POST /insert            — save previewed records to DB
  POST /extract/techno    — extract + save in one step (BSP-3-page-Tech.xlsx)
  POST /extract/oisco     — extract + save in one step (OISCO Excel)
  POST /inspect           — inspect file structure (sheet names, header rows)
"""

import sys
import tempfile
from pathlib import Path

from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from openpyxl import load_workbook

_TP_DIR = str(Path(__file__).parent / "techno_project")
if _TP_DIR not in sys.path:
    sys.path.insert(0, _TP_DIR)
_EX_DIR = str(Path(__file__).parent / "excel_extractors")
if _EX_DIR not in sys.path:
    sys.path.insert(0, _EX_DIR)

from bsp_extractor import BspTechnoExtractor            # noqa: E402
from bsp_oisco_extractor import BspOiscoExtractor       # noqa: E402
from db import init_db, merge_upsert_techno_data, enrich_techno_records_with_db  # noqa: E402

router = APIRouter(prefix="/api/bsp-techno", tags=["bsp-techno"])


def _validate_month(report_month: str):
    try:
        y, m = report_month.split("-")
        assert len(y) == 4 and 1 <= int(m) <= 12
    except Exception:
        raise HTTPException(
            status_code=400,
            detail="report_month must be in YYYY-MM format, e.g. '2026-05'",
        )


def _save_temp(file: UploadFile, content: bytes) -> Path:
    suffix = Path(file.filename or "upload.xlsx").suffix or ".xlsx"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tmp.write(content)
    tmp.close()
    return Path(tmp.name)


def _preview_response(records, report_month, source_file, source_type):
    total_params = sum(len(r["techno_json"].get("month", {})) for r in records)
    preview_records = [{"unit": r["unit"], "techno_json": r["techno_json"]} for r in records]
    # Attach current DB values so the UI can show DB-vs-extracted side by side
    # (month and cumulative) before the user confirms the insert.
    enrich_techno_records_with_db(preview_records, "BSP", report_month)
    return {
        "status":          "preview",
        "source_type":     source_type,
        "report_month":    report_month,
        "source_file":     source_file,
        "units_extracted": len(records),
        "total_params":    total_params,
        "records": preview_records,
    }


def _unlink(p: Path):
    try:
        p.unlink()
    except OSError:
        pass


# ── Preview ───────────────────────────────────────────────────────────────────

@router.post("/preview/techno")
async def preview_bsp_techno(
    file: UploadFile = File(...),
    report_month: str = Form(...),
):
    """Preview BSP-3-page-Tech.xlsx — does NOT write to DB."""
    _validate_month(report_month)
    tmp = _save_temp(file, await file.read())
    try:
        records = BspTechnoExtractor(str(tmp), report_month=report_month).extract()
        if not records:
            raise HTTPException(status_code=422,
                detail="No data extracted. Use /inspect to check the file structure, "
                       "then verify row numbers in bsp_techno_map.json.")
        return _preview_response(records, report_month, file.filename or "", "BSP-3-page-Tech.xlsx")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        _unlink(tmp)


@router.post("/preview/oisco")
async def preview_bsp_oisco(
    file: UploadFile = File(...),
    report_month: str = Form(...),
):
    """Preview BSP OISCO Excel — does NOT write to DB."""
    _validate_month(report_month)
    tmp = _save_temp(file, await file.read())
    try:
        records = BspOiscoExtractor(str(tmp), report_month=report_month).extract()
        if not records:
            raise HTTPException(status_code=422,
                detail="No data extracted. Use /inspect to check the file structure, "
                       "then verify row numbers in bsp_oisco_map.json.")
        return _preview_response(records, report_month, file.filename or "", "BSP-OISCO-Techno.xlsx")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        _unlink(tmp)


@router.post("/preview/flash-pdf")
async def preview_bsp_flash_pdf(
    file: UploadFile = File(...),
    report_month: str = Form(...),
):
    """Preview techno params from the BSP flash monthly PDF — does NOT write
    to DB.  The month printed in the PDF wins over the selected month (the
    response's report_month is the one /insert will save under)."""
    _validate_month(report_month)
    tmp = _save_temp(file, await file.read())
    try:
        from pdf_extractor_bsp_flash import extract_techno_records
        records = extract_techno_records(str(tmp), report_month=report_month)
        if not records:
            raise HTTPException(status_code=422,
                detail="No techno data extracted — is this a BSP flash monthly "
                       "PDF (flash-<mon>YY.pdf)?")
        # extract_techno_records resolves the month from the PDF cover
        resolved_month = records[0]["report_month"]
        return _preview_response(records, resolved_month, file.filename or "",
                                 "BSP Flash Monthly PDF")
    except HTTPException:
        raise
    except ValueError as ve:
        raise HTTPException(status_code=422, detail=str(ve))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        _unlink(tmp)


# ── Insert (after preview confirm) ───────────────────────────────────────────

@router.post("/insert")
async def insert_bsp_techno(payload: dict):
    """Save previewed BSP records to DB via merge-upsert.
    Body: { report_month, source_file, records: [{unit, techno_json}] }
    """
    report_month = payload.get("report_month", "")
    source_file  = payload.get("source_file",  "")
    records      = payload.get("records",       [])
    _validate_month(report_month)
    if not records:
        raise HTTPException(status_code=400, detail="No records to insert")
    from api_unified_techno import validate_units_for_plant
    validate_units_for_plant("BSP", (rec.get("unit", "") for rec in records))
    try:
        init_db()
        for rec in records:
            merge_upsert_techno_data(
                plant="BSP", report_month=report_month,
                unit=rec["unit"], new_techno_json=rec["techno_json"],
                source_file=source_file,
            )
        return {"status": "ok", "report_month": report_month, "units_saved": len(records)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Extract-and-save (one step) ───────────────────────────────────────────────

@router.post("/extract/techno")
async def extract_bsp_techno(
    file: UploadFile = File(...),
    report_month: str = Form(...),
):
    """Extract BSP-3-page-Tech.xlsx and immediately save (merge-upsert)."""
    _validate_month(report_month)
    tmp = _save_temp(file, await file.read())
    try:
        records = BspTechnoExtractor(str(tmp), report_month=report_month).extract()
        if not records:
            raise HTTPException(status_code=422,
                detail="No data extracted. Use POST /api/bsp-techno/inspect to check "
                       "sheet names and header rows, then fix bsp_techno_map.json.")
        init_db()
        for rec in records:
            merge_upsert_techno_data(
                plant="BSP", report_month=rec["report_month"],
                unit=rec["unit"], new_techno_json=rec["techno_json"],
                source_file=file.filename or "",
            )
        return {
            "status": "ok", "source_type": "BSP-3-page-Tech.xlsx",
            "report_month": report_month,
            "units_extracted": len(records),
            "units": [r["unit"] for r in records],
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        _unlink(tmp)


@router.post("/extract/oisco")
async def extract_bsp_oisco(
    file: UploadFile = File(...),
    report_month: str = Form(...),
):
    """Extract BSP OISCO Excel and immediately save (merge-upsert)."""
    _validate_month(report_month)
    tmp = _save_temp(file, await file.read())
    try:
        records = BspOiscoExtractor(str(tmp), report_month=report_month).extract()
        if not records:
            raise HTTPException(status_code=422,
                detail="No data extracted. Use POST /api/bsp-techno/inspect to check "
                       "sheet names and header rows, then fix bsp_oisco_map.json.")
        init_db()
        for rec in records:
            merge_upsert_techno_data(
                plant="BSP", report_month=rec["report_month"],
                unit=rec["unit"], new_techno_json=rec["techno_json"],
                source_file=file.filename or "",
            )
        return {
            "status": "ok", "source_type": "BSP-OISCO-Techno.xlsx",
            "report_month": report_month,
            "units_extracted": len(records),
            "units": [r["unit"] for r in records],
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        _unlink(tmp)


# ── Diagnostic ────────────────────────────────────────────────────────────────

@router.post("/inspect")
async def inspect_bsp_file(file: UploadFile = File(...)):
    """
    Upload any BSP Excel file and get back its structure:
    sheet names, rows 1-11 of the active sheet (all non-empty cells per row).
    Use this to confirm sheet name, header row layout, and first data row
    before fixing row numbers in bsp_techno_map.json or bsp_oisco_map.json.
    """
    tmp = _save_temp(file, await file.read())
    try:
        wb = load_workbook(str(tmp), data_only=True)
        sheets_out = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_preview = {}
            for r in range(1, 12):
                row_vals = []
                for c in range(1, 25):
                    v = ws.cell(r, c).value
                    if v is not None:
                        row_vals.append({"col": c, "val": str(v)[:80]})
                if row_vals:
                    rows_preview[f"row_{r}"] = row_vals
            sheets_out[sheet_name] = rows_preview
        return {"filename": file.filename, "sheet_names": wb.sheetnames, "sheets": sheets_out}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        _unlink(tmp)
