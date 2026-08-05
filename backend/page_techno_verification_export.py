"""
Excel / PDF export for Techno Verification (frontend: /reports/techno-verification)
— Reported vs Calculated cumulative for every MAJOR page (27) parameter, every
plant plus the SAIL rollup. Source: page_techno.generate_major_techno_verification.

Visual style mirrors page_techno_custom_export.py (blue header, light-blue
subheader, zebra rows, amber fill on deviating cells) for consistency with
the rest of the app; `render_pdf_bytes` is reused rather than duplicated.
"""
import io

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from page_techno_custom_export import render_pdf_bytes  # noqa: F401  (re-exported for main.py)

_SECTION_FILL = PatternFill("solid", fgColor="1A73E8")
_SECTION_FONT = Font(bold=True, color="FFFFFF", size=10)
_SUBHDR_FILL = PatternFill("solid", fgColor="E8F0FE")
_SUBHDR_FONT = Font(bold=True, color="174EA6", size=9)
_SAIL_FILL = PatternFill("solid", fgColor="F9AB00")
_SAIL_FONT = Font(bold=True, color="3C2F00", size=9)
_ZEBRA_FILL = PatternFill("solid", fgColor="F8F9FA")
_DEV_FILL = PatternFill("solid", fgColor="FBBC04")
_THIN = Side(style="thin", color="DADCE0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_SAIL_DEV_BORDER = Border(left=Side(style="medium", color="EA4335"), right=Side(style="medium", color="EA4335"),
                           top=Side(style="medium", color="EA4335"), bottom=Side(style="medium", color="EA4335"))


def _cell_str(v):
    if v is None or v == "":
        return ""
    return str(v)


def build_excel_bytes(data: dict, month: str) -> bytes:
    month_labels = data.get("month_labels") or []
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Techno Verification"[:31]

    ws.cell(row=1, column=1, value=data.get("title", "MAJOR TECHNO-ECONOMIC PARAMETERS — VERIFICATION")).font = Font(bold=True, size=13)
    ws.cell(row=2, column=1, value=f"Report month: {month} — {data.get('subtitle', '')}").font = Font(italic=True, size=9)

    cum_label = data.get("cum_label", "Cum")
    headers = (["Parameter / Plant", "Unit"] + month_labels
               + [f"{cum_label} (Reported)", f"{cum_label} (Calculated)"])
    total_cols = len(headers)

    row = 4
    for sec in data.get("sections", []):
        sc = ws.cell(row=row, column=1, value=sec.get("label", ""))
        sc.font = _SECTION_FONT
        sc.fill = _SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        for c in range(2, total_cols + 1):
            ws.cell(row=row, column=c).fill = _SECTION_FILL
        row += 1

        for c, h in enumerate(headers, start=1):
            hc = ws.cell(row=row, column=c, value=h)
            hc.font = _SUBHDR_FONT
            hc.fill = _SUBHDR_FILL
            hc.alignment = Alignment(horizontal="center")
            hc.border = _BORDER
        row += 1

        for idx, r in enumerate(sec.get("rows", [])):
            is_sail = r.get("label") == "SAIL"
            deviates = bool(r.get("deviation"))
            fill = _SAIL_FILL if is_sail else (_ZEBRA_FILL if idx % 2 == 1 else None)
            font = _SAIL_FONT if is_sail else Font(size=9)
            values = ([r.get("label", ""), r.get("unit", "")]
                      + (r.get("months") or [])
                      + [r.get("reported", ""), r.get("calculated", "")])
            calc_col = total_cols  # last column = Calculated
            for c, v in enumerate(values, start=1):
                vc = ws.cell(row=row, column=c, value=_cell_str(v))
                vc.font = font
                vc.border = _SAIL_DEV_BORDER if (deviates and is_sail) else _BORDER
                vc.alignment = Alignment(horizontal="left" if c == 1 else "right")
                if c == calc_col and deviates and not is_sail:
                    vc.fill = _DEV_FILL
                elif fill:
                    vc.fill = fill
            row += 1
        row += 1  # blank spacer row between sections

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    for i in range(3, total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 10
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf_html(data: dict, month: str) -> str:
    month_labels = data.get("month_labels") or []
    cum_label = data.get("cum_label", "Cum")
    headers = (["Parameter / Plant", "Unit"] + month_labels
               + [f"{cum_label} (Reported)", f"{cum_label} (Calculated)"])
    header_html = "".join(f"<th>{h}</th>" for h in headers)

    sections_html = []
    for sec in data.get("sections", []):
        body_rows = []
        for r in sec.get("rows", []):
            is_sail = r.get("label") == "SAIL"
            deviates = bool(r.get("deviation"))
            values = ([r.get("label", ""), r.get("unit", "")]
                      + (r.get("months") or [])
                      + [r.get("reported", ""), r.get("calculated", "")])
            cells = []
            last_idx = len(values) - 1
            for i, v in enumerate(values):
                cls = "plant" if i == 0 else ""
                if i == last_idx and deviates:
                    cls = (cls + " dev-sail" if is_sail else cls + " dev").strip()
                cells.append(f'<td class="{cls}">{_cell_str(v)}</td>')
            body_rows.append(f'<tr class="{"sail-row" if is_sail else ""}">{"".join(cells)}</tr>')
        sections_html.append(
            f'<div class="section-title">{sec.get("label","")} ({sec.get("unit","")})</div>'
            f'<table><thead><tr>{header_html}</tr></thead>'
            f'<tbody>{"".join(body_rows)}</tbody></table>'
        )

    return f"""<!doctype html>
<html><head><meta charset="utf-8">
<style>
  @page {{ size: A4 landscape; margin: 10mm 8mm; }}
  * {{ box-sizing: border-box; }}
  body {{ font-family: Arial, sans-serif; color: #202124; margin: 0; }}
  h1 {{ font-size: 14pt; margin: 0 0 2px 0; }}
  .subtitle {{ font-size: 9pt; color: #5f6368; margin: 0 0 8px 0; }}
  .section-title {{ background: #1a73e8; color: #fff; font-weight: 700; font-size: 9pt;
    padding: 4px 6px; margin-top: 10px; page-break-after: avoid; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 6.8pt; margin-bottom: 4px; }}
  thead {{ display: table-header-group; }}
  tr {{ page-break-inside: avoid; }}
  th, td {{ border: 1px solid #dadce0; padding: 2px 3px; text-align: right; white-space: nowrap; }}
  th {{ background: #e8f0fe; color: #174ea6; font-weight: 700; }}
  td.plant {{ text-align: left; font-weight: 700; }}
  tr:nth-child(even) td {{ background: #f8f9fa; }}
  tr.sail-row td {{ background: #f9ab00; color: #3c2f00; font-weight: 700; }}
  td.dev {{ background: #fbbc04 !important; color: #202124; font-weight: 700; }}
  td.dev-sail {{ box-shadow: inset 0 0 0 2px #ea4335; }}
</style>
</head>
<body>
  <h1>{data.get("title", "MAJOR TECHNO-ECONOMIC PARAMETERS — VERIFICATION")}</h1>
  <p class="subtitle">Report month: {month} — {data.get("subtitle", "")}</p>
  {''.join(sections_html)}
</body></html>"""
