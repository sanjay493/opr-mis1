"""
JPC monthly report — replicates "Report_format/MONTHEND/1JPC report input Jun26.xlsx":
4 sheets (Pmix report, spl stl Report (2), Sinter, MoU report), each showing a
selected month's Actual vs the CPLY (corresponding month, prior year) Actual
and % growth, built entirely from `production_table`.

Values not tracked anywhere in the DB (confirmed by checking against a real
month's data, not just grepping for the label) are left blank rather than
guessed — see the per-item comments below. The MoU/Var/%Ful. columns on the
"MoU report" sheet are always blank: this app has no MoU-target data source.
"""
import io
from datetime import date

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import db
from constants import FIVE_PLANTS as _5P

_MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
           "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _mlabel(ym: str) -> str:
    y, m = int(ym[:4]), int(ym[5:7])
    return f"{_MONTHS[m - 1]}'{str(y)[2:]}"


# ── DB access ────────────────────────────────────────────────────────────────

_FS_ALIAS = frozenset({"SSP", "VISL"})   # no dedicated "Finished Steel" row


def _one(cur, plant, item, month):
    if not item:
        return None
    cur.execute(
        "SELECT month_actual FROM production_table WHERE plant_name=? AND item_name=? AND report_month=?",
        (plant, item, month),
    )
    r = cur.fetchone()
    return r[0] if r and r[0] is not None else None


def _one_of(cur, plant, items, month):
    """First matching item_name that has a value — for concepts stored under
    different item_names depending which extractor version wrote them (e.g.
    ASP's Plates: 'PLATES' vs 'Plate Mill')."""
    for item in items:
        v = _one(cur, plant, item, month)
        if v is not None:
            return v
    return None


def _fs(cur, plant, month):
    """'Finished Steel', SSP/VISL falling back to Saleable Steel (matches
    page4.py's convention — those two plants have no dedicated FS row)."""
    v = _one(cur, plant, "Finished Steel", month)
    if v is None and plant in _FS_ALIAS:
        return _one(cur, plant, "Saleable Steel", month)
    return v


def _sum(cur, combos, month):
    """combos: [(plant, item), ...]. Sums whatever is found; None if nothing found."""
    total, found = 0.0, False
    for plant, item in combos:
        v = _one(cur, plant, item, month)
        if v is not None:
            total += v
            found = True
    return round(total, 3) if found else None


def _sub(a, b):
    """a - b, or None if either side is missing."""
    return round(a - b, 3) if a is not None and b is not None else None


def _gr(cur_v, prev_v):
    if cur_v is None or prev_v is None or prev_v == 0:
        return None
    return round((cur_v - prev_v) / abs(prev_v) * 100, 1)


# ── styling ──────────────────────────────────────────────────────────────────

_TITLE_FONT = Font(bold=True, size=12)
_UNIT_FONT = Font(italic=True, size=9)
_HDR_FONT = Font(bold=True, size=10)
_HDR_FILL = PatternFill("solid", fgColor="D9E6F5")
_PLANT_FONT = Font(bold=True, size=10)
_PLANT_FILL = PatternFill("solid", fgColor="F0F0F0")
_TOTAL_FONT = Font(bold=True, size=10)
_THIN = Side(style="thin", color="B0B0B0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _num_cell(ws, row, col, value, bold=False, pct=False):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
        c.number_format = "0.0" if pct else "0.000"
    c.border = _BORDER
    c.alignment = Alignment(horizontal="right")
    if bold:
        c.font = _TOTAL_FONT


def _label_cell(ws, row, col, text, indent=0, bold=False, fill=None):
    c = ws.cell(row=row, column=col, value=text)
    c.alignment = Alignment(horizontal="left", indent=indent)
    c.border = _BORDER
    if bold:
        c.font = _PLANT_FONT if fill else _TOTAL_FONT
    if fill:
        c.fill = fill


def _sheet_header(ws, title, month, cply_month, unit="'000T", label_col=1):
    ws.cell(row=1, column=label_col, value=title).font = _TITLE_FONT
    ws.cell(row=2, column=label_col + 3, value=f"Unit: {unit}").font = _UNIT_FONT
    for col, text in [(label_col + 1, _mlabel(month)),
                      (label_col + 2, _mlabel(cply_month)),
                      (label_col + 3, "% Gr. over")]:
        c = ws.cell(row=3, column=col, value=text)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="center")
    for col, text in [(label_col + 1, "Actual"), (label_col + 2, "Actual"),
                      (label_col + 3, _mlabel(cply_month))]:
        c = ws.cell(row=4, column=col, value=text)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="center")
    return 5   # first data row


def _write_row(ws, row, label, actual, cply, indent=1, bold=False, label_col=1):
    _label_cell(ws, row, label_col, label, indent=indent, bold=bold)
    _num_cell(ws, row, label_col + 1, actual, bold=bold)
    _num_cell(ws, row, label_col + 2, cply, bold=bold)
    _num_cell(ws, row, label_col + 3, _gr(actual, cply), bold=bold, pct=True)
    return row + 1


def _write_plant_hdr(ws, row, plant, label_col=1):
    for col in range(label_col, label_col + 4):
        ws.cell(row=row, column=col).fill = _PLANT_FILL
        ws.cell(row=row, column=col).border = _BORDER
    ws.cell(row=row, column=label_col, value=plant).font = _PLANT_FONT
    return row + 1


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Sheet 1: Pmix report ────────────────────────────────────────────────────

# Per-plant mill breakdown. Each entry: (label, plant, item_name_or_None).
# item_name=None → left blank: confirmed not tracked anywhere in this app's
# DB (checked against page_catwise_saleable.py's item catalog + a live query,
# not just a label grep) — will be filled in once a source is extracted.
_BSP_ITEMS = [
    ("WRM (Total)",           "BSP", "WIRERODS"),
    ("TMT coils(WRM)",        "BSP", "TMT COILS(WRM)"),
    ("Others (WRM)",          "BSP", "OTHERS(WRM)"),
    ("TMT Bars (MM)",         "BSP", "TMT BARS(MM)"),
    ("Lt Strls (MM)",         "BSP", "LT STRS(MM)"),
    ("Rounds & Bars (BRM)",   "BSP", "BARS&RODMILL"),
    ("Rail (RSM)",            "BSP", "RSM_RAIL"),
    ("Rail (URM)",            "BSP", "URM_RAIL"),
    ("Hy.Struls.",            "BSP", None),
    ("Plates",                "BSP", "PLATEMILL"),
    ("Semis Total",           "BSP", "Saleable Semis"),
    ("Saleable Steel",        "BSP", "Saleable Steel"),
]
_DSP_ITEMS = [
    ("Med Structurals (SM)",  "DSP", None),   # DB only has the combined MSM figure
    ("Med Structurals (MSM)", "DSP", "MSM"),
    ("Round & bars (MM)",     "DSP", "MM"),
    ("Wheel & Axle",          "DSP", "WAP"),
    ("Semis Total",           "DSP", "Saleable Semis"),
    ("Saleable steel",        "DSP", "Saleable Steel"),
]
_RSP_ITEMS = [
    ("Plate (PM)",            "RSP", "OPM Plate"),
    ("New Plate Mill",        "RSP", "NPM Plate"),
    ("HR Plates",             "RSP", None),
    ("HR Coils",              "RSP", None),
    ("New HR Plate",          "RSP", "HSM-2 HR Plate"),
    ("New HR Coils",          "RSP", "HSM-2 HR Coil (Sale)"),
    ("GP/GC Sheets",          "RSP", None),
    ("CRNO",                  "RSP", "CRNO Coils"),
    ("ERW Pipes",             "RSP", "ERW Pipes"),
    ("SW Pipes",              "RSP", "SW Pipes"),
    ("Semis",                 "RSP", None),
    ("Saleable Steel",        "RSP", "Saleable Steel"),
]
_BSL_ITEMS = [
    ("HR Plates",             "BSL", "HSM HR Plate"),
    ("HR Coils",              "BSL", "HSM HR Coil (Sale)"),
    ("HR Sheets",             "BSL", "HR Sheet"),
    ("CR Coils (old)",        "BSL", "CR I/II CR(Coil) Sale"),
    ("CR Coils (new)",        "BSL", "CR III CR(Coil) Sale"),
    ("CR Sheet (old)",        "BSL", "CR Sheets"),
    ("CR Sheet (new)",        "BSL", "New CR Sheet"),
    ("GP/GC (old)",           "BSL", "GP/GC"),
    ("GP/GC (new)",           "BSL", "GPC3"),
    ("Thick plate",           "BSL", "Thick Plate"),
    ("Semis",                 "BSL", "Saleable Semis"),
    ("Saleable Steel",        "BSL", "Saleable Steel"),
]
_ISP_ITEMS = [
    ("Round & Bars",          "ISP", "BARMILL"),
    ("Hy Structurals",        "ISP", "USMILL"),
    ("TMT Coils(WRM)",        "ISP", None),   # DB has only one combined WRM figure
    ("Wire Rod Coils(WRM)",   "ISP", "WRMILL"),
    ("WRM",                   "ISP", "WRMILL"),
    ("Semis",                 "ISP", "Saleable Semis"),
    ("Saleable Steel",        "ISP", "Saleable Steel"),
]
_PLANT_BLOCKS = [
    ("BSP", _BSP_ITEMS), ("DSP", _DSP_ITEMS), ("RSP", _RSP_ITEMS),
    ("BSL", _BSL_ITEMS), ("ISP", _ISP_ITEMS),
]

# SAIL-wide rollup — sum of the specific plant/item combos that make up each
# product category (cross-checked against the sample file's actual numbers;
# combos not present anywhere in the DB are simply left out of the sum, so
# these totals may run lower than a manually-compiled report until every
# plant's mill split is fully extracted).
_SAIL_ROLLUP = [
    ("Semis",             [("BSP", "Saleable Semis"), ("DSP", "Saleable Semis"),
                            ("BSL", "Saleable Semis"), ("ISP", "Saleable Semis")]),
    ("Wire Rods",         [("BSP", "WIRERODS"), ("ISP", "WRMILL")]),
    ("Round & Bars",      [("BSP", "TMT BARS(MM)"), ("BSP", "BARS&RODMILL"),
                            ("DSP", "MM"), ("ISP", "BARMILL")]),
    ("Lt. Structurals",   [("BSP", "LT STRS(MM)")]),
    ("Med. Structurals",  [("DSP", "MSM")]),
    ("Hy. Structurals",   [("ISP", "USMILL")]),
    ("Rail",              [("BSP", "RSM_RAIL"), ("BSP", "URM_RAIL")]),
    ("Wheel & Axle",      [("DSP", "WAP")]),
    ("Plate",             [("BSP", "PLATEMILL"), ("RSP", "OPM Plate"), ("RSP", "NPM Plate")]),
    ("Thick plate",       [("BSL", "Thick Plate")]),
    ("HR Plates",         [("RSP", "HSM-2 HR Plate"), ("BSL", "HSM HR Plate")]),
    ("HR Sheets",         [("BSL", "HR Sheet")]),
    ("HR Coils/Skelp",    [("RSP", "HSM-2 HR Coil (Sale)"), ("BSL", "HSM HR Coil (Sale)")]),
    ("CR Coils/Sheets",   [("BSL", "CR I/II CR(Coil) Sale"), ("BSL", "CR III CR(Coil) Sale"),
                            ("BSL", "CR Sheets"), ("BSL", "New CR Sheet")]),
    ("GP/GC",             [("BSL", "GP/GC"), ("BSL", "GPC3")]),
    ("CRNO",              [("RSP", "CRNO Coils")]),
    ("ERW pipes",         [("RSP", "ERW Pipes")]),
    ("SW Pipes",          [("RSP", "SW Pipes")]),
]

_FS_PLANTS = ["BSP", "DSP", "RSP", "BSL", "ISP"]


def compute_pmix_rows(cur, month):
    """One month's worth of Pmix report rows, in display order — shared by
    the single-month JPC sheet (called once per month + once per CPLY month,
    then zipped) and the FY-wide Pmix report (called once per FY month, then
    grouped into quarters/cumulative). Each row:
        {"kind": "header"|"data", "label": str, "value": float|None, "bold": bool}
    """
    rows = []

    def header(label):
        rows.append({"kind": "header", "label": label, "value": None, "bold": False})

    def data(label, value, bold=False):
        rows.append({"kind": "data", "label": label, "value": value, "bold": bold})

    for plant, items in _PLANT_BLOCKS:
        header(plant)
        for label, plant_code, item in items:
            data(label, _one(cur, plant_code, item, month), bold=label.lower().startswith("saleable"))

    header("SAIL")
    for label, combos in _SAIL_ROLLUP:
        data(label, _sum(cur, combos, month))
    sail_5pl_combo = [(p, "Saleable Steel") for p in _FS_PLANTS]
    data("Saleable steel 5 PL", _sum(cur, sail_5pl_combo, month), bold=True)

    header("Finished Steel")
    fs_vals = {}
    for plant in _FS_PLANTS:
        fs_vals[plant] = _one(cur, plant, "Finished Steel", month)
        data(plant, fs_vals[plant])
    fs5 = _sum(cur, [(p, "Finished Steel") for p in _FS_PLANTS], month)
    data("Finished Steel 5PL", fs5, bold=True)

    # ASP
    asp_sal = _one(cur, "ASP", "Saleable Steel", month)
    asp_fin = _fs(cur, "ASP", month)
    data("ASP (Total)", asp_sal, bold=True)
    data("ASP (Finished)", asp_fin)
    data("ASP (Semis)", _sub(asp_sal, asp_fin))

    # SSP — HRCS = Carbon Steel Production; Stainless = Saleable - Carbon
    ssp_sal = _one(cur, "SSP", "Saleable Steel", month)
    ssp_carb = _one(cur, "SSP", "Carbon Steel Production", month)
    data("SSP (Finished Carbon Steel)", ssp_carb)
    data("SSP (Finished Stainless Steel)", _sub(ssp_sal, ssp_carb))

    # VISL
    visl = _fs(cur, "VISL", month)
    data("VISL(Finished)", visl)

    sail_set = _FS_PLANTS + ["ASP", "SSP", "VISL"]
    fst = _sum(cur, [(p, "Finished Steel") for p in _FS_PLANTS], month)
    fst = (fst or 0) + (asp_fin or 0) + (ssp_sal or 0) + (visl or 0)
    data("Finished Steel (Total)", fst, bold=True)

    sal = _sum(cur, [(p, "Saleable Steel") for p in sail_set], month)
    data("Semis Total", _sub(sal, fst))
    data("Saleable Steel (Total)", sal, bold=True)

    return rows


def _build_pmix_sheet(ws, cur, month, cply_month):
    row = _sheet_header(ws, f"SAIL Product mix performance: {_mlabel(month)}", month, cply_month)

    m_rows = compute_pmix_rows(cur, month)
    c_rows = compute_pmix_rows(cur, cply_month)
    for m_row, c_row in zip(m_rows, c_rows):
        if m_row["kind"] == "header":
            row = _write_plant_hdr(ws, row, m_row["label"])
        else:
            row = _write_row(ws, row, m_row["label"], m_row["value"], c_row["value"], bold=m_row["bold"])

    _autosize(ws, [26, 13, 13, 13])
    ws.freeze_panes = "B5"


# ── Sheet 2: Special Steel plants (ASP/SSP/VISL) ────────────────────────────

def _build_special_steel_sheet(ws, cur, month, cply_month):
    row = _sheet_header(ws, "SAIL: PMIX Special Steel plants", month, cply_month)

    row = _write_plant_hdr(ws, row, "ASP")
    plates_m = _one_of(cur, "ASP", ["PLATES", "Plate Mill"], month)
    plates_c = _one_of(cur, "ASP", ["PLATES", "Plate Mill"], cply_month)
    others_m = _one(cur, "ASP", "FORGINGS", month)
    others_c = _one(cur, "ASP", "FORGINGS", cply_month)
    asp_sal_m = _one(cur, "ASP", "Saleable Steel", month)
    asp_sal_c = _one(cur, "ASP", "Saleable Steel", cply_month)
    asp_fin_m = _one(cur, "ASP", "Finished Steel", month)
    asp_fin_c = _one(cur, "ASP", "Finished Steel", cply_month)
    row = _write_row(ws, row, "Semis (billet, Bloom, Ingots)", None, None)  # no reliable source — see notes
    row = _write_row(ws, row, "Others (RCS, squares, Forgings)", others_m, others_c)
    row = _write_row(ws, row, "Plates", plates_m, plates_c)
    row = _write_row(ws, row, "Plant Total Saleable Steel", asp_sal_m, asp_sal_c, bold=True)
    row = _write_row(ws, row, "Plant Total Finished Alloy Steel", asp_fin_m, asp_fin_c)

    row = _write_plant_hdr(ws, row, "SSP")
    hrcs_m, hrcs_c = _one(cur, "SSP", "Carbon Steel Production", month), _one(cur, "SSP", "Carbon Steel Production", cply_month)
    hrss_m, hrss_c = _one(cur, "SSP", "HRSS Production", month), _one(cur, "SSP", "HRSS Production", cply_month)
    crss_m, crss_c = _one(cur, "SSP", "CRSS Production", month), _one(cur, "SSP", "CRSS Production", cply_month)
    hrap_m, hrap_c = _one(cur, "SSP", "NO1 Production", month), _one(cur, "SSP", "NO1 Production", cply_month)
    ssp_sal_m, ssp_sal_c = _one(cur, "SSP", "Saleable Steel", month), _one(cur, "SSP", "Saleable Steel", cply_month)
    row = _write_row(ws, row, "Semis (slabs)", None, None)  # no confirmed direct source
    row = _write_row(ws, row, "HRCS", hrcs_m, hrcs_c)
    row = _write_row(ws, row, "HRSS", hrss_m, hrss_c)
    row = _write_row(ws, row, "CRSS", crss_m, crss_c)
    row = _write_row(ws, row, "C.Q.", None, None)  # not tracked anywhere
    row = _write_row(ws, row, "HRAP", hrap_m, hrap_c)
    row = _write_row(ws, row, "Plant Total Saleable Steel", ssp_sal_m, ssp_sal_c, bold=True)
    row = _write_row(ws, row, "Plant Total Finished Steel", ssp_sal_m, ssp_sal_c, bold=True)

    row = _write_plant_hdr(ws, row, "VISL")
    visl_sal_m, visl_sal_c = _one(cur, "VISL", "Saleable Steel", month), _one(cur, "VISL", "Saleable Steel", cply_month)
    row = _write_row(ws, row, "Semis (Ingots, CC Billet/Bloom)", None, None)
    row = _write_row(ws, row, "Others (Flats, RCS, Rounds)", visl_sal_m, visl_sal_c)
    row = _write_row(ws, row, "Saleable Steel Production", visl_sal_m, visl_sal_c, bold=True)

    _autosize(ws, [32, 13, 13, 13])
    ws.freeze_panes = "B5"


# ── Sheet 3: Sinter ──────────────────────────────────────────────────────────

def _build_sinter_sheet(ws, cur, month, cply_month):
    row = _sheet_header(ws, f"Sinter performance: {_mlabel(month)}", month, cply_month)
    combo = [(p, "Total Sinter") for p in _5P]
    for plant in _5P:
        act, cply = _one(cur, plant, "Total Sinter", month), _one(cur, plant, "Total Sinter", cply_month)
        row = _write_row(ws, row, plant, act, cply)
    act, cply = _sum(cur, combo, month), _sum(cur, combo, cply_month)
    row = _write_row(ws, row, "SAIL", act, cply, bold=True)
    _autosize(ws, [14, 13, 13, 13])


# ── Sheet 4: MoU report (MoU/Var/%Ful columns always blank) ────────────────

_MOU_ITEMS = ["Hot Metal", "Total Crude Steel", "Saleable Steel", "Pig Iron", "Finished Steel"]
_MOU_LABELS = {"Hot Metal": "HOT METAL", "Total Crude Steel": "CRUDE STEEL",
               "Saleable Steel": "SALEABLE STEEL", "Pig Iron": "PIG IRON",
               "Finished Steel": "Finished Steel"}
# per-item plant row order (VISL only listed where it actually has data)
_MOU_PLANT_ROWS = {
    "Hot Metal":          _5P + ["5 Plants", "VISL", "SAIL"],
    "Total Crude Steel":  _5P + ["5 Plants", "ASP", "SSP", "VISL", "SAIL"],
    "Saleable Steel":     _5P + ["5 Plants", "ASP", "SSP", "VISL", "SAIL"],
    "Pig Iron":           _5P + ["5 Plants", "VISL", "SAIL"],
    "Finished Steel":     _5P + ["5 Plants", "ASP", "SSP", "VISL", "SAIL"],
}


def _mou_value(cur, db_item, plant, month):
    if plant == "5 Plants":
        # None of the 5 core plants need the SSP/VISL Finished-Steel alias.
        return _sum(cur, [(p, db_item) for p in _5P], month)
    if plant == "SAIL":
        sail_set = _5P + (["VISL"] if db_item in ("Hot Metal", "Pig Iron") else ["ASP", "SSP", "VISL"])
        total, found = 0.0, False
        for p in sail_set:
            v = _fs(cur, p, month) if db_item == "Finished Steel" else _one(cur, p, db_item, month)
            if v is not None:
                total += v
                found = True
        return round(total, 3) if found else None
    if db_item == "Finished Steel":
        return _fs(cur, plant, month)
    return _one(cur, plant, db_item, month)


def _build_mou_sheet(ws, cur, month, cply_month):
    ws.cell(row=1, column=1, value=f"Production Performance during {_mlabel(month)}").font = _TITLE_FONT
    ws.cell(row=2, column=9, value="w.r.t   MoU").font = _UNIT_FONT
    ws.cell(row=3, column=9, value="Unit:'000 T").font = _UNIT_FONT
    headers = ["Items", "Plant", "MoU", _mlabel(month), "", "", "", _mlabel(cply_month), "% Gr. over"]
    for col, text in enumerate(headers, start=1):
        if text:
            c = ws.cell(row=4, column=col, value=text)
            c.font = _HDR_FONT
            c.fill = _HDR_FILL
    sub = ["", "", "MoU", "Actual", "Var", "% Ful.", "Actual", _mlabel(cply_month)]
    for col, text in enumerate(sub, start=1):
        if text:
            c = ws.cell(row=5, column=col, value=text)
            c.font = _HDR_FONT

    row = 6
    for db_item in _MOU_ITEMS:
        first = True
        for plant in _MOU_PLANT_ROWS[db_item]:
            act = _mou_value(cur, db_item, plant, month)
            cply = _mou_value(cur, db_item, plant, cply_month)
            _label_cell(ws, row, 1, _MOU_LABELS[db_item] if first else "")
            _label_cell(ws, row, 2, plant, bold=(plant in ("5 Plants", "SAIL")))
            # col 3 MoU, col 4 Var, col 5 %Ful. — always blank (no MoU-target source)
            for col in (3, 5, 6):
                ws.cell(row=row, column=col).border = _BORDER
            _num_cell(ws, row, 4, act, bold=(plant in ("5 Plants", "SAIL")))
            _num_cell(ws, row, 7, cply, bold=(plant in ("5 Plants", "SAIL")))
            _num_cell(ws, row, 8, _gr(act, cply), pct=True)
            first = False
            row += 1

    _autosize(ws, [12, 10, 9, 11, 9, 9, 11, 11])


# ── entry point ──────────────────────────────────────────────────────────────

def build_jpc_workbook(month: str) -> "openpyxl.Workbook":
    """month: 'YYYY-MM'. Builds the full 4-sheet JPC report workbook."""
    cply_month = db.get_cply_month(month)

    conn = db.connect()
    cur = conn.cursor()
    try:
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Pmix report"
        _build_pmix_sheet(ws1, cur, month, cply_month)

        ws2 = wb.create_sheet("spl stl Report")
        _build_special_steel_sheet(ws2, cur, month, cply_month)

        ws3 = wb.create_sheet("Sinter")
        _build_sinter_sheet(ws3, cur, month, cply_month)

        ws4 = wb.create_sheet("MoU report")
        _build_mou_sheet(ws4, cur, month, cply_month)
    finally:
        conn.close()

    return wb


def build_jpc_report_bytes(month: str) -> bytes:
    wb = build_jpc_workbook(month)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
