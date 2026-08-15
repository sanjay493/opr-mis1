"""
One-off backfill: BSP/BSL/ISP Special Steel despatch for FY 2022-23
(April 2022 - March 2023), from the legacy consolidated workbook
Report_format/22-23 Special Steel.xlsx (one sheet per plant: 'ISP', 'BSL',
'BSP' — no RSP/DSP sheet, so those two plants aren't covered by this file).

Layout (confirmed by direct inspection, not re-derived at runtime beyond the
month-header row, since this is a single known legacy file, not a recurring
upload format):

  ISP  — row 6: col C 'ITEM', cols D..O = Apr'22..Mar'23 (one datetime header
         cell per month). Data rows 9-14: col C = item name, no grade split
         (matches the "TOTAL" quality_grade sentinel _gen_isp / the ISP OCR
         extractor already use). Only despatch figures, no order book.
         ISP's report page (_gen_isp in page_special_steel.py) only ever
         queries 6 hardcoded canonical mills (WR COIL/TMT COIL/TMT BAR/
         STRUCTURALS/150 BLT/200 BLM) — of the 5 legacy items, 3 map onto
         those by name; "Semis" and "Merchant Mill" don't correspond to any
         of the 6 and are dropped (confirmed with user rather than guessed).
             Hy. Strl. Mill -> STRUCTURALS
             WRM            -> WR COIL
             Bar Mill       -> TMT BAR
             Semis          -> (dropped, no canonical match)
             Merchant Mill  -> (dropped, no canonical match)
             TOTAL SPECIAL STEEL -> (grand total, not a data row)

  BSL  — row 6: col B 'ITEM', col C 'GRADE', cols D..O = Apr'22..Mar'23.
         Data rows 7+, col B forward-filled down blank continuation rows
         (grade rows only carry col C). Stops at the 'Total Special Quality
         products' grand-total row (row 42). Despatch only, no order book,
         no section split (matches the live BSL extractors' section="").
         Product names renamed onto the canonical set _gen_bsl / the BSL PDF
         extractor (excel_extractor_bsl.py) already use:
             HR PLATES       -> HR PLATE
             CR COILS/SHEETS -> CR COIL/SHEET/GP GC
             HR COIL, HR SHEET -> unchanged (already canonical)
         (no SLAB group present in this legacy file)

  BSP  — row 6: col C 'GRADE' (col B is the item/group, unlabeled in the
         header), cols E..P = Apr'22..Mar'23 (col D is a blank spacer
         column). Data rows 9-41, col B forward-filled. Product names here
         already match _gen_bsp's canonical group list exactly except one:
             HEAVY STRLS -> BRM Product
             Semis, Wire Rods, Merchant Products, Rails, Plates -> unchanged
         A stray row with grade literal 0 and no data (row 33) and rows with
         no despatch figure in any month are skipped (nothing to backfill).

Every row in the source file is a single despatch figure per month — order_qty
is always None (this workbook never carried an order book, only actuals).

Usage:
  python scripts/backfill_special_steel_2022_23.py            # dry-run: prints a diff, writes nothing
  python scripts/backfill_special_steel_2022_23.py --apply     # actually writes to special_steel_orders (live DB)

Dry-run is the default on purpose — this touches the live MySQL DB
(DB_ENGINE=mysql), so the diff should be reviewed before anything is written.
"""
import sys
import os
import argparse
import datetime as _dt

BACKEND_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BACKEND_DIR)

import db  # noqa: E402

SRC_FILE = os.path.join(BACKEND_DIR, "..", "Report_format", "22-23 Special Steel.xlsx")


def _month_columns(ws, header_row: int) -> list:
    """[(col, 'YYYY-MM'), ...] for every datetime cell in header_row, in column order."""
    out = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, _dt.datetime):
            out.append((c, f"{v.year}-{v.month:02d}"))
    return sorted(out)


def _row_values(ws, row: int, month_cols: list) -> dict:
    """{report_month: float or None} for one data row."""
    out = {}
    for c, month in month_cols:
        v = ws.cell(row, c).value
        out[month] = float(v) if isinstance(v, (int, float)) else None
    return out


def _parse_isp(wb) -> list:
    ws = wb["ISP"]
    month_cols = _month_columns(ws, header_row=6)
    canon = {"Hy. Strl. Mill": "STRUCTURALS", "WRM": "WR COIL", "Bar Mill": "TMT BAR"}
    rows = []
    sort_order = 0
    for r in range(9, 15):
        item = ws.cell(r, 3).value
        item = str(item).strip() if item is not None else ""
        product = canon.get(item)
        if not product:
            continue  # Semis / Merchant Mill (no canonical match) / TOTAL SPECIAL STEEL
        vals = _row_values(ws, r, month_cols)
        if all(v is None for v in vals.values()):
            continue
        sort_order += 1
        for month, val in vals.items():
            if val is None:
                continue
            rows.append({
                "report_month": month, "plant": "ISP", "product": product,
                "quality_grade": "TOTAL", "section": "", "sort_order": sort_order,
                "order_qty": None, "actual_despatch": val,
            })
    return rows


def _parse_bsl(wb) -> list:
    ws = wb["BSL"]
    month_cols = _month_columns(ws, header_row=6)
    canon = {"HR PLATES": "HR PLATE", "CR COILS/SHEETS": "CR COIL/SHEET/GP GC"}
    rows = []
    current_item = None
    sort_order = 0
    for r in range(7, ws.max_row + 1):
        item_raw = ws.cell(r, 2).value
        grade_raw = ws.cell(r, 3).value
        if item_raw is not None:
            item = str(item_raw).strip()
            if item.upper().startswith("TOTAL SPECIAL"):
                break  # grand total — end of table
            current_item = canon.get(item, item)
        grade = str(grade_raw).strip() if grade_raw is not None else ""
        if not grade or not current_item:
            continue
        vals = _row_values(ws, r, month_cols)
        if all(v is None for v in vals.values()):
            continue
        sort_order += 1
        for month, val in vals.items():
            if val is None:
                continue
            rows.append({
                "report_month": month, "plant": "BSL", "product": current_item,
                "quality_grade": grade, "section": "", "sort_order": sort_order,
                "order_qty": None, "actual_despatch": val,
            })
    return rows


def _parse_bsp(wb) -> list:
    ws = wb["BSP"]
    month_cols = _month_columns(ws, header_row=6)
    canon = {"HEAVY STRLS": "BRM Product"}
    rows = []
    current_item = None
    sort_order = 0
    for r in range(9, 42):
        item_raw = ws.cell(r, 2).value
        grade_raw = ws.cell(r, 3).value
        if item_raw is not None:
            current_item = canon.get(str(item_raw).strip(), str(item_raw).strip())
        grade = str(grade_raw).strip() if grade_raw not in (None, 0) else ""
        if not grade or not current_item:
            continue
        vals = _row_values(ws, r, month_cols)
        if all(v is None for v in vals.values()):
            continue
        sort_order += 1
        for month, val in vals.items():
            if val is None:
                continue
            rows.append({
                "report_month": month, "plant": "BSP", "product": current_item,
                "quality_grade": grade, "section": "", "sort_order": sort_order,
                "order_qty": None, "actual_despatch": val,
            })
    return rows


def current_db_values(plants, months):
    conn = db.connect()
    cur = conn.cursor()
    ph_p = ",".join("?" * len(plants))
    ph_m = ",".join("?" * len(months))
    cur.execute(
        f"SELECT plant_name, report_month, product, quality_grade, section, actual_despatch "
        f"FROM special_steel_orders WHERE plant_name IN ({ph_p}) AND report_month IN ({ph_m})",
        (*plants, *months),
    )
    out = {(p, m, prod, grade, sec): val for p, m, prod, grade, sec, val in cur.fetchall()}
    conn.close()
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true", help="actually write to the DB (default: dry-run)")
    args = ap.parse_args()

    import openpyxl
    wb = openpyxl.load_workbook(SRC_FILE, data_only=True)

    all_rows = _parse_isp(wb) + _parse_bsl(wb) + _parse_bsp(wb)
    months = sorted({r["report_month"] for r in all_rows})
    plants = sorted({r["plant"] for r in all_rows})
    print(f"Parsed {len(all_rows)} (plant, month, product, grade) despatch values "
          f"across {len(plants)} plants, {len(months)} months ({months[0]}..{months[-1]}).\n")

    cur_vals = current_db_values(plants, months)

    changed = 0
    for r in sorted(all_rows, key=lambda r: (r["plant"], r["report_month"], r["sort_order"])):
        key = (r["plant"], r["report_month"], r["product"], r["quality_grade"], r["section"])
        old = cur_vals.get(key)
        new = r["actual_despatch"]
        if old is None or round(float(old), 3) != round(float(new), 3):
            changed += 1
            print(f"  {r['plant']} {r['report_month']}  {r['product']:20s} {r['quality_grade']:30s} "
                  f"{str(old):>10s} -> {new:<10}")

    print(f"\n({changed} of {len(all_rows)} differ from current DB values)")

    if not args.apply:
        print(f"\nDRY RUN — nothing written. Re-run with --apply to write.")
        return

    by_plant_month = {}
    for r in all_rows:
        by_plant_month.setdefault((r["plant"], r["report_month"]), []).append(r)

    for (plant, month), rows in by_plant_month.items():
        db.clear_special_steel_orders(month, plant)
        for r in rows:
            db.save_special_steel_entry(
                month, plant, r["product"], r["quality_grade"], r["sort_order"],
                r["order_qty"], r["actual_despatch"], section=r["section"],
            )
    print(f"\nAPPLIED — {len(all_rows)} value(s) written to the DB across "
          f"{len(by_plant_month)} (plant, month) groups.")


if __name__ == "__main__":
    main()
