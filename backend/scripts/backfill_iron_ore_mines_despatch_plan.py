"""
One-off backfill for Iron Ore Mines Despatch PLAN (mine-level,
mines_despatch_plan_monthly — see db.get_mines_production_despatch_monthly /
get_iron_ore_group_rollup_monthly) from Report_format/iron ore.xlsx.

This is the "despatch data ... comes from the user separately later" promised
in backfill_iron_ore_mines_production.py's docstring. Only despatch PLAN is in
this workbook (qty_plan per report_month x mine x material x end_use — Plan
has no Rail/Road split, matching mines_despatch_plan_monthly's grain).
Despatch Actual and Booked Quantity are still not provided.

Sheet1 stacked blocks, all covering 12 months Apr'26-Mar'27:

  row  2  "Tailing sales despatch plan"       rows  3-5   TAILINGS  / SALES
  row  8  "Dump Fines sale Despatch plan"     rows  9-12  DUMP_FINES/ SALES
  row 15  "Pellets despatch to captive"       row  16     PELLETS   / CAPTIVE
  row 20  "Dump Despatch Captive"             row  21     DUMP_FINES/ CAPTIVE
  row 42  "conversion agent"                  rows 44-47  LUMP+FINES/ PELLET_CONV
  row 52  "despatch plan Sales lump & Fines"  rows 54-66  LUMP+FINES/ SALES
  row 69  "Lump Fines depatch plan captive"   rows 71-83  LUMP+FINES/ CAPTIVE

NOT read: the older "Lump Fines Despatch Captive" block at rows 25-38 — the
rows 71-83 block is its corrected revision (per direct instruction,
2026-08-27: ROWGHAT Sep'26 Fines -29 -> 1, BOLANI Nov'26 Fines 205 -> 195),
and rows 71-83 is used as the single source for LUMP/FINES -> CAPTIVE.

Layouts:
  - "narrow" (rows 2/8/15/20): one 12-wide row of months, cols B..M, dates on
    the title row itself.
  - "wide-hdr" (row 42): a (Lump, Fines) column pair per month (cols B..Y),
    month dates on the title row at B,D,F,..., a Lump/Fines sub-header below.
    Only the Fines column is populated (conversion agents take fines).
  - "wide" (rows 54-66, 71-83): same (Lump, Fines) pair-per-month grid but NO
    date row and NO Lump/Fines sub-header — months are assumed Apr'26..Mar'27
    left to right (MONTHS_APR26). "JHR TOTAL" / "ODISHA TOTAL" subtotal rows
    are skipped; "BARSUA rail" / "KALTA road" / "... GROUP" name variants map
    to the plain mine (Plan has no Rail/Road split).

Values in this workbook are already in '000 T (unlike
Report_format/iron ore production.xlsx which was raw tonnes) — magnitudes
match mines_production_monthly's '000 T figures for the same mines/materials,
so NO scaling is applied here.

Run once: python scripts/backfill_iron_ore_mines_despatch_plan.py [--apply]
Without --apply, only prints the diff against current DB values (dry run).
"""
import argparse
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import db

import openpyxl

WORKBOOK_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "Report_format", "iron ore.xlsx",
)

# This workbook is already in '000 T (see module docstring).
UNIT_SCALE = 1

MONTHS_APR26 = (
    [f"2026-{m:02d}" for m in range(4, 13)] + [f"2027-{m:02d}" for m in range(1, 4)]
)

MINE_NAME_TO_CODE = {
    "kiriburu": "KIRIBURU", "mburu": "MEGHAHATUBURU", "meghahatuburu": "MEGHAHATUBURU",
    "gua": "GUA", "manoharpur": "MANOHARPUR", "bolani": "BOLANI", "barsua": "BARSUA",
    "taldih": "TALDIH", "kalta": "KALTA", "rajhara": "RAJHARA",
    "dalli": "DALLI", "rowghat": "ROWGHAT",
}
SKIP_ROWS = {"jhr total", "odisha total", "total"}

# "narrow" blocks: (title_row, first_data_row, last_data_row, (material, end_use))
SINGLE_BLOCKS = [
    (2, 3, 5, ("TAILINGS", "SALES")),
    (8, 9, 12, ("DUMP_FINES", "SALES")),
    (15, 16, 16, ("PELLETS", "CAPTIVE")),
    (20, 21, 21, ("DUMP_FINES", "CAPTIVE")),
]
# "wide-hdr" blocks: (date_row, first_data_row, last_data_row, end_use)
WIDE_HDR_BLOCKS = [
    (42, 44, 47, "PELLET_CONV"),
]
# "wide" blocks (no date/sub-header row): (first_data_row, last_data_row, end_use)
WIDE_BLOCKS = [
    (54, 66, "SALES"),
    (71, 83, "CAPTIVE"),
]


def _month_str(dt) -> str:
    return f"{dt.year}-{dt.month:02d}"


def _mine_code(name):
    """Normalise a workbook row label to a mines_master code, or None for a
    subtotal row that should be skipped."""
    n = str(name).strip().lower().replace("'", "")
    for suffix in (" group", " rail", " road"):
        if n.endswith(suffix):
            n = n[: -len(suffix)].strip()
    if n in SKIP_ROWS:
        return None
    return MINE_NAME_TO_CODE[n]


def _num(v):
    if v is None or v == "":
        return None
    return float(v) * UNIT_SCALE


def build_entries():
    """-> {(report_month, mine_code): [{material_code, end_use_code, plan}, ...]}"""
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb["Sheet1"]
    out = {}  # (rm, mine_code) -> list

    def add(rm, mine_code, material_code, end_use_code, plan):
        if plan is None or mine_code is None:
            return
        out.setdefault((rm, mine_code), []).append({
            "material_code": material_code, "end_use_code": end_use_code, "plan": plan,
        })

    for title_row, first_row, last_row, (material_code, end_use_code) in SINGLE_BLOCKS:
        months = []
        col = 2
        while ws.cell(row=title_row, column=col).value is not None:
            months.append((col, _month_str(ws.cell(row=title_row, column=col).value)))
            col += 1
        for r in range(first_row, last_row + 1):
            name = ws.cell(row=r, column=1).value
            if name is None:
                continue
            mc = _mine_code(name)
            for col, rm in months:
                add(rm, mc, material_code, end_use_code, _num(ws.cell(row=r, column=col).value))

    for date_row, first_row, last_row, end_use_code in WIDE_HDR_BLOCKS:
        months = []
        col = 2
        while ws.cell(row=date_row, column=col).value is not None:
            months.append((col, _month_str(ws.cell(row=date_row, column=col).value)))
            col += 2
        for r in range(first_row, last_row + 1):
            name = ws.cell(row=r, column=1).value
            if name is None:
                continue
            mc = _mine_code(name)
            for col, rm in months:
                add(rm, mc, "LUMP", end_use_code, _num(ws.cell(row=r, column=col).value))
                add(rm, mc, "FINES", end_use_code, _num(ws.cell(row=r, column=col + 1).value))

    for first_row, last_row, end_use_code in WIDE_BLOCKS:
        for r in range(first_row, last_row + 1):
            name = ws.cell(row=r, column=1).value
            if name is None:
                continue
            mc = _mine_code(name)
            for i, rm in enumerate(MONTHS_APR26):
                col = 2 + 2 * i
                add(rm, mc, "LUMP", end_use_code, _num(ws.cell(row=r, column=col).value))
                add(rm, mc, "FINES", end_use_code, _num(ws.cell(row=r, column=col + 1).value))

    return out


def diff_and_apply(apply: bool):
    entries_by_rm_mine = build_entries()
    changed = 0
    for (rm, mine_code) in sorted(entries_by_rm_mine):
        entries = entries_by_rm_mine[(rm, mine_code)]
        current = db.get_mines_production_despatch_monthly(rm, mine_code)["despatch_plan"]
        to_apply = []
        for e in entries:
            cur_p = current.get(e["material_code"], {}).get(e["end_use_code"])
            new_p = e["plan"]
            if cur_p is None or abs(cur_p - new_p) > 0.001:
                print(f"  {rm} {mine_code:>14} {e['material_code']:>10} {e['end_use_code']:>11}: "
                      f"plan {cur_p!r} -> {new_p}")
                changed += 1
            to_apply.append(e)
        if apply:
            db.save_mines_production_despatch_monthly(rm, mine_code, [], [], to_apply)
    print(f"{changed} cell(s) changed" + (" (applied)" if apply else " (dry run)"))


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="Write changes to the DB (default: dry run)")
    args = parser.parse_args()
    diff_and_apply(args.apply)
