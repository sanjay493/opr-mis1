"""
One-off backfill for Iron Ore Mines Production (mine-level, mines_production_
monthly — see page_sail_mines.py's module docstring and db.py's
get_iron_ore_group_rollup_monthly) from Report_format/iron ore production.xlsx.

Sheet1 has two stacked blocks, both per-mine (11 mines, same order/names as
mines_master) x Lump/Fines:
  - "Iron ore Prduction Plan" (row 3 header dates, row 5-15 data): 12 months
    Apr'26-Mar'27 -> qty_plan.
  - "Iron ore Prduction  Actual" (row 24 header dates, row 26-36 data): 16
    months Apr'25-Jul'26 -> qty_actual (same Apr'25-Jul'26 span as coal's
    ACTUAL_MONTHS in backfill_mines_data.py, for the same CPLY/YTD reasons).
Each mine's row is a flat 2N-wide grid (Lump, Fines per month, in that column
order) with no gaps — read directly from the workbook rather than
transcribing by hand (unlike backfill_mines_data.py's hardcoded DATA, this
source is a single clean grid so direct reads avoid transcription error).

Despatch and Sales (Booked Quantity) are NOT in this workbook and are left
untouched here — per direct instruction (2026-08-26), those come from the
user separately later.

Values in the workbook are raw tonnes; mines_production_monthly (like every
other sail_mines-family table) stores '000 T (confirmed against the 4 months
of KIRIBURU data already entered manually via the entry form, which are
exactly this workbook's figures / 1000), so every value is scaled here.

Run once: python scripts/backfill_iron_ore_mines_production.py [--apply]
Without --apply, only prints the diff against current DB values (dry run).
"""
import argparse
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import db

import openpyxl

WORKBOOK_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "Report_format", "iron ore production.xlsx",
)

UNIT_SCALE = 1 / 1000

MINE_NAME_TO_CODE = {
    "Kiriburu": "KIRIBURU", "Meghahatuburu": "MEGHAHATUBURU", "Gua": "GUA",
    "Manoharpur": "MANOHARPUR", "Bolani": "BOLANI", "Barsua": "BARSUA",
    "Taldih": "TALDIH", "Kalta": "KALTA", "RAJHARA": "RAJHARA",
    "DALLI": "DALLI", "ROWGHAT": "ROWGHAT",
}

PLAN_HEADER_ROW, PLAN_FIRST_DATA_ROW, PLAN_LAST_DATA_ROW = 3, 5, 15
ACTUAL_HEADER_ROW, ACTUAL_FIRST_DATA_ROW, ACTUAL_LAST_DATA_ROW = 24, 26, 36


def _month_str(dt) -> str:
    return f"{dt.year}-{dt.month:02d}"


def _read_block(ws, header_row: int, first_row: int, last_row: int, field: str):
    """-> {(mine_code, material_code): {report_month: value}} for one block
    (field is 'actual' or 'plan', used only for the KeyError message)."""
    months = []
    col = 2
    while ws.cell(row=header_row, column=col).value is not None:
        months.append(_month_str(ws.cell(row=header_row, column=col).value))
        col += 2

    out = {}
    for r in range(first_row, last_row + 1):
        mine_name = ws.cell(row=r, column=1).value
        if mine_name is None:
            continue
        mine_code = MINE_NAME_TO_CODE[mine_name]
        col = 2
        for m in months:
            lump = ws.cell(row=r, column=col).value
            fines = ws.cell(row=r, column=col + 1).value
            out[(mine_code, "LUMP")] = out.get((mine_code, "LUMP"), {})
            out[(mine_code, "FINES")] = out.get((mine_code, "FINES"), {})
            out[(mine_code, "LUMP")][m] = (lump or 0) * UNIT_SCALE
            out[(mine_code, "FINES")][m] = (fines or 0) * UNIT_SCALE
            col += 2
    return out


def build_entries():
    """-> {(report_month, mine_code): [{material_code, actual, plan}, ...]}"""
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb["Sheet1"]
    plan = _read_block(ws, PLAN_HEADER_ROW, PLAN_FIRST_DATA_ROW, PLAN_LAST_DATA_ROW, "plan")
    actual = _read_block(ws, ACTUAL_HEADER_ROW, ACTUAL_FIRST_DATA_ROW, ACTUAL_LAST_DATA_ROW, "actual")

    by_key = {}  # (report_month, mine_code, material_code) -> {"actual":, "plan":}
    for (mine_code, material_code), by_month in actual.items():
        for rm, v in by_month.items():
            by_key.setdefault((rm, mine_code, material_code), {})["actual"] = v
    for (mine_code, material_code), by_month in plan.items():
        for rm, v in by_month.items():
            by_key.setdefault((rm, mine_code, material_code), {})["plan"] = v

    out = {}
    for (rm, mine_code, material_code), vals in by_key.items():
        out.setdefault((rm, mine_code), []).append({"material_code": material_code, **vals})
    return out


def diff_and_apply(apply: bool):
    entries_by_rm_mine = build_entries()
    changed = 0
    for (rm, mine_code) in sorted(entries_by_rm_mine):
        entries = entries_by_rm_mine[(rm, mine_code)]
        current = db.get_mines_production_despatch_monthly(rm, mine_code)["production"]
        to_apply = []
        for e in entries:
            cur = current.get(e["material_code"], {})
            cur_a, cur_p = cur.get("actual"), cur.get("plan")
            new_a, new_p = e.get("actual"), e.get("plan")
            a_changed = (new_a is not None) and (cur_a is None or abs(cur_a - new_a) > 0.001)
            p_changed = (new_p is not None) and (cur_p is None or abs(cur_p - new_p) > 0.001)
            if a_changed or p_changed:
                print(f"  {rm} {mine_code:>14} {e['material_code']:>6}: "
                      + (f"actual {cur_a!r}->{new_a} " if new_a is not None else "")
                      + (f"plan {cur_p!r}->{new_p}" if new_p is not None else ""))
                changed += 1
            # Preserve whichever field this batch doesn't carry (e.g.
            # actual-only for months outside the Apr'26-Mar'27 plan span).
            to_apply.append({
                "material_code": e["material_code"],
                "actual": new_a if new_a is not None else cur_a,
                "plan": new_p if new_p is not None else cur_p,
            })
        if apply:
            db.save_mines_production_despatch_monthly(rm, mine_code, to_apply, [], [])
    print(f"{changed} cell(s) changed" + (" (applied)" if apply else " (dry run)"))


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="Write changes to the DB (default: dry run)")
    args = parser.parse_args()
    diff_and_apply(args.apply)
