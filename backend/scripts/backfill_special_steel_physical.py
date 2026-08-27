"""
One-off seed for the "Special Steel Plants Physical Performance" report
(page_special_steel_physical.py) from
Report_format/Special Steel Production history comprehensive.xlsx, sheet
"Production".

Seeds four tables (see migrate_add_special_steel_physical.sql):
  - special_steel_phys_meta        : capacity, best-achieved actual+year, remark
                                     per (plant, series)
  - special_steel_phys_perf        : metric='actual' for FY 2014-15..2025-26,
                                     metric='plan' for 2025-26 (APP) & 2026-27 (ABP)
  - special_steel_phys_note        : the two free-text footnotes (FY 2026-27)
  - special_steel_ipt_requirement  : the 10-row IPT requirement list (FY 2026-27)

Rows reproduced (matching the PDF, not every row in the workbook — VISP's
defunct Hot Metal / Crude Steel rows are omitted):
  ASP  : Crude Steel, Saleable Steel
  SSP  : Crude Steel, Saleable Steel, Stainless Steel, Carbon steel
  VISP : Saleable Steel

The workbook is in Tonnes; every physical figure is stored in '000 T
(repo-wide convention) -> divided by 1000 here. The IPT requirement "Plan"
column is already in '000 T in the workbook and is stored as-is.

After seeding, the grid is maintained by hand via
/data-entry/special-steel-physical and /data-entry/special-steel-ipt — this
script is not a live sync (Crude/Saleable actuals could later be re-derived
from production_table FY-sums, but that is deliberately not done here).

Run once: python scripts/backfill_special_steel_physical.py [--apply]
Without --apply, only prints the diff against current DB values (dry run).
"""
import argparse
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import db

import openpyxl

WORKBOOK_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "Report_format", "Special Steel Production history comprehensive.xlsx",
)

KT = 1 / 1000.0  # workbook Tonnes -> stored '000 T

PLAN_PREV_FY = "2025-26"   # workbook col 30/31 = APP / Actual
PLAN_CUR_FY = "2026-27"    # workbook col 34 = ABP
NOTE_FY = "2026-27"
IPT_FY = "2026-27"

# (workbook row, plant, series, sort_order) — "Production" sheet.
GRID_ROWS = [
    (9,  "ASP",  "CRUDE",     1),
    (10, "ASP",  "SALEABLE",  2),
    (12, "SSP",  "CRUDE",     3),
    (13, "SSP",  "SALEABLE",  4),
    (14, "SSP",  "STAINLESS", 5),
    (15, "SSP",  "CARBON",    6),
    (19, "VISP", "SALEABLE",  7),
]
COL_CAPACITY, COL_BEST_ACTUAL, COL_BEST_YEAR = 4, 5, 6
COL_FY_FIRST, FY_FIRST_START = 7, 2014         # col 7 = FY 2014-15
COL_FY_LAST = 17                                # col 17 = FY 2024-25
COL_PREV_APP, COL_PREV_ACTUAL, COL_CUR_ABP = 30, 31, 34
COL_REMARK = 37

NOTE_ROWS = [(22, 3, 4), (23, None, 4)]         # (row, prefix_col, text_col)
IPT_HEADER_ROW, IPT_FIRST_ROW, IPT_LAST_ROW = 26, 27, 36
IPT_COL_ITEM, IPT_COL_FROM, IPT_COL_TO, IPT_COL_PLAN = 3, 4, 5, 6


def _fy_label(start_year: int) -> str:
    return f"{start_year}-{(start_year + 1) % 100:02d}"


def _num(v):
    if v is None or isinstance(v, str):
        return None
    return float(v)


def _clean(s):
    return " ".join(str(s).split()) if s is not None else None


def read_workbook():
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb["Production"]

    meta = {}   # (plant, series) -> {capacity_kt, best_actual_kt, best_year, remark, sort_order}
    perf = {}   # (fy, plant, series, metric) -> value_kt
    for row, plant, series, sort_order in GRID_ROWS:
        cap = _num(ws.cell(row=row, column=COL_CAPACITY).value)
        best_a = _num(ws.cell(row=row, column=COL_BEST_ACTUAL).value)
        best_y = _clean(ws.cell(row=row, column=COL_BEST_YEAR).value)
        remark = _clean(ws.cell(row=row, column=COL_REMARK).value)
        meta[(plant, series)] = {
            "capacity_kt": cap * KT if cap is not None else None,
            "best_actual_kt": best_a * KT if best_a is not None else None,
            "best_year": best_y,
            "remark": remark,
            "sort_order": sort_order,
        }
        for col in range(COL_FY_FIRST, COL_FY_LAST + 1):
            v = _num(ws.cell(row=row, column=col).value)
            if v is not None:
                fy = _fy_label(FY_FIRST_START + (col - COL_FY_FIRST))
                perf[(fy, plant, series, "actual")] = v * KT
        prev_app = _num(ws.cell(row=row, column=COL_PREV_APP).value)
        prev_act = _num(ws.cell(row=row, column=COL_PREV_ACTUAL).value)
        cur_abp = _num(ws.cell(row=row, column=COL_CUR_ABP).value)
        if prev_app is not None:
            perf[(PLAN_PREV_FY, plant, series, "plan")] = prev_app * KT
        if prev_act is not None:
            perf[(PLAN_PREV_FY, plant, series, "actual")] = prev_act * KT
        if cur_abp is not None:
            perf[(PLAN_CUR_FY, plant, series, "plan")] = cur_abp * KT

    notes = []
    for i, (row, prefix_col, text_col) in enumerate(NOTE_ROWS, start=1):
        text = _clean(ws.cell(row=row, column=text_col).value)
        if not text:
            continue
        prefix = _clean(ws.cell(row=row, column=prefix_col).value) if prefix_col else None
        notes.append((i, f"{prefix}: {text}" if prefix else text))

    ipt = []
    last_item = None
    for r in range(IPT_FIRST_ROW, IPT_LAST_ROW + 1):
        item = _clean(ws.cell(row=r, column=IPT_COL_ITEM).value) or last_item
        frm = _clean(ws.cell(row=r, column=IPT_COL_FROM).value)
        to = _clean(ws.cell(row=r, column=IPT_COL_TO).value)
        plan = _num(ws.cell(row=r, column=IPT_COL_PLAN).value)
        if not item or not frm or not to:
            continue
        last_item = item
        ipt.append({"item": item, "from_plant": frm, "to_plant": to,
                    "plan_kt": plan, "sort_order": r - IPT_FIRST_ROW + 1})

    return meta, perf, notes, ipt


def _fnum(v):
    return None if v is None else round(v, 6)


def diff_and_apply(apply: bool):
    meta, perf, notes, ipt = read_workbook()
    cur_meta, cur_perf = db.get_ss_phys_perf()
    cur_notes = {so: t for so, t in db.get_ss_phys_notes(NOTE_FY)}
    cur_ipt = {(r["item"], r["from_plant"], r["to_plant"]): r
               for r in db.get_ss_ipt_requirement(IPT_FY)}
    changed = 0

    for (plant, series), m in sorted(meta.items()):
        c = cur_meta.get((plant, series), {})
        for k in ("capacity_kt", "best_actual_kt", "best_year", "remark", "sort_order"):
            if _fnum(c.get(k)) != _fnum(m[k]) if k.endswith("_kt") else c.get(k) != m[k]:
                print(f"  meta {plant:>4} {series:<9} {k}: {c.get(k)!r} -> {m[k]!r}")
                changed += 1
                break

    for key in sorted(perf):
        fy, plant, series, metric = key
        new = perf[key]
        old = cur_perf.get(key)
        if _fnum(old) != _fnum(new):
            print(f"  perf {fy} {plant:>4} {series:<9} {metric:<6}: {old!r} -> {round(new, 3)}")
            changed += 1

    for so, text in notes:
        if cur_notes.get(so) != text:
            print(f"  note {so}: {cur_notes.get(so)!r} -> {text!r}")
            changed += 1

    for r in ipt:
        k = (r["item"], r["from_plant"], r["to_plant"])
        old = cur_ipt.get(k)
        if old is None or _fnum(old.get("plan_kt")) != _fnum(r["plan_kt"]) or old.get("sort_order") != r["sort_order"]:
            print(f"  ipt  {r['item']:<32} {r['from_plant']}->{r['to_plant']}: "
                  f"{(old or {}).get('plan_kt')!r} -> {r['plan_kt']}")
            changed += 1

    if apply:
        db.save_ss_phys_meta([
            {"plant": p, "series": s, **m} for (p, s), m in meta.items()
        ])
        db.save_ss_phys_perf([
            {"financial_year": fy, "plant": p, "series": s, "metric": mt, "value_kt": v}
            for (fy, p, s, mt), v in perf.items()
        ])
        db.save_ss_phys_notes(NOTE_FY, [{"sort_order": so, "note_text": t} for so, t in notes])
        db.save_ss_ipt_requirement(IPT_FY, [
            {**r, "orig_item": None, "orig_from_plant": None, "orig_to_plant": None} for r in ipt
        ])
        print(f"{changed} value(s) changed (applied)")
    else:
        print(f"{changed} value(s) changed (dry run)")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="Write to the DB (default: dry run)")
    args = parser.parse_args()
    diff_and_apply(args.apply)
