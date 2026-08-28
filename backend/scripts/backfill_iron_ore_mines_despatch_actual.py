"""
One-off backfill for Iron Ore Mines Despatch ACTUAL (mine-level,
mines_despatch_actual_monthly — see db.get_mines_production_despatch_monthly)
from Report_format/Actual iron ore.xlsx.

This is the "Despatch ACTUAL ... STILL not backfilled" gap noted in
backfill_iron_ore_mines_despatch_plan.py's docstring — the user provided this
separate workbook. Actual is per transport_mode (RAIL/ROAD), unlike Plan.
Grain written: (report_month, mine_code, material_code, transport_mode,
end_use_code) -> qty_actual.

Sheet1 stacked blocks, all covering 4 months Apr'26-Jul'26 (the workbook's
own extent — cols B..E on the "narrow" blocks, and B..I Lump/Fines pairs on
the "wide" blocks; block 7 has no header row so Apr-Jul is assumed the same
as every other block):

  rows  3-5   "Tailing sales despatch road"        TAILINGS  / ROAD / SALES
  rows  9-12  "Dump Fines sale Despatch Road"      DUMP_FINES/ ROAD / SALES
  row  16     "Pellets despatch to captive rail"   PELLETS   / RAIL / CAPTIVE
  row  21     "Dump Despatch Captive rail"         DUMP_FINES/ RAIL / CAPTIVE
  rows 28-38  "Lump Fines Despatch Captive rail"   LUMP+FINES/ RAIL / CAPTIVE  (11 mines)
  rows 44-47  "conversion agent rail"              FINES     / RAIL / PELLET_CONV (Lump col blank)
  rows 56-58  "actual despatch Sales lump & Fines" LUMP+FINES/ RAIL+ROAD / SALES

Block 7: a bare mine label ("BOLANI") is that mine's TOTAL sales despatch;
a "<mine> rail" label ("BARSUA  rail", "BOLANI rail") is the rail-only
slice. So RAIL = the "…rail" row, ROAD = total - rail per cell (per direct
instruction, 2026-08-28). BARSUA has only a rail row -> no road figure.

Values in this workbook are already '000 T (NO /1000 scaling) — confirmed
against the 4 KIRIBURU LUMP/FINES/RAIL/CAPTIVE rows already entered manually
via the entry form (2026-04/05), which match this workbook's rows 28 cells
B/C/D/E bit-for-bit.

Run:  python scripts/backfill_iron_ore_mines_despatch_actual.py [--apply]
Without --apply, only prints the diff against current DB values (dry run).
Idempotent; safe to re-run after the user updates the workbook.
"""
import argparse
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import db

import openpyxl

WORKBOOK_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "Report_format", "Actual iron ore.xlsx",
)

UNIT_SCALE = 1  # workbook is already '000 T

MONTHS = ["2026-04", "2026-05", "2026-06", "2026-07"]  # workbook extent

MINE_NAME_TO_CODE = {
    "kiriburu": "KIRIBURU", "mburu": "MEGHAHATUBURU", "meghahatuburu": "MEGHAHATUBURU",
    "gua": "GUA", "manoharpur": "MANOHARPUR", "bolani": "BOLANI", "barsua": "BARSUA",
    "taldih": "TALDIH", "kalta": "KALTA", "rajhara": "RAJHARA",
    "dalli": "DALLI", "rowghat": "ROWGHAT",
}
SKIP_ROWS = {"jhr total", "odisha total", "total"}

# "narrow" blocks: months in cols B..E, one value per (mine, month).
#   (first_data_row, last_data_row, material_code, transport_mode, end_use_code)
NARROW_BLOCKS = [
    (3, 5, "TAILINGS", "ROAD", "SALES"),
    (9, 12, "DUMP_FINES", "ROAD", "SALES"),
    (16, 16, "PELLETS", "RAIL", "CAPTIVE"),
    (21, 21, "DUMP_FINES", "RAIL", "CAPTIVE"),
]
# "wide" blocks: (Lump, Fines) column pair per month, cols B..I = Apr L, Apr F,
# May L, May F, ... . Fixed mode/end-use for the whole block.
#   (first_data_row, last_data_row, transport_mode, end_use_code, lump_ok)
WIDE_BLOCKS = [
    (28, 38, "RAIL", "CAPTIVE", True),       # block 5
    (44, 47, "RAIL", "PELLET_CONV", False),  # block 6 — Fines only
]
# block 7 ("actual despatch Sales lump & Fines", rows 56-58): same wide
# (Lump, Fines) pair-per-month grid, but a bare mine label ("BOLANI") is the
# TOTAL sales despatch and a "<mine> rail" label is the rail-only slice, so
# ROAD = total - rail per cell (per direct instruction, 2026-08-28). A mine
# with only a rail row (BARSUA) has no road despatch; one with only a total
# row would be treated as all road.
BLOCK7_ROWS = (56, 58)
BLOCK7_END_USE = "SALES"


def _mine_code_and_mode(name):
    """(mine_code | None, mode | None). None mode means 'not stated in the label'."""
    n = str(name).strip().lower().replace("'", "")
    mode = None
    for suffix, m in ((" rail", "RAIL"), (" road", "ROAD")):
        if n.endswith(suffix):
            n = n[: -len(suffix)].strip()
            mode = m
    for suffix in (" group",):
        if n.endswith(suffix):
            n = n[: -len(suffix)].strip()
    if n in SKIP_ROWS or n == "":
        return None, None
    return MINE_NAME_TO_CODE.get(n), mode


def _num(v):
    if v is None or v == "":
        return None
    return round(float(v) * UNIT_SCALE, 4)


def build_entries():
    """-> {(report_month, mine_code): {(material, mode, end_use): qty_actual}}"""
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb["Sheet1"]
    out = {}

    def add(rm, mine_code, material, mode, end_use, val):
        if val is None or mine_code is None or mode is None:
            return
        out.setdefault((rm, mine_code), {})[(material, mode, end_use)] = val

    for first_row, last_row, material, mode, end_use in NARROW_BLOCKS:
        for r in range(first_row, last_row + 1):
            mc, _ = _mine_code_and_mode(ws.cell(r, 1).value)
            for i, rm in enumerate(MONTHS):
                add(rm, mc, material, mode, end_use, _num(ws.cell(r, 2 + i).value))

    for first_row, last_row, block_mode, end_use, lump_ok in WIDE_BLOCKS:
        for r in range(first_row, last_row + 1):
            mc, _ = _mine_code_and_mode(ws.cell(r, 1).value)
            for i, rm in enumerate(MONTHS):
                lump = _num(ws.cell(r, 2 + 2 * i).value)
                fines = _num(ws.cell(r, 2 + 2 * i + 1).value)
                if lump_ok:
                    add(rm, mc, "LUMP", block_mode, end_use, lump)
                add(rm, mc, "FINES", block_mode, end_use, fines)

    # block 7 — pair each bare "total" row with its "<mine> rail" row
    def _row_cells(r):
        """{(month_i, 'LUMP'|'FINES'): value}"""
        d = {}
        for i in range(len(MONTHS)):
            d[(i, "LUMP")] = _num(ws.cell(r, 2 + 2 * i).value)
            d[(i, "FINES")] = _num(ws.cell(r, 2 + 2 * i + 1).value)
        return d

    total_by_mine, rail_by_mine = {}, {}
    for r in range(BLOCK7_ROWS[0], BLOCK7_ROWS[1] + 1):
        mc, row_mode = _mine_code_and_mode(ws.cell(r, 1).value)
        if mc is None:
            continue
        (rail_by_mine if row_mode == "RAIL" else total_by_mine)[mc] = _row_cells(r)

    for mc in set(total_by_mine) | set(rail_by_mine):
        total = total_by_mine.get(mc, {})
        rail = rail_by_mine.get(mc, {})
        for i, rm in enumerate(MONTHS):
            for material in ("LUMP", "FINES"):
                t = total.get((i, material))
                rl = rail.get((i, material))
                if rl is not None:
                    add(rm, mc, material, "RAIL", BLOCK7_END_USE, rl)
                if t is not None:
                    add(rm, mc, material, "ROAD", BLOCK7_END_USE, round(t - (rl or 0), 4))

    return out


def diff_and_apply(apply: bool):
    entries = build_entries()
    changed = new = 0
    for (rm, mine_code) in sorted(entries):
        cells = entries[(rm, mine_code)]
        current = db.get_mines_production_despatch_monthly(rm, mine_code)["despatch"]
        despatch_rows = []
        for (material, mode, end_use), val in sorted(cells.items()):
            cur = current.get(material, {}).get(mode, {}).get(end_use, {}).get("actual")
            if cur is None:
                new += 1
                tag = "NEW  "
            elif abs(cur - val) > 0.001:
                changed += 1
                tag = "CHG  "
            else:
                tag = "same "
            if tag != "same ":
                print(f"  {tag}{rm} {mine_code:>14} {material:>10} {mode:>4} {end_use:>11}: {cur!r} -> {val}")
            despatch_rows.append({"material_code": material, "transport_mode": mode,
                                  "end_use_code": end_use, "actual": val})
        if apply:
            db.save_mines_production_despatch_monthly(rm, mine_code, [], despatch_rows, [])
    total_cells = sum(len(v) for v in entries.values())
    print(f"\n{total_cells} workbook cells -> {new} new, {changed} changed, "
          f"{total_cells - new - changed} unchanged" + (" (APPLIED)" if apply else " (dry run)"))


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="Write to the DB (default: dry run)")
    args = parser.parse_args()
    diff_and_apply(args.apply)
