"""
One-off backfill: load RSP and BSL's legacy SMS-1 Ingot production history
from the flat Report_format/RSPBSL_INGOT.xlsx sheet into production_table.

Source file:
  D:\\opr-mis1\\Report_format\\RSPBSL_INGOT.xlsx

Layout: row 1 holds month headers (datetime, 1st of month) starting at
column C; row 2 is RSP's "SMS-1 Ingot" series, row 3 is BSL's. Every
historic cell (Apr 2020 through the last reported month) is populated,
including explicit 0 values (actual zero production that month, not a
missing entry) — trailing columns beyond the last reported month are
blank and are skipped.

Usage:
  python scripts/backfill_rspbsl_ingot.py            # dry-run: prints a diff, writes nothing
  python scripts/backfill_rspbsl_ingot.py --apply     # actually writes to production_table (live DB)

Dry-run is the default on purpose — this touches the live MySQL DB (DB_ENGINE=mysql), so the
diff should be reviewed before anything is written.
"""
import sys
import os
import argparse
from datetime import datetime as dt

BACKEND_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BACKEND_DIR)

import db  # noqa: E402

SOURCE_FILE = r"D:\opr-mis1\Report_format\RSPBSL_INGOT.xlsx"
ITEM_NAME = "SMS-1 Ingot"
PLANTS = {"RSP": 2, "BSL": 3}  # plant -> row number in the sheet


def extract_rows():
    """Return {(plant, report_month): value} for every populated historic cell."""
    import openpyxl
    wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
    ws = wb["Sheet1"]

    month_cols = []
    for col in range(3, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if isinstance(val, dt):
            month_cols.append((col, f"{val.year}-{val.month:02d}"))

    out = {}
    for plant, row in PLANTS.items():
        assert ws.cell(row=row, column=1).value == plant
        assert ws.cell(row=row, column=2).value == ITEM_NAME
        for col, report_month in month_cols:
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            out[(plant, report_month)] = round(float(val), 3)
    return out


def current_db_values():
    conn = db.connect()
    cur = conn.cursor()
    placeholders = ",".join("?" for _ in PLANTS)
    cur.execute(
        f"SELECT plant_name, report_month, month_actual FROM production_table "
        f"WHERE item_name=? AND plant_name IN ({placeholders})",
        (ITEM_NAME, *PLANTS.keys()),
    )
    out = {(plant, month): value for plant, month, value in cur.fetchall()}
    conn.close()
    return out


def upsert(plant, report_month, value):
    conn = db.connect()
    conn.execute("""
        INSERT INTO production_table (report_month, plant_name, item_name, month_actual)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(report_month, plant_name, item_name)
        DO UPDATE SET month_actual = excluded.month_actual
    """, (report_month, plant, ITEM_NAME, value))
    conn.commit()
    conn.close()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true", help="actually write to the DB (default: dry-run)")
    args = ap.parse_args()

    final = extract_rows()
    cur_vals = current_db_values()

    print(f"--- production_table ({ITEM_NAME}, RSP+BSL): {len(final)} (plant, month) values from source file ---")
    changed = 0
    conflicts = 0
    to_write = {}
    for (plant, month), new_val in sorted(final.items(), key=lambda x: (x[0][1], x[0][0])):
        old_val = cur_vals.get((plant, month))
        if old_val is None:
            changed += 1
            print(f"  {month}  {plant:4s}  {str(old_val):>10s} -> {new_val:<10}")
            to_write[(plant, month)] = new_val
        elif round(float(old_val), 3) != round(float(new_val), 3):
            conflicts += 1
            print(f"  {month}  {plant:4s}  {old_val:<10} <> {new_val:<10}  [CONFLICT — keeping existing DB value, not overwriting]")
        # else: unchanged, no print
    print(f"  ({changed} new value(s) to fill in, {conflicts} conflicting value(s) left untouched)\n")

    if not args.apply:
        print(f"DRY RUN — nothing written. {changed} value(s) would be written. Re-run with --apply to write.")
        return

    for (plant, month), val in to_write.items():
        upsert(plant, month, val)
    print(f"APPLIED — {changed} value(s) written to the DB. {conflicts} conflicting value(s) left untouched.")


if __name__ == "__main__":
    main()
