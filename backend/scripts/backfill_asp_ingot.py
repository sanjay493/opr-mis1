"""
One-off backfill: re-extract ASP's historical REP/Excel/FL production files
and the current ABP plan sheet, using the corrected extractors (FL's "ING"
row now maps to "Ingot Semis" instead of wrongly clobbering "Ingot Steel";
item names get folded onto their canonical form via main.normalize_item_name).

Source files (must exist on disk, not archived by the app itself):
  D:\\opr-mis1\\Report_format\\MONTHEND\\ASP\\   REP*/rep*/FL*/fl* actuals
  D:\\opr-mis1\\Report_format\\ABP\\ASPSSPVISL.xlsx   plan (ASP/SSP/VISL)

Usage:
  python scripts/backfill_asp_ingot.py            # dry-run: prints a diff, writes nothing
  python scripts/backfill_asp_ingot.py --apply     # actually writes to production_table /
                                                    # production_plan_table (live DB)

Dry-run is the default on purpose — this touches the live MySQL DB (DB_ENGINE=mysql), so the
diff should be reviewed before anything is written.
"""
import sys
import os
import glob
import argparse

BACKEND_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BACKEND_DIR)
sys.path.insert(0, os.path.join(BACKEND_DIR, "excel_extractors"))

import db  # noqa: E402
from main import normalize_item_name  # noqa: E402
from excel_extractors import pdf_extractor_asp  # noqa: E402
from excel_extractors import excel_extractor_asp  # noqa: E402
from excel_extractors import excel_extractor_asp_ssp_visl_plan  # noqa: E402

ARCHIVE_DIR = r"D:\opr-mis1\Report_format\MONTHEND\ASP"
ABP_FILE = r"D:\opr-mis1\Report_format\ABP\ASPSSPVISL.xlsx"
PLANT = "ASP"


def _classify(path):
    """Return 'rep_pdf' | 'rep_excel' | 'fl_pdf' | None (skip) for a file in the archive."""
    base = os.path.basename(path)
    lower = base.lower()
    ext = os.path.splitext(lower)[1]
    if lower.startswith("fl") and ext == ".pdf":
        return "fl_pdf"
    if lower.startswith("rep"):
        if ext == ".pdf":
            return "rep_pdf"
        if ext in (".xlsx", ".xls"):
            return "rep_excel"
    return None


def extract_file(path, kind):
    """Return (rows, error) where rows is a list of
    {item_name, value, status, report_month, source_file} — value is None for
    non-'ok' rows and those are skipped by the caller."""
    fname = os.path.basename(path)
    try:
        if kind == "rep_excel":
            result = excel_extractor_asp.extract_preview(path, "")
            month = result["month"]
            rows = [
                {"item_name": r["item_name"], "value": r["value"], "status": r["status"],
                 "report_month": month, "source_file": fname}
                for r in result["production_rows"]
            ]
        else:  # rep_pdf or fl_pdf — same extractor, auto-detects REP vs FL
            result = pdf_extractor_asp.extract_preview(path, "")
            rows = [
                {"item_name": r["item_name"], "value": r["value"], "status": r["status"],
                 "report_month": r.get("report_month", result["month"]), "source_file": fname}
                for r in result["production_rows"]
            ]
        return rows, None
    except Exception as exc:
        return [], str(exc)


def extract_plan_preview():
    """Read-only re-implementation of extract_and_save_excel_plan's row scan for ASP only,
    so the dry-run diff doesn't require writing to the DB. Mirrors that function's layout
    assumptions exactly (see excel_extractor_asp_ssp_visl_plan.py's own docstring)."""
    import openpyxl
    from datetime import datetime as dt

    wb = openpyxl.load_workbook(ABP_FILE, data_only=True)
    sheet_key = next((s for s in wb.sheetnames if "APP" in s.upper() or "PLAN" in s.upper()),
                      wb.sheetnames[0])
    ws = wb[sheet_key]

    month_cols = []
    col_idx = 2
    while True:
        val = ws.cell(row=1, column=col_idx + 1).value
        if not isinstance(val, dt):
            break
        month_cols.append((col_idx + 1, f"{val.year}-{val.month:02d}"))
        col_idx += 1

    rows = []
    current_plant = None
    for row_num in range(2, ws.max_row + 1):
        a_val = ws.cell(row=row_num, column=1).value
        b_val = ws.cell(row=row_num, column=2).value
        if a_val is not None:
            current_plant = str(a_val).strip()
        if current_plant != PLANT or b_val is None:
            continue
        item_name = normalize_item_name(str(b_val).strip(), current_plant)
        for col_num, report_month in month_cols:
            raw = ws.cell(row=row_num, column=col_num).value
            val = excel_extractor_asp_ssp_visl_plan.clean_val(raw)
            if val is not None:
                rows.append({"item_name": item_name, "value": round(val, 3),
                              "report_month": report_month, "source_file": "ASPSSPVISL.xlsx"})
    return rows


def current_db_values(table):
    conn = db.connect()
    cur = conn.cursor()
    cur.execute(f"SELECT item_name, report_month, month_actual FROM {table} WHERE plant_name=?",
                (PLANT,))
    out = {(item, month): value for item, month, value in cur.fetchall()}
    conn.close()
    return out


def upsert(table, report_month, item_name, value):
    conn = db.connect()
    conn.execute(f"""
        INSERT INTO {table} (report_month, plant_name, item_name, month_actual)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(report_month, plant_name, item_name)
        DO UPDATE SET month_actual = excluded.month_actual
    """, (report_month, PLANT, item_name, value))
    conn.commit()
    conn.close()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true", help="actually write to the DB (default: dry-run)")
    args = ap.parse_args()

    files = sorted(glob.glob(os.path.join(ARCHIVE_DIR, "*.pdf"))
                    + glob.glob(os.path.join(ARCHIVE_DIR, "*.xlsx"))
                    + glob.glob(os.path.join(ARCHIVE_DIR, "*.xls")))

    by_kind = {"rep_excel": [], "rep_pdf": [], "fl_pdf": []}
    for f in files:
        kind = _classify(f)
        if kind:
            by_kind[kind].append(f)

    print(f"Archive: {len(by_kind['rep_excel'])} REP Excel, {len(by_kind['rep_pdf'])} REP PDF, "
          f"{len(by_kind['fl_pdf'])} FL PDF files found.\n")

    # final[(item, report_month)] = (value, source_file) — REP Excel/PDF processed first
    # (Excel preferred when both exist for the same month), then FL last, matching the
    # order these were actually uploaded historically (see extraction_log).
    final = {}
    errors = []

    def apply_rows(rows):
        for r in rows:
            if r["status"] != "ok" or r["value"] is None:
                continue
            item = normalize_item_name(r["item_name"], PLANT)
            key = (item, r["report_month"])
            final[key] = (r["value"], r["source_file"])

    rep_excel_months = set()
    for f in by_kind["rep_excel"]:
        rows, err = extract_file(f, "rep_excel")
        if err:
            errors.append((f, err))
            continue
        if rows:
            rep_excel_months.add(rows[0]["report_month"])
        apply_rows(rows)

    for f in by_kind["rep_pdf"]:
        rows, err = extract_file(f, "rep_pdf")
        if err:
            errors.append((f, err))
            continue
        # Skip if the Excel version already covered this month (Excel has more items).
        if rows and rows[0]["report_month"] in rep_excel_months:
            continue
        apply_rows(rows)

    for f in by_kind["fl_pdf"]:
        rows, err = extract_file(f, "fl_pdf")
        if err:
            errors.append((f, err))
            continue
        apply_rows(rows)

    plan_rows = extract_plan_preview()
    plan_final = {}
    for r in plan_rows:
        plan_final[(r["item_name"], r["report_month"])] = (r["value"], r["source_file"])

    if errors:
        print(f"--- {len(errors)} file(s) failed to extract (skipped) ---")
        for f, err in errors:
            print(f"  {os.path.basename(f)}: {err}")
        print()

    cur_actual = current_db_values("production_table")
    cur_plan = current_db_values("production_plan_table")

    def print_diff(label, new_vals, cur_vals):
        print(f"--- {label}: {len(new_vals)} (item, month) values from source files ---")
        changed = 0
        for (item, month), (new_val, source) in sorted(new_vals.items(), key=lambda x: (x[0][1], x[0][0])):
            old_val = cur_vals.get((item, month))
            if old_val is None or round(float(old_val), 3) != round(float(new_val), 3):
                changed += 1
                print(f"  {month}  {item:22s}  {str(old_val):>10s} -> {new_val:<10}  [{source}]")
        print(f"  ({changed} of {len(new_vals)} differ from current DB values)\n")
        return changed

    n1 = print_diff("production_table (actuals)", final, cur_actual)
    n2 = print_diff("production_plan_table (plan)", plan_final, cur_plan)

    if not args.apply:
        print(f"DRY RUN — nothing written. {n1 + n2} total value(s) would change. "
              f"Re-run with --apply to write.")
        return

    for (item, month), (val, _src) in final.items():
        upsert("production_table", month, item, val)
    for (item, month), (val, _src) in plan_final.items():
        upsert("production_plan_table", month, item, val)
    print(f"APPLIED — {n1 + n2} value(s) written to the DB.")


if __name__ == "__main__":
    main()
