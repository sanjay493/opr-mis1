"""
Excel / PDF export for the ad-hoc Unit-wise Production Query
(frontend: /reports/production-query) — same data as POST /api/production-query,
aggregated month-wise, quarter-wise or year-wise (financial year) per request.

Table orientation matches the on-screen table: periods (months/quarters/FYs)
as rows, one APP + Actual column pair per selected plant/unit, plus a
Cumulative row summarizing the whole selected range.
"""
import io
import re

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

_MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

_RATE_ITEM_RE = re.compile(r"/day|/d\)", re.IGNORECASE)


def month_label(ym: str) -> str:
    y, m = ym.split("-")
    return f"{_MONTH_NAMES[int(m) - 1]}'{y[2:]}"


def is_rate_item(name: str) -> bool:
    """Items expressed as a daily rate — summing across a period is
    meaningless, average instead. Mirrors the frontend's isRateItem()."""
    return bool(_RATE_ITEM_RE.search(name)) or name.upper().startswith("COB#")


def _fy_start(month: str) -> int:
    y, m = int(month[:4]), int(month[5:7])
    return y if m >= 4 else y - 1


def _quarter_num(month: str) -> int:
    m = int(month[5:7])
    return {4: 1, 5: 1, 6: 1, 7: 2, 8: 2, 9: 2, 10: 3, 11: 3, 12: 3, 1: 4, 2: 4, 3: 4}[m]


def fy_label(fy_start: int) -> str:
    return f"{fy_start}-{(fy_start + 1) % 100:02d}"


def bucket_months(months: list, view: str) -> list:
    """Group a chronological list of 'YYYY-MM' strings into display periods.
    Returns [{"label": str, "months": [...]}, ...] in chronological order.
    view='month': one bucket per month. 'quarter'/'year': grouped by FY
    quarter / FY, relying on `months` already being sorted ascending."""
    if view == "month":
        return [{"label": month_label(m), "months": [m]} for m in months]

    if view == "quarter":
        def key(m):
            return (_fy_start(m), _quarter_num(m))
        def label(k):
            return f"Q{k[1]} {fy_label(k[0])}"
    else:  # year
        def key(m):
            return _fy_start(m)
        def label(k):
            return fy_label(k)

    buckets = []
    cur_key, cur_months = None, []
    for m in months:
        k = key(m)
        if k != cur_key:
            if cur_months:
                buckets.append({"label": label(cur_key), "months": cur_months})
            cur_key, cur_months = k, []
        cur_months.append(m)
    if cur_months:
        buckets.append({"label": label(cur_key), "months": cur_months})
    return buckets


def _period_value(values: dict, month_list: list, is_rate: bool):
    nums = [values.get(m) for m in month_list if values.get(m) is not None]
    if not nums:
        return None
    s = sum(nums)
    return round(s / len(nums) if is_rate else s, 3)


VIEW_TITLES = {"month": "Month-wise", "quarter": "Quarter-wise", "year": "Year-wise"}


# ── Excel ────────────────────────────────────────────────────────────────────

_HDR_FILL = PatternFill("solid", fgColor="1A73E8")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_SUBHDR_FILL = PatternFill("solid", fgColor="E8F0FE")
_SUBHDR_FONT = Font(bold=True, color="174EA6", size=9)
_PERIOD_FONT = Font(bold=True, size=9)
_CUM_FILL = PatternFill("solid", fgColor="E8F0FE")
_CUM_FONT = Font(bold=True, size=9, color="174EA6")
_THIN = Side(style="thin", color="DADCE0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_ZEBRA_FILL = PatternFill("solid", fgColor="F8F9FA")


def build_excel_bytes(data: dict, view: str) -> bytes:
    months = data["months"]
    series = data["series"]
    periods = bucket_months(months, view)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = VIEW_TITLES.get(view, "Month-wise")[:31]

    start_lbl = month_label(months[0]) if months else ""
    end_lbl = month_label(months[-1]) if months else ""
    title = f"Unit-wise Production Query — {VIEW_TITLES.get(view, view)} ({start_lbl} to {end_lbl})"
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
    ws.cell(row=2, column=1, value="Unit: '000T unless stated").font = Font(italic=True, size=9)

    top_row, sub_row = 4, 5
    ws.cell(row=top_row, column=1, value="Period").font = _HDR_FONT
    ws.cell(row=top_row, column=1).fill = _HDR_FILL
    ws.merge_cells(start_row=top_row, start_column=1, end_row=sub_row, end_column=1)

    for i, s in enumerate(series):
        col = 2 + i * 2
        ws.merge_cells(start_row=top_row, start_column=col, end_row=top_row, end_column=col + 1)
        c = ws.cell(row=top_row, column=col, value=f"{s['plant']} · {s['item']}")
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="center")
        ws.cell(row=top_row, column=col + 1).fill = _HDR_FILL
        for j, sub in enumerate(("APP", "Actual")):
            sc = ws.cell(row=sub_row, column=col + j, value=sub)
            sc.font = _SUBHDR_FONT
            sc.fill = _SUBHDR_FILL
            sc.alignment = Alignment(horizontal="center")

    total_cols = 1 + len(series) * 2
    row = sub_row + 1
    for idx, period in enumerate(periods):
        fill = _ZEBRA_FILL if idx % 2 == 1 else None
        pc = ws.cell(row=row, column=1, value=period["label"])
        pc.font = _PERIOD_FONT
        pc.border = _BORDER
        if fill:
            pc.fill = fill
        for i, s in enumerate(series):
            is_rate = is_rate_item(s["item"])
            plan_v = _period_value(s["plan"], period["months"], is_rate)
            act_v = _period_value(s["actual"], period["months"], is_rate)
            for j, v in enumerate((plan_v, act_v)):
                vc = ws.cell(row=row, column=2 + i * 2 + j)
                if v is not None:
                    vc.value = v
                    vc.number_format = "#,##0.000"
                vc.alignment = Alignment(horizontal="right")
                vc.border = _BORDER
                if fill:
                    vc.fill = fill
        row += 1

    # Cumulative row — aggregate over the whole selected range
    cc = ws.cell(row=row, column=1, value="Cumulative")
    cc.font = _CUM_FONT
    cc.fill = _CUM_FILL
    cc.border = _BORDER
    for i, s in enumerate(series):
        is_rate = is_rate_item(s["item"])
        plan_v = _period_value(s["plan"], months, is_rate)
        act_v = _period_value(s["actual"], months, is_rate)
        for j, v in enumerate((plan_v, act_v)):
            vc = ws.cell(row=row, column=2 + i * 2 + j)
            if v is not None:
                vc.value = v
                vc.number_format = "#,##0.000"
            vc.font = _CUM_FONT
            vc.fill = _CUM_FILL
            vc.alignment = Alignment(horizontal="right")
            vc.border = _BORDER

    ws.column_dimensions["A"].width = 16
    for i in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 11
    ws.freeze_panes = ws.cell(row=sub_row + 1, column=2).coordinate

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF ──────────────────────────────────────────────────────────────────────

def _fmt(v):
    if v is None:
        return "—"
    return f"{v:,.3f}".rstrip("0").rstrip(".") if v != int(v) else f"{int(v):,}"


def build_pdf_html(data: dict, view: str) -> str:
    months = data["months"]
    series = data["series"]
    periods = bucket_months(months, view)

    top_headers = "".join(
        f'<th colspan="2">{s["plant"]} &middot; {s["item"]}</th>' for s in series
    )
    sub_headers = "".join("<th>APP</th><th>Actual</th>" for _ in series)

    body_rows = []
    for period in periods:
        cells = []
        for s in series:
            is_rate = is_rate_item(s["item"])
            plan_v = _period_value(s["plan"], period["months"], is_rate)
            act_v = _period_value(s["actual"], period["months"], is_rate)
            cells.append(f"<td>{_fmt(plan_v)}</td><td>{_fmt(act_v)}</td>")
        body_rows.append(f'<tr><td class="period">{period["label"]}</td>{"".join(cells)}</tr>')

    cum_cells = []
    for s in series:
        is_rate = is_rate_item(s["item"])
        plan_v = _period_value(s["plan"], months, is_rate)
        act_v = _period_value(s["actual"], months, is_rate)
        avg_tag = ' <span class="avg">(avg)</span>' if is_rate and (plan_v is not None or act_v is not None) else ""
        cum_cells.append(f"<td>{_fmt(plan_v)}{avg_tag}</td><td>{_fmt(act_v)}{avg_tag}</td>")
    cum_row = f'<tr class="cum-row"><td class="period">Cumulative</td>{"".join(cum_cells)}</tr>'

    start_lbl = month_label(months[0]) if months else ""
    end_lbl = month_label(months[-1]) if months else ""

    return f"""<!doctype html>
<html><head><meta charset="utf-8">
<style>
  @page {{ size: A4 landscape; margin: 12mm 10mm; }}
  * {{ box-sizing: border-box; }}
  body {{ font-family: Arial, sans-serif; color: #202124; margin: 0; }}
  h1 {{ font-size: 14pt; margin: 0 0 2px 0; }}
  .subtitle {{ font-size: 9pt; color: #5f6368; margin: 0 0 10px 0; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 7.3pt; }}
  thead {{ display: table-header-group; }}
  tr {{ page-break-inside: avoid; }}
  th, td {{ border: 1px solid #dadce0; padding: 3px 5px; text-align: right; white-space: nowrap; }}
  th {{ background: #1a73e8; color: #ffffff; font-weight: 700; }}
  thead tr:first-child th {{ font-size: 7.6pt; }}
  th:first-child, td.period {{ text-align: left; }}
  td.period {{ font-weight: 700; }}
  tr:nth-child(even):not(.cum-row) {{ background: #f8f9fa; }}
  tr.cum-row td {{ background: #e8f0fe; color: #174ea6; font-weight: 700; border-top: 2px solid #1a73e8; }}
  .avg {{ font-weight: 400; color: #5f6368; font-size: 6.5pt; }}
</style>
</head>
<body>
  <h1>Unit-wise Production Query ({VIEW_TITLES.get(view, view)})</h1>
  <p class="subtitle">{start_lbl} to {end_lbl} &middot; Unit: '000T unless stated</p>
  <table>
    <thead>
      <tr><th>Period</th>{top_headers}</tr>
      <tr><th></th>{sub_headers}</tr>
    </thead>
    <tbody>{''.join(body_rows)}{cum_row}</tbody>
  </table>
</body></html>"""


def render_pdf_bytes(html: str) -> bytes:
    """Synchronous — call via a threadpool executor from async code."""
    from playwright.sync_api import sync_playwright
    with sync_playwright() as pw:
        browser = pw.chromium.launch()
        page = browser.new_page()
        page.set_content(html, wait_until="domcontentloaded")
        pdf_bytes = page.pdf(format="A4", landscape=True, print_background=True,
                             margin={"top": "12mm", "right": "10mm", "bottom": "12mm", "left": "10mm"})
        browser.close()
    return pdf_bytes
