"""
Excel / PDF export for the "Special Steel — Order & Actual Despatch" FY
report (frontend: /reports/special-steel-fy) — same data as
GET /api/special-steel-fy, one plant-grouped table: Order Qty / Actual
Despatch | Apr..Mar | Total.
"""
import io

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

_MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

_ROW_LABELS = [("order", "Order Qty"), ("actual", "Actual Despatch")]


def month_label(ym: str) -> str:
    y, m = ym.split("-")
    return f"{_MONTH_NAMES[int(m) - 1]}'{y[2:]}"


def row_total(values: dict, months: list):
    nums = [values.get(m) for m in months if values.get(m) is not None]
    if not nums:
        return None
    return sum(nums)


# ── Excel ────────────────────────────────────────────────────────────────────

_HDR_FILL = PatternFill("solid", fgColor="1A73E8")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_PLANT_FILL = PatternFill("solid", fgColor="174EA6")
_SAIL_FILL = PatternFill("solid", fgColor="0B3D91")
_PLANT_FONT = Font(bold=True, color="FFFFFF", size=11)
_ITEM_FONT = Font(bold=True, size=9)
_TOTAL_FONT = Font(bold=True, size=9, color="174EA6")
_THIN = Side(style="thin", color="DADCE0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_ZEBRA_FILL = PatternFill("solid", fgColor="F8F9FA")


def build_excel_bytes(data: dict) -> bytes:
    months = data["months"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Special Steel"

    title = f"Special Steel — Order & Actual Despatch — FY {data['fy_label']}"
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
    ws.cell(row=2, column=1, value="Unit: T").font = Font(italic=True, size=9)

    header_row = 4
    ws.cell(row=header_row, column=1, value="Plant").font = _HDR_FONT
    ws.cell(row=header_row, column=1).fill = _HDR_FILL
    for i, m in enumerate(months, start=2):
        c = ws.cell(row=header_row, column=i, value=month_label(m))
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="center")
    total_col = len(months) + 2
    c = ws.cell(row=header_row, column=total_col, value="Total")
    c.font = _HDR_FONT
    c.fill = _HDR_FILL
    c.alignment = Alignment(horizontal="center")

    row = header_row + 1
    for plant in data["plants"]:
        is_sail = plant["plant"] == "SAIL"
        ws.cell(row=row, column=1, value=plant["plant"]).font = _PLANT_FONT
        for col in range(1, total_col + 1):
            ws.cell(row=row, column=col).fill = _SAIL_FILL if is_sail else _PLANT_FILL
        ws.row_dimensions[row].height = 20
        row += 1

        for idx, (key, label) in enumerate(_ROW_LABELS):
            values = plant[key]
            fill = _ZEBRA_FILL if idx % 2 == 1 else None
            name_cell = ws.cell(row=row, column=1, value=label)
            name_cell.font = _ITEM_FONT
            name_cell.border = _BORDER
            if fill:
                name_cell.fill = fill
            for i, m in enumerate(months, start=2):
                v = values.get(m)
                vc = ws.cell(row=row, column=i)
                if v is not None:
                    vc.value = round(v, 3)
                    vc.number_format = "#,##0.000"
                vc.alignment = Alignment(horizontal="right")
                vc.border = _BORDER
                if fill:
                    vc.fill = fill
            total = row_total(values, months)
            tc = ws.cell(row=row, column=total_col)
            if total is not None:
                tc.value = round(total, 3)
                tc.number_format = "#,##0.000"
            tc.font = _TOTAL_FONT
            tc.alignment = Alignment(horizontal="right")
            tc.border = _BORDER
            if fill:
                tc.fill = fill
            row += 1

    ws.column_dimensions["A"].width = 24
    for i in range(2, total_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 11
    ws.freeze_panes = ws.cell(row=header_row + 1, column=2).coordinate

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF ──────────────────────────────────────────────────────────────────────

def _fmt(v):
    if v is None:
        return "—"
    return f"{v:,.3f}".rstrip("0").rstrip(".") if v != int(v) else f"{int(v):,}"


def build_pdf_html(data: dict) -> str:
    months = data["months"]

    col_headers = "".join(f"<th>{month_label(m)}</th>" for m in months)

    body_rows = []
    for plant in data["plants"]:
        is_sail = plant["plant"] == "SAIL"
        row_cls = "plant-row sail-row" if is_sail else "plant-row"
        body_rows.append(
            f'<tr class="{row_cls}"><td colspan="{len(months) + 2}">{plant["plant"]}</td></tr>'
        )
        for key, label in _ROW_LABELS:
            values = plant[key]
            cells = "".join(f"<td>{_fmt(values.get(m))}</td>" for m in months)
            total = row_total(values, months)
            body_rows.append(
                f'<tr><td class="item">{label}</td>{cells}'
                f'<td class="total">{_fmt(total)}</td></tr>'
            )

    return f"""<!doctype html>
<html><head><meta charset="utf-8">
<style>
  @page {{ size: A4 landscape; margin: 12mm 10mm; }}
  * {{ box-sizing: border-box; }}
  body {{ font-family: Arial, sans-serif; color: #202124; margin: 0; }}
  h1 {{ font-size: 14pt; margin: 0 0 2px 0; }}
  .subtitle {{ font-size: 9pt; color: #5f6368; margin: 0 0 10px 0; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 8pt; }}
  thead {{ display: table-header-group; }}
  tr {{ page-break-inside: avoid; }}
  th, td {{ border: 1px solid #dadce0; padding: 4px 6px; text-align: right; white-space: nowrap; }}
  th {{ background: #1a73e8; color: #ffffff; font-weight: 700; }}
  th:first-child, td.item {{ text-align: left; }}
  td.item {{ font-weight: 600; }}
  td.total {{ font-weight: 700; color: #174ea6; }}
  tr.plant-row td {{ background: #174ea6; color: #ffffff; font-weight: 800;
                      text-align: left; font-size: 10pt; letter-spacing: 0.03em; }}
  tr.sail-row td {{ background: #0b3d91; }}
  tr:nth-child(even):not(.plant-row) {{ background: #f8f9fa; }}
</style>
</head>
<body>
  <h1>Special Steel — Order &amp; Actual Despatch</h1>
  <p class="subtitle">Month-wise, plant-wise — FY {data['fy_label']} · Unit: T</p>
  <table>
    <thead><tr><th>Plant</th>{col_headers}<th>Total</th></tr></thead>
    <tbody>{''.join(body_rows)}</tbody>
  </table>
</body></html>"""


def render_pdf_bytes(html: str) -> bytes:
    """Synchronous — call via a threadpool executor from async code."""
    from playwright.sync_api import sync_playwright
    with sync_playwright() as pw:
        browser = pw.chromium.launch()
        page = browser.new_page()
        page.set_content(html, wait_until="domcontentloaded")
        pdf_bytes = page.pdf(format="A4", landscape=True, print_background=True,
                             margin={"top": "12mm", "right": "10mm", "bottom": "12mm", "left": "10mm"})
        browser.close()
    return pdf_bytes
