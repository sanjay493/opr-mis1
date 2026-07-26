"""
Excel / PDF export for the "Month-wise Production" FY report
(frontend: /reports/production-fy) — same data as GET /api/production-fy,
one plant-grouped table: Item | Apr..Mar | Total.
"""
import io
import re

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

_MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

_RATE_ITEM_RE = re.compile(r"/day|/d\)", re.IGNORECASE)


def month_label(ym: str) -> str:
    y, m = ym.split("-")
    return f"{_MONTH_NAMES[int(m) - 1]}'{y[2:]}"


def is_rate_item(name: str) -> bool:
    """Items expressed as a daily rate — summing across a year is meaningless,
    average instead. Mirrors the frontend's isRateItem()."""
    return bool(_RATE_ITEM_RE.search(name)) or name.upper().startswith("COB#")


def row_total(item_name: str, values: dict, months: list):
    nums = [values.get(m) for m in months if values.get(m) is not None]
    if not nums:
        return None
    s = sum(nums)
    return s / len(nums) if is_rate_item(item_name) else s


# ── Excel ────────────────────────────────────────────────────────────────────

_HDR_FILL = PatternFill("solid", fgColor="1A73E8")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_PLANT_FILL = PatternFill("solid", fgColor="1A73E8")
_PLANT_FONT = Font(bold=True, color="FFFFFF", size=11)
_ITEM_FONT = Font(bold=True, size=9)
_TOTAL_FONT = Font(bold=True, size=9, color="174EA6")
_THIN = Side(style="thin", color="DADCE0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_ZEBRA_FILL = PatternFill("solid", fgColor="F8F9FA")


def build_excel_bytes(data: dict, mode: str) -> bytes:
    months = data["months"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Actual" if mode == "actual" else "Plan"

    title = f"Month-wise Production ({'Actual' if mode == 'actual' else 'Plan'}) — FY {data['fy_label']}"
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
    ws.cell(row=2, column=1, value="Unit: '000T unless stated").font = Font(italic=True, size=9)

    header_row = 4
    ws.cell(row=header_row, column=1, value="Item").font = _HDR_FONT
    ws.cell(row=header_row, column=1).fill = _HDR_FILL
    for i, m in enumerate(months, start=2):
        c = ws.cell(row=header_row, column=i, value=month_label(m))
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="center")
    total_col = len(months) + 2
    c = ws.cell(row=header_row, column=total_col, value="Total")
    c.font = _HDR_FONT
    c.fill = _HDR_FILL
    c.alignment = Alignment(horizontal="center")

    row = header_row + 1
    for plant in data["plants"]:
        items = [it for it in plant["items"] if any(it[mode].get(m) is not None for m in months)]
        if not items:
            continue
        ws.cell(row=row, column=1, value=plant["plant"]).font = _PLANT_FONT
        for col in range(1, total_col + 1):
            ws.cell(row=row, column=col).fill = _PLANT_FILL
        ws.row_dimensions[row].height = 20
        row += 1

        for idx, it in enumerate(items):
            values = it[mode]
            fill = _ZEBRA_FILL if idx % 2 == 1 else None
            name_cell = ws.cell(row=row, column=1, value=it["item_name"])
            name_cell.font = _ITEM_FONT
            name_cell.border = _BORDER
            if fill:
                name_cell.fill = fill
            for i, m in enumerate(months, start=2):
                v = values.get(m)
                vc = ws.cell(row=row, column=i)
                if v is not None:
                    vc.value = round(v, 3)
                    vc.number_format = "#,##0.000"
                vc.alignment = Alignment(horizontal="right")
                vc.border = _BORDER
                if fill:
                    vc.fill = fill
            total = row_total(it["item_name"], values, months)
            tc = ws.cell(row=row, column=total_col)
            if total is not None:
                tc.value = round(total, 3)
                tc.number_format = "#,##0.000"
            tc.font = _TOTAL_FONT
            tc.alignment = Alignment(horizontal="right")
            tc.border = _BORDER
            if fill:
                tc.fill = fill
            row += 1

    ws.column_dimensions["A"].width = 30
    for i in range(2, total_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 11
    ws.freeze_panes = ws.cell(row=header_row + 1, column=2).coordinate

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF ──────────────────────────────────────────────────────────────────────

def _fmt(v):
    if v is None:
        return "—"
    return f"{v:,.3f}".rstrip("0").rstrip(".") if v != int(v) else f"{int(v):,}"


def build_pdf_html(data: dict, mode: str) -> str:
    months = data["months"]
    mode_label = "Actual" if mode == "actual" else "Plan"

    col_headers = "".join(f"<th>{month_label(m)}</th>" for m in months)

    body_rows = []
    for plant in data["plants"]:
        items = [it for it in plant["items"] if any(it[mode].get(m) is not None for m in months)]
        if not items:
            continue
        body_rows.append(
            f'<tr class="plant-row"><td colspan="{len(months) + 2}">{plant["plant"]}</td></tr>'
        )
        for it in items:
            values = it[mode]
            cells = "".join(f"<td>{_fmt(values.get(m))}</td>" for m in months)
            total = row_total(it["item_name"], values, months)
            avg_tag = " <span class='avg'>(avg)</span>" if total is not None and is_rate_item(it["item_name"]) else ""
            body_rows.append(
                f'<tr><td class="item">{it["item_name"]}</td>{cells}'
                f'<td class="total">{_fmt(total)}{avg_tag}</td></tr>'
            )

    return f"""<!doctype html>
<html><head><meta charset="utf-8">
<style>
  @page {{ size: A4 landscape; margin: 12mm 10mm; }}
  * {{ box-sizing: border-box; }}
  body {{ font-family: Arial, sans-serif; color: #202124; margin: 0; }}
  h1 {{ font-size: 14pt; margin: 0 0 2px 0; }}
  .subtitle {{ font-size: 9pt; color: #5f6368; margin: 0 0 10px 0; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 7.3pt; }}
  thead {{ display: table-header-group; }}
  tr {{ page-break-inside: avoid; }}
  th, td {{ border: 1px solid #dadce0; padding: 3px 5px; text-align: right; white-space: nowrap; }}
  th {{ background: #1a73e8; color: #ffffff; font-weight: 700; }}
  th:first-child, td.item {{ text-align: left; }}
  td.item {{ font-weight: 600; }}
  td.total {{ font-weight: 700; color: #174ea6; }}
  tr.plant-row td {{ background: #1a73e8; color: #ffffff; font-weight: 800;
                      text-align: left; font-size: 9pt; letter-spacing: 0.03em; }}
  tr:nth-child(even):not(.plant-row) {{ background: #f8f9fa; }}
  .avg {{ font-weight: 400; color: #5f6368; font-size: 6.5pt; }}
</style>
</head>
<body>
  <h1>Month-wise Production ({mode_label})</h1>
  <p class="subtitle">All plants, all items — FY {data['fy_label']} · Unit: '000T unless stated</p>
  <table>
    <thead><tr><th>Item</th>{col_headers}<th>Total</th></tr></thead>
    <tbody>{''.join(body_rows)}</tbody>
  </table>
</body></html>"""


def render_pdf_bytes(html: str) -> bytes:
    """Synchronous — call via a threadpool executor from async code."""
    from playwright.sync_api import sync_playwright
    with sync_playwright() as pw:
        browser = pw.chromium.launch()
        page = browser.new_page()
        page.set_content(html, wait_until="domcontentloaded")
        pdf_bytes = page.pdf(format="A4", landscape=True, print_background=True,
                             margin={"top": "12mm", "right": "10mm", "bottom": "12mm", "left": "10mm"})
        browser.close()
    return pdf_bytes
