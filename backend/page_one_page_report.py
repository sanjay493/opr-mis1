"""
SAIL "1 page report" (Report_format/1 page report for Jun26.xlsx format) —
combines 4 tables into one workbook for a selected report_month:

  A. Sales             — sail_sales_table (data_json, verbatim archive —
                          extracted from the external Sales report, see
                          excel_extractors/sail_sales_stock_extractor.py;
                          every figure here — including %Ful/CPLY/Growth/
                          cumulative — is read back exactly as reported,
                          never recomputed, since the source department's
                          own cumulative figures are provisional and don't
                          reliably equal summing each month's own Actual;
                          sail_sales_note_table carries the asterisked
                          remark that sometimes follows the table, e.g.
                          "*Jul25 & Apr-Jul25 fig incl NSL sales: 98 & 482
                          respectively", reprinted verbatim underneath)
  B. Production         — production_table / production_plan_table
                          (Hot Metal / Crude Steel / Saleable Steel, same
                          source as page4.py's Crude Steel Production page)
  C. Techno-Economic     — reuses page_techno.generate_summary_te_table()'s
     Parameters            SAIL rows (Coke Rate / SEC / BF Productivity)
  D. Stock - 8 Plants    — sail_stock_snapshot_table (extracted alongside
                          Sales from the same external report)

Tables A and D depend entirely on having been extracted+saved via
sail_sales_stock_extractor.py for the relevant months/dates; anything not
yet uploaded renders blank rather than guessed.
"""
import calendar
from datetime import date

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Color
from openpyxl.utils import get_column_letter

import db
from constants import FIVE_PLANTS as _5P
from page_techno import generate_summary_te_table

_MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
           "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _mlabel(ym: str) -> str:
    y, m = int(ym[:4]), int(ym[5:7])
    return f"{_MONTHS[m - 1]}'{str(y)[2:]}"


def _fy_start(ym: str) -> int:
    y, m = int(ym[:4]), int(ym[5:7])
    return y if m >= 4 else y - 1


def _cum_label(month: str) -> str:
    fy = _fy_start(month)
    return f"Apr-{_mlabel(month)}"


def _gr(cur_v, prev_v):
    """Fraction (0.18, not 18) — cells use Excel's '0%' number format for
    display, same convention as the verbatim Sales figures (see
    sail_sales_stock_extractor.py)."""
    if cur_v is None or prev_v is None or prev_v == 0:
        return None
    return (cur_v - prev_v) / abs(prev_v)


def _pct_ful(actual, abp):
    """Fraction (0.81, not 81) — see _gr."""
    if actual is None or abp is None or abp == 0:
        return None
    return actual / abp


# ── Table A: Sales ───────────────────────────────────────────────────────

_SALES_ITEMS = [
    "Lp Sales", "Fp Sales", "Pet Sales", "Total : Home Sales",
    "Sales By Special Steel Plants", "Exports", "Total Cmo Sales",
    "Plant Sales", "Total Sales",
]


def _sales_data(cur, item, month):
    """Raw data_json for one (report_month, item) — every field exactly as
    the source department reported it, no computation. None if not extracted."""
    import json
    cur.execute("SELECT data_json FROM sail_sales_table WHERE report_month=? AND item_name=?", (month, item))
    r = cur.fetchone()
    if not r or not r[0]:
        return {}
    data = r[0] if isinstance(r[0], dict) else json.loads(r[0])
    return data or {}


def _sum_months(getter, months):
    total, found = 0.0, False
    for m in months:
        v = getter(m)
        if v is not None:
            total += v
            found = True
    return round(total, 3) if found else None


def _build_sales_rows(cur, month):
    """Table A rows, read back verbatim — no CPLY/cumulative lookups or
    summation across months; this month's own report already carries
    everything (see sail_sales_stock_extractor.py's module docstring)."""
    rows = []
    for item in _SALES_ITEMS:
        d = _sales_data(cur, item, month)
        rows.append({
            "label": item, "bold": item.upper().startswith("TOTAL"),
            "m_abp": d.get("month_abp"), "m_act": d.get("month_actual"),
            "m_ful": d.get("month_ful"), "m_cply": d.get("month_cply"),
            "m_growth": d.get("month_growth"),
            "c_abp": d.get("till_month_abp"), "c_act": d.get("till_month_actual"),
            "c_ful": d.get("till_month_ful"), "c_cply": d.get("till_month_cply"),
            "c_growth": d.get("till_month_growth"),
        })
    return rows


def _sales_note(cur, month):
    """Asterisked remark extracted alongside Table A (e.g. NSL-inclusion
    caveat on the CPLY figures) — reprinted verbatim under the table."""
    cur.execute("SELECT note FROM sail_sales_note_table WHERE report_month=?", (month,))
    r = cur.fetchone()
    return r[0] if r and r[0] else None


# ── Table B: Production ─────────────────────────────────────────────────

def _p_one(cur, table, plant, item, month):
    tbl = "production_table" if table == "act" else "production_plan_table"
    cur.execute(f"SELECT month_actual FROM {tbl} WHERE report_month=? AND plant_name=? AND item_name=?",
                (month, plant, item))
    r = cur.fetchone()
    return r[0] if r and r[0] is not None else None


def _p_sum(cur, table, plants, item, month):
    if not plants:
        return None
    tbl = "production_table" if table == "act" else "production_plan_table"
    phs = ",".join("?" for _ in plants)
    cur.execute(f"SELECT SUM(month_actual) FROM {tbl} WHERE report_month=? AND plant_name IN ({phs}) AND item_name=?",
                [month] + list(plants) + [item])
    r = cur.fetchone()
    return r[0] if r and r[0] is not None else None


def _p_cum_one(cur, table, plant, item, months):
    return _sum_months(lambda mo: _p_one(cur, table, plant, item, mo), months)


def _p_cum_sum(cur, table, plants, item, months):
    return _sum_months(lambda mo: _p_sum(cur, table, plants, item, mo), months)


def _prod_row(label, m_abp, m_act, m_cply, c_abp, c_act, c_cply, indent=1, bold=False):
    return {"label": label, "indent": indent, "bold": bold,
            "m_abp": m_abp, "m_act": m_act, "m_cply": m_cply,
            "c_abp": c_abp, "c_act": c_act, "c_cply": c_cply}


def _build_production_rows(cur, month, cply_month, ytd_months, cply_ytd_months):
    rows = []

    def plant_block(item, plants, extra_plants=None, extra_label=None):
        block = []
        for p in plants:
            block.append(_prod_row(
                f"-{p}",
                _p_one(cur, "plan", p, item, month), _p_one(cur, "act", p, item, month),
                _p_one(cur, "act", p, item, cply_month),
                _p_cum_one(cur, "plan", p, item, ytd_months), _p_cum_one(cur, "act", p, item, ytd_months),
                _p_cum_one(cur, "act", p, item, cply_ytd_months)))
        block.append(_prod_row(
            "Total: 5 PLANTS",
            _p_sum(cur, "plan", plants, item, month), _p_sum(cur, "act", plants, item, month),
            _p_sum(cur, "act", plants, item, cply_month),
            _p_cum_sum(cur, "plan", plants, item, ytd_months), _p_cum_sum(cur, "act", plants, item, ytd_months),
            _p_cum_sum(cur, "act", plants, item, cply_ytd_months), bold=True))
        sail_set = list(plants) + list(extra_plants or [])
        if extra_plants and extra_label:
            block.append(_prod_row(
                extra_label,
                _p_sum(cur, "plan", extra_plants, item, month), _p_sum(cur, "act", extra_plants, item, month),
                _p_sum(cur, "act", extra_plants, item, cply_month),
                _p_cum_sum(cur, "plan", extra_plants, item, ytd_months), _p_cum_sum(cur, "act", extra_plants, item, ytd_months),
                _p_cum_sum(cur, "act", extra_plants, item, cply_ytd_months)))
        block.append(_prod_row(
            "SAIL (Total)",
            _p_sum(cur, "plan", sail_set, item, month), _p_sum(cur, "act", sail_set, item, month),
            _p_sum(cur, "act", sail_set, item, cply_month),
            _p_cum_sum(cur, "plan", sail_set, item, ytd_months), _p_cum_sum(cur, "act", sail_set, item, ytd_months),
            _p_cum_sum(cur, "act", sail_set, item, cply_ytd_months), bold=True))
        return block

    rows.append({"label": "HOT METAL", "header": True})
    rows += plant_block("Hot Metal", _5P, extra_plants=["VISL"])

    rows.append(_prod_row(
        "CRUDE STEEL",
        _p_sum(cur, "plan", _5P + ["ASP", "SSP", "VISL"], "Total Crude Steel", month),
        _p_sum(cur, "act", _5P + ["ASP", "SSP", "VISL"], "Total Crude Steel", month),
        _p_sum(cur, "act", _5P + ["ASP", "SSP", "VISL"], "Total Crude Steel", cply_month),
        _p_cum_sum(cur, "plan", _5P + ["ASP", "SSP", "VISL"], "Total Crude Steel", ytd_months),
        _p_cum_sum(cur, "act", _5P + ["ASP", "SSP", "VISL"], "Total Crude Steel", ytd_months),
        _p_cum_sum(cur, "act", _5P + ["ASP", "SSP", "VISL"], "Total Crude Steel", cply_ytd_months),
        indent=0, bold=True))

    rows.append({"label": "SALEABLE STEEL", "header": True})
    for p in _5P:
        rows.append(_prod_row(
            f"-{p}",
            _p_one(cur, "plan", p, "Saleable Steel", month), _p_one(cur, "act", p, "Saleable Steel", month),
            _p_one(cur, "act", p, "Saleable Steel", cply_month),
            _p_cum_one(cur, "plan", p, "Saleable Steel", ytd_months), _p_cum_one(cur, "act", p, "Saleable Steel", ytd_months),
            _p_cum_one(cur, "act", p, "Saleable Steel", cply_ytd_months)))
    rows.append(_prod_row(
        "Total: 5 PLANTS",
        _p_sum(cur, "plan", _5P, "Saleable Steel", month), _p_sum(cur, "act", _5P, "Saleable Steel", month),
        _p_sum(cur, "act", _5P, "Saleable Steel", cply_month),
        _p_cum_sum(cur, "plan", _5P, "Saleable Steel", ytd_months), _p_cum_sum(cur, "act", _5P, "Saleable Steel", ytd_months),
        _p_cum_sum(cur, "act", _5P, "Saleable Steel", cply_ytd_months), bold=True))
    ssp_plants = ["ASP", "SSP", "VISL"]
    rows.append(_prod_row(
        "SPECIAL STEEL PLANTS",
        _p_sum(cur, "plan", ssp_plants, "Saleable Steel", month), _p_sum(cur, "act", ssp_plants, "Saleable Steel", month),
        _p_sum(cur, "act", ssp_plants, "Saleable Steel", cply_month),
        _p_cum_sum(cur, "plan", ssp_plants, "Saleable Steel", ytd_months), _p_cum_sum(cur, "act", ssp_plants, "Saleable Steel", ytd_months),
        _p_cum_sum(cur, "act", ssp_plants, "Saleable Steel", cply_ytd_months)))
    sail_set = _5P + ssp_plants
    rows.append(_prod_row(
        "SAIL (Total)",
        _p_sum(cur, "plan", sail_set, "Saleable Steel", month), _p_sum(cur, "act", sail_set, "Saleable Steel", month),
        _p_sum(cur, "act", sail_set, "Saleable Steel", cply_month),
        _p_cum_sum(cur, "plan", sail_set, "Saleable Steel", ytd_months), _p_cum_sum(cur, "act", sail_set, "Saleable Steel", ytd_months),
        _p_cum_sum(cur, "act", sail_set, "Saleable Steel", cply_ytd_months), bold=True))

    return rows


# ── Table C: Techno-Economic Parameters ─────────────────────────────────

_TECHNO_WANTED = ["Coke Rate", "Specific Energy Consumption", "BF Productivity"]
_TECHNO_DISPLAY = {
    "Coke Rate": "COKE RATE (Kg/THM)",
    "Specific Energy Consumption": "ENERGY CONSUMPTION (G Cal/TCS)",
    "BF Productivity": "BF PRODUCTIVITY (T/CuM/Day)",
}
_TECHNO_FMT = {
    "Coke Rate": "0",
    "Specific Energy Consumption": "0.00",
    "BF Productivity": "0.00",
}
# Coke Rate / Energy Consumption are consumption-rate params — lower is
# better. BF Productivity is a throughput param — higher is better. The
# "% imp." sign must flip accordingly, or a real productivity gain would
# print as a negative "improvement".
_TECHNO_HIGHER_IS_BETTER = {
    "Coke Rate": False,
    "Specific Energy Consumption": False,
    "BF Productivity": True,
}


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _build_techno_rows(month):
    te = generate_summary_te_table(month)
    by_param = {r["parameter"]: r for r in te}
    rows = []
    for param in _TECHNO_WANTED:
        r = by_param.get(param, {})
        vals = r.get("values", [])
        abp = _num(vals[0]) if len(vals) > 0 else None
        m_act = _num(vals[1]) if len(vals) > 1 else None
        m_cply = _num(vals[2]) if len(vals) > 2 else None
        c_act = _num(vals[3]) if len(vals) > 3 else None
        c_cply = _num(vals[4]) if len(vals) > 4 else None
        # "% imp." sign depends on whether the param is lower-is-better
        # (Coke Rate, Energy Consumption) or higher-is-better (BF
        # Productivity) — see _TECHNO_HIGHER_IS_BETTER. Values are fractions
        # (0.03, not 3) — cells use Excel's '0%' format, same convention as
        # _gr/_pct_ful.
        sign = 1 if _TECHNO_HIGHER_IS_BETTER[param] else -1
        m_imp = sign * (m_act - m_cply) / m_cply if m_act is not None and m_cply else None
        c_imp = sign * (c_act - c_cply) / c_cply if c_act is not None and c_cply else None
        rows.append({
            "label": _TECHNO_DISPLAY[param], "fmt": _TECHNO_FMT[param], "abp": abp,
            "m_act": m_act, "m_cply": m_cply, "m_imp": m_imp,
            "c_act": c_act, "c_cply": c_cply, "c_imp": c_imp,
        })
    return rows


# ── Table D: Stock - 8 Plants ────────────────────────────────────────────

_STOCK_ITEMS = ["Plants", "Stockyards", "Stock In Transit", "Total"]


def _stock_reference_dates(as_on: date):
    """Column dates to show: April 1st of the last 5 FYs, the H1 midpoint
    (Sep 1) of the FY immediately before the current one, this FY's start,
    previous month, and the current 'as on' date — matches the sample
    file's own column selection (1.4.21..1.4.25, 1.9.25, 1.4.26, 1.6.26,
    1.7.26 for a Jun'26 report). Adjust here if a different history window
    is wanted in future."""
    fy = as_on.year if as_on.month >= 4 else as_on.year - 1
    dates = [date(fy - i, 4, 1) for i in range(5, 0, -1)]
    dates.append(date(fy - 1, 9, 1))
    dates.append(date(fy, 4, 1))
    prev_month_end = as_on.replace(day=1)
    py, pm = (prev_month_end.year, prev_month_end.month - 1) if prev_month_end.month > 1 else (prev_month_end.year - 1, 12)
    dates.append(date(py, pm, 1))
    dates.append(as_on)
    # de-dup while preserving order (as_on may coincide with an FY-start etc.)
    seen, out = set(), []
    for d in dates:
        if d not in seen:
            seen.add(d)
            out.append(d)
    return sorted(out)


def _stock_value(cur, item, snap_date):
    cur.execute("SELECT value FROM sail_stock_snapshot_table WHERE snapshot_date=? AND item_name=?",
                (snap_date.isoformat(), item))
    r = cur.fetchone()
    return r[0] if r and r[0] is not None else None


def _build_stock_table(cur, month):
    y, m = int(month[:4]), int(month[5:7])
    as_on = date(y + (1 if m == 12 else 0), (m % 12) + 1, 1)
    dates = _stock_reference_dates(as_on)
    fy_start_date = date(as_on.year if as_on.month >= 4 else as_on.year - 1, 4, 1)

    rows = []
    for item in _STOCK_ITEMS:
        vals = [_stock_value(cur, item, d) for d in dates]
        fy_start_val = _stock_value(cur, item, fy_start_date)
        var = round(vals[-1] - fy_start_val, 3) if vals and vals[-1] is not None and fy_start_val is not None else None
        rows.append({"label": item, "values": vals, "var": var, "bold": item == "Total"})
    return {"dates": dates, "fy_start_date": fy_start_date, "rows": rows}


# ── Excel writer ─────────────────────────────────────────────────────────
#
# Styling mirrors Report_format/1 page report for <Mon><YY>.xlsx cell-for-
# cell (column layout, row layout, fonts, fill colors, borders, number
# formats) so the generated workbook is visually indistinguishable from the
# department's own template:
#   col A = anchor/margin, col B = item label, col C = spacer,
#   cols D-M = the 10 ABP/Actual/%Ful/CPLY/%Gr (month block + cum block)
#   data columns for Tables A & B; Table C uses D-J (7 cols), Table D uses
#   D onward (one per snapshot date) + a trailing "Var. w.r.t." column.
# Row numbers match the template's own row numbers (title at row 2, Table A
# at rows 4-15, Table B at rows 17-36, Table C at rows 40-43, Table D at
# rows 45-49, closing note at row 51) so a reader familiar with the
# original immediately recognizes the layout.
#
# Fill colors are the template's own theme colors (Book Antiqua/Office
# 2007-2010 theme: dk2=1F497D for header bands, accent3=9BBB59 for
# grand-total highlight rows), not flat hex guesses, so they render as the
# exact same shade Excel would compute.

_FONT_TITLE = Font(name="Arial", size=24, bold=True)
_FONT_ASOF = Font(name="Arial", size=12, bold=True)
_FONT_UNIT = Font(name="Arial", size=14, italic=True)
_FONT_HDR = Font(name="Book Antiqua", size=14, bold=True)        # group/anchor/column headers, item labels
_FONT_DATA = Font(name="Book Antiqua", size=16, bold=True)       # data values
_FONT_NOTE = Font(name="Times New Roman", size=16)               # "*...incl NSL..." remark
_FONT_FINAL_NOTE = Font(name="Book Antiqua", size=14, bold=True)  # "Note: All figures are provisional"

_HDR_FILL = PatternFill("solid", fgColor=Color(theme=3, tint=0.7999816888943144))
_TOTAL_LABEL_FILL = PatternFill("solid", fgColor=Color(theme=6, tint=0.5999938962981048))
_TOTAL_DATA_FILL = PatternFill("solid", fgColor=Color(theme=6, tint=0.7999816888943144))

_THIN = Side(style="thin", color="000000")
_MEDIUM = Side(style="medium", color="000000")
_BOX_THIN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FMT_INT = "0"
_FMT_PCT = "0%"

# Data column order shared by Table A (Sales) and Table B (Production) —
# month block then cum block, cols D-M (4-13).
_METRIC_COLS = [
    ("abp", _FMT_INT), ("act", _FMT_INT), ("ful", _FMT_PCT),
    ("cply", _FMT_INT), ("growth", _FMT_PCT),
]


def _hdr_cell(ws, row, col, text, merge_to=None, wrap=False):
    c = ws.cell(row=row, column=col, value=text)
    c.font = _FONT_HDR
    c.fill = _HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    c.border = _BOX_THIN
    if merge_to:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
    return c


def _anchor_cell(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = _FONT_HDR
    c.fill = _HDR_FILL
    c.border = _BOX_THIN
    return c


def _item_label(ws, row, text, indent=0, total=False, merge_to=None):
    c = ws.cell(row=row, column=2, value=text)
    c.font = _FONT_HDR
    c.alignment = Alignment(horizontal="left", indent=indent, wrap_text=bool(merge_to))
    if total:
        c.fill = _TOTAL_LABEL_FILL
    if merge_to:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=merge_to)
    return c


def _data_cell(ws, row, col, value, fmt=_FMT_INT, total=False):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
        c.number_format = fmt
    c.font = _FONT_DATA
    c.alignment = Alignment(horizontal="center")
    c.border = _BOX_THIN
    if total:
        c.fill = _TOTAL_DATA_FILL
    return c


def _write_metric_block(ws, row, m_abp, m_act, m_ful, m_cply, m_growth,
                         c_abp, c_act, c_ful, c_cply, c_growth, total=False):
    values = [m_abp, m_act, m_ful, m_cply, m_growth, c_abp, c_act, c_ful, c_cply, c_growth]
    for i, (v, (_, fmt)) in enumerate(zip(values, _METRIC_COLS * 2)):
        _data_cell(ws, row, 4 + i, v, fmt=fmt, total=total)


def _group_header_row(ws, row, month_label, cply_label, cum_label, cply_cum_label):
    """The 2-row header above Tables A & B: this row groups D:F under the
    report month, I:K under the cumulative period, with the CPLY/growth
    columns (G, H, L, M) standing alone alongside each group."""
    _hdr_cell(ws, row, 4, month_label, merge_to=6)
    _hdr_cell(ws, row, 7, cply_label)
    _hdr_cell(ws, row, 8, "Growth over")
    _hdr_cell(ws, row, 9, cum_label, merge_to=11)
    _hdr_cell(ws, row, 12, cply_cum_label)
    _hdr_cell(ws, row, 13, "Growth over")


def _column_header_row(ws, row, anchor_text):
    _anchor_cell(ws, row, anchor_text)
    for col, text in zip(range(4, 14),
                          ["ABP", "Actual", "% Ful", "Actual", "% Gr.",
                           "ABP", "Actual", "% Ful", "Actual", "% Gr."]):
        _hdr_cell(ws, row, col, text)


_COL_WIDTHS = {
    "A": 6.11, "B": 38.44, "C": 33.0, "D": 10.44, "E": 12.44, "F": 13.0,
    "G": 12.44, "H": 17.11, "I": 19.0, "J": 12.33, "K": 14.44, "L": 19.0, "M": 22.0,
}


def build_one_page_workbook(report_month: str):
    cply_month = db.get_cply_month(report_month)
    ytd_months = db.get_ytd_months(report_month)
    cply_ytd_months = [db.get_cply_month(m) for m in ytd_months]

    conn = db.connect()
    cur = conn.cursor()
    try:
        sales_rows = _build_sales_rows(cur, report_month)
        sales_note = _sales_note(cur, report_month)
        prod_rows = _build_production_rows(cur, report_month, cply_month, ytd_months, cply_ytd_months)
        techno_rows = _build_techno_rows(report_month)
        stock = _build_stock_table(cur, report_month)
    finally:
        conn.close()

    y, m = int(report_month[:4]), int(report_month[5:7])
    as_on = date(y + (1 if m == 12 else 0), (m % 12) + 1, 1)

    cply_label = _mlabel(cply_month)
    cply_cum_label = _cum_label(cply_month)
    cum_label = _cum_label(report_month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{_mlabel(report_month)} PS CMO"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    for letter, width in _COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    # Row 1: "#As on <date>"; Row 2: title (merged A:H); Row 3: unit
    ws.cell(row=1, column=12, value="#As on ").font = _FONT_ASOF
    ws.cell(row=1, column=12).alignment = Alignment(horizontal="right")
    c = ws.cell(row=1, column=13, value=as_on)
    c.font = _FONT_ASOF
    c.number_format = "d-mmm-yy"
    ws.merge_cells("A2:H2")
    c = ws.cell(row=2, column=1,
                value=f"PERFORMANCE: {_mlabel(report_month)} and {_cum_label(report_month)}")
    c.font = _FONT_TITLE
    c.alignment = Alignment(horizontal="center")
    ws.cell(row=3, column=13, value="Unit:'000 T").font = _FONT_UNIT
    ws.row_dimensions[2].height = 37.5
    ws.row_dimensions[3].height = 19.5

    row = 4
    _group_header_row(ws, row, _mlabel(report_month), cply_label, cum_label, cply_cum_label)
    row += 1
    _column_header_row(ws, row, "A.    SALES ['000 T]")
    row += 1

    for r in sales_rows:
        # Every value here — including %Ful/Growth — is written back exactly
        # as the source department reported it; nothing is computed (see
        # sail_sales_stock_extractor.py). %Ful/Growth are stored as raw
        # fractions (0.81, not 81), so they're formatted as Excel percentages
        # rather than rescaled.
        _item_label(ws, row, r["label"], total=r["bold"])
        _write_metric_block(ws, row, r["m_abp"], r["m_act"], r["m_ful"], r["m_cply"], r["m_growth"],
                             r["c_abp"], r["c_act"], r["c_ful"], r["c_cply"], r["c_growth"], total=r["bold"])
        row += 1

    if sales_note:
        c = ws.cell(row=row, column=3, value=sales_note)
        c.font = _FONT_NOTE
        c.alignment = Alignment(horizontal="left", vertical="top")
        row += 1

    row += 1  # blank spacer (template row 16)
    _group_header_row(ws, row, _mlabel(report_month), cply_label, cum_label, cply_cum_label)
    row += 1
    _column_header_row(ws, row, "B.    PRODUCTION  ['000 T]")
    row += 1

    for r in prod_rows:
        if r.get("header"):
            c = ws.cell(row=row, column=2, value=r["label"])
            c.font = _FONT_HDR
            row += 1
            continue
        total = r["label"] == "SAIL (Total)"
        _item_label(ws, row, r["label"], indent=r.get("indent", 1), total=total)
        m_ful = _pct_ful(r["m_act"], r["m_abp"])
        m_growth = _gr(r["m_act"], r["m_cply"])
        c_ful = _pct_ful(r["c_act"], r["c_abp"])
        c_growth = _gr(r["c_act"], r["c_cply"])
        _write_metric_block(ws, row, r["m_abp"], r["m_act"], m_ful, r["m_cply"], m_growth,
                             r["c_abp"], r["c_act"], c_ful, r["c_cply"], c_growth, total=total)
        row += 1

    row += 3  # 3 blank spacer rows (template rows 37-39)
    _anchor_cell(ws, row, "C.    TECHNO-ECONOMIC PARAMETERS")
    for col, text in zip(range(4, 11), ["ABP", _mlabel(report_month), cply_label, "% imp.",
                                         cum_label, cply_cum_label, "% imp."]):
        _hdr_cell(ws, row, col, text)
    row += 1
    for r in techno_rows:
        _item_label(ws, row, r["label"], merge_to=3)
        _data_cell(ws, row, 4, r["abp"], fmt=r["fmt"])
        _data_cell(ws, row, 5, r["m_act"], fmt=r["fmt"])
        _data_cell(ws, row, 6, r["m_cply"], fmt=r["fmt"])
        _data_cell(ws, row, 7, r["m_imp"], fmt=_FMT_PCT)
        _data_cell(ws, row, 8, r["c_act"], fmt=r["fmt"])
        _data_cell(ws, row, 9, r["c_cply"], fmt=r["fmt"])
        _data_cell(ws, row, 10, r["c_imp"], fmt=_FMT_PCT)
        row += 1

    row += 1  # blank spacer (template row 44)
    n_date_cols = len(stock["dates"])
    var_col = 4 + n_date_cols
    _anchor_cell(ws, row, "D.    STOCK  - 8 PLANTS ['000 T]")
    for i, d in enumerate(stock["dates"]):
        _hdr_cell(ws, row, 4 + i, d.strftime("%d.%m.%y"))
    fy_d = stock["fy_start_date"]
    _hdr_cell(ws, row, var_col, f"Var. w.r.t. {fy_d.day}.{fy_d.month}.{fy_d.strftime('%y')}", wrap=True)
    row += 1
    for r in stock["rows"]:
        _item_label(ws, row, r["label"], total=False)
        for i, v in enumerate(r["values"]):
            _data_cell(ws, row, 4 + i, v, fmt=_FMT_INT)
        _data_cell(ws, row, var_col, r["var"], fmt=_FMT_INT)
        row += 1

    row += 1  # blank spacer (template row 50)
    ws.cell(row=row, column=2, value="Note: All figures are provisional").font = _FONT_FINAL_NOTE

    for r in range(4, row + 1):
        ws.row_dimensions[r].height = 21

    ws.freeze_panes = "D6"
    ws.print_area = f"A1:{get_column_letter(var_col)}{row}"

    return wb


def build_one_page_report_bytes(report_month: str) -> bytes:
    import io
    wb = build_one_page_workbook(report_month)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
