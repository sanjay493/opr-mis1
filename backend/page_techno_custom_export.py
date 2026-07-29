"""
Excel / PDF export for the Techno Custom Report (frontend: /reports/techno-custom,
under External Reports) — both its modes:

  Standard mode      — page-27-style table (FY-3/FY-2/FY-1/Target/YTD
                        months/CPLY/Cum/Cum-CPLY), filtered to user-picked
                        plants/params. Source: page_techno.generate_major_techno_from_db.
  Custom Period mode — one column per user-defined period (quarter/half/
                        custom range), computed by techno_period.build_period_report.

Visual style mirrors page_production_query_export.py (blue header, light-blue
subheader, zebra rows, `#,##0.000`-style number formatting) for consistency
with the rest of the app; `render_pdf_bytes` is imported from that module
directly rather than duplicated, since it has zero business logic (pure
HTML-string-in, PDF-bytes-out).
"""
import io

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import page_production_query_export as _ppqe

render_pdf_bytes = _ppqe.render_pdf_bytes

_HDR_FILL = PatternFill("solid", fgColor="1A73E8")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_SUBHDR_FILL = PatternFill("solid", fgColor="E8F0FE")
_SUBHDR_FONT = Font(bold=True, color="174EA6", size=9)
_SECTION_FILL = PatternFill("solid", fgColor="1A73E8")
_SECTION_FONT = Font(bold=True, color="FFFFFF", size=10)
_SAIL_FILL = PatternFill("solid", fgColor="F9AB00")
_SAIL_FONT = Font(bold=True, color="3C2F00", size=9)
_ZEBRA_FILL = PatternFill("solid", fgColor="F8F9FA")
_THIN = Side(style="thin", color="DADCE0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _cell_str(v):
    """Standard-mode row values are already pre-formatted strings (from
    page_techno._fmt_param) — pass through, blank for None/empty."""
    if v is None or v == "":
        return ""
    return str(v)


# ── Standard mode ───────────────────────────────────────────────────────────

def filter_major_techno(data: dict, plants=None, params=None) -> dict:
    """Filter generate_major_techno_from_db's raw section/row shape to the
    user-picked plants (plant codes and/or "SAIL") and parameter labels.
    None/falsy = no filtering on that axis."""
    plant_set = set(plants) if plants else None
    param_set = set(params) if params else None
    sections = []
    for sec in data.get("sections", []):
        if param_set is not None and sec.get("label") not in param_set:
            continue
        rows = sec.get("rows", [])
        if plant_set is not None:
            rows = [r for r in rows if r.get("label") in plant_set]
        if not rows:
            continue
        sections.append({**sec, "rows": rows})
    return {**data, "sections": sections}


def build_standard_excel_bytes(data: dict, month: str) -> bytes:
    month_labels = data.get("month_labels") or []
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Techno Custom Report"[:31]

    ws.cell(row=1, column=1, value="Techno Custom Report — Standard (Page 27 style)").font = Font(bold=True, size=13)
    ws.cell(row=2, column=1, value=f"Report month: {month}").font = Font(italic=True, size=9)

    headers = (["Plant", "Unit", f"FY {data.get('fy3_label','')}", f"FY {data.get('fy2_label','')}",
                f"FY {data.get('fy1_label','')}", data.get("target_label", "Target")]
               + month_labels
               + [data.get("cply_label", "CPLY"), data.get("cum_label", "Cum"),
                  data.get("cum_cply_label", "Cum-CPLY")])
    total_cols = len(headers)

    row = 4
    for sec in data.get("sections", []):
        sc = ws.cell(row=row, column=1, value=sec.get("label", ""))
        sc.font = _SECTION_FONT
        sc.fill = _SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        for c in range(2, total_cols + 1):
            ws.cell(row=row, column=c).fill = _SECTION_FILL
        row += 1

        for c, h in enumerate(headers, start=1):
            hc = ws.cell(row=row, column=c, value=h)
            hc.font = _SUBHDR_FONT
            hc.fill = _SUBHDR_FILL
            hc.alignment = Alignment(horizontal="center")
            hc.border = _BORDER
        row += 1

        for idx, r in enumerate(sec.get("rows", [])):
            is_sail = r.get("label") == "SAIL"
            fill = _SAIL_FILL if is_sail else (_ZEBRA_FILL if idx % 2 == 1 else None)
            font = _SAIL_FONT if is_sail else Font(size=9)
            values = ([r.get("label", ""), r.get("unit", ""), r.get("fy3", ""), r.get("fy2", ""),
                       r.get("fy1", ""), r.get("target", "")]
                      + (r.get("months") or [])
                      + [r.get("cply", ""), r.get("cum", ""), r.get("cum_cply", "")])
            for c, v in enumerate(values, start=1):
                vc = ws.cell(row=row, column=c, value=_cell_str(v))
                vc.font = font
                vc.border = _BORDER
                vc.alignment = Alignment(horizontal="left" if c == 1 else "right")
                if fill:
                    vc.fill = fill
            row += 1
        row += 1  # blank spacer row between sections

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    for i in range(3, total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 9
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_standard_pdf_html(data: dict, month: str) -> str:
    month_labels = data.get("month_labels") or []
    headers = (["Plant", "Unit", f"FY {data.get('fy3_label','')}", f"FY {data.get('fy2_label','')}",
                f"FY {data.get('fy1_label','')}", data.get("target_label", "Target")]
               + month_labels
               + [data.get("cply_label", "CPLY"), data.get("cum_label", "Cum"),
                  data.get("cum_cply_label", "Cum-CPLY")])
    header_html = "".join(f"<th>{h}</th>" for h in headers)

    sections_html = []
    for sec in data.get("sections", []):
        body_rows = []
        for r in sec.get("rows", []):
            is_sail = r.get("label") == "SAIL"
            values = ([r.get("label", ""), r.get("unit", ""), r.get("fy3", ""), r.get("fy2", ""),
                       r.get("fy1", ""), r.get("target", "")]
                      + (r.get("months") or [])
                      + [r.get("cply", ""), r.get("cum", ""), r.get("cum_cply", "")])
            cells = "".join(
                f'<td class="{"plant" if i == 0 else ""}">{_cell_str(v)}</td>'
                for i, v in enumerate(values))
            body_rows.append(f'<tr class="{"sail-row" if is_sail else ""}">{cells}</tr>')
        sections_html.append(
            f'<div class="section-title">{sec.get("label","")}</div>'
            f'<table><thead><tr>{header_html}</tr></thead>'
            f'<tbody>{"".join(body_rows)}</tbody></table>'
        )

    return f"""<!doctype html>
<html><head><meta charset="utf-8">
<style>
  @page {{ size: A4 landscape; margin: 10mm 8mm; }}
  * {{ box-sizing: border-box; }}
  body {{ font-family: Arial, sans-serif; color: #202124; margin: 0; }}
  h1 {{ font-size: 14pt; margin: 0 0 2px 0; }}
  .subtitle {{ font-size: 9pt; color: #5f6368; margin: 0 0 8px 0; }}
  .section-title {{ background: #1a73e8; color: #fff; font-weight: 700; font-size: 9pt;
    padding: 4px 6px; margin-top: 10px; page-break-after: avoid; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 6.6pt; margin-bottom: 4px; }}
  thead {{ display: table-header-group; }}
  tr {{ page-break-inside: avoid; }}
  th, td {{ border: 1px solid #dadce0; padding: 2px 3px; text-align: right; white-space: nowrap; }}
  th {{ background: #e8f0fe; color: #174ea6; font-weight: 700; }}
  td.plant {{ text-align: left; font-weight: 700; }}
  tr:nth-child(even) td {{ background: #f8f9fa; }}
  tr.sail-row td {{ background: #f9ab00; color: #3c2f00; font-weight: 700; }}
</style>
</head>
<body>
  <h1>Techno Custom Report — Standard (Page 27 style)</h1>
  <p class="subtitle">Report month: {month}</p>
  {''.join(sections_html)}
</body></html>"""


# ── Custom Period mode ──────────────────────────────────────────────────────

def build_period_excel_bytes(data: dict) -> bytes:
    periods = data.get("periods", [])
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Techno Custom Period"[:31]

    ws.cell(row=1, column=1, value="Techno Custom Report — Custom Period").font = Font(bold=True, size=13)
    ws.cell(row=2, column=1, value="* = production-weight data incomplete for this period; simple average shown").font = Font(italic=True, size=9)

    total_cols = 2 + len(periods)
    row = 4
    for sec in data.get("sections", []):
        sc = ws.cell(row=row, column=1, value=f'{sec.get("parameter","")} ({sec.get("unit","")})' if sec.get("unit") else sec.get("parameter", ""))
        sc.font = _SECTION_FONT
        sc.fill = _SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        for c in range(2, total_cols + 1):
            ws.cell(row=row, column=c).fill = _SECTION_FILL
        row += 1

        headers = ["Plant"] + periods
        for c, h in enumerate(headers, start=1):
            hc = ws.cell(row=row, column=c, value=h)
            hc.font = _SUBHDR_FONT
            hc.fill = _SUBHDR_FILL
            hc.alignment = Alignment(horizontal="center")
            hc.border = _BORDER
        row += 1

        for idx, r in enumerate(sec.get("rows", [])):
            is_sail = r.get("plant") == "SAIL"
            fill = _SAIL_FILL if is_sail else (_ZEBRA_FILL if idx % 2 == 1 else None)
            font = _SAIL_FONT if is_sail else Font(size=9)
            pc = ws.cell(row=row, column=1, value=r.get("plant", ""))
            pc.font = font
            pc.border = _BORDER
            if fill:
                pc.fill = fill
            for c, label in enumerate(periods, start=2):
                cell_data = (r.get("values") or {}).get(label) or {}
                disp = cell_data.get("display", "")
                if cell_data.get("method_used") in ("average",) and cell_data.get("warnings"):
                    disp = f"{disp}*" if disp else disp
                vc = ws.cell(row=row, column=c, value=disp)
                vc.font = font
                vc.border = _BORDER
                vc.alignment = Alignment(horizontal="right")
                if fill:
                    vc.fill = fill
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 12
    for i in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_period_pdf_html(data: dict) -> str:
    periods = data.get("periods", [])
    header_html = "<th>Plant</th>" + "".join(f"<th>{p}</th>" for p in periods)

    sections_html = []
    for sec in data.get("sections", []):
        body_rows = []
        for r in sec.get("rows", []):
            is_sail = r.get("plant") == "SAIL"
            cells = [f'<td class="plant">{r.get("plant","")}</td>']
            for label in periods:
                cell_data = (r.get("values") or {}).get(label) or {}
                disp = cell_data.get("display", "") or "—"
                if cell_data.get("method_used") == "average" and cell_data.get("warnings"):
                    disp = f"{disp}*"
                cells.append(f"<td>{disp}</td>")
            body_rows.append(f'<tr class="{"sail-row" if is_sail else ""}">{"".join(cells)}</tr>')
        title = f'{sec.get("parameter","")} ({sec.get("unit","")})' if sec.get("unit") else sec.get("parameter", "")
        sections_html.append(
            f'<div class="section-title">{title}</div>'
            f'<table><thead><tr>{header_html}</tr></thead>'
            f'<tbody>{"".join(body_rows)}</tbody></table>'
        )

    return f"""<!doctype html>
<html><head><meta charset="utf-8">
<style>
  @page {{ size: A4 landscape; margin: 12mm 10mm; }}
  * {{ box-sizing: border-box; }}
  body {{ font-family: Arial, sans-serif; color: #202124; margin: 0; }}
  h1 {{ font-size: 14pt; margin: 0 0 2px 0; }}
  .subtitle {{ font-size: 9pt; color: #5f6368; margin: 0 0 8px 0; }}
  .section-title {{ background: #1a73e8; color: #fff; font-weight: 700; font-size: 9.5pt;
    padding: 4px 6px; margin-top: 10px; page-break-after: avoid; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 8pt; margin-bottom: 4px; }}
  thead {{ display: table-header-group; }}
  tr {{ page-break-inside: avoid; }}
  th, td {{ border: 1px solid #dadce0; padding: 3px 6px; text-align: right; white-space: nowrap; }}
  th {{ background: #e8f0fe; color: #174ea6; font-weight: 700; }}
  td.plant {{ text-align: left; font-weight: 700; }}
  tr:nth-child(even) td {{ background: #f8f9fa; }}
  tr.sail-row td {{ background: #f9ab00; color: #3c2f00; font-weight: 700; }}
</style>
</head>
<body>
  <h1>Techno Custom Report — Custom Period</h1>
  <p class="subtitle">* = production-weight data incomplete for this period; simple average shown</p>
  {''.join(sections_html)}
</body></html>"""
