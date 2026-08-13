"""
Excel / PDF export for the Large BF Benchmarking comparison page
(frontend: /reports/bf-benchmark). Takes the same dict api_bf_benchmark.
build_compare() returns — {params, sail_bfs, periods} — so the export
always matches what's on screen.

Layout: one row per parameter (2 leading columns: Parameter, Unit; Working
Volume — a per-BF constant — is just the first such row, per
BF_BENCHMARK_PARAMS' own ordering), then one column per furnace, grouped
under a 4-level header: Period (FY or Month) -> Company (SAIL / JSW / TATA /
...) -> Location (BSP / RSP / Vijaynagar / ...) -> Furnace. Each `period`
in the input already carries every furnace's row (SAIL and non-SAIL
together for FY periods; SAIL-only for Month periods, since non-SAIL BFs
only ever publish FY-level figures) — this module just groups those rows by
company/location for the header and doesn't otherwise touch the numbers.

render_pdf_bytes is imported from page_production_query_export.py directly
rather than duplicated, since it has zero business logic (pure
HTML-string-in, PDF-bytes-out).
"""
import io

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import page_production_query_export as _ppqe

render_pdf_bytes = _ppqe.render_pdf_bytes

_HDR_FILL = PatternFill("solid", fgColor="1A73E8")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=9)
_SUBHDR_FILL = PatternFill("solid", fgColor="E8F0FE")
_SUBHDR_FONT = Font(bold=True, color="174EA6", size=8)
_EXT_SUBHDR_FILL = PatternFill("solid", fgColor="FCE8E6")
_EXT_SUBHDR_FONT = Font(bold=True, color="A50E0E", size=8)
_EXT_FILL = PatternFill("solid", fgColor="FEF7F6")
_ZEBRA_FILL = PatternFill("solid", fgColor="F8F9FA")
_THIN = Side(style="thin", color="DADCE0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# BF Productivity and O2 Enrichment read to 2 decimal places; every other
# param (Working Volume, rates, HBT, etc.) is a whole-number figure — same
# rule the frontend's fmt() applies (see reports/bf-benchmark/page.js).
_TWO_DECIMAL_KEYS = {"bf_productivity", "o2_enrichment"}


def _fmt(v, key=None):
    if v is None:
        return ""
    if isinstance(v, float):
        if key in _TWO_DECIMAL_KEYS:
            return round(v, 2)
        return int(round(v))  # plain int, not e.g. 397.0, for Excel/HTML display
    return v


def _group_rows(rows):
    """rows -> [(company, [(location, [row, ...]), ...]), ...], preserving
    each company's/location's first-seen order (not sorted alphabetically —
    SAIL's 3 plants and a non-SAIL company's furnaces should stay in the
    order the request listed them)."""
    companies = []
    by_company = {}
    for r in rows:
        c = r["company"]
        if c not in by_company:
            by_company[c] = []
            companies.append(c)
        by_company[c].append(r)
    grouped = []
    for c in companies:
        locations = []
        by_loc = {}
        for r in by_company[c]:
            loc = r["location"]
            if loc not in by_loc:
                by_loc[loc] = []
                locations.append(loc)
            by_loc[loc].append(r)
        grouped.append((c, [(loc, by_loc[loc]) for loc in locations]))
    return grouped


def _flat_rows(period):
    """period['rows'] flattened back out in the same Company -> Location
    order _group_rows produces, so header and body columns line up."""
    out = []
    for _company, locations in _group_rows(period["rows"]):
        for _location, rs in locations:
            out.extend(rs)
    return out


def build_excel_bytes(data: dict) -> bytes:
    params = data.get("params", [])
    periods = data.get("periods", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BF Benchmarking"[:31]

    ws.cell(row=1, column=1, value="Large BF Benchmarking — Comparative Performance Analysis").font = Font(bold=True, size=13)

    r1, r2, r3, r4 = 3, 4, 5, 6  # Period | Company | Location | Furnace
    ws.cell(row=r1, column=1, value="Techno Parameter").font = _HDR_FONT
    ws.cell(row=r1, column=1).fill = _HDR_FILL
    ws.merge_cells(start_row=r1, start_column=1, end_row=r4, end_column=1)
    ws.cell(row=r1, column=2, value="Unit").font = _HDR_FONT
    ws.cell(row=r1, column=2).fill = _HDR_FILL
    ws.merge_cells(start_row=r1, start_column=2, end_row=r4, end_column=2)

    col = 3
    period_rows = {}  # period key -> flattened furnace rows, for the body loop
    for period in periods:
        grouped = _group_rows(period["rows"])
        width = sum(len(rs) for _c, locs in grouped for _l, rs in locs)
        if width == 0:
            continue
        period_start = col
        for company, locations in grouped:
            comp_start = col
            for location, rs in locations:
                loc_start = col
                is_ext = rs[0]["is_external"]
                sub_fill = _EXT_SUBHDR_FILL if is_ext else _SUBHDR_FILL
                sub_font = _EXT_SUBHDR_FONT if is_ext else _SUBHDR_FONT
                for r in rs:
                    fc = ws.cell(row=r4, column=col, value=r["label"])
                    fc.font = sub_font; fc.fill = sub_fill; fc.alignment = Alignment(horizontal="center"); fc.border = _BORDER
                    col += 1
                lc = ws.cell(row=r3, column=loc_start, value=location)
                lc.font = sub_font; lc.fill = sub_fill; lc.alignment = Alignment(horizontal="center")
                if col - 1 > loc_start:
                    ws.merge_cells(start_row=r3, start_column=loc_start, end_row=r3, end_column=col - 1)
            cc = ws.cell(row=r2, column=comp_start, value=company)
            cc.font = _HDR_FONT if company == "SAIL" else _EXT_SUBHDR_FONT
            cc.fill = _HDR_FILL if company == "SAIL" else _EXT_SUBHDR_FILL
            cc.alignment = Alignment(horizontal="center")
            if col - 1 > comp_start:
                ws.merge_cells(start_row=r2, start_column=comp_start, end_row=r2, end_column=col - 1)
        pc = ws.cell(row=r1, column=period_start, value=period["label"])
        pc.font = _HDR_FONT; pc.fill = _HDR_FILL; pc.alignment = Alignment(horizontal="center")
        if col - 1 > period_start:
            ws.merge_cells(start_row=r1, start_column=period_start, end_row=r1, end_column=col - 1)
        for c in range(period_start, col):
            ws.cell(row=r1, column=c).fill = _HDR_FILL
        period_rows[period["key"]] = _flat_rows(period)

    total_cols = col - 1
    row = r4 + 1
    for p in params:
        key, label, unit = p["key"], p["label"], p.get("unit", "")
        lc = ws.cell(row=row, column=1, value=label); lc.border = _BORDER
        uc = ws.cell(row=row, column=2, value=unit); uc.border = _BORDER
        fill = _ZEBRA_FILL if row % 2 == 0 else None
        if fill:
            lc.fill = fill; uc.fill = fill

        c = 3
        for period in periods:
            for r in period_rows.get(period["key"], []):
                if not r["has_data"]:
                    vc = ws.cell(row=row, column=c, value="—")
                    vc.alignment = Alignment(horizontal="center")
                else:
                    vc = ws.cell(row=row, column=c, value=_fmt(r["values"].get(key), key))
                    vc.alignment = Alignment(horizontal="right")
                vc.border = _BORDER
                if r["is_external"]:
                    vc.fill = _EXT_FILL
                elif fill:
                    vc.fill = fill
                c += 1
        row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 10
    for i in range(3, total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 10
    ws.freeze_panes = f"C{r4 + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf_html(data: dict) -> str:
    params = data.get("params", [])
    periods = data.get("periods", [])

    row1, row2, row3, row4 = ['<th rowspan="4">Parameter</th><th rowspan="4">Unit</th>'], [], [], []
    period_rows = {}
    for period in periods:
        grouped = _group_rows(period["rows"])
        width = sum(len(rs) for _c, locs in grouped for _l, rs in locs)
        if width == 0:
            continue
        row1.append(f'<th colspan="{width}">{period["label"]}</th>')
        for company, locations in grouped:
            comp_width = sum(len(rs) for _l, rs in locations)
            comp_cls = "" if company == "SAIL" else ' class="ext-hdr"'
            row2.append(f'<th colspan="{comp_width}"{comp_cls}>{company}</th>')
            for location, rs in locations:
                loc_cls = "" if rs[0]["is_external"] is False else ' class="ext-hdr"'
                row3.append(f'<th colspan="{len(rs)}"{loc_cls}>{location}</th>')
                for r in rs:
                    row4.append(f'<th{loc_cls}>{r["label"]}</th>')
        period_rows[period["key"]] = _flat_rows(period)

    body_rows = []
    for p in params:
        key, label, unit = p["key"], p["label"], p.get("unit", "")
        cells = [f'<td class="param">{label}</td><td class="unit">{unit}</td>']
        for period in periods:
            for r in period_rows.get(period["key"], []):
                ext_cls = ' class="ext"' if r["is_external"] else ""
                if not r["has_data"]:
                    cells.append(f'<td{ext_cls}>—</td>')
                else:
                    v = _fmt(r["values"].get(key), key)
                    cells.append(f'<td{ext_cls}>{v if v != "" else "—"}</td>')
        body_rows.append(f"<tr>{''.join(cells)}</tr>")

    return f"""<!doctype html>
<html><head><meta charset="utf-8">
<style>
  @page {{ size: A4 landscape; margin: 10mm 8mm; }}
  * {{ box-sizing: border-box; }}
  body {{ font-family: Arial, sans-serif; color: #202124; margin: 0; }}
  h1 {{ font-size: 13pt; margin: 0 0 8px 0; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 7pt; }}
  thead {{ display: table-header-group; }}
  tr {{ page-break-inside: avoid; }}
  th, td {{ border: 1px solid #dadce0; padding: 2px 4px; text-align: right; white-space: nowrap; }}
  th {{ background: #e8f0fe; color: #174ea6; font-weight: 700; text-align: center; }}
  th.ext-hdr {{ background: #fce8e6; color: #a50e0e; }}
  td.param {{ text-align: left; font-weight: 700; }}
  td.unit {{ text-align: left; color: #5f6368; }}
  td.ext {{ background: #fef7f6; }}
  tr:nth-child(even) td {{ background: #f8f9fa; }}
  tr:nth-child(even) td.ext {{ background: #fdeceb; }}
</style>
</head>
<body>
  <h1>Large BF Benchmarking — Comparative Performance Analysis</h1>
  <table>
    <thead>
      <tr>{''.join(row1)}</tr>
      <tr>{''.join(row2)}</tr>
      <tr>{''.join(row3)}</tr>
      <tr>{''.join(row4)}</tr>
    </thead>
    <tbody>{''.join(body_rows)}</tbody>
  </table>
</body></html>"""
