"""
SAIL Pmix FY report — Report_format/pmix.xlsx format ("Pmix'26-27 ACT" sheet):
year-wise, month-wise Product Mix Performance for a selected financial year
(Apr-Mar), with 4 quarterly sum columns and a full-year cumulative column.

Reuses page_jpc_report.compute_pmix_rows() row-by-row (called once per FY
month) so this stays in sync with the JPC report's Pmix sheet — same plant
blocks, SAIL rollup, Finished Steel / ASP / SSP / VISL / Totals section,
same blanks where nothing is tracked in the DB.

Four extra rows beyond compute_pmix_rows(), specific to this template:
  Finished Steel (conversion)          — SAIL/"Conversion" (page4.py's
                                          existing SAIL-incl.-conversion item)
  Finished Steel including Conversion  — Finished Steel (Total) + Conversion
  RSP HSM-2                            — RSP/"HSM-2 Total HR Coil"
  BSL HSM                              — BSL/"HSM Total HR Coil"
(The sample file's own quarter/cum formulas for the last two rows are
broken — copy-paste leftovers referencing the wrong rows — so quarters/cum
here are computed properly as sums of that row's own months instead of
copying the same mistake.)
"""
import io

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import db
from constants import FIVE_PLANTS as _5P
from page_jpc_report import compute_pmix_rows, _one, _one_of, _sum, _fs, _mou_value

_MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
           "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

_QTR_LABELS = ["1st Qtr", "IInd Qtr", "IIIrd Qtr", "IVth Qtr"]


def _mlabel(ym: str) -> str:
    y, m = int(ym[:4]), int(ym[5:7])
    return f"{_MONTHS[m - 1]}'{str(y)[2:]}"


def _fy_months(fy_start: int):
    return [f"{fy_start}-{m:02d}" for m in range(4, 13)] + \
           [f"{fy_start + 1}-{m:02d}" for m in range(1, 4)]


def _sum_opt(vals):
    nums = [v for v in vals if v is not None]
    return round(sum(nums), 3) if nums else None


def _avg_opt(vals):
    nums = [v for v in vals if v is not None]
    return round(sum(nums) / len(nums), 3) if nums else None


_TITLE_FONT = Font(bold=True, size=13)
_UNIT_FONT = Font(italic=True, size=9)
_HDR_FONT = Font(bold=True, size=9)
_HDR_FILL = PatternFill("solid", fgColor="D9E6F5")
_PLANT_FONT = Font(bold=True, size=10)
_PLANT_FILL = PatternFill("solid", fgColor="F0F0F0")
_BOLD = Font(bold=True, size=9)
_THIN = Side(style="thin", color="B0B0B0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _num_cell(ws, row, col, value, bold=False):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
        c.number_format = "0.000"
    c.border = _BORDER
    c.alignment = Alignment(horizontal="right")
    if bold:
        c.font = _BOLD


def _label_cell(ws, row, col, text, bold=False):
    c = ws.cell(row=row, column=col, value=text)
    c.alignment = Alignment(horizontal="left")
    c.border = _BORDER
    if bold:
        c.font = _BOLD


def _plant_hdr(ws, row, label, n_cols):
    for col in range(1, n_cols + 1):
        ws.cell(row=row, column=col).fill = _PLANT_FILL
        ws.cell(row=row, column=col).border = _BORDER
    ws.cell(row=row, column=1, value=label).font = _PLANT_FONT


# ── Sheet 1: "ACT <FY>" — plant-wise monthwise production summary,
# matching Report_format/PMix26-27.xlsx's "ACT 26-27" sheet ────────────────

# (label, db_item, plant_rows) — plant_rows uses "5 Plants" as the
# aggregate-row key (resolved to "SAIL - 5PL" for display) so _mou_value()
# from page_jpc_report.py (already built for the MoU report sheet) can be
# reused as-is: same rollup rules (Hot Metal/Pig Iron pull in VISL only;
# Crude Steel/Saleable Steel/Finished Steel also pull in ASP/SSP/VISL).
_ACT_ROWS_5PL_ONLY = _5P + ["5 Plants"]
_ACT_ROWS_FULL = _5P + ["5 Plants", "ASP", "SSP", "VISL", "SAIL"]
_ACT_ROWS_HM_PI = _5P + ["5 Plants", "VISL", "SAIL"]

_ACT_ITEMS = [
    ("Oven Pushing\n(nos./day)", "Oven Pushing (nos/day)", _ACT_ROWS_5PL_ONLY, True),
    ("Sinter",                   "Total Sinter",            _ACT_ROWS_5PL_ONLY, False),
    ("Hot Metal",                "Hot Metal",               _ACT_ROWS_HM_PI,    False),
    ("Crude Steel",              "Total Crude Steel",       _ACT_ROWS_FULL,     False),
    ("Pig Iron",                 "Pig Iron",                _ACT_ROWS_HM_PI,    False),
    ("Saleable Steel",           "Saleable Steel",          _ACT_ROWS_FULL,     False),
    ("Saleable Production",      None,                      _ACT_ROWS_FULL,     False),   # = Saleable Steel + Pig Iron
    ("Finished Steel Production","Finished Steel",          _ACT_ROWS_FULL,     False),
]

_ACT_PLANT_LABEL = {"5 Plants": "SAIL - 5PL"}


def _act_plant_value(cur, label, db_item, plant, month):
    if label.startswith("Oven Pushing"):
        # Two raw spellings exist across plants/months ("Oven Pushing (nos/day)"
        # vs "Oven Pushing(nos/d)") — production_table isn't normalized at the
        # row level, only main.py's /api/production-fy query normalizes it.
        if plant == "5 Plants":
            # Summed across plants (a combined ovens/day figure for the same
            # month is meaningful), unlike the quarter/cum columns below
            # which average a single row's own rate across months.
            total, found = 0.0, False
            for p in _5P:
                v = _one_of(cur, p, ["Oven Pushing (nos/day)", "Oven Pushing(nos/d)"], month)
                if v is not None:
                    total += v
                    found = True
            return round(total, 3) if found else None
        return _one_of(cur, plant, ["Oven Pushing (nos/day)", "Oven Pushing(nos/d)"], month)
    if db_item is None:   # Saleable Production = Saleable Steel + Pig Iron
        ss = _mou_value(cur, "Saleable Steel", plant, month)
        pi = _mou_value(cur, "Pig Iron", plant, month)
        if ss is None and pi is None:
            return None
        return round((ss or 0) + (pi or 0), 3)
    return _mou_value(cur, db_item, plant, month)


def build_act_fy_sheet(ws, cur, fy_start: int, months: list):
    fy_label = f"{fy_start}-{str(fy_start + 1)[2:]}"
    n_cols = 2 + 12 + 4 + 1   # item + plant + 12 months + 4 quarters + cum

    ws.cell(row=1, column=1, value="Operations Directorate").font = Font(size=10)
    ws.cell(row=2, column=1, value=f"Actual Monthwise Production: {fy_label}").font = _TITLE_FONT
    ws.cell(row=2, column=n_cols, value="Unit: '000 T (rates: nos/day)").font = _UNIT_FONT

    header_row = 3
    for col, text in ((1, "Items"), (2, "Plant")):
        c = ws.cell(row=header_row, column=col, value=text)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
    for i, m in enumerate(months, start=3):
        c = ws.cell(row=header_row, column=i, value=_mlabel(m))
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="center")
    for i, label in enumerate(_QTR_LABELS, start=15):
        c = ws.cell(row=header_row, column=i, value=label)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=header_row, column=19, value=f"{fy_label}\nCum")
    c.font = _HDR_FONT
    c.fill = _HDR_FILL
    c.alignment = Alignment(horizontal="center", wrap_text=True)

    row = header_row + 1
    for item_label, db_item, plant_rows, is_rate in _ACT_ITEMS:
        first = True
        for plant in plant_rows:
            display_plant = _ACT_PLANT_LABEL.get(plant, plant)
            vals = [_act_plant_value(cur, item_label, db_item, plant, m) for m in months]
            _label_cell(ws, row, 1, item_label if first else "")
            _label_cell(ws, row, 2, display_plant, bold=(plant in ("5 Plants", "SAIL")))
            for i, v in enumerate(vals, start=3):
                _num_cell(ws, row, i, v, bold=(plant in ("5 Plants", "SAIL")))
            quarters = [vals[0:3], vals[3:6], vals[6:9], vals[9:12]]
            agg = _avg_opt if is_rate else _sum_opt
            q_aggs = [agg(q) for q in quarters]
            for i, qa in enumerate(q_aggs, start=15):
                _num_cell(ws, row, i, qa, bold=(plant in ("5 Plants", "SAIL")))
            cum = _avg_opt(vals) if is_rate else _sum_opt(q_aggs)
            _num_cell(ws, row, 19, cum, bold=(plant in ("5 Plants", "SAIL")))
            row += 1
            first = False

    # Conversion / SAIL incl. conversion — SAIL-level only, appended after
    # the Finished Steel Production block (matches the sample's layout).
    conversion_vals = [_one(cur, "SAIL", "Conversion", m) for m in months]
    fs_sail_vals = [_mou_value(cur, "Finished Steel", "SAIL", m) for m in months]
    inc_conv_vals = [
        None if (fs is None and cv is None) else round((fs or 0) + (cv or 0), 3)
        for fs, cv in zip(fs_sail_vals, conversion_vals)
    ]
    for label, vals, bold in (("Conversion", conversion_vals, False),
                              ("SAIL inc. conversion", inc_conv_vals, True)):
        _label_cell(ws, row, 1, "")
        _label_cell(ws, row, 2, label, bold=bold)
        for i, v in enumerate(vals, start=3):
            _num_cell(ws, row, i, v, bold=bold)
        quarters = [vals[0:3], vals[3:6], vals[6:9], vals[9:12]]
        q_sums = [_sum_opt(q) for q in quarters]
        for i, qs in enumerate(q_sums, start=15):
            _num_cell(ws, row, i, qs, bold=bold)
        _num_cell(ws, row, 19, _sum_opt(q_sums), bold=bold)
        row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    for i in range(3, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 10
    ws.freeze_panes = "C4"


def build_pmix_fy_workbook(fy_start: int):
    fy_label = f"{fy_start}-{str(fy_start + 1)[2:]}"
    months = _fy_months(fy_start)
    n_cols = 1 + 12 + 4 + 1   # label + 12 months + 4 quarters + cum

    conn = db.connect()
    cur = conn.cursor()
    try:
        rows_by_month = {m: compute_pmix_rows(cur, m) for m in months}
        conversion = {m: _one(cur, "SAIL", "Conversion", m) for m in months}
        rsp_hsm2 = {m: _one(cur, "RSP", "HSM-2 Total HR Coil", m) for m in months}
        bsl_hsm = {m: _one(cur, "BSL", "HSM Total HR Coil", m) for m in months}

        wb = openpyxl.Workbook()
        ws_act = wb.active
        ws_act.title = f"ACT {fy_label}"
        build_act_fy_sheet(ws_act, cur, fy_start, months)
    finally:
        conn.close()

    n_rows = len(rows_by_month[months[0]])
    fst_by_month = {}
    for m in months:
        fst_row = next((r for r in rows_by_month[m] if r["label"] == "Finished Steel (Total)"), None)
        fst_by_month[m] = fst_row["value"] if fst_row else None

    ws = wb.create_sheet(f"Pmix'{fy_label[2:]} ACT")

    ws.cell(row=1, column=1, value="Operations Directorate").font = Font(size=10)
    ws.cell(row=2, column=1, value=f"Product mix Performance: {fy_label}").font = _TITLE_FONT
    ws.cell(row=2, column=n_cols, value="Unit: '000 T").font = _UNIT_FONT

    header_row = 3
    ws.cell(row=header_row, column=1, value="Item").font = _HDR_FONT
    ws.cell(row=header_row, column=1).fill = _HDR_FILL
    for i, m in enumerate(months, start=2):
        c = ws.cell(row=header_row, column=i, value=_mlabel(m))
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="center")
    for i, label in enumerate(_QTR_LABELS, start=14):
        c = ws.cell(row=header_row, column=i, value=label)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=header_row, column=18, value="Cum")
    c.font = _HDR_FONT
    c.fill = _HDR_FILL
    c.alignment = Alignment(horizontal="center")

    def write_data_row(label, values_by_month, bold=False):
        nonlocal row
        _label_cell(ws, row, 1, label, bold=bold)
        vals = [values_by_month.get(m) for m in months]
        for i, v in enumerate(vals, start=2):
            _num_cell(ws, row, i, v, bold=bold)
        quarters = [vals[0:3], vals[3:6], vals[6:9], vals[9:12]]
        q_sums = [_sum_opt(q) for q in quarters]
        for i, qs in enumerate(q_sums, start=14):
            _num_cell(ws, row, i, qs, bold=bold)
        _num_cell(ws, row, 18, _sum_opt(q_sums), bold=bold)
        row += 1

    row = header_row + 1
    for row_idx in range(n_rows):
        spec = rows_by_month[months[0]][row_idx]
        if spec["kind"] == "header":
            _plant_hdr(ws, row, spec["label"], n_cols)
            row += 1
        else:
            values_by_month = {m: rows_by_month[m][row_idx]["value"] for m in months}
            write_data_row(spec["label"], values_by_month, bold=spec["bold"])

    write_data_row("Finished Steel (conversion)", conversion)
    fsic = {m: (fst_by_month[m] or 0) + (conversion[m] or 0)
            if fst_by_month[m] is not None or conversion[m] is not None else None
            for m in months}
    write_data_row("Finished Steel including Conversion", fsic, bold=True)
    write_data_row("RSP HSM-2", rsp_hsm2)
    write_data_row("BSL HSM", bsl_hsm)

    ws.column_dimensions["A"].width = 30
    for i in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 10
    ws.freeze_panes = "B4"

    return wb


def build_pmix_fy_report_bytes(fy_start: int) -> bytes:
    wb = build_pmix_fy_workbook(fy_start)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
