"""
RSP Month-End Techno Extractor — tentative for-the-month techno values from
the RSP Daily Morning Report (Report_format/MONTHEND/RSP31052026.xlsx style).

When generated on the last day of a month, the report's "Cum"/"ToDate"
columns hold the for-the-month values. Layout (single sheet):

  A1  'RSP Daily Morning Report'          — file marker
  A2  ' For the Date -: 31.05.2026'       — report date (verified vs selection)

  Production block (~rows 49-53): header row has 'Furnace' in column I;
  furnace rows labelled I / IV / V / TOT in column I:
      K = HM Production Cum (t)     O = BF Productivity Cum

  Techno block (~rows 56-60): header row has 'Furnace' in column B;
  furnace rows labelled I / IV / V / TOT in column B:
      E = Coke Rate Cum    J = CDI/PCI Cum      L = Sinter% Cum
      N = Nut Coke Cum     R = Pellet% Cum      V = O2 Enrich% Cum
      X = Hot Blast Temp Cum
      Slag Rate Cum — column located from the header ('SlagRt (Cu...', col Y)

  SMS blocks: anchor 'SMS1' / 'SMS-II' in column B; rows labelled
  'Consm HM' / 'ScrapRate' in column C, ToDate value in column E.
  TMI is derived as HM consumption + scrap consumption.

  Specific energy consumption: 'EnergyCons' header in column N; the 'G.Cal'
  row below it carries the month rate in column P → General unit.

Param keys match what the RSP technopara upload already stores (cdi,
o2_enrichment, hot_blast_temp, …) so the final technopara extraction
cleanly replaces these tentative values later.
"""

import re
import sys
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional

sys.path.insert(0, str(Path(__file__).parent))
from dsp_mcr_techno_extractor import McrMonthMismatch, _clean_val  # noqa: E402

# furnace label (column I / column B) → unit name
_FURNACE_LABELS = {"I": "BF-1", "IV": "BF-4", "V": "BF-5", "TOT": "BF_Shop"}

# production block: (1-based column, param key)
_PRODUCTION_COLS = [(11, "production"), (15, "bf_productivity")]   # K, O
_PROD_LABEL_COL = 9    # column I

# techno block: (1-based column, param key) — all "Cum" columns
_TECHNO_COLS = [
    (5,  "coke_rate"),          # E
    (10, "cdi"),                # J  (report header says PCI — same parameter)
    (12, "sinter_in_burden"),   # L
    (14, "nut_coke_rate"),      # N
    (18, "pellet_in_burden"),   # R
    (22, "o2_enrichment"),      # V (matches rsp_technopara_sections.py's key
                                #    name, so the final technopara upload's
                                #    merge_upsert_techno_data call replaces this
                                #    tentative value instead of leaving both
                                #    keys side by side)
    (24, "hot_blast_temp"),     # X
]
_TECHNO_LABEL_COL = 2   # column B

# SMS blocks: (column-B anchor, unit)
_SMS_ANCHORS = [("SMS1", "SMS-1"), ("SMS-II", "SMS-2")]
_SMS_LABELS = {"CONSM HM": "specific_hm_consumption",
               "SCRAPRATE": "specific_scrap_consumption"}
_SMS_LABEL_COL = 3   # column C
_SMS_VALUE_COL = 5   # column E — 'ToDate'


def _load_grid(file_path: str) -> List[List]:
    """First sheet as a list of row lists (.xlsx via openpyxl, .xls via xlrd)."""
    with open(file_path, "rb") as f:
        magic = f.read(4)
    if magic == b"\xd0\xcf\x11\xe0":
        import xlrd
        sh = xlrd.open_workbook(file_path).sheet_by_index(0)
        return [[sh.cell_value(r, c) for c in range(sh.ncols)]
                for r in range(sh.nrows)]
    if magic[:2] == b"PK":
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        grid = [list(row) for row in wb[wb.sheetnames[0]].iter_rows(values_only=True)]
        wb.close()
        return grid
    raise ValueError("Unrecognised file format — expected the RSP Morning "
                     "Report Excel file (.xlsx/.xls).")


def _cell(grid, row_1b, col_1b):
    if row_1b < 1 or row_1b > len(grid):
        return None
    row = grid[row_1b - 1]
    if col_1b < 1 or col_1b > len(row):
        return None
    v = row[col_1b - 1]
    return None if v == "" else v


class RspMonthendTechnoExtractor:
    """
    extract() returns the standard techno record list:
        [{"plant": "RSP", "report_month": "YYYY-MM", "unit": str,
          "techno_json": {"month": {...}, "till_month": {}}}]

    After extract():  .report_date, .warnings
    """

    def __init__(self, file_path: str, report_month: str = ""):
        self.file_path = file_path
        self.report_month = report_month
        self.report_date: Optional[date] = None
        self.warnings: List[str] = []

    def _verify_month(self):
        file_month = f"{self.report_date.year}-{self.report_date.month:02d}"
        if self.report_month and self.report_month != file_month:
            raise McrMonthMismatch(
                f"This RSP Morning Report is dated "
                f"{self.report_date.strftime('%d.%m.%Y')} (month {file_month}) "
                f"but you selected {self.report_month}. Select the matching "
                "month or upload the correct file."
            )
        self.report_month = file_month

    def _find_label_row(self, grid, col_1b, text, start=1, end=None):
        """1-based row whose cell at col equals text (stripped, upper)."""
        end = end or len(grid)
        for r in range(start, min(end, len(grid)) + 1):
            v = _cell(grid, r, col_1b)
            if v is not None and str(v).strip().upper() == text.upper():
                return r
        return None

    def extract(self) -> List[Dict]:
        grid = _load_grid(self.file_path)

        if "RSP DAILY MORNING REPORT" not in str(_cell(grid, 1, 1) or "").upper():
            raise ValueError(
                "File not recognised — cell A1 must contain "
                "'RSP Daily Morning Report'."
            )

        m = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", str(_cell(grid, 2, 1) or ""))
        if not m:
            raise ValueError(
                "Cannot read the report date from cell A2 — expected "
                "'For the Date -: DD.MM.YYYY'."
            )
        d, mo, y = (int(g) for g in m.groups())
        self.report_date = date(y, mo, d)
        self._verify_month()

        units: Dict[str, Dict] = {}

        def put(unit, key, value):
            units.setdefault(unit, {"month": {}, "till_month": {}})
            units[unit]["month"][key] = value

        # ── Production block — header row has 'Furnace' in column I ──────────
        hdr = self._find_label_row(grid, _PROD_LABEL_COL, "Furnace")
        if hdr is None:
            self.warnings.append("Production block ('Furnace' header in column I) not found — skipped.")
        else:
            for r in range(hdr + 1, min(hdr + 8, len(grid)) + 1):
                label = str(_cell(grid, r, _PROD_LABEL_COL) or "").strip().upper()
                unit = _FURNACE_LABELS.get(label)
                if unit:
                    for col, key in _PRODUCTION_COLS:
                        put(unit, key, _clean_val(_cell(grid, r, col)))

        # ── Techno block — header row has 'Furnace' in column B ──────────────
        hdr = self._find_label_row(grid, _TECHNO_LABEL_COL, "Furnace")
        if hdr is None:
            self.warnings.append("Techno block ('Furnace' header in column B) not found — skipped.")
        else:
            # Slag-rate cumulative column is found from its header ('SlagRt (Cu…')
            slag_col = None
            hdr_row = grid[hdr - 1]
            for c, v in enumerate(hdr_row, start=1):
                if isinstance(v, str) and v.strip().upper().startswith("SLAGRT"):
                    slag_col = c
                    break
            if slag_col is None:
                self.warnings.append("Slag Rate column ('SlagRt (Cum)' header) not found — skipped.")

            for r in range(hdr + 1, min(hdr + 8, len(grid)) + 1):
                label = str(_cell(grid, r, _TECHNO_LABEL_COL) or "").strip().upper()
                unit = _FURNACE_LABELS.get(label)
                if unit:
                    for col, key in _TECHNO_COLS:
                        put(unit, key, _clean_val(_cell(grid, r, col)))
                    if slag_col:
                        put(unit, "slag_rate", _clean_val(_cell(grid, r, slag_col)))

        # ── SMS blocks — 'SMS1' / 'SMS-II' anchors in column B ───────────────
        for anchor, unit in _SMS_ANCHORS:
            arow = self._find_label_row(grid, _TECHNO_LABEL_COL, anchor)
            if arow is None:
                self.warnings.append(f"SMS block '{anchor}' not found in column B — skipped.")
                continue
            for r in range(arow, min(arow + 10, len(grid)) + 1):
                label = str(_cell(grid, r, _SMS_LABEL_COL) or "").strip().upper()
                key = _SMS_LABELS.get(label)
                if key:
                    put(unit, key, _clean_val(_cell(grid, r, _SMS_VALUE_COL)))
            # TMI = HM consumption + scrap consumption
            mvals = units.get(unit, {}).get("month", {})
            hm, scrap = mvals.get("specific_hm_consumption"), mvals.get("specific_scrap_consumption")
            if hm is not None and scrap is not None:
                put(unit, "tmi", round(hm + scrap, 4))

        # ── Specific energy consumption → General ────────────────────────────
        erow = self._find_label_row(grid, 14, "EnergyCons")   # column N
        sec = None
        if erow:
            for r in range(erow + 1, min(erow + 5, len(grid)) + 1):
                if str(_cell(grid, r, 14) or "").strip().upper() == "G.CAL":
                    sec = _clean_val(_cell(grid, r, 16))      # column P — 'Rate'
                    break
        if sec is not None:
            put("General", "specific_energy_consumption", sec)
        else:
            self.warnings.append(
                "Specific energy consumption ('EnergyCons' / 'G.Cal' rows, column P) not found — skipped.")

        records = []
        for unit, tj in units.items():
            if any(v is not None for v in tj["month"].values()):
                records.append({
                    "plant": "RSP",
                    "report_month": self.report_month,
                    "unit": unit,
                    "techno_json": tj,
                })
        if not records:
            raise ValueError(
                "No techno values found in the RSP Morning Report — verify the "
                "file contents."
            )
        return records
