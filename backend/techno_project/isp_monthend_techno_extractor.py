"""
ISP Month-End Techno Extractor — tentative for-the-month techno values from
ISP's "MORNING REPORT" workbook (Report_format/MONTHEND/MORNING REPORT.xlsx
style), sheet 'DAILYREPORT1'.

When generated on the last day of the month, the report's "तिथि तक" ("to
date" / cumulative-so-far) column holds the for-the-month value.

  Report date: J5 (day, e.g. 31) + K5 (month/year as an Excel date, day
  irrelevant, e.g. 01-MAY-2026) → combined into the full report date.

  Blast-furnace techno block (fixed rows, column F label / column J value):
      J67 Coke Rate            J68 Nut Coke Rate      J69 CDI Rate
      J70 Fuel Rate             J71 Sinter in Burden   J72 Pellet in Burden
      J73 BF Productivity       J74 BF Slag Rate       J75 Si in Hot Metal
      J76 Hot Blast Temperature J78 O2 Enrichment in BF
  → all stored under unit 'BF-5' (ISP's only furnace).

  SMS block: J81 Gross Hotmetal Consumption, J82 Scrap Consumption
  → unit 'SMS'.

  J85 Specific Energy Consumption → unit 'General'.

Rows are fixed (per the report's stable layout) but each is verified against
its expected column-F label before reading column J, so a shifted layout
produces a warning instead of a silently wrong value.

Param keys match what the ISP technopara upload already stores (coke_rate,
bf_productivity, silicon_in_hm, …) so the final technopara extraction cleanly
replaces these tentative values later.
"""

import sys
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional

sys.path.insert(0, str(Path(__file__).parent))
from dsp_mcr_techno_extractor import McrMonthMismatch, _clean_val  # noqa: E402

_SHEET_NAME = "DAILYREPORT1"

_LABEL_COL = 6    # column F
_VALUE_COL = 10   # column J — "तिथि तक" (to date / cum)

# (row, canonical param key, expected column-F label substring — upper case)
_BF_PARAMS = [
    (67, "coke_rate",              "COKE RATE"),
    (68, "nut_coke_rate",          "NUT COKE RATE"),
    (69, "cdi",                    "CDI RATE"),
    (70, "fuel_rate",              "FUEL RATE"),
    (71, "sinter_in_burden",       "SINTER IN BURDEN"),
    (72, "pellet_in_burden",       "PELLET IN BURDEN"),
    (73, "bf_productivity",        "BF PRODUCTIVITY"),
    (74, "slag_rate",              "BF SLAG RATE"),
    (75, "silicon_in_hm",          "SI IN HOTMETAL"),
    (76, "hot_blast_temp",         "HOT BLAST TEMPERATURE"),
    (78, "o2_enrichment",          "O2 ENRICHMENT"),
]
_BF_UNIT = "BF-5"

_SMS_PARAMS = [
    (81, "specific_hm_consumption",    "GROSS HOTMETAL CONSUMPTION"),
    (82, "specific_scrap_consumption", "SCRAP CONSUMPTION"),
]
_SMS_UNIT = "SMS"

_GENERAL_PARAMS = [
    (85, "specific_energy_consumption", "SP.ENERGY CONSUMPTION"),
]
_GENERAL_UNIT = "General"

_ALL_BLOCKS = [
    (_BF_PARAMS, _BF_UNIT),
    (_SMS_PARAMS, _SMS_UNIT),
    (_GENERAL_PARAMS, _GENERAL_UNIT),
]


def _load_sheet(file_path: str, sheet_name: str):
    with open(file_path, "rb") as f:
        magic = f.read(4)
    if magic == b"\xd0\xcf\x11\xe0":
        import xlrd
        wb = xlrd.open_workbook(file_path)
        sh = wb.sheet_by_name(sheet_name)
        grid = [[sh.cell_value(r, c) for c in range(sh.ncols)]
                for r in range(sh.nrows)]
        return grid, wb.datemode
    if magic[:2] == b"PK":
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Workbook has no '{sheet_name}' sheet.")
        grid = [list(row) for row in wb[sheet_name].iter_rows(values_only=True)]
        wb.close()
        return grid, None
    raise ValueError("Unrecognised file format — expected the ISP MORNING "
                     "REPORT Excel file (.xlsx/.xls).")


def _cell(grid, row_1b, col_1b):
    if row_1b < 1 or row_1b > len(grid):
        return None
    row = grid[row_1b - 1]
    if col_1b < 1 or col_1b > len(row):
        return None
    v = row[col_1b - 1]
    return None if v == "" else v


class IspMonthendTechnoExtractor:
    """
    extract() returns the standard techno record list:
        [{"plant": "ISP", "report_month": "YYYY-MM", "unit": str,
          "techno_json": {"month": {...}, "till_month": {}}}]

    After extract():  .report_date, .warnings
    """

    def __init__(self, file_path: str, report_month: str = ""):
        self.file_path = file_path
        self.report_month = report_month
        self.report_date: Optional[date] = None
        self.warnings: List[str] = []
        self._xls_datemode = None

    def _verify_month(self):
        file_month = f"{self.report_date.year}-{self.report_date.month:02d}"
        if self.report_month and self.report_month != file_month:
            raise McrMonthMismatch(
                f"This ISP Morning Report is dated "
                f"{self.report_date.strftime('%d.%m.%Y')} (month {file_month}) "
                f"but you selected {self.report_month}. Select the matching "
                "month or upload the correct file."
            )
        self.report_month = file_month

    def _read_report_date(self, grid):
        day_v = _cell(grid, 5, 10)   # J5
        ym_v = _cell(grid, 5, 11)    # K5
        try:
            day = int(float(day_v))
        except (TypeError, ValueError):
            raise ValueError(
                "Cannot read the report day from cell J5 of "
                f"'{_SHEET_NAME}' — expected a day number (e.g. 31)."
            )
        try:
            if self._xls_datemode is not None:
                import xlrd
                y, m, *_ = xlrd.xldate_as_tuple(float(ym_v), self._xls_datemode)
            else:  # openpyxl gives a datetime directly
                y, m = ym_v.year, ym_v.month
            self.report_date = date(y, m, day)
        except Exception:
            raise ValueError(
                f"Cannot read the report month from cell K5 of "
                f"'{_SHEET_NAME}' (got {ym_v!r}) — verify this is the ISP "
                "MORNING REPORT workbook."
            )

    def extract(self) -> List[Dict]:
        grid, self._xls_datemode = _load_sheet(self.file_path, _SHEET_NAME)

        self._read_report_date(grid)
        self._verify_month()

        units: Dict[str, Dict] = {}

        def put(unit, key, value):
            units.setdefault(unit, {"month": {}, "till_month": {}})
            units[unit]["month"][key] = value

        for params, unit in _ALL_BLOCKS:
            for row, key, expected_label in params:
                label = str(_cell(grid, row, _LABEL_COL) or "").strip().upper()
                if expected_label not in label:
                    self.warnings.append(
                        f"Row {row}: expected label containing "
                        f"'{expected_label}' in column F, found {label!r} — "
                        f"'{key}' skipped."
                    )
                    continue
                put(unit, key, _clean_val(_cell(grid, row, _VALUE_COL)))

        records = []
        for unit, tj in units.items():
            if any(v is not None for v in tj["month"].values()):
                records.append({
                    "plant": "ISP",
                    "report_month": self.report_month,
                    "unit": unit,
                    "techno_json": tj,
                })
        if not records:
            raise ValueError(
                "No techno values found in the ISP MORNING REPORT — verify "
                "the file contents."
            )
        return records
