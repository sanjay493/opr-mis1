import json
import re
from datetime import datetime, time
from pathlib import Path
from openpyxl import load_workbook
from typing import Dict, List, Optional

_MONTH_ABBRS = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
_MONTH_NUM_TO_ABBR = {
    1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun',
    7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'
}
_MONTH_ABBR_TO_NUM = {
    'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
    'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
}
_NEXT_YEAR_MONTHS = {'Jan', 'Feb', 'Mar'}

# dry_coal_charge_oven's denominator ("Num of Ovens Pushed (COB#10/11)", the
# MONTHLY TOTAL oven-push count) is simply absent as its own row in some
# older report templates — they only carry a per-day average instead, under
# various wordings across vintages. Per SAIL convention (confirmed against
# the canonical template: Average Pushing (COB#10)=98, Num of Ovens Pushed
# (COB#10)=2940, April=30 days -> 98*30=2940 exactly; and independently
# against Mar'24Summarized Monthly Report.xlsx's own wording, "No.of Ovens
# Pushed (COB#10)"=99.0645..., March=31 days -> 3071, reproducing the
# known-correct dry_coal_charge_oven of 17.2 for that file too):
#   Num of Ovens Pushed = <daily average> * days in month
_OVEN_COUNT_FALLBACK_ALIASES = {
    "num of ovens pushed (cob#10)": ["Average Pushing (COB#10)", "No.of Ovens Pushed (COB#10)"],
    "num of ovens pushed (cob#11)": ["Average Pushing (COB#11)", "Nos.of Ovens Pushed (COB#11)"],
}


def _fy_month_sequence(upto_month: str) -> List[str]:
    """April..`upto_month` (inclusive) in FY order, as 'YYYY-MM' strings.

    e.g. upto_month='2026-06' -> ['2026-04','2026-05','2026-06'];
    upto_month='2026-03' -> ['2025-04', ..., '2026-03'] (a full FY, since
    March belongs to the FY that started the previous April)."""
    y, m = upto_month.split('-')
    y, m = int(y), int(m)
    fy_start_year = y if m >= 4 else y - 1
    upto_idx = _MONTH_ABBRS.index(_MONTH_NUM_TO_ABBR[m])
    months = []
    for abbr in _MONTH_ABBRS[:upto_idx + 1]:
        num = _MONTH_ABBR_TO_NUM[abbr]
        yr = fy_start_year + 1 if abbr in _NEXT_YEAR_MONTHS else fy_start_year
        months.append(f"{yr}-{num:02d}")
    return months


class IspTechnoExtractor:
    def __init__(self, excel_file: str, report_month: str = None):
        """
        Args:
            excel_file: Path to the ISP monthly report Excel file.
            report_month: Report month in YYYY-MM format (e.g. "2026-03").
        """
        self.excel_file = Path(excel_file)
        self.report_month = report_month
        self.workbook = None
        self.hardcoded_map = self._load_hardcoded_map()
        self.row_labels = self._load_row_labels()
        self.expr_row_labels = self._load_expr_row_labels()
        self.month_col = None
        self.header_row = None

    def _load_hardcoded_map(self) -> Dict:
        """Load sheet-wise parameter mapping."""
        map_path = Path(__file__).parent / "isp_technopara_map.json"
        try:
            with open(map_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except FileNotFoundError:
            print(f"Error: isp_technopara_map.json not found at {map_path}!")
            return {}

    def _load_row_labels(self) -> Dict:
        """Load the companion {sheet: {unit: {param_key: expected column-B
        label}}} file used to verify/self-heal isp_technopara_map.json's
        hardcoded row numbers against future report-template row shifts.
        Covers only the simple (non-expression) row specs — see
        _verified_row(). Missing file/entries just disable verification,
        never break extraction."""
        labels_path = Path(__file__).parent / "isp_technopara_row_labels.json"
        try:
            with open(labels_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except FileNotFoundError:
            return {}

    def _load_expr_row_labels(self) -> Dict:
        """Load {sheet: {unit: {param_key: {row_num_str: expected column-B
        label}}}} — per-row verification for the individual rows referenced
        inside an expression spec (e.g. '17/8'), since a whole expression
        can't be checked against a single label the way a simple row spec
        can. See isp_technopara_expr_row_labels.json's _comment. Missing
        file/entries just fall back to the unit's blanket offset, never
        break extraction."""
        labels_path = Path(__file__).parent / "isp_technopara_expr_row_labels.json"
        try:
            with open(labels_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except FileNotFoundError:
            return {}

    @staticmethod
    def _norm_label(s) -> str:
        return re.sub(r"\s+", " ", str(s or "")).strip().lower()

    @staticmethod
    def _safe_print(msg: str) -> None:
        """print(), tolerant of consoles that can't encode a label's
        non-ASCII characters (confirmed: Windows cp1252 raises
        UnicodeEncodeError on the '₂' in 'Total (NH4)₂SO4' — a real
        isp_technopara_row_labels.json entry). Unlike a bare print() inside
        the per-param try/except (which safely re-prints the exception's
        own ASCII description, never the raw label), these diagnostic
        messages embed the raw label text themselves, so they need their
        own fallback rather than crashing the whole extraction over a
        logging line."""
        try:
            print(msg)
        except UnicodeEncodeError:
            print(msg.encode("ascii", errors="replace").decode("ascii"))

    @staticmethod
    def _is_simple_row_spec(row_spec) -> bool:
        """True for a plain row number (int, or a numeric string with no
        arithmetic operators) — the specs _resolve_unit_rows can check
        directly against a single expected label."""
        if isinstance(row_spec, int):
            return True
        if isinstance(row_spec, str):
            return not any(op in row_spec for op in ['+', '-', '/', '*', '(', ')'])
        return False

    def _resolve_unit_rows(self, ws, sheet_name: str, unit_name: str, unit_params: Dict) -> Dict[str, int]:
        """Resolve every row this unit's params need by searching the WHOLE
        sheet for each expected column-B label, rather than starting from
        the hardcoded row and searching a fixed window outward. This is what
        makes extraction robust to real report-template drift: (a) shifts
        bigger than any fixed window (confirmed in a real file: USM shifted
        -51 rows), and (b) a single sheet having more than one shift zone at
        once (confirmed: COKE OVENS' oven-count block shifted -4 while its
        coke-quality block shifted -7 in the same file, in the same sheet)
        — every row is found independently by its own label, never inferred
        from a neighbour's shift or a single blanket per-unit offset.

        A label with more than one occurrence in the sheet (e.g. COB-old and
        COB-new both have a 'Sp Heat Cons' row) is disambiguated by picking
        the occurrence closest to this unit's OTHER already-resolved rows
        (falling back to the configured row as the anchor before anything
        else has resolved) — a unit's real rows always sit in one
        contiguous block, so its own resolved rows are a reliable anchor for
        the ambiguous ones.

        Returns {param_key: row} for simple specs and
        {f"{param_key}#{row_token}": row} for each row referenced inside an
        expression spec (isp_technopara_map.json / isp_technopara_row_labels
        .json / isp_technopara_expr_row_labels.json respectively). A key
        with no configured expected label is simply absent — callers fall
        back to the configured row / a blanket unit offset for those."""
        row_labels = self.row_labels.get(sheet_name, {}).get(unit_name, {})
        expr_labels = self.expr_row_labels.get(sheet_name, {}).get(unit_name, {})
        if not row_labels and not expr_labels:
            return {}

        sheet_rows = [
            (r, self._norm_label(ws.cell(row=r, column=2).value))
            for r in range(1, ws.max_row + 1)
        ]
        sheet_rows = [(r, t) for r, t in sheet_rows if t]

        targets = []  # (key, [alias, ...], configured_row)
        for param_key, row_spec in unit_params.items():
            if self._is_simple_row_spec(row_spec):
                expected = row_labels.get(param_key)
                if expected:
                    aliases = expected if isinstance(expected, list) else [expected]
                    targets.append((param_key, aliases, int(row_spec)))
            elif isinstance(row_spec, str):  # expression spec
                per_row = expr_labels.get(param_key, {})
                for tok in re.findall(r'\d+', row_spec):
                    expected = per_row.get(tok)
                    if expected:
                        aliases = expected if isinstance(expected, list) else [expected]
                        targets.append((f"{param_key}#{tok}", aliases, int(tok)))

        def find_candidates(aliases):
            # Exact matches (whole normalized cell text equals an alias)
            # always win over mere substring hits — a short label like
            # 'Coke' is a substring of 'Nut Coke' too, and confirmed in a
            # real file: 'Nut Coke' happened to sit exactly at coke_rate's
            # configured row, so the nearest-to-configured-row tiebreak
            # picked it as a false "exact-row" match even though it isn't
            # the same parameter at all. Only fall back to substring
            # matching (for legitimate label-text drift between file
            # vintages) when nothing matches exactly anywhere in the sheet.
            norms = [self._norm_label(a) for a in aliases]
            exact = [r for r, text in sheet_rows if text in norms]
            if exact:
                return exact
            return [r for r, text in sheet_rows if any(n in text for n in norms)]

        # A short/generic label ('Coke', 'CDI') can substring-match many
        # unrelated rows sheet-wide, not just its own two (e.g. one per
        # COB-old/COB-new block). Disambiguating those against a
        # cross-parameter centroid was tried and is provably wrong — it let
        # one contaminated resolution (e.g. a distant 'BF Coke' match) drag
        # the anchor for every other ambiguous param toward it. Each param's
        # OWN configured row — where the un-shifted map already puts it, a
        # few rows away at most in every real file seen — is a far more
        # reliable anchor than any other resolved param, so always pick
        # whichever candidate sits nearest to the CONFIGURED row itself.
        resolved: Dict[str, int] = {}
        for key, aliases, configured_row in targets:
            cands = find_candidates(aliases)
            if not cands:
                has_fallback = any(
                    self._norm_label(a) in _OVEN_COUNT_FALLBACK_ALIASES for a in aliases
                )
                if has_fallback:
                    self._safe_print(
                        f"Info: '{sheet_name}/{unit_name}/{key}' expected label "
                        f"{aliases!r} not found — will use the daily-average "
                        f"fallback instead (see _oven_count_fallback_value)")
                else:
                    self._safe_print(
                        f"Warning: '{sheet_name}/{unit_name}/{key}' expected label "
                        f"{aliases!r} not found anywhere in '{sheet_name}' — using "
                        f"configured row {configured_row} unverified")
                continue
            best = min(cands, key=lambda r: abs(r - configured_row))
            if best != configured_row:
                extra = f", {len(cands)} label matches in sheet" if len(cands) > 1 else ""
                self._safe_print(
                    f"Info: '{sheet_name}/{unit_name}/{key}' row shifted "
                    f"{configured_row} -> {best} (label {aliases!r}{extra})")
            resolved[key] = best

        return resolved

    @staticmethod
    def _unit_offset_from_map(unit_params: Dict, unit_row_map: Dict[str, int]) -> Optional[int]:
        """Last-resort blanket shift for list specs and any expression-row
        token _resolve_unit_rows couldn't resolve by label (no expected
        label configured, or not found anywhere in the sheet) — derived from
        whichever simple specs DID resolve. Returns None (no correction) if
        nothing resolved, or if the resolved simple specs disagree on the
        shift amount (a real possibility — see _resolve_unit_rows — safer to
        leave those specific unresolved rows unshifted than guess wrong)."""
        deltas = set()
        for param_key, row_spec in unit_params.items():
            if not IspTechnoExtractor._is_simple_row_spec(row_spec) or param_key not in unit_row_map:
                continue
            deltas.add(unit_row_map[param_key] - int(row_spec))
        if len(deltas) == 1:
            return next(iter(deltas))
        return None

    def _oven_count_fallback_value(self, ws, sheet_name: str, unit_name: str,
                                    param_key: str, token: str, col: int, days: float) -> Optional[float]:
        """Fallback for dry_coal_charge_oven's denominator token when its
        primary label ('Num of Ovens Pushed (COB#10/11)') isn't found
        anywhere in the sheet — see _OVEN_COUNT_FALLBACK_ALIASES's comment
        for the formula and how it was verified. Returns the computed
        monthly-total oven-push count, or None if no alias matches either."""
        expected = self.expr_row_labels.get(sheet_name, {}).get(unit_name, {}).get(param_key, {}).get(token)
        if not expected:
            return None
        primary_labels = expected if isinstance(expected, list) else [expected]
        fallback_aliases = []
        for label in primary_labels:
            fallback_aliases.extend(_OVEN_COUNT_FALLBACK_ALIASES.get(self._norm_label(label), []))
        if not fallback_aliases:
            return None
        norms = [self._norm_label(a) for a in fallback_aliases]
        for r in range(1, ws.max_row + 1):
            if self._norm_label(ws.cell(row=r, column=2).value) in norms:
                avg = self._get_cell_value(ws, r, col)
                if avg is not None:
                    self._safe_print(
                        f"Info: '{sheet_name}/{unit_name}/{param_key}#{token}' — 'Num of "
                        f"Ovens Pushed' row absent, using row {r} (daily average) * "
                        f"{days:g} days = {float(avg) * days!r}")
                    return float(avg) * days
        return None

    @staticmethod
    def _clean_value(val):
        """Convert value to JSON-serializable format."""
        if isinstance(val, time):
            return val.strftime("%H:%M:%S")
        if isinstance(val, datetime):
            return val.time().strftime("%H:%M:%S")
        if isinstance(val, str) and not val.strip():
            return None
        _bad = {"#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#N/A", "#NULL!", "#NUM!", "-", "--", None, ""}
        if val in _bad:
            return None
        return val

    def open_workbook(self):
        self.workbook = load_workbook(self.excel_file, data_only=True)
        print(f"Opened workbook with sheets: {self.workbook.sheetnames[:10]}...")

    def _find_month_column(self, ws):
        """Detect month column in given worksheet."""
        # Try rows 3, 2, 4, 5 to find headers
        for row_num in [3, 2, 4, 5]:
            header = [str(c.value).strip() if c.value else "" for c in ws[row_num]]
            has_months = any(
                abbr in h or f"{abbr}'" in h
                for h in header
                for abbr in _MONTH_ABBRS
            )
            if has_months:
                self.header_row = row_num
                return row_num, header
        return None, []

    def _get_cum_column_offset(self, month_num: int) -> int:
        """
        Get cumulative column offset based on month.
        ISP Excel pattern (verified against the full-year column header row,
        e.g. Apr/May/2M/Jun/QTR-1/Jul/4M/Aug/5M/Sep/QRT-2/H1/Oct/7M/Nov/8M/
        Dec/QRT-3/...):
          - Apr: NO offset - there is no separate "1-month cumulative" column
            in this template (it would just duplicate April's own ACT
            column), so till_month must read from the same column as month.
          - Most other months: cum is 2 columns ahead (the next periodic
            cumulative marker - 2M, QTR-1, 4M, 5M, 7M, 8M - sits right after
            that month's own pair).
          - Sep & Dec: cum is 4 columns ahead, skipping a nearer quarter-only
            marker (QRT-2/QRT-3) to reach the true YTD figure (H1/9M).
          - Mar: cum is 6 columns ahead (the annual figure).
        """
        if month_num == 4:  # April - FY's first month
            return 0
        if month_num == 3:  # March
            return 6
        elif month_num in [9, 12]:  # September, December
            return 4
        else:  # May, June, July, Aug, Oct, Nov, Jan, Feb
            return 2

    def _get_cell_value(self, ws, row_num: int, col_num: int):
        """Get value from cell, handling None/error values."""
        try:
            val = ws.cell(row_num, col_num + 1).value
            if val is None or val in {"#DIV/0!", "#VALUE!", "-", "--", ""}:
                return None
            if isinstance(val, (int, float)):
                return val
            return None
        except Exception:
            return None

    def _evaluate_row_expression(self, ws, expression: str, month_col: int, value_lookup=None) -> float:
        """Evaluate expressions like '5+6', '5/days', '(5+6)/days'.

        `value_lookup(row_num, col, days) -> value`, if given, resolves each
        row-number token to its VALUE directly (not just a corrected row) —
        needed because some tokens have no usable row at all in a given
        file and must instead be computed from a different row entirely
        (see _oven_count_fallback_value). Falls back to a plain cell read
        at the literal token row when no lookup is given."""
        try:
            # Get days in month from row 2
            days_val = ws.cell(2, month_col + 1).value
            days = float(days_val) if days_val and isinstance(days_val, (int, float)) else 30

            # Replace 'days' with actual value
            expr = expression.replace("days", str(days))

            # Parse row references (numbers)
            import re
            def get_row_value(match):
                row_num = int(match.group(1))
                if value_lookup:
                    val = value_lookup(row_num, month_col, days)
                else:
                    val = self._get_cell_value(ws, row_num, month_col)
                return str(val) if val is not None else "0"

            expr = re.sub(r'(\d+)(?![\d\.])', get_row_value, expr)

            # Safely evaluate the expression
            result = eval(expr)
            return float(result) if result else None

        except Exception as e:
            print(f"Error evaluating expression '{expression}': {e}")
            return None

    def _get_actual_column(self, ws, month_col: int) -> int:
        """
        ISP format: Columns are merged with Plan/Actual sub-columns.
        Check row 4 to find the Actual (ACT) column.
        Usually month_col = Plan, month_col+1 = Actual
        """
        try:
            row4 = [str(c.value).strip().upper() if c.value else "" for c in ws[4]]

            # Check if month_col is Plan and month_col+1 is Actual
            if month_col < len(row4) and month_col + 1 < len(row4):
                curr = row4[month_col]
                next_col = row4[month_col + 1]

                if "ACT" in next_col or "ACT" in curr:
                    # Use whichever is ACT
                    if "ACT" in next_col:
                        return month_col + 1
                    else:
                        return month_col
        except Exception as e:
            print(f"Warning detecting actual column: {e}")

        # Fallback: assume next column is actual
        return month_col + 1

    def _get_parameter_multiplier(self, sheet_name: str, param_key: str) -> float:
        """
        Get unit multiplier for specific sheet/parameter combinations.
        Used to normalize units across different sheets.

        ISP Mills (BM, USM, WRM): specific_heat_consumption needs *1000
        COKE OVENS: Sp Heat Cons is in 10^6 kcal/t -> *1000 gives kcal/kg DC
        """
        multipliers = {
            "BM": {"specific_heat_consumption": 1000},
            "USM": {"specific_heat_consumption": 1000},
            "WRM": {"specific_heat_consumption": 1000},
            "COKE OVENS": {"specific_heat_coke_ovens": 1000, "bf_coke_yield": 100},
            # "Gross Coal to Hot Metal Ratio" (row 15 of "Maj Techno Summ") is
            # reported in Kg/THM; SAIL's convention for this parameter is a
            # small ~0.8-1.0 ratio (raw Kg/THM ÷ 1000), matching the
            # historical stored values. The map used to read this via a
            # "75/1000" string expression against row 75 of B-FCE, but that
            # row is absent in some files (verified: present in the Mar'26
            # reference file, missing in May'26, where row 75 is "Calendar
            # Hours" instead) — and even where present, the generic
            # row-expression evaluator has no way to tell "1000" is a
            # literal constant apart from a (non-existent) row 1000, so it
            # always silently evaluated to a division-by-zero. Later moved to
            # a dedicated "Coal to Hot Metal" sheet's row 20 (also stable, but
            # that sheet doesn't exist in some older archival files, and its
            # till_month/cumulative column carries a #REF! error baked into
            # the source template for the FY's first month). "Maj Techno
            # Summ" row 15 ("Gross Coal to Hot Metal Ratio") carries the
            # identical value, is present in every sample file checked
            # (2016-17 through 2026-27), and the /1000 still happens here
            # rather than in a row-expression string.
            "Maj Techno Summ": {"coal_to_hm": 0.001},
        }

        if sheet_name in multipliers and param_key in multipliers[sheet_name]:
            return multipliers[sheet_name][param_key]
        return 1.0

    def _extract_from_sheet(self, sheet_name: str, unit_name: str, unit_params: Dict) -> Dict:
        """Extract techno data from a single sheet for both month and till_month."""
        if sheet_name not in self.workbook.sheetnames:
            print(f"Sheet '{sheet_name}' not found")
            return None

        ws = self.workbook[sheet_name]
        header_row, header = self._find_month_column(ws)

        if header_row is None:
            print(f"Cannot find month headers in sheet '{sheet_name}'")
            return None

        # Find month column for report_month
        month_col = None
        if self.report_month:
            try:
                month_num = int(self.report_month.split('-')[1])
                target_abbr = _MONTH_NUM_TO_ABBR.get(month_num)
                for i, h in enumerate(header):
                    if target_abbr in h:
                        month_col = i
                        break
            except (ValueError, IndexError):
                pass

        if month_col is None:
            # Use last filled month
            for abbr in reversed(_MONTH_ABBRS):
                for i, h in enumerate(header):
                    if abbr in h:
                        month_col = i
                        break
                if month_col is not None:
                    break

        if month_col is None:
            print(f"Cannot find month column in sheet '{sheet_name}'")
            return None

        # Get the actual (ACT) column - ISP has Plan/Actual pairs
        month_col = self._get_actual_column(ws, month_col)
        print(f"Using actual column: {month_col}")

        # Calculate cumulative column based on ISP pattern
        cum_col = None
        try:
            month_num = int(self.report_month.split('-')[1])
            cum_offset = self._get_cum_column_offset(month_num)
            cum_col = month_col + cum_offset
            print(f"Cumulative offset for month {month_num}: +{cum_offset} -> column {cum_col}")
        except Exception as e:
            print(f"Warning calculating cum_col: {e}")

        print(f"Extracting from sheet '{sheet_name}', month_col={month_col}, cum_col={cum_col}")

        # Resolve every row this unit's params need by searching the WHOLE
        # sheet for each expected column-B label — see _resolve_unit_rows's
        # docstring for why this replaces a fixed-window nearby search.
        unit_row_map = self._resolve_unit_rows(ws, sheet_name, unit_name, unit_params)
        # Last-resort blanket shift, used only for list specs and any
        # expression-row token _resolve_unit_rows couldn't resolve by label
        # (no expected label configured, or not found anywhere in the sheet).
        unit_offset = self._unit_offset_from_map(unit_params, unit_row_map) or 0

        # Extract parameters from this sheet
        data = {"month": {}, "till_month": {}}

        for param_key, row_spec in unit_params.items():
            try:
                # Determine row number(s) to read
                if isinstance(row_spec, list):
                    # Average of several rows (e.g. 3 converters' utilisation
                    # figures) — a plain list of row numbers, not a string
                    # expression, so there's no ambiguity between "this digit
                    # sequence is a row number" and "this one is a literal
                    # divisor" (see _evaluate_row_expression's docstring).
                    shifted_rows = [r + unit_offset for r in row_spec]
                    def _avg_rows(col):
                        if col is None:
                            return None
                        vals = []
                        for r in shifted_rows:
                            row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
                            v = self._clean_value(row[col]) if col < len(row) else None
                            if v is not None:
                                vals.append(float(v))
                        return sum(vals) / len(vals) if vals else None
                    month_val = _avg_rows(month_col)
                    till_val = _avg_rows(cum_col)
                elif isinstance(row_spec, str) and any(op in row_spec for op in ['+', '-', '/', '*', '(', ')']):
                    # Expression - evaluate for both columns. Each row token
                    # is looked up individually in unit_row_map (keyed
                    # "param_key#token"); a token with no registered per-row
                    # label falls back to the unit's blanket offset unless a
                    # dedicated computed fallback applies (see
                    # _oven_count_fallback_value).
                    def value_lookup(rn, col, days, pk=param_key):
                        row = unit_row_map.get(f"{pk}#{rn}")
                        if row is not None:
                            return self._get_cell_value(ws, row, col)
                        computed = self._oven_count_fallback_value(
                            ws, sheet_name, unit_name, pk, str(rn), col, days)
                        if computed is not None:
                            return computed
                        return self._get_cell_value(ws, rn + unit_offset, col)
                    month_val = self._evaluate_row_expression(ws, row_spec, month_col, value_lookup)
                    till_val = self._evaluate_row_expression(ws, row_spec, cum_col, value_lookup) if cum_col else None
                else:
                    # Simple row number (int, or numeric string)
                    row_num = unit_row_map.get(param_key, int(row_spec))
                    row = list(ws.iter_rows(
                        min_row=row_num, max_row=row_num, values_only=True
                    ))[0]
                    month_val = row[month_col] if month_col < len(row) else None
                    till_val = row[cum_col] if cum_col and cum_col < len(row) else None

                # Clean values
                month_val = self._clean_value(month_val)
                till_val = self._clean_value(till_val)

                # Apply unit multipliers if needed
                multiplier = self._get_parameter_multiplier(sheet_name, param_key)
                if multiplier != 1.0:
                    if month_val is not None:
                        month_val = float(month_val) * multiplier
                    if till_val is not None:
                        till_val = float(till_val) * multiplier
                    print(f"  Applied multiplier {multiplier}x to {param_key}")

                # Store both month and till_month
                data["month"][param_key] = month_val
                data["till_month"][param_key] = till_val

            except Exception as e:
                print(f"Warning: Could not read '{row_spec}' for {param_key} in sheet '{sheet_name}': {e}")

        return data

    def extract(self) -> List[Dict]:
        """Extract techno data from all mapped sheets."""
        self.open_workbook()

        # Auto-detect report month if not provided
        if not self.report_month:
            # Try to find month from first available sheet
            for sheet_name in list(self.hardcoded_map.keys())[:1]:
                if sheet_name in self.workbook.sheetnames:
                    ws = self.workbook[sheet_name]
                    _, header = self._find_month_column(ws)
                    if header:
                        for h in header:
                            for abbr in _MONTH_ABBRS:
                                if abbr in h:
                                    try:
                                        year_part = h.split("'")[-1]
                                        if year_part.isdigit() and len(year_part) == 2:
                                            year = 2000 + int(year_part)
                                        else:
                                            year = 2026
                                    except:
                                        year = 2026
                                    month_num = _MONTH_ABBR_TO_NUM.get(abbr, 4)
                                    self.report_month = f"{year}-{month_num:02d}"
                                    print(f"Auto-detected report month: {self.report_month}")
                                    break
                            if self.report_month:
                                break

        if not self.report_month:
            self.report_month = "2026-03"
            print(f"Using default report month: {self.report_month}")

        return self._extract_records_for_month()

    def _extract_records_for_month(self) -> List[Dict]:
        """Process every mapped sheet for self.report_month (already open,
        already set). Shared body for extract() and extract_for_month()."""
        records = []
        print(f"\n--- Starting ISP Techno Extraction ---\n")

        # Process each sheet in the mapping
        for sheet_name, sheet_units in self.hardcoded_map.items():
            print(f"\nProcessing sheet: '{sheet_name}'")

            for unit_name, unit_params in sheet_units.items():
                data = self._extract_from_sheet(sheet_name, unit_name, unit_params)

                if data and any(v is not None for v in data["month"].values()):
                    records.append({
                        "report_month": self.report_month,
                        "plant": "ISP",
                        "unit": unit_name,
                        "techno_json": data,
                    })
                    print(f"  OK Extracted: {unit_name}")
                else:
                    print(f"  -- No data for: {unit_name}")

        print(f"\nExtraction Completed. Total Records: {len(records)}")
        return records

    def extract_for_month(self, report_month: str) -> List[Dict]:
        """Extract this workbook's data for one specific FY month, without
        re-running auto-detection. Used by extract_available_months() to
        pull several months out of one cumulative workbook."""
        if self.workbook is None:
            self.open_workbook()
        self.report_month = report_month
        return self._extract_records_for_month()

    def extract_available_months(self, upto_month: str) -> Dict:
        """Extract every FY month from April through `upto_month` that this
        workbook's header row actually has a column for — lets one upload
        (e.g. the FY-end March closing file, or any later cumulative file)
        backfill/refresh every earlier month in one pass, since ISP revises
        earlier months' figures in-place before the FY closes.

        Returns {"months": [report_month...], "skipped_months": [...],
        "records_by_month": {report_month: [records...]}}. A candidate month
        the file simply doesn't have a column for (e.g. an older single-month
        upload) is skipped, not an error."""
        if self.workbook is None:
            self.open_workbook()

        candidate_months = _fy_month_sequence(upto_month)

        ref_sheet_name = next((s for s in self.hardcoded_map if s in self.workbook.sheetnames), None)
        if ref_sheet_name is None:
            raise ValueError(
                "No mapped sheet found in this workbook — verify this is "
                "the ISP Summarized Monthly Report."
            )
        ws = self.workbook[ref_sheet_name]
        _, header = self._find_month_column(ws)
        if not header:
            raise ValueError(f"Cannot find month headers in sheet '{ref_sheet_name}'.")

        included, skipped = [], []
        for rm in candidate_months:
            abbr = _MONTH_NUM_TO_ABBR[int(rm.split('-')[1])]
            if any(abbr in h for h in header):
                included.append(rm)
            else:
                skipped.append(rm)

        records_by_month: Dict[str, List[Dict]] = {}
        for rm in included:
            try:
                records_by_month[rm] = self.extract_for_month(rm)
            except Exception as e:
                print(f"Warning: could not extract {rm} from this file: {e}")
                skipped.append(rm)

        return {
            "months": list(records_by_month.keys()),
            "skipped_months": skipped,
            "records_by_month": records_by_month,
        }
