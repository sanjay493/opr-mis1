import calendar
import re
import sys
from datetime import datetime, time
from pathlib import Path
from typing import Dict, List, Optional

sys.path.insert(0, str(Path(__file__).parent))
from rsp_row_scan import find_p18_sheet, find_month_cum_columns, norm_label  # noqa: E402  (path set above)
from rsp_technopara_parser import clean_technopara_sheet  # noqa: E402
from rsp_technopara_sections import (  # noqa: E402
    SECTION_UNITS, PARAM_ALIASES, FURNACE_BLOCKS, PARAM_UNIT_FILTERS, DAILY_AVG_PARAMS,
)

# ---------------------------------------------------------------------------
# Month/year validation - the sheet has no single "this is the report month"
# cell (every FY month's header column exists as a fixed template regardless
# of how far the year has actually progressed, same "growing FY table"
# pattern as RSP's production page-9 sheet), so the report_month a caller
# passes in was previously trusted blindly with no cross-check at all.
# ---------------------------------------------------------------------------

_MONTH_ORDER = ["04", "05", "06", "07", "08", "09", "10", "11", "12", "01", "02", "03"]
_TECHNO_MONTH_DETECT_LABEL = "coke rate"
_TECHNO_MONTH_DETECT_LABEL2 = "shop"

# RSP's own sheet reports these two coke-oven byproduct yields at 1/10th the
# scale every other plant (and every manually-entered RSP value already in
# the DB) uses for the same unit — confirmed by comparing the sheet's own
# raw cell (~3.2, ~0.2-0.3) against manually-entered RSP figures for other
# months (~30-32, ~1-3), which are consistent with other plants' Crude Tar
# Yield/Ammonium Sulphate Yield range. Scaled here, once, centrally, rather
# than leaving every caller to remember to do it.
_PARAM_SCALE = {
    "crude_tar_yield": 10,
    "ammonium_sulphate_yield": 10,
}


def _techno_clean_value(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s in ("", "-", "--", "#DIV/0!", "#VALUE!"):
            return None
        try:
            return float(s)
        except ValueError:
            return None
    if isinstance(v, (int, float)):
        return v
    return None


def _detect_actual_report_month(ws) -> Optional[str]:
    """Which month's own column is the true LAST one (in FY order, scanned
    across the FULL year, never capped at the selected month) with a real
    non-zero value in a reliably-always-populated row ("Coke Rate SHOP") -
    the same technique used for RSP's production sheet. Scanning the whole
    FY rather than stopping at the selected month is what lets this catch
    mismatches in BOTH directions: a selected month later than the file's
    real content (later columns are genuinely blank) and — the case a capped
    scan would miss — a selected month earlier than the file's real content
    (real data exists past the selected column, since a capped scan would
    only ever "confirm" whatever was selected)."""
    target_row = None
    for r in range(1, ws.max_row + 1):
        for c in (1, 2):
            label = norm_label(ws.cell(r, c).value)
            if _TECHNO_MONTH_DETECT_LABEL in label and _TECHNO_MONTH_DETECT_LABEL2 in label:
                target_row = r
                break
        if target_row:
            break
    if target_row is None:
        return None

    last_found = None
    for m in _MONTH_ORDER:
        month_col, _cum_col = find_month_cum_columns(ws, m)
        if month_col is None:
            continue
        v = _techno_clean_value(ws.cell(target_row, month_col).value)
        if v:
            last_found = m
    return last_found


def _detect_fy_start_year(ws, max_row: int = 6, max_col: int = 400) -> Optional[int]:
    """FY start year (e.g. 2026 for '2026-27') from the sheet's own 'Norm
    <YYYY>-<YY>' header cell (e.g. 'Norm 2026-27')."""
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                m = re.search(r'norm\s+(\d{4})\s*-\s*\d{2,4}', v.lower())
                if m:
                    return int(m.group(1))
    return None


def _assert_techno_month_year_match(ws, report_month: str, month_num: str) -> None:
    """Raise ValueError if the uploaded file's own content disagrees with the
    user-selected month/year - never blocks upload just because a signal
    couldn't be detected at all, only when detection and selection actively
    disagree."""
    year = int(report_month[:4])

    detected_month = _detect_actual_report_month(ws)
    if detected_month and detected_month != month_num:
        raise ValueError(
            f"Month mismatch: the uploaded file's techno data appears to be for "
            f"month {detected_month}, but you selected month {month_num} "
            f"({report_month}). Please select the matching month, or upload the "
            f"correct file."
        )

    fy_start_year = _detect_fy_start_year(ws)
    if fy_start_year:
        expected_fy_start = year if int(month_num) >= 4 else year - 1
        if fy_start_year != expected_fy_start:
            raise ValueError(
                f"Year mismatch: the uploaded file is for FY {fy_start_year}-"
                f"{str(fy_start_year + 1)[2:]}, but your selected month/year "
                f"({report_month}) implies FY {expected_fy_start}-"
                f"{str(expected_fy_start + 1)[2:]}. Please verify the uploaded "
                f"file matches the selected period."
            )


def _ytd_days(year_i: int, month_i: int) -> int:
    """Total calendar days from April 1 of the financial year to end of report
    month (RSP's FY starts in April)."""
    fy_year = year_i if month_i >= 4 else year_i - 1
    total, y, m = 0, fy_year, 4
    while True:
        total += calendar.monthrange(y, m)[1]
        if y == year_i and m == month_i:
            break
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return total


def _clean_value(val):
    """Convert a raw cell value to a JSON-serializable number/None."""
    if isinstance(val, time):
        return None
    if isinstance(val, datetime):
        return None
    if isinstance(val, str):
        s = val.strip()
        if not s or s in {"#DIV/0!", "#VALUE!", "-", "--"}:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    if isinstance(val, (int, float)):
        return val
    return None


# Pre-normalize the registry once at import time so every lookup at runtime
# is a plain dict hit — no repeated re-normalization of the same keys.
_SECTION_UNITS_NORM = {norm_label(k): v for k, v in SECTION_UNITS.items()}
_PARAM_ALIASES_NORM = {norm_label(k): v for k, v in PARAM_ALIASES.items()}
_FURNACE_BLOCKS_NORM = {norm_label(k): v for k, v in FURNACE_BLOCKS.items()}
_PARAM_UNIT_FILTERS_NORM = {norm_label(k): norm_label(v) for k, v in PARAM_UNIT_FILTERS.items()}
_PROBE_LABELS = list(PARAM_ALIASES.keys()) + list(SECTION_UNITS.keys())


class TechnoExtractor:
    def __init__(self, excel_file: str, report_month: str = None):
        """
        Args:
            excel_file: Path to the RSP technopara Excel file.
            report_month: Report month in YYYY-MM format (e.g. "2026-05").
        """
        self.excel_file = Path(excel_file)
        self.report_month = report_month
        self.workbook = None
        self.ws = None

    def open_workbook(self):
        from openpyxl import load_workbook
        self.workbook = load_workbook(self.excel_file, data_only=True)

    def load_sheet(self):
        if self.workbook is None:
            self.open_workbook()
        sheet = find_p18_sheet(self.workbook.sheetnames)
        if sheet is None:
            raise Exception(
                "No page-1-8-style sheet found in workbook (expected a sheet name "
                "starting with 'page1-8'/'page-1-9', e.g. 'page-1-8', "
                "'PAGE-1-8 & 11,12', 'PAGE-1-9')."
            )
        self.ws = self.workbook[sheet]
        print(f"Loaded sheet: {sheet}")

    def _walk(self, df) -> Dict[str, Dict]:
        """Single top-to-bottom pass over the cleaned sheet. Tracks the
        current unit section and resolves every row via SECTION_UNITS /
        FURNACE_BLOCKS / PARAM_ALIASES — never by stored row number."""
        units: Dict[str, Dict] = {}

        def _set(unit, param_key, month_val, cum_val):
            if unit is None or param_key is None:
                return
            scale = _PARAM_SCALE.get(param_key)
            if scale:
                if isinstance(month_val, (int, float)):
                    month_val = round(month_val * scale, 4)
                if isinstance(cum_val, (int, float)):
                    cum_val = round(cum_val * scale, 4)
            slot = units.setdefault(unit, {"month": {}, "till_month": {}})
            # First occurrence wins — a couple of labels legitimately repeat
            # later in the sheet (e.g. "Make-Up Water Cons." appears twice,
            # once per unit-of-measure variant; only the first is wanted).
            if param_key in slot["month"]:
                return
            slot["month"][param_key] = month_val
            slot["till_month"][param_key] = cum_val

        current_unit = None
        n = len(df)
        for i in range(n):
            label_raw = df.at[i, "label"]
            label_norm = norm_label(label_raw)
            if not label_norm:
                continue

            if label_norm in _SECTION_UNITS_NORM:
                current_unit = _SECTION_UNITS_NORM[label_norm]
                continue

            if label_norm in _FURNACE_BLOCKS_NORM:
                param_key, offsets = _FURNACE_BLOCKS_NORM[label_norm]
                for offset, unit in offsets:
                    j = i + offset
                    if j < n:
                        mv = _clean_value(df.at[j, "month_val"])
                        cv = _clean_value(df.at[j, "cum_val"])
                        _set(unit, param_key, mv, cv)
                continue

            alias = _PARAM_ALIASES_NORM.get(label_norm)
            if alias is None:
                continue

            required_unit_str = _PARAM_UNIT_FILTERS_NORM.get(label_norm)
            if required_unit_str is not None:
                unit_str = norm_label(df.at[i, "unit_str"])
                if required_unit_str not in unit_str:
                    continue

            if isinstance(alias, tuple):
                unit, param_key = alias
            else:
                if current_unit is None:
                    continue
                unit, param_key = current_unit, alias

            mv = _clean_value(df.at[i, "month_val"])
            cv = _clean_value(df.at[i, "cum_val"])
            _set(unit, param_key, mv, cv)

        return units

    def extract(self) -> List[Dict]:
        self.load_sheet()

        month_num = None
        if self.report_month:
            try:
                month_num = f"{int(self.report_month.split('-')[1]):02d}"
            except (ValueError, IndexError):
                pass
        if not month_num:
            raise Exception(
                "report_month is required (YYYY-MM) — the technopara sheet has "
                "no reliable way to auto-detect which month column is current."
            )

        _assert_techno_month_year_match(self.ws, self.report_month, month_num)

        df = clean_technopara_sheet(self.ws, month_num, _PROBE_LABELS)

        year_i, month_i = int(self.report_month[:4]), int(month_num)
        days_in_month = calendar.monthrange(year_i, month_i)[1]
        ytd_days = _ytd_days(year_i, month_i)

        units = self._walk(df)

        records = []
        print("\n--- Starting RSP Techno Extraction ---\n")
        for unit_name, techno in units.items():
            for param_key in DAILY_AVG_PARAMS:
                mv = techno["month"].get(param_key)
                cv = techno["till_month"].get(param_key)
                if isinstance(mv, (int, float)):
                    techno["month"][param_key] = round(mv / days_in_month, 1)
                if isinstance(cv, (int, float)):
                    techno["till_month"][param_key] = round(cv / ytd_days, 1)

            if any(v is not None for v in techno["month"].values()):
                records.append({
                    "report_month": self.report_month,
                    "plant": "RSP",
                    "unit": unit_name,
                    "techno_json": techno,
                })
                print(f"  Extracted: {unit_name}")

        print(f"\nExtraction Completed. Total Records: {len(records)}")
        return records
