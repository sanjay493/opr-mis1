"""
BSP OISCO Excel extractor.

Sheet layout (Sheet1):
  Row 3, Col C : title containing month, e.g. "TECHNO ECONOMIC PARAMETERS_MAY'25"
  Row 6         : month headers — e.g. APR | MAY | CUM  (CUM = data_col + 1)
  Col C (3)     : parameter label
  Col D (4)     : unit

Row numbers come from bsp_oisco_map.json (editable without touching Python).
Records are stored in techno_data with plant='BSP'.
"""

import json
import re
from pathlib import Path
from openpyxl import load_workbook
from typing import Dict, List, Optional, Tuple

_MONTH_ABBR_TO_NUM: Dict[str, int] = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}
_MONTH_NUM_TO_ABBR = {v: k for k, v in _MONTH_ABBR_TO_NUM.items()}

_HEADER_ROW = 6     # row containing month abbreviation headers in OISCO file
_BAD = {"#DIV/0!", "#VALUE!", "-", "--", None, ""}

_MAP_PATH = Path(__file__).parent / "bsp_oisco_map.json"


def _load_map() -> Dict:
    try:
        with open(_MAP_PATH, "r", encoding="utf-8") as f:
            raw = json.load(f)
        return {k: v for k, v in raw.items() if not k.startswith("_")}
    except FileNotFoundError:
        raise FileNotFoundError(f"bsp_oisco_map.json not found at {_MAP_PATH}")


def _clean(v) -> Optional[float]:
    if v in _BAD:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in _BAD:
        return None
    try:
        return float(s.replace(",", ""))
    except (ValueError, TypeError):
        return None


def _detect_month_from_title(ws) -> Optional[int]:
    """Parse month number from title cell R3/C3, e.g. '…PARAMETERS_MAY'25'."""
    raw = str(ws.cell(3, 3).value or "").upper()
    m = re.search(r"_([A-Z]{3})'?\d{2}", raw)
    if m:
        return _MONTH_ABBR_TO_NUM.get(m.group(1))
    for abbr, num in _MONTH_ABBR_TO_NUM.items():
        if abbr in raw:
            return num
    return None


def _find_data_and_cum_col(ws, target_month_num: int) -> Tuple[int, int]:
    """
    Scan row 6 for the month abbreviation matching target_month_num.
    Returns (data_col, cum_col) as 1-based column indices.
    CUM is always the column immediately after the data column.
    """
    expected = _MONTH_NUM_TO_ABBR.get(target_month_num, "").upper()
    max_col = ws.max_column or 30
    for c in range(1, max_col + 5):
        v = str(ws.cell(_HEADER_ROW, c).value or "").strip().upper()
        if v == expected:
            return c, c + 1

    # Build helpful diagnostics
    found = [
        str(ws.cell(_HEADER_ROW, c).value or "").strip()
        for c in range(1, max_col + 1)
        if ws.cell(_HEADER_ROW, c).value
    ]
    raise ValueError(
        f"Month header '{expected}' not found in row {_HEADER_ROW}. "
        f"Found headers: {found}. Check that the file matches the selected month."
    )


class BspOiscoExtractor:
    def __init__(self, excel_file: str, report_month: Optional[str] = None):
        self.excel_file = Path(excel_file)
        self.report_month = report_month
        self._map = _load_map()

    def extract(self) -> List[Dict]:
        wb = load_workbook(self.excel_file, data_only=True)

        ws = None
        for name in wb.sheetnames:
            if name.strip().lower() in ("sheet1", "sheet 1"):
                ws = wb[name]
                break
        if ws is None:
            ws = wb.active

        # Resolve report month
        if self.report_month:
            y_str, m_str = self.report_month.split("-")
            month_num = int(m_str)
            report_month = self.report_month
        else:
            month_num = _detect_month_from_title(ws)
            if month_num is None:
                raise ValueError("Cannot detect month from title cell. Pass report_month explicitly.")
            import datetime
            report_month = f"{datetime.date.today().year}-{month_num:02d}"

        data_col, cum_col = _find_data_and_cum_col(ws, month_num)

        # Check if file month matches selected month (warn if extracting prior month)
        file_month_num = _detect_month_from_title(ws)
        if file_month_num and file_month_num != month_num:
            print(
                f"[BSP-OiscoExtractor] Warning: file title month={file_month_num} "
                f"but extracting month={month_num}. "
                "Cumulative in file is till file month, not extraction month."
            )

        records = []
        for unit_name, params in self._map.items():
            techno = {"month": {}, "till_month": {}}
            for param_key, row_num in params.items():
                try:
                    month_val = ws.cell(row_num, data_col).value
                    till_val  = ws.cell(row_num, cum_col).value
                    techno["month"][param_key]     = _clean(month_val)
                    techno["till_month"][param_key] = _clean(till_val)
                except Exception as e:
                    print(f"[BSP-OiscoExtractor] Warning: row {row_num} / {param_key}: {e}")

            if any(v is not None for v in techno["month"].values()):
                records.append({
                    "report_month": report_month,
                    "plant":        "BSP",
                    "unit":         unit_name,
                    "techno_json":  techno,
                })
                print(f"  Extracted: {unit_name}")

        print(f"[BSP-OiscoExtractor] {len(records)} units extracted for {report_month}")
        return records
