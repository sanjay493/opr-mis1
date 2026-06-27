"""
BSP 3-page-Tech.xlsx extractor.

Sheet layout (Sheet1):
  Row 3, Col A : month name, e.g. "MAY"
  Row 4, Cols D-O (1-based cols 4-15): APR … MAR
  Row 4, Col P (col 16): cumulative header ("ACTUAL" or "CUM")
  Col layout is FIXED — April always = col 4, May = col 5, … Mar = col 15, Cum = col 16

Row numbers come from bsp_techno_map.json (editable without touching Python).
Records are stored in techno_data with plant='BSP'.
"""

import json
from pathlib import Path
from openpyxl import load_workbook
from typing import Dict, List, Optional

_MONTH_NUM_TO_ABBR = {
    1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun',
    7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec',
}
_MONTH_ABBR_TO_NUM = {v: k for k, v in _MONTH_NUM_TO_ABBR.items()}

# Fixed column index (1-based) for each month in the BSP 3-page-tech format
_MONTH_NUM_TO_COL: Dict[int, int] = {
    4: 4,  5: 5,  6: 6,  7: 7,  8: 8,  9: 9,
    10: 10, 11: 11, 12: 12, 1: 13, 2: 14, 3: 15,
}
_CUM_COL = 16   # Column P = always cumulative (Apr → report month)

_BAD = {"#DIV/0!", "#VALUE!", "-", "--", None, ""}

_MAP_PATH = Path(__file__).parent / "bsp_techno_map.json"


def _load_map() -> Dict:
    try:
        with open(_MAP_PATH, "r", encoding="utf-8") as f:
            raw = json.load(f)
        return {k: v for k, v in raw.items() if not k.startswith("_")}
    except FileNotFoundError:
        raise FileNotFoundError(f"bsp_techno_map.json not found at {_MAP_PATH}")


def _clean(v) -> Optional[float]:
    if v in _BAD:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in _BAD:
        return None
    try:
        return float(s.replace(",", ""))
    except (ValueError, TypeError):
        return None


def _detect_report_month(ws, report_month: Optional[str]) -> str:
    """Return 'YYYY-MM' — prefer explicit parameter, fall back to cell A3 + FY guess."""
    if report_month:
        return report_month

    raw = str(ws.cell(3, 1).value or "").strip().upper()
    mon_num = None
    for abbr, num in _MONTH_ABBR_TO_NUM.items():
        if abbr.upper() in raw:
            mon_num = num
            break
    if mon_num is None:
        raise ValueError("Cannot detect month from cell A3. Pass report_month explicitly.")

    # Guess year from today — caller should always pass report_month anyway
    import datetime
    today = datetime.date.today()
    # If detected month ≥ April, assume current FY start year; else next calendar year
    year = today.year
    return f"{year}-{mon_num:02d}"


class BspTechnoExtractor:
    def __init__(self, excel_file: str, report_month: Optional[str] = None):
        self.excel_file = Path(excel_file)
        self.report_month = report_month
        self._map = _load_map()

    def extract(self) -> List[Dict]:
        wb = load_workbook(self.excel_file, data_only=True)

        # Find Sheet1
        ws = None
        for name in wb.sheetnames:
            if name.strip().lower() in ("sheet1", "sheet 1"):
                ws = wb[name]
                break
        if ws is None:
            ws = wb.active

        report_month = _detect_report_month(ws, self.report_month)
        month_num = int(report_month.split("-")[1])

        if month_num not in _MONTH_NUM_TO_COL:
            raise ValueError(f"Month {month_num} not in fixed column map")
        month_col = _MONTH_NUM_TO_COL[month_num]
        cum_col = _CUM_COL

        # Verify col header in row 4 matches expected month abbreviation
        expected_abbr = _MONTH_NUM_TO_ABBR.get(month_num, "").upper()
        actual_hdr = str(ws.cell(4, month_col).value or "").strip().upper()
        if actual_hdr != expected_abbr:
            print(
                f"[BSP-TechnoExtractor] Warning: expected '{expected_abbr}' in row 4 col {month_col}, "
                f"found '{actual_hdr}'. Verify bsp_techno_map.json row numbers."
            )

        records = []
        for unit_name, params in self._map.items():
            techno = {"month": {}, "till_month": {}}
            for param_key, row_num in params.items():
                try:
                    row_data = list(ws.iter_rows(
                        min_row=row_num, max_row=row_num, values_only=True
                    ))[0]
                    month_val = row_data[month_col - 1] if month_col - 1 < len(row_data) else None
                    till_val  = row_data[cum_col - 1]   if cum_col  - 1 < len(row_data) else None
                    techno["month"][param_key]     = _clean(month_val)
                    techno["till_month"][param_key] = _clean(till_val)
                except Exception as e:
                    print(f"[BSP-TechnoExtractor] Warning: row {row_num} / {param_key}: {e}")

            if any(v is not None for v in techno["month"].values()):
                records.append({
                    "report_month": report_month,
                    "plant":        "BSP",
                    "unit":         unit_name,
                    "techno_json":  techno,
                })
                print(f"  Extracted: {unit_name}")

        print(f"[BSP-TechnoExtractor] {len(records)} units extracted for {report_month}")
        return records
