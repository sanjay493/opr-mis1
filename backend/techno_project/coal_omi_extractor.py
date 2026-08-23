"""
"Coal OMI" Excel Extractor — pulls per-plant coking coal consumption and
SAIL-level receipt/consumption/stock figures from EMD's monthly
"Coal OMI - <Mon><YY>.xlsx" workbook (2 sheets, OIS-1 and OIS-2).

This is a separate, higher-precision source for the same 4 coal-consumption
keys (indigenous_pcc, indigenous_mcc, imported_hard_coal, imported_soft_coal)
that coal_co2_epi_extractor.py's PDF/docx path already writes to
techno_data (unit="General") — that older path reads whole numbers off a
PDF table and never populates till_month for these keys. This extractor
reads the workbook's actual decimal cell values directly, and till_month is
computed by summing DB-stored monthly values via techno_cumulative.py
(see plant_and_sail_techno_json / SAIL/till-month handling in
api_coal_omi_techno.py, which calls this module).

extract_ois1_detail reads OIS-1's full as-printed row (Total Coking Coal,
CDI Coal, and both blend% column groups, not just the 4 raw quantities) for
the "Consumption of Coking Coal and CDI Coal" display page — see its own
docstring below.

Sheet layouts (verified against 4 real files, Apr-Jul'26 — identical row
positions across all 4, safe to hardcode, but every read still verifies the
expected label text at that position first and raises ValueError naming the
actual cell contents on a mismatch, rather than silently trusting a
hardcoded position against a differently-laid-out file):

OIS-1 ("Consumption of Coking Coal and CDI Coal - <Mon>'YY") — per plant
  (BSP/DSP/RSP/BSL/ISP) plus a SAIL row, two consecutive rows each: the
  report month's own row (col C = e.g. "Jul'26"), then an "Apr-<Mon>'YY"
  FY-cumulative row directly below it. Column B holds the plant/SAIL label
  on the first of the two rows only. Value columns: D=PCC, E=MCC (Indigenous
  Coking Coal), G=Hard, H=Soft (Imported Coking Coal) — all '000 T. The
  sheet's own Blend% columns (N-S) are intentionally NOT read here — that's
  already computed correctly from these 4 raw quantities by
  page_key_parameters.py's _coal_blend_pct, once till_month has data; CDI
  Coal (col L) is out of scope for this feature.

OIS-2 ("Receipt, Consumption and Stocks of Coking Coal at Plants during
  <Mon>'YY") — SAIL-level only (no plant breakdown):
    Row 5/6/7 = Indigenous/Imported/Total: col D=Receipt Plan (TPD),
      E=Receipt Actual (TPD), I=Consumption Actual ('000 T), J=Consumption
      Average (TPD).
    Rows 10-13 and 15-18 = TWO fixed Category/Indigenous/Imported/Total
      stock blocks, not one ever-growing table: row 10's rolling series of
      month-end date cells (starting at column D, with a blank gap column
      roughly every 3rd column - an artifact of how the sheet is
      copy-pasted forward each month) holds roughly the first 6-7 months of
      its window; once that fills, a second identically-shaped block at row
      15 continues the SAME rolling window with later months (confirmed
      continuous and non-overlapping - including across a calendar-year
      wrap - across ~30 real sample files spanning 2022-2026; never a 3rd
      block). Which block holds report_month's own current-month column
      depends entirely on how far the FY has progressed by the time this
      file was produced (e.g. a Nov report's own column can land in row 15,
      not row 10) - _find_stock_cell/_extract_stock_history always scan
      both blocks together as one continuous sequence, never just the
      first. The YEAR on these date cells is unreliable (confirmed off by
      one in the sample data) and is never trusted, only the MONTH NUMBER,
      matched by walking outward from whichever column report_month's own
      month number reliably identifies. Rows 11/12/13 (block 1) or 16/17/18
      (block 2) = Indigenous/Imported/Total stock ('000 T) under whichever
      column.
    Rows 37+ (seen in some files: stale leftover data from an unrelated
    template, old dates, a completely different per-coal-type receipt
    breakdown) are never read.

Run as a script to dry-extract a folder of these files without touching the
DB:
    python coal_omi_extractor.py "D:\\opr-mis1\\Report_format\\Coal_co2"
"""
import re
import sys
from datetime import datetime
from pathlib import Path

PLANTS = ["BSP", "DSP", "RSP", "BSL", "ISP"]

COAL_KEY_UNITS = {
    "indigenous_pcc":     "'000 T",
    "indigenous_mcc":     "'000 T",
    "imported_hard_coal": "'000 T",
    "imported_soft_coal": "'000 T",
}

_OIS1_ROWS = {
    "BSP": (6, 7), "DSP": (9, 10), "RSP": (12, 13),
    "BSL": (15, 16), "ISP": (19, 20), "SAIL": (22, 23),
}
_OIS1_COLS = {
    "indigenous_pcc": 4, "indigenous_mcc": 5,
    "imported_hard_coal": 7, "imported_soft_coal": 8,
}
_OIS1_PLANT_COL = 2   # B
_OIS1_MONTH_COL = 3   # C

# Full as-printed row — every quantity + blend% column on OIS-1, verified
# against a live workbook (Coal OMI - Jul26.xlsx): D=PCC E=MCC F=Indigenous
# Total, G=Hard H=Soft I=Imported Total, J=Total Coking Coal, L=CDI Coal
# (K is a spacer column), N-P=Indigenous PCC/MCC/Total %, Q-S=Imported
# Hard/Soft/Total % (M is a spacer column before the Blend% block). Used
# for the "Consumption of Coking Coal and CDI Coal" display page, which
# renders these values verbatim rather than recomputing totals/blend% —
# see extract_ois1_detail.
_OIS1_DETAIL_QTY_COLS = {
    "pcc": 4, "mcc": 5, "indigenous_total": 6,
    "hard": 7, "soft": 8, "imported_total": 9,
    "total_coking_coal": 10, "cdi_coal": 12,
}
_OIS1_DETAIL_PCT_COLS = {
    "pcc_pct": 14, "mcc_pct": 15, "indigenous_total_pct": 16,
    "hard_pct": 17, "soft_pct": 18, "imported_total_pct": 19,
}

_OIS2_ROW = {"indigenous": 5, "imported": 6, "total": 7}
_OIS2_RECEIPT_PLAN_COL = 4   # D
_OIS2_RECEIPT_ACTUAL_COL = 5  # E
_OIS2_CONSUMPTION_ACTUAL_COL = 9   # I
_OIS2_CONSUMPTION_AVG_COL = 10  # J
_OIS2_STOCK_SCAN_COLS = range(4, 15)  # D..N — generous, blank cells skipped

# The sheet's month-wise stock table is a fixed 2-row-block layout, not a
# single ever-growing one: row 10's block holds roughly the first 6-7 months
# of its rolling window, then a second block starting at row 15 (same
# Category/Indigenous/Imported/Total shape) picks up wherever row 10 left
# off and continues the same rolling window. Confirmed continuous and
# non-overlapping (including across a calendar-year wrap, e.g. one sample
# file's row 10 ends at month 6 and row 15 continues 7,8,...,12,1) across
# every sample file checked (~30, spanning 2022-2026) — never a 3rd block.
# A report month can land in EITHER block depending on how far the FY has
# progressed (e.g. a Nov report's own current-month column is in row 15,
# not row 10, once row 10's ~7-month window has filled) - both blocks are
# always scanned together as one continuous chronological sequence, never
# just the first one.
_OIS2_STOCK_BLOCKS = [
    {"header_row": 10, "rows": {"indigenous": 11, "imported": 12, "total": 13}},
    {"header_row": 15, "rows": {"indigenous": 16, "imported": 17, "total": 18}},
]

_MONTH_ABBR = {1: "jan", 2: "feb", 3: "mar", 4: "apr", 5: "may", 6: "jun",
               7: "jul", 8: "aug", 9: "sep", 10: "oct", 11: "nov", 12: "dec"}


def mlabel_from_report_month(report_month: str) -> str:
    """"2026-07" -> "Jul'26" """
    year, mon_num = report_month.split("-")
    return f"{_MONTH_ABBR[int(mon_num)].capitalize()}'{year[-2:]}"


def cum_mlabel_from_report_month(report_month: str) -> str:
    """"2026-07" -> "Apr-Jul'26" (FY-cumulative row header on OIS-1)."""
    year, mon_num = report_month.split("-")
    mon_num = int(mon_num)
    fy_year = int(year) if mon_num >= 4 else int(year) - 1
    if mon_num == 4:
        return f"Apr'{str(fy_year)[-2:]}"
    return f"Apr-{mlabel_from_report_month(report_month)}"


def _cum_label_matches(found, expected: str) -> bool:
    """True if a sheet's own FY-cumulative row label means the same thing as
    expected (from cum_mlabel_from_report_month), tolerating cosmetic
    variance seen across older workbooks: extra/missing spaces around the
    dash ("Apr - Jul'25" vs "Apr-Jul'25"), and an explicit FY-start year
    tacked onto "Apr" when the range crosses a calendar year ("Apr'24 -
    Jan'25" vs "Apr-Jan'25"). Whitespace and the optional "'YY" right after
    "Apr" are stripped from both sides before comparing — the target
    month/year token itself still has to match exactly, so this can only
    accept labels that were already unambiguously correct, never a
    genuinely wrong month."""
    if not isinstance(found, str):
        return False
    norm = lambda s: re.sub(r"^Apr'\d{2}-", "Apr-", re.sub(r"\s+", "", s))
    return norm(found) == norm(expected)


def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def extract_ois1(path, report_month: str) -> dict:
    """-> {plant_or_SAIL: {"month": {4 keys}, "report_cumulative": {4 keys}}}
    for PLANTS + ["SAIL"]."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    if "OIS-1" not in wb.sheetnames:
        raise ValueError(f"'OIS-1' sheet not found — sheets present: {wb.sheetnames}")
    ws = wb["OIS-1"]

    mlabel = mlabel_from_report_month(report_month)
    cum_mlabel = cum_mlabel_from_report_month(report_month)

    # April is the FY's first month, so its cumulative trivially equals its
    # own month value — no separate cumulative row is meaningful. Confirmed
    # against a real sample file that April's own "cumulative" row is in
    # fact mislabeled ("2026-27" instead of "Apr'26") and holds stale
    # leftover values (matching what later becomes the FOLLOWING month's
    # Apr-<mon> cumulative, not April's own) — a sheet-template artifact,
    # not real April data. Skip reading/validating that row entirely for
    # April rather than trusting it.
    is_april = report_month.endswith("-04")

    out = {}
    for plant, (month_row, cum_row) in _OIS1_ROWS.items():
        plant_cell = ws.cell(month_row, _OIS1_PLANT_COL).value
        if plant_cell != plant:
            raise ValueError(
                f"OIS-1 row {month_row} col B expected '{plant}', found {plant_cell!r} — "
                f"sheet layout doesn't match what this extractor expects."
            )
        month_label = ws.cell(month_row, _OIS1_MONTH_COL).value
        if month_label != mlabel:
            raise ValueError(
                f"OIS-1 row {month_row} col C expected '{mlabel}' (from selected report_month "
                f"{report_month}), found {month_label!r} — check the selected month matches "
                f"this file."
            )

        month_vals = {k: _num(ws.cell(month_row, c).value) for k, c in _OIS1_COLS.items()}

        if is_april:
            cum_vals = dict(month_vals)
        else:
            cum_label = ws.cell(cum_row, _OIS1_MONTH_COL).value
            if not _cum_label_matches(cum_label, cum_mlabel):
                raise ValueError(
                    f"OIS-1 row {cum_row} col C expected '{cum_mlabel}', found {cum_label!r}."
                )
            cum_vals = {k: _num(ws.cell(cum_row, c).value) for k, c in _OIS1_COLS.items()}

        out[plant] = {"month": month_vals, "report_cumulative": cum_vals}

    return out


def extract_ois1_detail(path, report_month: str) -> dict:
    """-> {plant_or_SAIL: {"month": {...14 fields, label}, "till_month":
    {...14 fields, label}}} for PLANTS + ["SAIL"] — every quantity and
    blend% column on OIS-1, read verbatim (blend% converted from the
    sheet's raw 0-1 fraction to a 0-100 percentage, rounded to 1dp to match
    what's printed). "till_month" (not "ytd") to match
    db.merge_upsert_techno_data's hardcoded ("month", "till_month") period
    pair — it merges only those two keys, so a different key name would
    silently vanish on any re-insert that merges into an existing row.
    Powers the "Consumption of Coking Coal and CDI Coal" display page,
    which renders this as-is with no recomputation — including the SAIL
    row, which is the sheet's own printed row rather than a sum of the 5
    plants (unlike extract_ois1/_build_plant_records's validation-focused
    path, this one has no cross-check to fail on a mismatch).
    Reuses the same row/column positions and April-has-no-YTD-row handling
    as extract_ois1 — see that function's docstring for the sheet-layout
    verification rationale."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    if "OIS-1" not in wb.sheetnames:
        raise ValueError(f"'OIS-1' sheet not found — sheets present: {wb.sheetnames}")
    ws = wb["OIS-1"]

    mlabel = mlabel_from_report_month(report_month)
    cum_mlabel = cum_mlabel_from_report_month(report_month)
    is_april = report_month.endswith("-04")

    def _row_dict(row, label):
        d = {"label": label}
        for k, c in _OIS1_DETAIL_QTY_COLS.items():
            d[k] = _num(ws.cell(row, c).value)
        for k, c in _OIS1_DETAIL_PCT_COLS.items():
            v = _num(ws.cell(row, c).value)
            d[k] = round(v * 100, 1) if v is not None else None
        return d

    out = {}
    for plant, (month_row, cum_row) in _OIS1_ROWS.items():
        plant_cell = ws.cell(month_row, _OIS1_PLANT_COL).value
        if plant_cell != plant:
            raise ValueError(
                f"OIS-1 row {month_row} col B expected '{plant}', found {plant_cell!r}."
            )
        month_label = ws.cell(month_row, _OIS1_MONTH_COL).value
        if month_label != mlabel:
            raise ValueError(
                f"OIS-1 row {month_row} col C expected '{mlabel}', found {month_label!r}."
            )
        month_detail = _row_dict(month_row, mlabel)

        if is_april:
            till_month_detail = dict(month_detail)
        else:
            cum_label = ws.cell(cum_row, _OIS1_MONTH_COL).value
            if not _cum_label_matches(cum_label, cum_mlabel):
                raise ValueError(
                    f"OIS-1 row {cum_row} col C expected '{cum_mlabel}', found {cum_label!r}."
                )
            till_month_detail = _row_dict(cum_row, cum_mlabel)

        out[plant] = {"month": month_detail, "till_month": till_month_detail}

    return out


def _populated_stock_cells(ws):
    """-> [(block_idx, col, month_num), ...] for every populated date cell
    across BOTH stock blocks (see _OIS2_STOCK_BLOCKS), block 0's columns
    first then block 1's — confirmed to always be one continuous
    chronological sequence (including across a calendar-year wrap) with no
    overlap between the two blocks, across ~30 real sample files spanning
    2022-2026. The YEAR on each date cell is unreliable (confirmed
    off-by-one in real files) - only the MONTH NUMBER is ever trusted."""
    out = []
    for block_idx, block in enumerate(_OIS2_STOCK_BLOCKS):
        for col in _OIS2_STOCK_SCAN_COLS:
            v = ws.cell(block["header_row"], col).value
            if isinstance(v, datetime):
                out.append((block_idx, col, v.month))
    return out


def _find_stock_cell(ws, report_month: str):
    """Scan both stock blocks (a report month can land in either one,
    depending on how far this rolling-window sheet has progressed - see
    _OIS2_STOCK_BLOCKS) for whichever populated date cell's MONTH NUMBER
    matches report_month's, ignoring the unreliable year entirely.
    Returns (block_idx, col), or None if no match."""
    target_month = int(report_month[5:7])
    matches = [(b, c) for b, c, m in _populated_stock_cells(ws) if m == target_month]
    if not matches:
        return None
    # If more than one matches (shouldn't normally happen within one file),
    # prefer the last one — the more recently-added column is more likely
    # to be this file's own intended entry for the report month.
    return matches[-1]


def _stock_at_block_col(ws, block_idx, col) -> dict:
    rows = _OIS2_STOCK_BLOCKS[block_idx]["rows"]
    return {
        "indigenous": _num(ws.cell(rows["indigenous"], col).value),
        "imported": _num(ws.cell(rows["imported"], col).value),
        "total": _num(ws.cell(rows["total"], col).value),
    }


def _extract_stock_history(ws, report_month: str) -> dict:
    """-> {"YYYY-MM": {"indigenous","imported","total"}, ...} for EVERY
    populated stock column across both blocks, not just the one matching
    report_month — a single upload's sheet is a rolling multi-month view,
    so extracting all of them lets one upload backfill several FY months
    at once instead of only report_month's own point.

    Absolute calendar months are derived by walking outward (across the
    block boundary, if needed) from the report_month-matching cell (found
    via _find_stock_cell), decrementing/incrementing one real month per
    populated column as it steps left/right, stopping at the first
    populated column whose own month number doesn't match the expected
    next step (a real gap/anomaly, not blindly extrapolated past — same
    "stay silent rather than guess" rule the rest of this module
    follows)."""
    populated = _populated_stock_cells(ws)
    anchor_cell = _find_stock_cell(ws, report_month)
    if anchor_cell is None:
        return {}
    anchor_idx = populated.index((*anchor_cell, int(report_month[5:7])))

    anchor_y, anchor_m = int(report_month[:4]), int(report_month[5:7])
    result = {report_month: _stock_at_block_col(ws, *anchor_cell)}

    y, m, idx = anchor_y, anchor_m, anchor_idx
    while idx > 0:
        prev_block, prev_col, prev_month_num = populated[idx - 1]
        m -= 1
        if m == 0:
            m, y = 12, y - 1
        if prev_month_num != m:
            break
        result[f"{y}-{m:02d}"] = _stock_at_block_col(ws, prev_block, prev_col)
        idx -= 1

    y, m, idx = anchor_y, anchor_m, anchor_idx
    while idx < len(populated) - 1:
        next_block, next_col, next_month_num = populated[idx + 1]
        m += 1
        if m == 13:
            m, y = 1, y + 1
        if next_month_num != m:
            break
        result[f"{y}-{m:02d}"] = _stock_at_block_col(ws, next_block, next_col)
        idx += 1

    return result


def extract_ois2(path, report_month: str) -> dict:
    """-> {"receipt": {"indigenous"|"imported"|"total": {"plan","actual"}},
           "consumption": {..same 3 keys..: {"actual","avg"}},
           "stock": {"indigenous","imported","total", "as_of_month"},
           "stock_history": {"YYYY-MM": {"indigenous","imported","total"}, ...}}
    All SAIL-level (no plant breakdown in this sheet). "stock" is
    report_month's own single point (kept for the (A)/(B)/current-stock
    displays); "stock_history" is every month this same upload's sheet
    happens to also carry data for — see _extract_stock_history."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    if "OIS-2" not in wb.sheetnames:
        raise ValueError(f"'OIS-2' sheet not found — sheets present: {wb.sheetnames}")
    ws = wb["OIS-2"]

    receipt, consumption = {}, {}
    for category, row in _OIS2_ROW.items():
        label = ws.cell(row, 3).value  # col C
        if not label or category.lower() not in str(label).lower():
            raise ValueError(
                f"OIS-2 row {row} col C expected an '{category}' label, found {label!r}."
            )
        receipt[category] = {
            "plan": _num(ws.cell(row, _OIS2_RECEIPT_PLAN_COL).value),
            "actual": _num(ws.cell(row, _OIS2_RECEIPT_ACTUAL_COL).value),
        }
        consumption[category] = {
            "actual": _num(ws.cell(row, _OIS2_CONSUMPTION_ACTUAL_COL).value),
            "avg": _num(ws.cell(row, _OIS2_CONSUMPTION_AVG_COL).value),
        }

    stock_cell = _find_stock_cell(ws, report_month)
    if stock_cell is None:
        stock = {"indigenous": None, "imported": None, "total": None, "as_of_month": None}
    else:
        stock = {
            **_stock_at_block_col(ws, *stock_cell),
            # Report the report_month's own 1st, not the (unreliable) year
            # actually stored in the cell — the month number is all that
            # was trustworthy about that cell in the first place.
            "as_of_month": f"{report_month}-01",
        }

    stock_history = _extract_stock_history(ws, report_month)

    return {"receipt": receipt, "consumption": consumption, "stock": stock, "stock_history": stock_history}


def extract_coal_omi(path, report_month: str) -> dict:
    """-> {"ois1": extract_ois1(...), "ois2": extract_ois2(...),
    "ois1_detail": extract_ois1_detail(...)}"""
    return {
        "ois1": extract_ois1(path, report_month),
        "ois2": extract_ois2(path, report_month),
        "ois1_detail": extract_ois1_detail(path, report_month),
    }


_FNAME_RE = re.compile(r"Coal OMI\s*-\s*([A-Za-z]{3})(\d{2})\.xlsx$", re.IGNORECASE)
_MONTH_NUM = {v: k for k, v in _MONTH_ABBR.items()}


def report_month_from_filename(fname: str):
    """"Coal OMI - Jul26.xlsx" -> "2026-07", or None if unrecognized."""
    m = _FNAME_RE.search(fname)
    if not m:
        return None
    mon, yy = m.group(1).lower(), int(m.group(2))
    if mon not in _MONTH_NUM:
        return None
    return f"{2000 + yy}-{_MONTH_NUM[mon]:02d}"


if __name__ == "__main__":
    import json as _json
    folder_arg = sys.argv[1] if len(sys.argv) > 1 else r"D:\opr-mis1\Report_format\Coal_co2"
    for path in sorted(Path(folder_arg).glob("Coal OMI*.xlsx")):
        report_month = report_month_from_filename(path.name)
        if not report_month:
            print(f"skip (unrecognized filename): {path.name}")
            continue
        print(f"=== {path.name} -> {report_month} ===")
        try:
            blob = extract_coal_omi(str(path), report_month)
            print(_json.dumps(blob, indent=2))
        except ValueError as e:
            print(f"EXTRACTION ERROR: {e}")
