import json
from pathlib import Path
from openpyxl import load_workbook
from typing import Dict, List


class TechnoExtractor:
    def __init__(self, excel_file: str):
        self.excel_file = Path(excel_file)
        self.workbook = None
        self.ws = None
        self.report_month = None
        self.month_col = None
        self.cum_col = None
        self.hardcoded_map = self.load_hardcoded_map()

    def load_hardcoded_map(self) -> Dict:
        try:
            with open("hardcoded_map.json", "r", encoding="utf-8") as f:
                return json.load(f)
        except FileNotFoundError:
            print("Error: hardcoded_map.json not found!")
            return {}

    def open_workbook(self):
        self.workbook = load_workbook(self.excel_file, data_only=True)

    def load_sheet(self):
        if self.workbook is None:
            self.open_workbook()
        for sheet in self.workbook.sheetnames:
            norm = sheet.lower().replace(" ", "").replace("-", "")
            if norm in ["page18", "page1-8", "page-1-8"]:
                self.ws = self.workbook[sheet]
                print(f"✅ Loaded sheet: {sheet}")
                return
        raise Exception("Sheet not found.")

    def detect_month_column(self):
        header = [str(c.value).strip() if c.value else "" for c in self.ws[3]]
        months = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]

        for month in months:
            if month not in header:
                continue
            col = header.index(month)
            valid = sum(1 for r in range(5, 40) if isinstance(self.ws.cell(r + 1, col + 1).value, (int, float)))
            if valid > 5:
                self.month_col = col
                break

        if self.month_col is None:
            raise Exception("Cannot detect report month")
        self.cum_col = header.index("Cum.")

    def detect_report_month(self):
        try:
            fy = str(self.ws["AM2"].value)
            start_year = int(fy.split("-")[0])
        except:
            start_year = 2026
        # Assuming current month is Apr for now
        self.report_month = f"{start_year}04"

    def extract(self) -> List[Dict]:
        self.load_sheet()
        self.detect_month_column()
        self.detect_report_month()

        records = []

        print("\n--- Starting Hardcoded Extraction ---\n")

        for unit_name, params in self.hardcoded_map.items():
            techno = {"month": {}, "till_month": {}}

            for param_key, row_num in params.items():
                try:
                    row = list(self.ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]

                    month_val = row[self.month_col] if self.month_col < len(row) else None
                    till_val = row[self.cum_col] if self.cum_col and self.cum_col < len(row) else None

                    if month_val in ("#DIV/0!", "#VALUE!", "-", "--", None):
                        month_val = None
                    if till_val in ("#DIV/0!", "#VALUE!", "-", "--", None):
                        till_val = None

                    techno["month"][param_key] = month_val
                    techno["till_month"][param_key] = till_val
                except Exception as e:
                    print(f"Warning: Could not read row {row_num} for {param_key}")

            if techno["month"]:
                records.append({
                    "report_month": self.report_month,
                    "plant": "RSP",
                    "unit": unit_name,
                    "techno_json": techno
                })
                print(f"  Extracted: {unit_name}")

        print(f"\n✅ Extraction Completed. Total Records: {len(records)}")
        return records