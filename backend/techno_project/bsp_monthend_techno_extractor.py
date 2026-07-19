"""
BSP Month-End Techno Extractor — tentative for-the-month techno values from
BSP's two month-end reports. The file type is auto-detected, so both are
uploaded through the same row on the techno data-entry page:

1. BSP MIS-2 ("BSP MIS 2_coff_print*.xls", first sheet)
   - Report date in row 2:  'Date:' | day | month name | year (e.g. 31 MAY 2026)
   - Blast-furnace block around rows 48-55: furnace labels in column B
     (BF-4 / BF-6 / BF-7 / BF-8 / 'BF 1-8(TOTAL)' = shop), each parameter as
     an ON DT./CUM column pair — the CUM column is the for-the-month value:
       D=Production(t)  H=CDI  J=Sinter%  L=Pellet%  N=Nut Coke  P=Productivity
       R=O2 enrichment%

2. BSP PPC MIS ("BSP PPC MIS*.xls", sheets S1/S2 — same workbook the BSP
   production upload uses)
   - Report date: S1!N1 (Excel date serial, e.g. 30.06.2026)
   - S1 coke-rate block: furnace labels in column P (BF-4 / BF-6 / BF-7 /
     MAHAMAYA = BF-8 / 'BF-1 to 8' = shop), month value in column U ("cum")
   - S2 SMS blocks: 'STEEL MELTING SHOP-2' / 'SHOP-3' sections, rows labelled
     Hot Metal / Scrap / Total in column D, month value in column G ("Cum")
   - S2!L67 — 'Energy Rate (Gcal/TCS) … ondt/cuml' row, column L is the
     for-the-month value → specific_energy_consumption, unit 'General'

Rows are located by label, not fixed row numbers, so shifts between report
versions don't break extraction. Values go into techno_json["month"] only —
cumulative is computed separately via the shared rules.
"""

import sys
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional

sys.path.insert(0, str(Path(__file__).parent))
from dsp_mcr_techno_extractor import McrMonthMismatch, _clean_val  # noqa: E402

_MONTH_NAME_TO_NUM = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}

# ── MIS-2 blast-furnace block ────────────────────────────────────────────────
# label (column A or B, whichever holds it — has drifted between eras) →
# unit name (BF-1 / BF-5 are shut and not extracted)
_MIS2_UNIT_LABELS = {
    "BF-4": "BF-4",
    "BF-6": "BF-6",
    "BF-7": "BF-7",
    "BF-8": "BF-8",
    "BF 1-8": "BF_Shop",   # prefix match — actual label is 'BF 1-8(TOTAL)'
}
# canonical param key → header-text substrings (matched against the row
# containing 'FURNACES', case-insensitive). The CUM value sits one column to
# the right of wherever the descriptive header text is found — this holds in
# every era seen, even though the header's own column position drifts (the
# ~2018-19 reports track Coal-Tar injection instead of Sinter%/Pellet%/BF
# productivity/O2 enrichment, shifting everything after 'cdi' rightward by a
# different amount than current-era files). Matching by label text — instead
# of trusting a fixed column offset — avoids silently reading the wrong
# parameter's value when the column layout doesn't match what's expected.
_MIS2_PARAM_LABELS = [
    ("production",       ["PRODUCTION"]),
    ("cdi",              ["CDI COAL RATE"]),   # not 'CDI COAL INJ.' (tonnage)
    ("sinter_in_burden", ["SINTER"]),
    ("pellet_in_burden", ["PALLET", "PELLET"]),  # report spells it 'Pallets'
    ("nut_coke_rate",    ["NUT COKE"]),
    ("bf_productivity",  ["PRODUCTIVITY"]),
    ("o2_enrichment",    ["OXYGEN"]),
]

# ── PPC MIS S1 coke-rate block ───────────────────────────────────────────────
# column-P label (exact, upper) → unit name
_PPC_COKE_LABELS = {
    "BF-4":      "BF-4",
    "BF-6":      "BF-6",
    "BF-7":      "BF-7",
    "MAHAMAYA":  "BF-8",
    "BF-1 TO 8": "BF_Shop",   # NOT 'BF-1 to 7' (partial-shop row)
}
_PPC_LABEL_COL = 16   # column P (1-based)
_PPC_COKE_COL = 21    # column U — "cum"

# ── PPC MIS S2 SMS blocks ────────────────────────────────────────────────────
# column-D label → canonical param key (same keys the 3-page-Tech upload uses)
_PPC_SMS_PARAMS = {
    "HOT METAL": "specific_hm_consumption",
    "SCRAP":     "specific_scrap_consumption",
    "TOTAL":     "tmi",
}
_PPC_SMS_LABEL_COL = 4   # column D
_PPC_SMS_VALUE_COL = 7   # column G — "Cum"
_PPC_SMS_ANCHORS = [("STEEL MELTING SHOP-2", "SMS-2"),
                    ("STEEL MELTING SHOP-3", "SMS-3")]

# ── PPC MIS S2 specific energy consumption (fixed cell) ──────────────────────
_PPC_SEC_ROW = 67
_PPC_SEC_LABEL_COL = 5    # column E — 'Energy Rate (Gcal/TCS) ... ondt/cuml'
_PPC_SEC_LABEL_TEXT = "ENERGY RATE"
_PPC_SEC_VALUE_COL = 12   # column L — "cuml"


def _cell(grid, row_1b, col_1b):
    if row_1b < 1 or row_1b > len(grid):
        return None
    row = grid[row_1b - 1]
    if col_1b < 1 or col_1b > len(row):
        return None
    v = row[col_1b - 1]
    return None if v == "" else v


class BspMonthendTechnoExtractor:
    """
    Extract for-the-month techno parameters from either BSP month-end report.

    extract() returns the standard techno record list:
        [{"plant": "BSP", "report_month": "YYYY-MM", "unit": str,
          "techno_json": {"month": {...}, "till_month": {}}}]

    After extract():  .report_date, .warnings, .detected_type
    """

    def __init__(self, file_path: str, report_month: str = ""):
        self.file_path = file_path
        self.report_month = report_month
        self.report_date: Optional[date] = None
        self.warnings: List[str] = []
        self.detected_type: str = ""

    # -- workbook loading (binary .xls via xlrd, .xlsx via openpyxl) ----------

    def _load_sheets(self) -> Dict[str, List[List]]:
        with open(self.file_path, "rb") as f:
            magic = f.read(4)

        if magic == b"\xd0\xcf\x11\xe0":
            import xlrd
            wb = xlrd.open_workbook(self.file_path)
            self._xls_datemode = wb.datemode
            return {
                sh.name: [[sh.cell_value(r, c) for c in range(sh.ncols)]
                          for r in range(sh.nrows)]
                for sh in wb.sheets()
            }

        if magic[:2] == b"PK":
            import openpyxl
            wb = openpyxl.load_workbook(self.file_path, data_only=True, read_only=True)
            self._xls_datemode = None
            sheets = {
                name: [list(row) for row in wb[name].iter_rows(values_only=True)]
                for name in wb.sheetnames
            }
            wb.close()
            return sheets

        raise ValueError(
            "Unrecognised file format — expected a BSP MIS-2 or PPC MIS Excel "
            "file (.xls/.xlsx)."
        )

    def _verify_month(self):
        file_month = f"{self.report_date.year}-{self.report_date.month:02d}"
        if self.report_month and self.report_month != file_month:
            raise McrMonthMismatch(
                f"This BSP report is dated {self.report_date.strftime('%d.%m.%Y')} "
                f"(month {file_month}) but you selected {self.report_month}. "
                "Select the matching month or upload the correct file."
            )
        self.report_month = file_month

    # -- MIS-2 -----------------------------------------------------------------

    def _extract_mis2(self, grid) -> Dict[str, Dict]:
        # Date from the row containing 'Date:' → day | month name | year
        self.report_date = None
        for r in range(1, 7):
            row = grid[r - 1] if r <= len(grid) else []
            for c, v in enumerate(row):
                if isinstance(v, str) and v.strip().rstrip(":").upper() == "DATE":
                    rest = [x for x in row[c + 1:] if x not in ("", None)][:3]
                    try:
                        day = int(float(rest[0]))
                        mon = _MONTH_NAME_TO_NUM[str(rest[1]).strip().upper()[:3]]
                        year = int(float(rest[2]))
                        self.report_date = date(year, mon, day)
                    except (IndexError, KeyError, ValueError, TypeError):
                        pass
                    break
            if self.report_date:
                break
        if not self.report_date:
            raise ValueError(
                "Cannot read the report date from the MIS-2 header (row 2 "
                "should contain 'Date:' day MONTH year, e.g. Date: 31 MAY 2026)."
            )
        self._verify_month()

        # Locate the furnace-block header row ('FURNACES', in column A or B —
        # has drifted between eras) and, on that same row, find each param's
        # CUM column by its descriptive header text (see _MIS2_PARAM_LABELS).
        header_row = None
        for r in range(1, len(grid) + 1):
            for c in (1, 2):
                if str(_cell(grid, r, c) or "").strip().upper() == "FURNACES":
                    header_row = r
                    break
            if header_row is not None:
                break
        if header_row is None:
            raise ValueError("Cannot find the 'FURNACES' header row in the MIS-2 report.")

        header = grid[header_row - 1]
        param_cols: Dict[str, int] = {}   # canonical key -> 1-based CUM column
        for key, substrings in _MIS2_PARAM_LABELS:
            for c, v in enumerate(header, start=1):
                text = str(v or "").strip().upper()
                if text and any(s in text for s in substrings):
                    param_cols[key] = c + 1   # CUM sits one column right of the label
                    break

        # Furnace labels sit in whichever of column A/B actually holds them
        # for this era.
        label_col = None
        for c in (2, 1):
            for r in range(header_row + 1, min(header_row + 20, len(grid)) + 1):
                v = str(_cell(grid, r, c) or "").strip().upper()
                if any(v.startswith(p.upper()) for p in _MIS2_UNIT_LABELS):
                    label_col = c
                    break
            if label_col is not None:
                break
        if label_col is None:
            raise ValueError("Cannot find the furnace label column in the MIS-2 report.")

        units: Dict[str, Dict] = {}
        found = set()
        for r in range(header_row + 1, len(grid) + 1):
            label = str(_cell(grid, r, label_col) or "").strip()
            for prefix, unit in _MIS2_UNIT_LABELS.items():
                if label.upper().startswith(prefix.upper()):
                    # exact furnace labels must not prefix-match each other
                    if prefix.startswith("BF-") and label.upper() != prefix.upper():
                        continue
                    month = {
                        key: _clean_val(_cell(grid, r, col))
                        for key, col in param_cols.items()
                    }
                    # Some older reports carry a blank placeholder row for a
                    # furnace not yet commissioned (production '-', not a
                    # real 0) ahead of — or instead of — the real data row;
                    # skip it so a later genuine row (or 'not found') wins.
                    if "production" in month and month["production"] is None:
                        continue
                    found.add(unit)
                    units[unit] = {"month": month, "till_month": {}}
        for prefix, unit in _MIS2_UNIT_LABELS.items():
            if unit not in found:
                self.warnings.append(f"MIS-2: furnace row '{prefix}' not found — skipped.")
        return units

    # -- PPC MIS ---------------------------------------------------------------

    def _extract_ppc(self, sheets) -> Dict[str, Dict]:
        s1 = sheets.get("S1") or sheets.get("s1")
        s2 = sheets.get("S2") or sheets.get("s2")
        if s1 is None:
            raise ValueError("PPC MIS workbook has no 'S1' sheet.")

        # Date lives at N1 (col 14) from ~2021-Mar onward, but older (~2012-2020)
        # reports have it one column to the left (M1, col 13) — scan the same
        # plausible-column window the production extractor uses (see
        # excel_extractor_bsp._detect_ppc_date) instead of trusting a fixed cell.
        n1 = None
        for c in range(10, 17):
            v = _cell(s1, 1, c)
            if isinstance(v, float) and 25000 < v < 55000:
                n1 = v
                break
            if isinstance(v, date):
                n1 = v
                break
        try:
            if n1 is None:
                raise TypeError
            if self._xls_datemode is not None:
                import xlrd
                y, m, d, *_ = xlrd.xldate_as_tuple(float(n1), self._xls_datemode)
                self.report_date = date(y, m, d)
            else:  # openpyxl gives a datetime directly
                self.report_date = n1.date() if hasattr(n1, "date") else n1
            if not isinstance(self.report_date, date):
                raise TypeError
        except Exception:
            raise ValueError(
                f"Cannot read the report date from S1 row 1 (got {n1!r}) — "
                "verify this is the BSP PPC MIS workbook."
            )
        self._verify_month()

        units: Dict[str, Dict] = {}

        def put(unit, key, value):
            units.setdefault(unit, {"month": {}, "till_month": {}})
            units[unit]["month"][key] = value

        # S1 — coke rate per furnace (column P label, column U cum value)
        found = set()
        for r in range(1, len(s1) + 1):
            label = str(_cell(s1, r, _PPC_LABEL_COL) or "").strip().upper()
            unit = _PPC_COKE_LABELS.get(label)
            if unit and unit not in found:
                found.add(unit)
                put(unit, "coke_rate", _clean_val(_cell(s1, r, _PPC_COKE_COL)))
        for label, unit in _PPC_COKE_LABELS.items():
            if unit not in found:
                self.warnings.append(f"PPC S1: coke-rate row '{label}' not found — skipped.")

        # S2 — SMS-2 / SMS-3 blocks (column D label, column G cum value)
        if s2 is None:
            self.warnings.append("PPC MIS workbook has no 'S2' sheet — SMS data skipped.")
            return units
        for anchor_text, unit in _PPC_SMS_ANCHORS:
            anchor_row = None
            for r in range(1, len(s2) + 1):
                if any(isinstance(v, str) and anchor_text in v.upper()
                       for v in s2[r - 1]):
                    anchor_row = r
                    break
            if anchor_row is None:
                self.warnings.append(f"PPC S2: '{anchor_text}' section not found — skipped.")
                continue
            remaining = dict(_PPC_SMS_PARAMS)
            for r in range(anchor_row + 1, min(anchor_row + 10, len(s2)) + 1):
                label = str(_cell(s2, r, _PPC_SMS_LABEL_COL) or "").strip().upper()
                key = remaining.pop(label, None)
                if key:
                    put(unit, key, _clean_val(_cell(s2, r, _PPC_SMS_VALUE_COL)))
                if not remaining:
                    break
            for label in remaining:
                self.warnings.append(
                    f"PPC S2 {unit}: row '{label.title()}' not found — skipped.")

        # Specific energy consumption — fixed cell S2!L67 ('Energy Rate' row)
        sec_label = str(_cell(s2, _PPC_SEC_ROW, _PPC_SEC_LABEL_COL) or "").strip().upper()
        if _PPC_SEC_LABEL_TEXT in sec_label:
            sec = _clean_val(_cell(s2, _PPC_SEC_ROW, _PPC_SEC_VALUE_COL))
            if sec is not None:
                put("General", "specific_energy_consumption", sec)
            else:
                self.warnings.append(
                    "PPC S2: specific energy consumption (cell L67) is blank — skipped.")
        else:
            self.warnings.append(
                f"PPC S2: expected 'Energy Rate' label in E{_PPC_SEC_ROW} — "
                f"found {sec_label!r}; specific energy consumption skipped.")
        return units

    # -- main ------------------------------------------------------------------

    def extract(self) -> List[Dict]:
        sheets = self._load_sheets()
        names_upper = {n.upper() for n in sheets}

        if {"S1", "S2"} & names_upper:
            self.detected_type = "BSP PPC MIS"
            units = self._extract_ppc(sheets)
        else:
            first = next(iter(sheets.values()))
            header = " ".join(
                str(v) for v in (first[1] if len(first) > 1 else []) if v
            ).upper()
            if "BSP MIS-2" not in header:
                raise ValueError(
                    "File not recognised as a BSP month-end report — expected "
                    "the PPC MIS workbook (sheets S1/S2) or MIS-2 ('BSP MIS-2' "
                    "in row 2)."
                )
            self.detected_type = "BSP MIS-2"
            units = self._extract_mis2(first)

        records = []
        for unit, tj in units.items():
            if any(v is not None for v in tj["month"].values()):
                records.append({
                    "plant": "BSP",
                    "report_month": self.report_month,
                    "unit": unit,
                    "techno_json": tj,
                })
        if not records:
            raise ValueError(
                f"No techno values found in the {self.detected_type or 'BSP'} "
                "report — verify the file contents."
            )
        return records
