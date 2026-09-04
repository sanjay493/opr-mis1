"""
RSP Blast-Furnace "GLANCE" Extractor — final/authoritative monthly BF techno
values from the RSP BF Department's own workbook (Report_format/Monthly/
RSP/GLANCE -<Mon>'<YY>.xlsx). This is a DIFFERENT file from the plant-wide
RSP Technopara sheet (rsp_technopara_extractor.py's "page1-8" workbook) —
the BF Department's own file carries more BF detail (furnace-wise
availability/utilisation, carbon rate, flue-dust rate, hot-metal/slag
chemistry) than the plant-wide sheet's Blast-Furnace section, so it's
treated as an equally "final" (not tentative) source and merged into the
same techno_data rows via merge_upsert_techno_data — same standing as the
Technopara sheet itself, just contributing a different/overlapping set of
BF-unit parameters.

Two sheet families are read:

  DETAIL — one row-block per parameter (label in column A), one column per
    FY month (APR..MAR, header row 1) plus a running "YEAR" column. Below
    each block's header row sit that parameter's furnace rows, labelled
    F1/F4/F5/SH in column A (not every block has all four — e.g. "TAP TIME
    DUR." only has F1/F4/F5). This is the primary source: it gives "month"
    for ANY already-filled FY month (Apr..whatever's latest), which is what
    makes backfilling an earlier month possible. Its YEAR column, however,
    is a running total that always reflects whichever month is LATEST in
    the file — not a per-month cumulative — so it's only trustworthy as
    "till_month" when the selected report month IS that latest month
    (checked via _detect_latest_month below); for an earlier backfilled
    month, till_month is left alone.

  <Mon><YY> (e.g. "JUL26") — one sheet per month, a fixed single-month
    report (SL.NO/DESCRIPTION in columns A/B, FCE-I/FCE-IV/FCE-V/TOTAL in
    columns C-F). No cumulative column at all. Only used to supply the
    handful of parameters DETAIL doesn't carry at all (Availability %,
    Utilisation %, Carbon Rate, Flue Dust Rate) — month value only. Sheet
    name is derived from the selected report_month (e.g. 2026-07 ->
    "JUL26"); if that month's sheet doesn't exist yet (a future month not
    filled in by the BF Department), these extra params are simply skipped
    with a warning rather than failing the whole extraction.

Param keys:
  Already tracked for RSP BF units via the Technopara sheet (this extractor
  UPDATES them with the BF Department's own, typically more precise,
  figures): bf_productivity, coke_rate, cdi, nut_coke_rate, fuel_rate,
  sinter_in_burden, pellet_in_burden, slag_rate, o2_enrichment,
  hot_blast_temp, silicon_in_hm. (The Technopara sheet only carries
  sinter_in_burden/pellet_in_burden/slag_rate at BF_Shop level — this
  source gives them per-furnace too.)

  Not tracked anywhere for RSP before (ADDED here, canonical names matching
  the same keys BSL/ISP already use for the identical measurement, so nothing
  new to teach the rest of the app): sulphur_in_hm, avg_hot_metal_temperature,
  slag_al2o3, slag_mgo, slag_basicity, co_co2_ratio, carbon_rate,
  f_dust_rate. Two genuinely new keys with no prior-plant precedent:
  furnace_availability / furnace_utilisation (paired, BF-department specific;
  furnace_availability at least mirrors ISP's own "furnace_availability").
"""

import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

sys.path.insert(0, str(Path(__file__).parent))
from rsp_row_scan import norm_label  # noqa: E402
from dsp_mcr_techno_extractor import _clean_val  # noqa: E402

# ---------------------------------------------------------------------------
# DETAIL sheet
# ---------------------------------------------------------------------------

_MONTH_ORDER = ["04", "05", "06", "07", "08", "09", "10", "11", "12", "01", "02", "03"]
_MONTH_ABBR_TO_NUM = {
    "APR": "04", "MAY": "05", "JUN": "06", "JUL": "07", "AUG": "08", "SEP": "09",
    "OCT": "10", "NOV": "11", "DEC": "12", "JAN": "01", "FEB": "02", "MAR": "03",
}
_MONTH_NUM_TO_ABBR3 = {v: k for k, v in _MONTH_ABBR_TO_NUM.items()}
_MONTH_NUM_TO_FULL = {
    "01": "JANUARY", "02": "FEBRUARY", "03": "MARCH", "04": "APRIL", "05": "MAY",
    "06": "JUNE", "07": "JULY", "08": "AUGUST", "09": "SEPTEMBER", "10": "OCTOBER",
    "11": "NOVEMBER", "12": "DECEMBER",
}

_FURNACE_ROW_LABELS = {"f1": "BF-1", "f4": "BF-4", "f5": "BF-5", "sh": "BF_Shop"}

# DETAIL column-A header (normalized) -> param_key
_DETAIL_PARAMS = {
    "hot metal prodn":        "production",
    "productivity on w.v.":   "bf_productivity",
    "hot blast temp":         "hot_blast_temp",
    "coke rate":               "coke_rate",
    "cdi rate":                "cdi",
    "nut coke rate":           "nut_coke_rate",
    "fuel rate":                "fuel_rate",
    "% sinter in burd (dry)": "sinter_in_burden",
    "% pellet in burd (dry)": "pellet_in_burden",
    "slag rate":                "slag_rate",
    "slag alumina":            "slag_al2o3",
    "slag mgo":                 "slag_mgo",
    "slag basicity":           "slag_basicity",
    "% si in hot metal":      "silicon_in_hm",
    "% s in hot metal":       "sulphur_in_hm",
    "hm temperature":         "avg_hot_metal_temperature",
    "% oxygen enrichment":    "o2_enrichment",
    "co / co2 ratio":         "co_co2_ratio",
}
# Reliably-always-populated row used to auto-detect the latest FY month the
# file actually has data for (same technique rsp_technopara_extractor.py
# uses for the plant-wide sheet).
_LATEST_MONTH_DETECT_HEADER = "hot metal prodn"

# ---------------------------------------------------------------------------
# <Mon><YY> single-month report sheet
# ---------------------------------------------------------------------------

_MONTH_SHEET_FURNACE_COLS = {3: "BF-1", 4: "BF-4", 5: "BF-5", 6: "BF_Shop"}
_MONTH_SHEET_LABEL_COL = 2   # column B

# column-B label (normalized) -> param_key — params NOT available in DETAIL.
_MONTH_SHEET_PARAMS = {
    "availability (%)":               "furnace_availability",
    "utilisation (%)":                "furnace_utilisation",
    "carbon rate":                     "carbon_rate",
    "flue dust recovered (kg/thm)":   "f_dust_rate",
}


def _find_month_year_columns(ws, header_row: int = 1, max_col: int = 40) -> Tuple[Dict[str, int], Optional[int]]:
    """DETAIL's header row 1: ITEMS/ABP/APR/MAY/.../MAR/YEAR. Returns
    ({month_num: col}, year_col)."""
    month_col: Dict[str, int] = {}
    year_col = None
    for c in range(1, max_col + 1):
        v = ws.cell(header_row, c).value
        if v is None:
            continue
        s = str(v).strip().upper()
        if s in _MONTH_ABBR_TO_NUM and s not in month_col:
            month_col[_MONTH_ABBR_TO_NUM[s]] = c
        elif s == "YEAR" and year_col is None:
            year_col = c
    return month_col, year_col


def _detect_latest_month(ws, header_row: int) -> Optional[str]:
    """Last FY month (Apr..Mar order) with a non-zero value in the 'SH'
    (BF_Shop) row of the Hot Metal Prodn block — the file's real latest
    month, independent of which columns merely exist as blank template."""
    sh_row = None
    for r in range(header_row + 1, header_row + 6):
        if norm_label(ws.cell(r, 1).value) == "sh":
            sh_row = r
            break
    if sh_row is None:
        return None
    month_col, _year_col = _find_month_year_columns(ws)
    last_found = None
    for m in _MONTH_ORDER:
        col = month_col.get(m)
        if col is None:
            continue
        v = _clean_val(ws.cell(sh_row, col).value)
        if v:
            last_found = m
    return last_found


def _read_furnace_block(ws, header_row: int, month_col: int, year_col: Optional[int]) -> Dict[str, Tuple]:
    """Rows immediately below a DETAIL header row, labelled F1/F4/F5/SH in
    column A (variable length — stops at the first row whose label isn't
    one of those). Returns {unit: (month_val, year_val)}."""
    out: Dict[str, Tuple] = {}
    r = header_row + 1
    while True:
        label = norm_label(ws.cell(r, 1).value)
        unit = _FURNACE_ROW_LABELS.get(label)
        if unit is None:
            break
        mv = _clean_val(ws.cell(r, month_col).value)
        yv = _clean_val(ws.cell(r, year_col).value) if year_col else None
        out[unit] = (mv, yv)
        r += 1
    return out


def _assert_fy_year_match(ws, report_month: str, month_num: str, month_col: Dict[str, int]) -> None:
    """Raise ValueError if DETAIL's own FY-year header (row 2, the numeric
    calendar year under the April column) disagrees with the selected
    report_month's implied FY — never blocks upload just because the year
    header couldn't be found, only on an active disagreement."""
    apr_col = month_col.get("04")
    if not apr_col:
        return
    fy_year_cell = ws.cell(2, apr_col).value
    if not isinstance(fy_year_cell, (int, float)):
        return
    fy_start_year = int(fy_year_cell)
    year = int(report_month[:4])
    expected_fy_start = year if int(month_num) >= 4 else year - 1
    if fy_start_year != expected_fy_start:
        raise ValueError(
            f"Year mismatch: the GLANCE workbook's DETAIL sheet is for FY "
            f"{fy_start_year}-{str(fy_start_year + 1)[2:]}, but your selected "
            f"month/year ({report_month}) implies FY {expected_fy_start}-"
            f"{str(expected_fy_start + 1)[2:]}. Please verify the uploaded "
            f"file matches the selected period."
        )


def _extract_detail(ws, report_month: str, month_num: str, units: Dict[str, Dict], warnings: List[str]) -> None:
    month_col, year_col = _find_month_year_columns(ws)
    if month_num not in month_col:
        raise ValueError(
            f"DETAIL sheet has no column for month {month_num} of {report_month} "
            "— verify the uploaded GLANCE workbook covers this month."
        )
    _assert_fy_year_match(ws, report_month, month_num, month_col)
    mcol = month_col[month_num]

    latest_header_row = None
    seen_headers = set()
    max_row = ws.max_row
    for r in range(1, max_row + 1):
        label = norm_label(ws.cell(r, 1).value)
        if not label or label in _FURNACE_ROW_LABELS:
            continue
        param_key = _DETAIL_PARAMS.get(label)
        if label == _LATEST_MONTH_DETECT_HEADER and latest_header_row is None:
            latest_header_row = r
        if param_key is None or param_key in seen_headers:
            continue
        seen_headers.add(param_key)
        block = _read_furnace_block(ws, r, mcol, year_col)
        for unit, (mv, yv) in block.items():
            slot = units.setdefault(unit, {"month": {}, "till_month": {}})
            slot["month"][param_key] = mv

    if latest_header_row is not None:
        latest_month = _detect_latest_month(ws, latest_header_row)
        if latest_month == month_num and year_col:
            # Running YEAR column is only a valid cumulative when the
            # selected month IS the file's latest filled month. Second pass
            # over the same headers (cheap — DETAIL is a small sheet) to
            # fill till_month from that column.
            seen2 = set()
            for r in range(1, max_row + 1):
                label = norm_label(ws.cell(r, 1).value)
                if not label or label in _FURNACE_ROW_LABELS:
                    continue
                param_key = _DETAIL_PARAMS.get(label)
                if param_key is None or param_key in seen2:
                    continue
                seen2.add(param_key)
                block = _read_furnace_block(ws, r, mcol, year_col)
                for unit, (mv, yv) in block.items():
                    if yv is not None:
                        units.setdefault(unit, {"month": {}, "till_month": {}})["till_month"][param_key] = yv
        elif latest_month and latest_month != month_num:
            warnings.append(
                f"DETAIL's cumulative (YEAR) column reflects {latest_month} "
                f"(the file's latest month), not the selected {month_num} — "
                "till_month left unchanged for this upload."
            )


def _month_sheet_name(report_month: str, month_num: str) -> str:
    year = int(report_month[:4])
    return f"{_MONTH_NUM_TO_ABBR3[month_num]}{str(year)[-2:]}"


def _verify_month_sheet_title(ws, report_month: str, month_num: str) -> None:
    title = str(ws.cell(4, 1).value or "").strip().upper()
    expected_name = _MONTH_NUM_TO_FULL[month_num]
    year = report_month[:4]
    if expected_name not in title or year not in title:
        raise ValueError(
            f"The '{ws.title}' sheet's own title ('{ws.cell(4, 1).value}') "
            f"doesn't match the selected month {report_month} — verify the "
            "uploaded GLANCE workbook matches the selected period."
        )


def _extract_month_sheet(wb, report_month: str, month_num: str, units: Dict[str, Dict], warnings: List[str]) -> None:
    sheet_name = _month_sheet_name(report_month, month_num)
    if sheet_name not in wb.sheetnames:
        warnings.append(
            f"No '{sheet_name}' sheet in the workbook — Availability, "
            "Utilisation, Carbon Rate and Flue Dust Rate were not extracted "
            "for this month (only available from the BF Department's own "
            "per-month report sheet)."
        )
        return
    ws = wb[sheet_name]
    _verify_month_sheet_title(ws, report_month, month_num)

    seen = set()
    max_row = ws.max_row
    for r in range(1, max_row + 1):
        label = norm_label(ws.cell(r, _MONTH_SHEET_LABEL_COL).value)
        if not label:
            continue
        param_key = _MONTH_SHEET_PARAMS.get(label)
        if param_key is None or param_key in seen:
            continue
        seen.add(param_key)
        for col, unit in _MONTH_SHEET_FURNACE_COLS.items():
            mv = _clean_val(ws.cell(r, col).value)
            if mv is None:
                continue
            slot = units.setdefault(unit, {"month": {}, "till_month": {}})
            slot["month"][param_key] = mv


class RspBfGlanceExtractor:
    """
    extract() returns the standard techno record list:
        [{"plant": "RSP", "report_month": "YYYY-MM", "unit": str,
          "techno_json": {"month": {...}, "till_month": {...}}}]

    After extract(): .warnings
    """

    def __init__(self, file_path: str, report_month: str = ""):
        self.file_path = file_path
        self.report_month = report_month
        self.warnings: List[str] = []

    def extract(self) -> List[Dict]:
        from openpyxl import load_workbook

        if not self.report_month or not re.match(r"^\d{4}-\d{2}$", self.report_month):
            raise ValueError("report_month is required (YYYY-MM).")
        month_num = self.report_month.split("-")[1]
        if month_num not in _MONTH_NUM_TO_ABBR3:
            raise ValueError(f"Invalid month in report_month: {self.report_month}")

        wb = load_workbook(self.file_path, data_only=True)
        if "DETAIL" not in wb.sheetnames:
            raise ValueError(
                "No 'DETAIL' sheet in the workbook — expected the RSP BF "
                "Department's GLANCE workbook (GLANCE -<Mon>'<YY>.xlsx)."
            )

        units: Dict[str, Dict] = {}
        _extract_detail(wb["DETAIL"], self.report_month, month_num, units, self.warnings)
        _extract_month_sheet(wb, self.report_month, month_num, units, self.warnings)
        wb.close()

        records = []
        for unit, tj in units.items():
            if any(v is not None for v in tj["month"].values()) or any(v is not None for v in tj["till_month"].values()):
                records.append({
                    "plant": "RSP",
                    "report_month": self.report_month,
                    "unit": unit,
                    "techno_json": tj,
                })
        if not records:
            raise ValueError(
                "No BF techno values found in the GLANCE workbook for "
                f"{self.report_month} — verify the file and selected month."
            )
        return records
