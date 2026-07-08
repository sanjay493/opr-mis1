"""
DSP MCR Techno Extractor — month-end MCR report, techno page (mcr1_*.xlsx).

The MCR is a daily management control report; when generated on the last day
of a month its "Todate" columns hold the (tentative) for-the-month values.
This extractor reads only those month values — cumulative (till_month) is
left empty and can be computed later via the shared cumulative rules.

Layout (see Report_format/MONTHEND/mcr1_.xlsx):
  C1                    — report date "DD.MM.YYYY" (month-end date)
  Column E              — parameter labels for the BF and SMS blocks
  Columns J,K,L,M       — BF "Todate" values for BF2, BF3, BF4, Shop
  Columns F,G           — SMS "Ondate"/"Todate" values (G = Todate)

Rows are located by matching the label in column E (startswith, case
insensitive) so small row shifts between report versions don't break
extraction; a missing label produces a warning, not a crash.

Note on SMS cells: the request referenced J29/J30/J31, but in the MCR sheet
column J of the SMS block holds "No. of Heats" per casting machine — the
for-the-month (Todate) figures for Heat Size / HM Cons / Scrap Cons live in
column G (F=Ondate, G=Todate). Column G is used here.
"""

import calendar
import re
import zipfile
from datetime import date, datetime
from typing import Dict, List, Optional


class McrMonthMismatch(ValueError):
    """Report date in C1 does not match the user-selected month."""


# (column-E label prefix, canonical param key). Keys follow page_techno.py's
# preferred alias names (hot_blast_temp, specific_hm_consumption, …) — the
# same convention the manual techno entry uses — so MCR values are seen by
# the report even when older manual/PDF values exist under fallback aliases.
_BF_PARAMS = [
    ("BF PRODUCTIVITY",      "bf_productivity"),
    ("BF COKE RATE",         "coke_rate"),
    ("CDI RATE",             "cdi"),
    ("NUT COKE RATE",        "nut_coke_rate"),
    ("FUEL RATE",            "fuel_rate"),
    ("SINTER IN BURDEN",     "sinter_in_burden"),
    ("PELLET IN BURDEN",     "pellet_in_burden"),
    ("HOT BLAST TEMP",       "hot_blast_temp"),
    ("SLAG RATE",            "slag_rate"),
    ("SILICON IN HOT METAL", "silicon_in_hm"),
    ("SULPHUR IN HOT METAL", "sulphur_in_hm"),
]

_SMS_PARAMS = [
    ("HEAT SIZE",           "average_heat_weight"),
    ("HOT METAL CONS RATE", "specific_hm_consumption"),
    ("SCRAP CONS RATE",     "specific_scrap_consumption"),
]

# BF block "Todate" columns (1-based): J, K, L, M
_BF_UNIT_COLS = [("BF-2", 10), ("BF-3", 11), ("BF-4", 12), ("BF_Shop", 13)]

_LABEL_COL = 5        # column E
_SMS_TODATE_COL = 7   # column G

_MAX_ROWS = 80
_MAX_COLS = 14


def _clean_val(val) -> Optional[float]:
    if val is None:
        return None
    try:
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip().upper()
        if s in ("NAN", "###", "-", "#DIV/0!", ""):
            return None
        return float(s)
    except (ValueError, TypeError):
        return None


def _load_grid(file_path: str) -> List[List]:
    """Return the first sheet as a list of rows (lists of raw cell values).

    Accepts real .xlsx workbooks and DSP's tab-separated text files that
    carry an .xls extension. Binary Excel 97-2003 files are rejected.
    """
    with open(file_path, "rb") as f:
        magic = f.read(4)

    if magic == b"\xd0\xcf\x11\xe0":
        raise ValueError(
            "Binary XLS format is not supported for the DSP MCR techno page. "
            "Upload the .xlsx file (or the tab-separated MCR text file)."
        )

    if magic[:2] == b"PK" and zipfile.is_zipfile(file_path):
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        grid = [
            [c for c in row]
            for row in ws.iter_rows(min_row=1, max_row=_MAX_ROWS,
                                    min_col=1, max_col=_MAX_COLS,
                                    values_only=True)
        ]
        wb.close()
        return grid

    # Fall back to tab-separated text (same convention as the MCR-I report)
    with open(file_path, encoding="utf-8", errors="replace") as f:
        return [line.rstrip("\r\n").split("\t") for line in f]


class DspMcrTechnoExtractor:
    """
    Extract for-the-month techno parameters from the DSP MCR techno page.

    extract() returns the standard techno record list:
        [{"plant": "DSP", "report_month": "YYYY-MM", "unit": str,
          "techno_json": {"month": {...}, "till_month": {}}}]

    After extract():
      .report_date  — the date parsed from C1
      .warnings     — list of non-fatal extraction warnings
    """

    def __init__(self, file_path: str, report_month: str = ""):
        self.file_path = file_path
        self.report_month = report_month
        self.report_date: Optional[date] = None
        self.warnings: List[str] = []

    # -- helpers -------------------------------------------------------------

    def _cell(self, grid, row_1b: int, col_1b: int):
        if row_1b < 1 or row_1b > len(grid):
            return None
        row = grid[row_1b - 1]
        if col_1b < 1 or col_1b > len(row):
            return None
        return row[col_1b - 1]

    def _find_row(self, grid, label_prefix: str) -> Optional[int]:
        """1-based row whose column-E label starts with label_prefix."""
        for i, row in enumerate(grid, start=1):
            v = row[_LABEL_COL - 1] if len(row) >= _LABEL_COL else None
            if v is None:
                continue
            if str(v).strip().upper().startswith(label_prefix):
                return i
        return None

    def _parse_report_date(self, grid) -> date:
        raw = self._cell(grid, 1, 3)  # C1
        if isinstance(raw, (datetime, date)):
            return raw.date() if isinstance(raw, datetime) else raw
        m = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", str(raw or ""))
        if not m:
            raise ValueError(
                f"Cannot read the report date from cell C1 (got {raw!r}). "
                "Expected DD.MM.YYYY, e.g. 30.06.2026 — verify this is the "
                "DSP MCR techno page."
            )
        d, mo, y = (int(g) for g in m.groups())
        return date(y, mo, d)

    def _verify_month(self):
        """Raise McrMonthMismatch if C1's month differs from the selection."""
        file_month = f"{self.report_date.year}-{self.report_date.month:02d}"
        if self.report_month and self.report_month != file_month:
            raise McrMonthMismatch(
                f"This MCR report is dated {self.report_date.strftime('%d.%m.%Y')} "
                f"(month {file_month}) but you selected {self.report_month}. "
                "Select the matching month or upload the correct file."
            )
        self.report_month = file_month

    # -- main ----------------------------------------------------------------

    def extract(self) -> List[Dict]:
        grid = _load_grid(self.file_path)

        head = str(self._cell(grid, 1, 1) or "")
        if "DAILY MANAGEMENT CON" not in head.upper():
            raise ValueError(
                "File does not look like a DSP MCR report — cell A1 must "
                "start with 'DAILY MANAGEMENT CON…'."
            )

        self.report_date = self._parse_report_date(grid)
        self._verify_month()

        # A mid-month MCR only covers month-to-date, not the full month
        last_day = calendar.monthrange(self.report_date.year, self.report_date.month)[1]
        if self.report_date.day != last_day:
            self.warnings.append(
                f"Report date {self.report_date.strftime('%d.%m.%Y')} is not the "
                f"last day of the month — 'Todate' values cover only part of "
                f"{self.report_month}."
            )

        units: Dict[str, Dict] = {}

        def put(unit: str, key: str, value: Optional[float]):
            units.setdefault(unit, {"month": {}, "till_month": {}})
            units[unit]["month"][key] = value

        # BF block — Todate columns J..M = BF-2, BF-3, BF-4, BF_Shop
        for label, key in _BF_PARAMS:
            row = self._find_row(grid, label)
            if row is None:
                self.warnings.append(f"Label '{label}' not found in column E — skipped.")
                continue
            for unit, col in _BF_UNIT_COLS:
                put(unit, key, _clean_val(self._cell(grid, row, col)))

        # SMS block — Todate is column G (F=Ondate, G=Todate)
        for label, key in _SMS_PARAMS:
            row = self._find_row(grid, label)
            if row is None:
                self.warnings.append(f"Label '{label}' not found in column E — skipped.")
                continue
            put("SMS", key, _clean_val(self._cell(grid, row, _SMS_TODATE_COL)))

        # Drop units where every value came back empty
        records = []
        for unit, tj in units.items():
            if any(v is not None for v in tj["month"].values()):
                records.append({
                    "plant": "DSP",
                    "report_month": self.report_month,
                    "unit": unit,
                    "techno_json": tj,
                })

        if not records:
            raise ValueError(
                "No techno values found in the MCR report — verify this is the "
                "techno page (BF / SMS blocks with 'Todate' columns)."
            )
        return records
