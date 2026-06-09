import re
import openpyxl
import logging
import sqlite3
import os
from datetime import datetime
from typing import Optional

logger = logging.getLogger("excel_extractor")

DB_PATH = os.path.join(os.path.dirname(__file__), "backend", "mis_reports.db")

MONTH_NAMES = {
    "01": "January", "02": "February", "03": "March", "04": "April",
    "05": "May",     "06": "June",     "07": "July",  "08": "August",
    "09": "September","10": "October", "11": "November","12": "December"
}
MONTHS_MAP = {v: k for k, v in MONTH_NAMES.items()}


def clean_val(val) -> Optional[float]:
    if val is None or str(val).strip().lower() in ("nan", "###", "-", "#div/0!", ""):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _parse_report_month(report_month: str):
    """Returns (db_report_month_yyyymm, month_num_str) from 'YYYY-MM' or legacy 'Month Year'."""
    if len(report_month) == 7 and report_month[4] == "-":
        return report_month, report_month[5:7]
    parts = report_month.split()
    m_name, y_str = parts[0], parts[1]
    m_num = MONTHS_MAP.get(m_name, "01")
    return f"{y_str}-{m_num}", m_num


def extract_and_save_excel(file_path: str, report_month: str = "", source_file_name: str = "") -> bool:
    """
    Dispatcher: auto-detects ISP file type by sheet name.

    Supported file types:
      • Morning Report (.xlsx)       — sheet 'DAILYREPORT1'. Month auto-detected from K5.
      • Final Monthly Report (.xlsx) — sheet 'Maj Production Summ'. Month set manually.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames
        logger.info(f"ISP: loading file. Sheets: {sheet_names}")

        if "DAILYREPORT1" in sheet_names:
            return _extract_morning_report(wb, source_file_name)

        if "Maj Production Summ" in sheet_names:
            return _extract_monthly_report(wb, report_month, source_file_name)

        raise ValueError(
            "Uploaded ISP file does not match any known format. "
            "Expected sheet 'DAILYREPORT1' (Morning Report) or "
            "'Maj Production Summ' (Final Monthly Report)."
        )
    except ValueError as ve:
        logger.error(f"ISP validation error: {ve}")
        raise
    except Exception as e:
        logger.error(f"ISP extraction error: {e}")
        return False


# ---------------------------------------------------------------------------
# Extractor 1 — ISP Morning Report (month-end daily, sheet: DAILYREPORT1)
# ---------------------------------------------------------------------------

def _extract_morning_report(wb, source_file_name: str) -> bool:
    """
    Extracts cumulative monthly production from ISP Morning Report (sheet 'DAILYREPORT1').

    Date is auto-detected from K5 (Python datetime object). Column E = Monthly Rate.

    Cell map — sheet 'DAILYREPORT1':
      E9:       COB#10              — monthly avg, no unit conversion
      E10:      COB#11              — monthly avg, no unit conversion
      E11:      Oven Pushing(nos/d) — monthly avg nos/day, no unit conversion
      E12:      SP-1                — tonnes → /1000
      E13:      SP-2                — tonnes → /1000
      E14:      Total Sinter        — tonnes → /1000
      E16:      Hot Metal           — tonnes → /1000
      E18+E20:  Pig Iron            — derived sum, tonnes → /1000
      E30:      CCM-1&2             — tonnes → /1000
      E31:      CCM-3               — tonnes → /1000
      E32:      Total Crude Steel   — tonnes → /1000
      E33:      WRMILL              — tonnes → /1000
      E34:      BARMILL             — tonnes → /1000
      E35:      USMILL              — tonnes → /1000
      E36:      Finished Steel      — tonnes → /1000
      E37:      Saleable 150 Billets — tonnes → /1000
      E38:      200 Blooms          — tonnes → /1000
      E37+E38:  Saleable Semis      — derived sum, tonnes → /1000
      E39:      Saleable Steel      — tonnes → /1000
    """
    import sys
    sys.path.insert(0, os.path.dirname(__file__))
    import db

    ws = wb["DAILYREPORT1"]

    # Date from K5 (Excel stores it as a Python datetime)
    k5_raw = ws["K5"].value
    if isinstance(k5_raw, datetime):
        m_num = str(k5_raw.month).zfill(2)
        year  = str(k5_raw.year)
        db_report_month = f"{year}-{m_num}"
    elif k5_raw:
        date_match = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', str(k5_raw))
        if date_match:
            _d, m_num, year = date_match.groups()
            db_report_month = f"{year}-{m_num}"
        else:
            raise ValueError(
                f"Cannot parse date from K5: {repr(k5_raw)}. "
                "Expected a datetime value or DD.MM.YYYY string."
            )
    else:
        raise ValueError("Cell K5 is empty — cannot determine report month.")

    logger.info(f"ISP Morning Report: month auto-detected from K5 → {db_report_month}")

    NO_CONVERT = {"COB#10", "COB#11", "Oven Pushing(nos/d)"}

    # Single-cell items: item_name → cell address
    production_cells = {
        "COB#10":                "E9",
        "COB#11":                "E10",
        "Oven Pushing(nos/d)":   "E11",
        "SP-1":                  "E12",
        "SP-2":                  "E13",
        "Total Sinter":          "E14",
        "Hot Metal":             "E16",
        "CCM-1&2":               "E30",
        "CCM-3":                 "E31",
        "Total Crude Steel":     "E32",
        "WRMILL":                "E33",
        "BARMILL":               "E34",
        "USMILL":                "E35",
        "Finished Steel":        "E36",
        "Saleable 150 Billets":  "E37",
        "200 Blooms":            "E38",
        "Saleable Steel":        "E39",
    }

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    vals_extracted = 0

    def _save(item_name, val):
        nonlocal vals_extracted
        if val is not None:
            vals_extracted += 1
            if item_name not in NO_CONVERT:
                val = round(val / 1000.0, 3)
        cursor.execute("""
            INSERT INTO production_table (report_month, plant_name, item_name, month_actual)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(report_month, plant_name, item_name)
            DO UPDATE SET month_actual = excluded.month_actual
        """, (db_report_month, "ISP", item_name, val))

    for item_name, cell in production_cells.items():
        _save(item_name, clean_val(ws[cell].value))

    # Derived: Pig Iron = E18 + E20
    pig_e18 = clean_val(ws["E18"].value)
    pig_e20 = clean_val(ws["E20"].value)
    if pig_e18 is not None or pig_e20 is not None:
        _save("Pig Iron", (pig_e18 or 0.0) + (pig_e20 or 0.0))

    # Derived: Saleable Semis = E37 + E38 (CC Billets + CC Blooms)
    sem_e37 = clean_val(ws["E37"].value)
    sem_e38 = clean_val(ws["E38"].value)
    if sem_e37 is not None or sem_e38 is not None:
        _save("Saleable Semis", (sem_e37 or 0.0) + (sem_e38 or 0.0))

    if vals_extracted == 0:
        raise ValueError(
            "No numeric data found at the expected cell locations in sheet 'DAILYREPORT1'. "
            "Please verify this is the correct ISP Morning Report file."
        )

    conn.commit()
    conn.close()

    db.log_extraction(
        plant="ISP",
        report_month=db_report_month,
        file_name=source_file_name,
        sheet_name="DAILYREPORT1",
        source_type="Daily Morning Report",
        items_extracted=vals_extracted,
    )
    logger.info(f"ISP Morning Report extraction done: {vals_extracted} values saved for {db_report_month}.")
    return True


# ---------------------------------------------------------------------------
# Extractor 2 — ISP Final Monthly Report (sheet: Maj Production Summ)
# ---------------------------------------------------------------------------

def _extract_monthly_report(wb, report_month: str, source_file_name: str) -> bool:
    """
    Extracts production data from ISP Final Monthly Report (sheet 'Maj Production Summ').
    Month must be provided via report_month ('Month Year' or 'YYYY-MM').

    Cell map: items are in fixed rows; month selects the column via col_map.
    """
    import sys
    sys.path.insert(0, os.path.dirname(__file__))
    import db

    if not report_month:
        raise ValueError(
            "report_month is required for ISP Final Monthly Report. "
            "Set the month selector before uploading."
        )

    db_report_month, month_num = _parse_report_month(report_month)

    col_map = {
        "04": "F",  "05": "H",  "06": "L",  "07": "P",
        "08": "T",  "09": "X",  "10": "AD", "11": "AH",
        "12": "AL", "01": "AR", "02": "AV", "03": "AZ",
    }
    col = col_map.get(month_num)
    if not col:
        raise ValueError(f"Month column mapping not found for month code '{month_num}'.")

    ws = wb["Maj Production Summ"]

    NO_CONVERT = {"COB#10", "COB#11", "Oven Pushing(nos/d)"}

    # item_name → row number (column is dynamic per month via col_map)
    production_cells = {
        "COB#10":                6,
        "COB#11":                7,
        "Oven Pushing(nos/d)":   8,
        "Total Sinter":          16,
        "Hot Metal":             17,
        "Pig Iron":              26,
        "CCM-1&2":               19,
        "CCM-3":                 20,
        "Total Crude Steel":     18,
        "WRMILL":                30,
        "BARMILL":               31,
        "USMILL":                32,
        "Finished Steel":        33,
        "Saleable 150 Billets":  34,
        "200 Blooms":            35,
        "Saleable Semis":        36,
        "Saleable Steel":        37,
    }

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    vals_extracted = 0

    for item_name, row_num in production_cells.items():
        cell_ref = f"{col}{row_num}"
        val = clean_val(ws[cell_ref].value)
        if val is not None:
            vals_extracted += 1
            if item_name not in NO_CONVERT:
                val = round(val / 1000.0, 3)
        cursor.execute("""
            INSERT INTO production_table (report_month, plant_name, item_name, month_actual)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(report_month, plant_name, item_name)
            DO UPDATE SET month_actual = excluded.month_actual
        """, (db_report_month, "ISP", item_name, val))

    if vals_extracted == 0:
        raise ValueError(
            "No numeric data could be extracted from 'Maj Production Summ'. "
            "Please verify the contents of the Excel file."
        )

    conn.commit()
    conn.close()

    db.log_extraction(
        plant="ISP",
        report_month=db_report_month,
        file_name=source_file_name,
        sheet_name="Maj Production Summ",
        source_type="Final Monthly Report",
        items_extracted=vals_extracted,
    )
    logger.info(f"ISP Monthly Report extraction done: {vals_extracted} values saved for {db_report_month}.")
    return True
