import openpyxl
import logging
import sqlite3
import os
from typing import Optional

logger = logging.getLogger("excel_extractor_plan")

DB_PATH = os.path.join(os.path.dirname(__file__), "backend", "mis_reports.db")


def clean_val(val) -> Optional[float]:
    if val is None or str(val).strip().lower() in ("nan", "###", "-", "#div/0!"):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def extract_and_save_excel_plan(file_path: str, financial_year: str) -> bool:
    """
    Extracts BSL ABP plan data for all 12 months.
    Expected file: BSLABP26-27.xlsx (or similar), sheet 'PLAN SUMMARY'.

    Layout: months in rows, items in columns.
      Row 10 = APR, 11 = MAY, 12 = JUN,  (13 = Q1 skip)
      Row 14 = JUL, 15 = AUG, 16 = SEP,  (17 = Q2 skip)
      Row 18 = OCT, 19 = NOV, 20 = DEC,  (21 = Q3 skip)
      Row 22 = JAN, 23 = FEB, 24 = MAR,  (25 = Q4 skip, 26 = Total skip)

    Columns (values already in '000 T per sheet header; B is nos/day):
      B = Oven Pushing (nos/d)   C = Total Sinter        D = Hot Metal
      E = Total Crude Steel      F = SMS-1 CCM-1         G = SMS-2 CCM-1&2
      H = HSM Total HR Coil      J = CRC&S(1&2)          K = CRC(3)
      L = CR Saleable            M = Saleable Steel       N = Pig Iron
      O = Saleable Semis         P = HSM HR Coil (Sale)  Q = HSM HR Plate
      R = HR Sheet

    Finished Steel is derived: Saleable Steel (M) − Saleable Semis (O).
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames
        logger.info(f"BSL Plan: loading file. Sheets: {sheet_names}")

        sheet_key = None
        for candidate in ("PLAN SUMMARY", "Plan Summary", "plan summary",
                          "Monthwise", "MONTHWISE", "Sheet1", "sheet1"):
            if candidate in sheet_names:
                sheet_key = candidate
                break

        if not sheet_key:
            raise ValueError(
                f"BSL Plan Excel missing expected sheet. "
                f"Found: {sheet_names}. Expected 'PLAN SUMMARY'."
            )

        logger.info(f"BSL Plan: using sheet '{sheet_key}'")

        if "-" in financial_year:
            year_val = int(financial_year.split("-")[0])
        else:
            year_val = int(financial_year)

        # (row_number) → (full_month_name, year_offset)
        # Quarter rows (13, 17, 21, 25) and total row (26) are intentionally absent.
        row_month_map = {
            10: ("April",     0),
            11: ("May",       0),
            12: ("June",      0),
            14: ("July",      0),
            15: ("August",    0),
            16: ("September", 0),
            18: ("October",   0),
            19: ("November",  0),
            20: ("December",  0),
            22: ("January",   1),
            23: ("February",  1),
            24: ("March",     1),
        }

        # (column_letter) → (db_item_name, nos_per_day)
        # nos_per_day=True → store as-is; False → round to 3 dp (already '000 T)
        col_item_map = {
            "B": ("Oven Pushing(nos/d)",  True),
            "C": ("Total Sinter",         False),
            "D": ("Hot Metal",            False),
            "E": ("Total Crude Steel",    False),
            "F": ("SMS-1 CCM-1",          False),
            "G": ("SMS-2 CCM-1&2",        False),
            "H": ("HSM Total HR Coil",    False),
            "J": ("CRC&S(1&2)",           False),
            "K": ("CRC(3)",               False),
            "L": ("CR Saleable",          False),
            "M": ("Saleable Steel",       False),
            "N": ("Pig Iron",             False),
            "O": ("Saleable Semis",       False),
            "P": ("HSM HR Coil (Sale)",   False),
            "Q": ("HSM HR Plate",         False),
            "R": ("HR Sheet",             False),
        }

        ws = wb[sheet_key]
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        vals_extracted = 0

        for row_num, (m_name, yr_off) in row_month_map.items():
            db_report_month = f"{m_name} {year_val + yr_off}"
            raw_row = {}

            for col_letter, (db_item, is_rate) in col_item_map.items():
                raw_val = ws[f"{col_letter}{row_num}"].value
                val = clean_val(raw_val)
                raw_row[db_item] = val

                if val is not None:
                    vals_extracted += 1
                    stored = val if is_rate else round(val, 3)
                else:
                    stored = None

                cursor.execute(
                    """
                    INSERT INTO production_plan_table (report_month, plant_name, item_name, month_actual)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT(report_month, plant_name, item_name)
                    DO UPDATE SET month_actual = excluded.month_actual
                    """,
                    (db_report_month, "BSL", db_item, stored),
                )

            # Derived: Finished Steel = Saleable Steel − Saleable Semis
            sal_steel = raw_row.get("Saleable Steel")
            sal_semis = raw_row.get("Saleable Semis")
            finished = (
                round(sal_steel - sal_semis, 3)
                if sal_steel is not None and sal_semis is not None
                else None
            )
            cursor.execute(
                """
                INSERT INTO production_plan_table (report_month, plant_name, item_name, month_actual)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(report_month, plant_name, item_name)
                DO UPDATE SET month_actual = excluded.month_actual
                """,
                (db_report_month, "BSL", "Finished Steel", finished),
            )

        if vals_extracted == 0:
            raise ValueError(
                "No numeric data extracted from BSL Plan Excel. "
                "Verify sheet name ('PLAN SUMMARY'), row numbers (10–24), "
                "and column mapping (B–R) against the actual file."
            )

        conn.commit()
        conn.close()
        logger.info(
            f"BSL Plan extraction done: {vals_extracted} values + 12 Finished Steel "
            f"derived rows for FY {financial_year}."
        )
        return True

    except ValueError as ve:
        logger.error(f"BSL Plan validation error: {ve}")
        raise
    except Exception as e:
        logger.error(f"BSL Plan extraction error: {e}")
        return False
