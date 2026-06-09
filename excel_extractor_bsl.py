import re
import openpyxl
import logging
import sqlite3
import os
from datetime import datetime
from typing import Optional

logger = logging.getLogger("excel_extractor")

DB_PATH = os.path.join(os.path.dirname(__file__), "backend", "mis_reports.db")

MONTH_NAMES = {
    "01": "January", "02": "February", "03": "March", "04": "April",
    "05": "May",     "06": "June",     "07": "July",  "08": "August",
    "09": "September","10": "October", "11": "November","12": "December"
}
MONTH_NUMS = {v: k for k, v in MONTH_NAMES.items()}


def clean_val(val) -> Optional[float]:
    if val is None or str(val).strip().lower() in ("nan", "###", "-", "#div/0!", ""):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def extract_and_save_excel(file_path: str, report_month: str = "", source_file_name: str = "") -> bool:
    """
    Dispatcher: auto-detects BSL file type by sheet name and calls the correct extractor.

    Supported file types:
      • DPR Mail (.xlsx)         — sheet 'DPR'  (month-end daily production report)
      • Final Monthly Report     — (commented out, to be implemented)
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames
        logger.info(f"BSL: loading file. Sheets: {sheet_names}")

        if "DPR" in sheet_names:
            return _extract_dpr_report(wb, source_file_name)

        raise ValueError(
            "Uploaded BSL file does not match any known format. "
            "Expected sheet 'DPR' (DPR Mail month-end report)."
        )
    except ValueError as ve:
        logger.error(f"BSL validation error: {ve}")
        raise ve
    except Exception as e:
        logger.error(f"BSL extraction error: {e}")
        return False


# ---------------------------------------------------------------------------
# Extractor 1 — BSL DPR Mail (month-end daily production report, sheet: DPR)
# ---------------------------------------------------------------------------

def _extract_dpr_report(wb, source_file_name: str) -> bool:
    """
    Extracts cumulative production data from the BSL DPR Mail report (sheet 'DPR').

    Date is auto-detected from cell O1 (stored as a Python datetime by Excel).

    Cell map — sheet 'DPR':
      Oven Pushing(nos/d)  P6    — nos/day average, no unit conversion
      Total Sinter         P7    — tonnes → /1000
      Hot Metal            P8
      Pig Iron             E30
      SMS-1 CCM-1          P10
      SMS-2 CCM-1&2        P11
      Total Crude Steel    P12
      HSM Total HR Coil    P14
      HSM HR Coil (Sale)   E7
      HSM HR Plate         E8
      HR Sheet             E9
      CRC&S(1&2)           E10
      CRC(3)               E11
      GP/GC                E12
      GPC3                 E13
      CRSALE               E29
      Saleable Steel       P31
      Saleable Semis       E32
      Finished Steel       P31 − E32  (derived: saleable steel minus semis)
    """
    import sys
    sys.path.insert(0, os.path.dirname(__file__))
    import db

    ws = wb["DPR"]

    # Auto-detect month from O1 (Excel stores it as a Python datetime object)
    o1_raw = ws["O1"].value
    if isinstance(o1_raw, datetime):
        m_num = str(o1_raw.month).zfill(2)
        year  = str(o1_raw.year)
        db_report_month = f"{year}-{m_num}"
    elif o1_raw:
        # Fallback: try to parse date string formats DD.MM.YYYY or YYYY-MM-DD
        date_match = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", str(o1_raw))
        if date_match:
            _d, m_num, year = date_match.groups()
            db_report_month = f"{year}-{m_num}"
        else:
            raise ValueError(
                f"Cannot parse date from cell O1: {repr(o1_raw)}. "
                "Expected a date value (DD.MM.YYYY or Excel date)."
            )
    else:
        raise ValueError("Cell O1 is empty — cannot determine report month.")

    logger.info(f"BSL DPR: month auto-detected from O1 → {db_report_month}")

    NO_CONVERT = {"Oven Pushing(nos/d)"}

    # Simple cells: item_name → cell address
    production_cells = {
        "Oven Pushing(nos/d)": "P6",
        "Total Sinter":        "P7",
        "Hot Metal":           "P8",
        "Pig Iron":            "E30",
        "SMS-1 CCM-1":         "P10",
        "SMS-2 CCM-1&2":       "P11",
        "Total Crude Steel":   "P12",
        "HSM Total HR Coil":   "P14",
        "HSM HR Coil (Sale)":  "E7",
        "HSM HR Plate":        "E8",
        "HR Sheet":            "E9",
        "CRC&S(1&2)":          "E10",
        "CRC(3)":              "E11",
        "GP/GC":               "E12",
        "GPC3":                "E13",
        "CRSALE":              "E29",
        "Saleable Steel":      "P31",
        "Saleable Semis":      "E32",
    }

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    vals_extracted = 0

    def _save(item_name, val):
        nonlocal vals_extracted
        if val is not None:
            vals_extracted += 1
            if item_name not in NO_CONVERT:
                val = round(val / 1000.0, 3)
        cursor.execute("""
            INSERT INTO production_table (report_month, plant_name, item_name, month_actual)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(report_month, plant_name, item_name)
            DO UPDATE SET month_actual = excluded.month_actual
        """, (db_report_month, "BSL", item_name, val))

    # Extract all mapped cells
    for item_name, cell in production_cells.items():
        _save(item_name, clean_val(ws[cell].value))

    # Derived: Finished Steel = Saleable Steel (P31) − Saleable Semis (E32)
    sal_steel = clean_val(ws["P31"].value)
    sal_semis  = clean_val(ws["E32"].value)
    if sal_steel is not None and sal_semis is not None:
        finished = round((sal_steel - sal_semis) / 1000.0, 3)
        cursor.execute("""
            INSERT INTO production_table (report_month, plant_name, item_name, month_actual)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(report_month, plant_name, item_name)
            DO UPDATE SET month_actual = excluded.month_actual
        """, (db_report_month, "BSL", "Finished Steel", finished))
        vals_extracted += 1
    elif sal_steel is not None:
        _save("Finished Steel", sal_steel)   # best-effort if semis missing

    if vals_extracted == 0:
        raise ValueError(
            "No numeric data found at the expected cell locations in sheet 'DPR'. "
            "Please verify this is the correct BSL DPR Mail file."
        )

    conn.commit()
    conn.close()

    db.log_extraction(
        plant="BSL",
        report_month=db_report_month,
        file_name=source_file_name,
        sheet_name="DPR",
        source_type="DPR Mail (Month-End)",
        items_extracted=vals_extracted,
    )
    logger.info(f"BSL DPR extraction done: {vals_extracted} values saved for {db_report_month}.")
    return True


# ---------------------------------------------------------------------------
# Extractor 2 — BSL Final Monthly Report (to be implemented)
# ---------------------------------------------------------------------------

# def _extract_monthly_report(wb, report_month: str, source_file_name: str) -> bool:
#     """Extracts production data from BSL final monthly consolidated report."""
#     pass
